"""
invoice_module.py
-----------------
Invoice generation module for Revenue Tracker.
Handles: USD invoice, INR+CGST/SGST invoice, INR+IGST invoice,
         PDF download, email sending, payment reminders, DSP statement.
Linked from app.py via: from invoice_module import render_invoice_module
"""

import streamlit as st
import pandas as pd
import sqlite3
import io
import os
import sys
import smtplib
import ssl
from invoice_subnav import render_invoice_subnav
from datetime import datetime, date
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from num2words import num2words

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, Image as RLImage
)
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.lib import utils as rl_utils

# ===============================
# SMTP CREDENTIALS — loaded once from st.secrets (secrets.toml / Streamlit Cloud)
# Falls back to session_state so the SMTP expander still works as manual override.
# ===============================

def _get_smtp_creds():
    """
    Returns (host, port, user, password) from st.secrets if configured,
    otherwise reads from st.session_state (manual entry via expander).
    Priority: secrets.toml → Streamlit Cloud Secrets → session_state input
    """
    try:
        _s = st.secrets["email"]
        return (
            str(_s.get("smtp_host", "smtpout.secureserver.net")),
            int(_s.get("smtp_port", 465)),
            str(_s.get("sender",    "")),
            str(_s.get("password",  "")),
        )
    except Exception:
        # secrets not configured — fall back to whatever was typed in expander
        return (
            st.session_state.get("smtp_host", "smtpout.secureserver.net"),
            int(st.session_state.get("smtp_port", 465)),
            st.session_state.get("smtp_user", "finance@peakmyads.com"),
            st.session_state.get("smtp_pass", ""),
        )

def _smtp_configured_via_secrets():
    """True if credentials are loaded from secrets (no manual entry needed)."""
    try:
        _s = st.secrets["email"]
        return bool(_s.get("password", ""))
    except Exception:
        return False

# ─────────────────────────────────────────────
# DB PATH (mirrors app.py logic)
# ─────────────────────────────────────────────
if os.environ.get("TRACKER_BASE_DIR"):
    BASE_DIR = os.environ["TRACKER_BASE_DIR"]
elif getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

DB_PATH = os.path.join(BASE_DIR, "tracker.db")
LOGO_PATH = os.path.join(BASE_DIR, "peakads_logo.png")
SIGN_PATH = os.path.join(BASE_DIR, "sign.jpg")

def get_db():
    return sqlite3.connect(DB_PATH, check_same_thread=False)


# ─────────────────────────────────────────────
# DUE DATE CALCULATION  (Invoice Date + Terms - 1 day)
# ─────────────────────────────────────────────
def compute_due_date(invoice_date: date, payment_terms: str) -> str:
    """
    invoice_date : Python date object
    payment_terms: e.g. 'Net 60', '60', '45 Days', etc.
    Returns: DD-MMM-YYYY  e.g. 01-Jun-2026
    Logic: due = invoice_date + days - 1
    """
    import re
    days = 60  # default
    m = re.search(r'\d+', str(payment_terms))
    if m:
        days = int(m.group())
    from datetime import timedelta
    due = invoice_date + timedelta(days=days - 1)
    return due.strftime("%d-%b-%Y")

# ─────────────────────────────────────────────
# ENSURE invoice_details TABLE EXISTS
# ─────────────────────────────────────────────
def init_invoice_table():
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS invoice_details (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            invoice_number      TEXT UNIQUE,
            invoice_date        TEXT,
            dsp_name            TEXT,
            month               TEXT,
            invoice_type        TEXT,
            amount              REAL,
            tax_amount          REAL,
            total_amount        REAL,
            currency            TEXT,
            due_date            TEXT,
            status              TEXT DEFAULT 'Draft',
            sent_at             TEXT,
            created_at          TEXT,
            igst_amount         REAL DEFAULT 0,
            cgst_amount         REAL DEFAULT 0,
            sgst_amount         REAL DEFAULT 0,
            taxable_amount      REAL DEFAULT 0,
            is_deleted          INTEGER DEFAULT 0,
            is_credit_note      INTEGER DEFAULT 0,
            credit_note_ref     TEXT DEFAULT '',
            credit_note_reason  TEXT DEFAULT '',
            fx_rate             REAL DEFAULT 0,
            month_breakdown     TEXT DEFAULT ''
        )
    """)
    for col, defval in [
        ("igst_amount",        "REAL DEFAULT 0"),
        ("cgst_amount",        "REAL DEFAULT 0"),
        ("sgst_amount",        "REAL DEFAULT 0"),
        ("taxable_amount",     "REAL DEFAULT 0"),
        ("is_deleted",         "INTEGER DEFAULT 0"),
        ("is_credit_note",     "INTEGER DEFAULT 0"),
        ("credit_note_ref",    "TEXT DEFAULT ''"),
        ("credit_note_reason", "TEXT DEFAULT ''"),
        ("fx_rate",            "REAL DEFAULT 0"),
        ("month_breakdown",    "TEXT DEFAULT ''"),
    ]:
        try:
            conn.execute(f"ALTER TABLE invoice_details ADD COLUMN {col} {defval}")
        except Exception:
            pass
    conn.commit()
    conn.close()

init_invoice_table()

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
COMPANY = {
    "name":    "PEAKADS LLP",
    "address": "Unit 830, REGUS, WTC,\nFountain Road, Kharadi,\nPune MH 411014 India",
    "email":   "finance@peakmyads.com",
    "website": "www.peakmyads.com",
    "gstin":   "27ABHFP9304J1ZW",
    "account_name":   "PEAKADS LLP",
    "account_number": "925020027425952",
    "ifsc":    "UTIB0000871",
    "swift":   "AXISINBB871",
    "bank":    "AXIS BANK LTD.",
    "paypal":  "vaishali.peakmyads@gmail.com",
    "payoneer":  "vaishali.peakmyads@gmail.com",
}

def _regenerate_invoice_pdf(row):
    """
    Regenerate the PDF for a stored invoice/CN row dict.
    Matches the exact parameter signatures of each generate_* function.
    """
    import json as _rj, math as _rm

    # ── Extract stored fields ─────────────────────────────────────────────
    inv_no   = str(row.get("invoice_number",""))
    inv_date = str(row.get("invoice_date",""))
    inv_type = str(row.get("invoice_type","USD"))
    dsp_name = str(row.get("dsp_name",""))
    month    = str(row.get("month",""))
    currency = str(row.get("currency","USD")).upper()
    is_cn    = int(row.get("is_credit_note",0) or 0) == 1
    fx_rate  = float(row.get("fx_rate",0) or 0)
    due_date = str(row.get("due_date","") or "")
    taxable  = float(row.get("taxable_amount",0) or 0)
    amount   = float(row.get("amount",0) or 0)  # USD amount

    # Parse breakdown JSON safely
    _bd_raw = row.get("month_breakdown","")
    _bd = []
    if _bd_raw and not (isinstance(_bd_raw, float) and _rm.isnan(_bd_raw)):
        s = str(_bd_raw).strip()
        if s.startswith("["):
            try:
                _bd = _rj.loads(s)
            except Exception:
                pass

    partner = get_partner_info(dsp_name)
    desc    = ""   # terms/description not stored; use empty string

    # ── Credit Note ───────────────────────────────────────────────────────
    if is_cn:
        igst = float(row.get("igst_amount",0) or 0)
        cgst = float(row.get("cgst_amount",0) or 0)
        sgst = float(row.get("sgst_amount",0) or 0)
        total= float(row.get("total_amount",0) or 0)
        return generate_credit_note_pdf(
            inv_no, inv_date,
            str(row.get("credit_note_ref","")),
            str(row.get("credit_note_reason","")),
            partner, month, total, currency,
            taxable=taxable, cgst=cgst, sgst=sgst, igst=igst,
            fx_rate=fx_rate
        )

    # ── Multi-month invoices (breakdown has > 1 entry) ────────────────────
    if len(_bd) > 1:
        if inv_type == "USD":
            # month_data: list of dicts {month, amount}
            _md = [{"month": b.get("month",""),
                    "amount": float(b.get("usd_amount",0) or 0)}
                   for b in _bd]
            return generate_multi_month_usd_invoice(
                inv_no, inv_date, due_date, desc, partner, _md, desc
            )
        elif inv_type == "INR_CGST_SGST":
            # month_data: list of dicts {month, usd_amount, fx_rate}
            _md = [{"month":      b.get("month",""),
                    "usd_amount": float(b.get("usd_amount",0) or 0),
                    "fx_rate":    float(b.get("fx_rate", fx_rate) or fx_rate)}
                   for b in _bd]
            return generate_multi_month_inr_cgst_sgst_invoice(
                inv_no, inv_date, due_date, desc, partner, _md, desc
            )
        elif inv_type == "INR_IGST":
            _md = [{"month":      b.get("month",""),
                    "usd_amount": float(b.get("usd_amount",0) or 0),
                    "fx_rate":    float(b.get("fx_rate", fx_rate) or fx_rate)}
                   for b in _bd]
            return generate_multi_month_inr_igst_invoice(
                inv_no, inv_date, due_date, desc, partner, _md, desc
            )

    # ── Single-month invoices ─────────────────────────────────────────────
    if inv_type == "USD":
        # generate_usd_invoice(inv_no, inv_date, due_date, terms, partner, month_label, amount, description)
        return generate_usd_invoice(
            inv_no, inv_date, due_date, desc, partner, month, amount, desc
        )
    elif inv_type == "INR_CGST_SGST":
        # generate_inr_cgst_sgst_invoice(..., usd_amount, fx_rate, description)
        # usd_amount = taxable / fx_rate if fx_rate known, else stored amount
        usd_amt = round(taxable / fx_rate, 4) if fx_rate > 1 else amount
        return generate_inr_cgst_sgst_invoice(
            inv_no, inv_date, due_date, desc, partner, month,
            usd_amt, fx_rate, desc
        )
    elif inv_type == "INR_IGST":
        usd_amt = round(taxable / fx_rate, 4) if fx_rate > 1 else amount
        return generate_inr_igst_invoice(
            inv_no, inv_date, due_date, desc, partner, month,
            usd_amt, fx_rate, desc
        )

    raise ValueError(f"Unknown invoice_type: {inv_type}")

def _current_fy_label():
    """Return FY label like '25-26' based on today's date."""
    fy_start = datetime.today().year if datetime.today().month >= 4 else datetime.today().year - 1
    return f"{str(fy_start)[2:]}-{str(fy_start+1)[2:]}"

def generate_invoice_number(conn):
    """Auto-generate next invoice number like PMA/25-26/0128
    CN records are excluded so invoice serial never mixes with credit note serial."""
    fy_label = _current_fy_label()
    rows = conn.execute(
        "SELECT invoice_number FROM invoice_details "
        "WHERE invoice_number LIKE ? AND is_credit_note=0",
        (f"PMA/{fy_label}/%",)
    ).fetchall()
    next_num = len(rows) + 1
    return f"PMA/{fy_label}/{next_num:04d}"

def generate_credit_note_number(conn):
    """Auto-generate next credit-note number like CN/25-26/0001.
    Completely separate sequence from invoice numbers; resets each FY on 1 Apr."""
    fy_label = _current_fy_label()
    rows = conn.execute(
        "SELECT invoice_number FROM invoice_details "
        "WHERE invoice_number LIKE ? AND is_credit_note=1",
        (f"CN/{fy_label}/%",)
    ).fetchall()
    next_num = len(rows) + 1
    return f"CN/{fy_label}/{next_num:04d}"

def amount_in_words(amount, currency="USD"):
    try:
        rupees = int(amount)
        paise  = round((amount - rupees) * 100)
        if currency == "USD":
            dollars = int(amount)
            cents   = round((amount - dollars) * 100)
            w = num2words(dollars, lang='en').title()
            if cents:
                w += f" and {num2words(cents, lang='en').title()} Cents"
            return f"United States Dollar {w} Only"
        else:
            w = num2words(rupees, lang='en').title()
            if paise:
                w += f" and {num2words(paise, lang='en').title()} Paise"
            return f"Indian Rupee {w} Only"
    except:
        return ""

def get_partner_info(dsp_name):
    """Fetch partner details from partner_list by short name."""
    conn = get_db()
    df = pd.read_sql(
        'SELECT * FROM partner_list WHERE "short_name"=? OR "Short Name using in Bidscube"=?',
        conn, params=(dsp_name, dsp_name)
    )
    conn.close()
    if df.empty:
        return {}
    r = df.iloc[0]
    return {
        "legal_name":    r.get("legal_name") or r.get("Legal Entity Name", dsp_name),
        "address":       r.get("address") or r.get("Registered Address", ""),
        "country":       r.get("country") or r.get("Country", ""),
        "gstin":         r.get("gstin") or r.get("GSTIN", ""),
        "payment_terms": r.get("payment_terms") or r.get("Payment Terms", "Net 60"),
        "email1":        r.get("email1") or r.get("Email 1", ""),
        "email2":        r.get("email2") or r.get("Email 2", ""),
        "finance_email": r.get("finance_email") or r.get("Finance Email", ""),
    }

def determine_invoice_type(partner_info, manual_override=None):
    """Determine invoice type from country/entity."""
    if manual_override:
        return manual_override
    country = str(partner_info.get("country", "")).lower()
    address = str(partner_info.get("address", "")).lower()
    if "india" not in country and "in)" not in country:
        return "USD"
    # Indian entity — check state
    if "maharashtra" in address or "pune" in address or "mumbai" in address:
        return "INR_CGST_SGST"
    return "INR_IGST"

# ─────────────────────────────────────────────
# PDF GENERATORS
# ─────────────────────────────────────────────
BRAND_BLUE  = colors.HexColor("#003366")
BRAND_LIGHT = colors.HexColor("#e8eef4")
TEXT_DARK   = colors.HexColor("#1a1a1a")

def _base_styles():
    styles = getSampleStyleSheet()
    normal = ParagraphStyle("INV_normal", fontSize=9,  leading=13, textColor=TEXT_DARK)
    bold   = ParagraphStyle("INV_bold",   fontSize=9,  leading=13, textColor=TEXT_DARK, fontName="Helvetica-Bold")
    small  = ParagraphStyle("INV_small",  fontSize=7.5,leading=11, textColor=colors.grey)
    title  = ParagraphStyle("INV_title",  fontSize=22, textColor=BRAND_BLUE, fontName="Helvetica-Bold")
    sub    = ParagraphStyle("INV_sub",    fontSize=11, textColor=BRAND_BLUE, fontName="Helvetica-Bold")
    right  = ParagraphStyle("INV_right",  fontSize=9,  leading=13, alignment=TA_RIGHT)
    return normal, bold, small, title, sub, right

def _invoice_table_style(header_row=0):
    return TableStyle([
        ("BACKGROUND",   (0, header_row), (-1, header_row), BRAND_BLUE),
        ("TEXTCOLOR",    (0, header_row), (-1, header_row), colors.white),
        ("FONTNAME",     (0, header_row), (-1, header_row), "Helvetica-Bold"),
        ("FONTSIZE",     (0, header_row), (-1, header_row), 9),
        ("ALIGN",        (0, header_row), (-1, header_row), "CENTER"),
        ("ROWBACKGROUNDS", (0, header_row+1), (-1, -1), [colors.white, BRAND_LIGHT]),
        ("FONTSIZE",     (0, header_row+1), (-1, -1), 9),
        ("GRID",         (0, header_row), (-1, -1), 0.4, colors.HexColor("#cccccc")),
        ("LEFTPADDING",  (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING",   (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
        ("VALIGN",       (0, 0), (-1, -1), "TOP"),
        ("ALIGN",        (2, 1), (-1, -1), "RIGHT"),   # right-align all numeric cols in data rows
    ])

def _build_header(story, normal, bold, small, title, sub,
                  invoice_number, invoice_date, due_date, terms,
                  partner_info, invoice_type, balance_due_str=""):
    """Build top header section of invoice.
    Zone 1 : Logo (left) | TAX INVOICE (page-centre)
    Zone 2 : Balance Due line
    Zone 3 : Company info (left) | Invoice meta + Bill To (right)
    All zones are exactly 180 mm wide to align with the item table.
    """
    # ── Zone 1: Logo left | TAX INVOICE centred on page ────────────────────
    logo_cell = ""
    if os.path.exists(LOGO_PATH):
        try:
            logo_cell = RLImage(LOGO_PATH, width=40*mm, height=14*mm)
        except Exception:
            logo_cell = Paragraph(COMPANY["name"], title)
    else:
        logo_cell = Paragraph(COMPANY["name"], title)

    title_c = ParagraphStyle("INV_title_c", fontSize=22, textColor=BRAND_BLUE,
                              fontName="Helvetica-Bold", alignment=TA_CENTER)
    # [45 | 90 | 45] = 180 mm; centre of title cell = 45+45 = 90 mm = true page centre
    logo_row = [[logo_cell, Paragraph("TAX INVOICE", title_c), ""]]
    logo_tbl = Table(logo_row, colWidths=[45*mm, 90*mm, 45*mm])
    logo_tbl.setStyle(TableStyle([
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING",   (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 0),
    ]))
    story.append(logo_tbl)

    # ── Zone 2: Balance Due ─────────────────────────────────────────────────
    if balance_due_str:
        story.append(Spacer(1, 6))
        bal_right = ParagraphStyle("INV_bal_r", fontSize=9, fontName="Helvetica-Bold",
                                    alignment=TA_RIGHT, textColor=TEXT_DARK, leading=13)
        story.append(Paragraph(balance_due_str, bal_right))

    story.append(Spacer(1, 8))

    # ── Zone 3: Company (left 85 mm) | Invoice meta + Bill To (right 95 mm) ─
    company_text = (
        f"<b>{COMPANY['name']}</b><br/>"
        f"{COMPANY['address'].replace(chr(10), '<br/>')}<br/>"
        f"{COMPANY['email']}<br/>{COMPANY['website']}"
    )
    if invoice_type != "USD":
        company_text += f"<br/><b>GSTIN</b> {COMPANY['gstin']}"

    bill_to_lines = [f"<b>{partner_info.get('legal_name','')}</b>"]
    addr = partner_info.get("address", "")
    if addr:
        bill_to_lines.append(addr.replace("\n", "<br/>"))
    if partner_info.get("country"):
        bill_to_lines.append(partner_info["country"])
    if invoice_type != "USD" and partner_info.get("gstin"):
        bill_to_lines.append(f"GSTIN {partner_info['gstin']}")
    bill_to_text = "<br/>".join(bill_to_lines)

    right_para = Paragraph(
        f"<b>Invoice# {invoice_number}</b><br/>"
        f"<b>Invoice Date :</b> {invoice_date}<br/>"
        f"<b>Terms :</b> {terms}<br/>"
        f"<b>Due Date :</b> {due_date}<br/><br/>"
        f"<b>Bill To</b><br/>{bill_to_text}",
        normal
    )
    info_data = [[Paragraph(company_text, normal), right_para]]
    info_tbl = Table(info_data, colWidths=[85*mm, 95*mm])
    info_tbl.setStyle(TableStyle([
        ("VALIGN",       (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING",   (0, 0), (-1, -1), 0),
    ]))
    story.append(info_tbl)
    story.append(Spacer(1, 10))

def _build_bank_terms(story, normal, bold, small):
    story.append(Spacer(1, 8))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.lightgrey))
    story.append(Spacer(1, 4))
    bank_text = (
        f"<b>Bank Details</b><br/>"
        f"Account Name: {COMPANY['account_name']}&nbsp;&nbsp;&nbsp;<br/>"
        f"Account Number: {COMPANY['account_number']}&nbsp;&nbsp;&nbsp;<br/>"
        f"IFSC: {COMPANY['ifsc']}<br/>"
        f"SWIFT: {COMPANY['swift']}&nbsp;&nbsp;&nbsp;<br/>"
        f"Bank: {COMPANY['bank']}&nbsp;&nbsp;&nbsp;<br/>"
        f"PayPal: {COMPANY['paypal']}<br/>"
        f"Payoneer: {COMPANY['payoneer']}"
    )
    story.append(Paragraph(bank_text, small))
    story.append(Spacer(1, 6))
    terms_text = (
        "<b>Terms &amp; Conditions</b><br/>"
        "1. Please ensure payment is made by the due date via international wire transfer.<br/>"
        "2. You may also make payment via PayPal | PayPal ID: vaishali.peakmyads@gmail.com<br/>"
        "3. You may also make payment via Payoneer | Payoneer ID: vaishali.peakmyads@gmail.com"
    )

    story.append(Paragraph(terms_text, small))
    story.append(Spacer(1, 6))
    story.append(Paragraph("Thanks for your business!", small))
    story.append(Spacer(1, 10))
    if os.path.exists(SIGN_PATH):
        try:
            _sign = RLImage(SIGN_PATH, width=28*mm, height=28*mm)
            _sign.hAlign = "LEFT"
            story.append(_sign)
        except Exception:
            pass
    story.append(Spacer(1, 4))
    story.append(Paragraph("Vaishali W<br/>Authorized Signature", small))


def generate_usd_invoice(invoice_number, invoice_date, due_date, terms,
                          partner_info, month_label, amount, description):
    """Generate USD invoice PDF (like ADOKUT format)."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             leftMargin=15*mm, rightMargin=15*mm,
                             topMargin=15*mm, bottomMargin=15*mm)
    normal, bold, small, title, sub, right = _base_styles()
    story = []

    _build_header(story, normal, bold, small, title, sub,
                  invoice_number, invoice_date, due_date, terms, partner_info, "USD",
                  balance_due_str=f"Balance Due: <b>${amount:,.2f}</b>")

    # Line items table
    headers = ["Sr.", "Item & Description", "Amount ($)", "Total ($)"]
    rows = [headers, [
        "1",
        Paragraph(f"Impressions of Online Media Advertising (oRTB)<br/><i>FOR {month_label.upper()} (AFTER OFFSET)</i><br/>", normal),
        f"${amount:,.2f}",
        f"${amount:,.2f}"
    ]]
    col_w = [10*mm, 116*mm, 27*mm, 27*mm]
    t = Table(rows, colWidths=col_w, repeatRows=1)
    t.setStyle(_invoice_table_style())
    story.append(t)
    story.append(Spacer(1, 4))

    # Total row
    total_data = [["", "", "Total", f"${amount:,.2f}"]]
    tot = Table(total_data, colWidths=col_w)
    tot.setStyle(TableStyle([
        ("FONTNAME",     (2,0), (-1,0), "Helvetica-Bold"),
        ("ALIGN",        (2,0), (-1,0), "RIGHT"),
        ("LINEABOVE",    (0,0), (-1,0), 0.5, colors.grey),
    ]))
    story.append(tot)
    story.append(Spacer(1, 4))
    story.append(Paragraph(f"<b>Total In Words:</b> {amount_in_words(amount, 'USD')}", normal))

    _build_bank_terms(story, normal, bold, small)
    doc.build(story)
    buf.seek(0)
    return buf


def generate_inr_cgst_sgst_invoice(invoice_number, invoice_date, due_date, terms,
                                    partner_info, month_label, usd_amount,
                                    fx_rate, description):
    """Generate INR invoice with CGST+SGST (same state — Maharashtra)."""
    taxable  = round(usd_amount * fx_rate, 2)
    cgst     = round(taxable * 0.09, 2)
    sgst     = round(taxable * 0.09, 2)
    total    = round(taxable + cgst + sgst, 2)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             leftMargin=15*mm, rightMargin=15*mm,
                             topMargin=15*mm, bottomMargin=15*mm)
    normal, bold, small, title, sub, right = _base_styles()
    story = []

    _build_header(story, normal, bold, small, title, sub,
                  invoice_number, invoice_date, due_date, terms, partner_info, "INR_CGST_SGST",
                  balance_due_str=f"Balance Due: <b>Rs.{total:,.2f}</b>")

    if partner_info.get("address"):
        story.append(Paragraph(f"Place Of Supply: Maharashtra (27)", normal))
        story.append(Spacer(1, 4))

    headers = ["Sr.", "Item & Description", "HSN/SAC", "Taxable Amount", "CGST", "SGST", "Total (Rs.)"]
    desc = (f"Impressions of Online Media Advertising (oRTB) INR<br/>"
            f"<i>FOR {month_label.upper()} (AFTER OFFSET)</i><br/>"
            f"${usd_amount:,.2f} * INR {fx_rate:.2f}")
    rows = [headers, [
        "1",
        Paragraph(desc, normal),
        "998361",
        f"Rs.{taxable:,.2f}",
        f"Rs.{cgst:,.2f}\n9%",
        f"Rs.{sgst:,.2f}\n9%",
        f"Rs.{total:,.2f}"
    ]]
    col_w = [8*mm, 70*mm, 18*mm, 26*mm, 18*mm, 18*mm, 22*mm]
    t = Table(rows, colWidths=col_w, repeatRows=1)
    t.setStyle(_invoice_table_style())
    story.append(t)
    story.append(Spacer(1, 4))

    total_data = [["", "", "", "", "", "Total", f"Rs.{total:,.2f}"]]
    tot = Table(total_data, colWidths=col_w)
    tot.setStyle(TableStyle([
        ("FONTNAME",     (5,0), (-1,0), "Helvetica-Bold"),
        ("ALIGN",        (5,0), (-1,0), "RIGHT"),
        ("LINEABOVE",    (0,0), (-1,0), 0.5, colors.grey),
    ]))
    story.append(tot)
    story.append(Spacer(1, 4))
    story.append(Paragraph(f"<b>Total In Words:</b> {amount_in_words(total, 'INR')}", normal))

    _build_bank_terms(story, normal, bold, small)
    doc.build(story)
    buf.seek(0)
    return buf


def generate_inr_igst_invoice(invoice_number, invoice_date, due_date, terms,
                               partner_info, month_label, usd_amount,
                               fx_rate, description):
    """Generate INR invoice with IGST (outside Maharashtra)."""
    taxable = round(usd_amount * fx_rate, 2)
    igst    = round(taxable * 0.18, 2)
    total   = round(taxable + igst, 2)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             leftMargin=15*mm, rightMargin=15*mm,
                             topMargin=15*mm, bottomMargin=15*mm)
    normal, bold, small, title, sub, right = _base_styles()
    story = []

    _build_header(story, normal, bold, small, title, sub,
                  invoice_number, invoice_date, due_date, terms, partner_info, "INR_IGST",
                  balance_due_str=f"Balance Due: <b>Rs.{total:,.2f}</b>")

    story.append(Paragraph("Place Of Supply: As per billing address", normal))
    story.append(Spacer(1, 4))

    headers = ["Sr.", "Item & Description", "HSN/SAC", "Taxable Amount", "IGST", "Total (Rs.)"]
    desc = (f"Impressions of Online Media Advertising (oRTB) INR<br/>"
            f"<i>FOR {month_label.upper()} (AFTER OFFSET)</i><br/>"
            f"${usd_amount:,.2f} * INR {fx_rate:.2f}")
    rows = [headers, [
        "1",
        Paragraph(desc, normal),
        "998361",
        f"Rs.{taxable:,.2f}",
        f"Rs.{igst:,.2f}\n18%",
        f"Rs.{total:,.2f}"
    ]]
    col_w = [8*mm, 77*mm, 18*mm, 30*mm, 22*mm, 25*mm]
    t = Table(rows, colWidths=col_w, repeatRows=1)
    t.setStyle(_invoice_table_style())
    story.append(t)
    story.append(Spacer(1, 4))

    total_data = [["", "", "", "", "Total", f"Rs.{total:,.2f}"]]
    tot = Table(total_data, colWidths=col_w)
    tot.setStyle(TableStyle([
        ("FONTNAME",     (4,0), (-1,0), "Helvetica-Bold"),
        ("ALIGN",        (4,0), (-1,0), "RIGHT"),
        ("LINEABOVE",    (0,0), (-1,0), 0.5, colors.grey),
    ]))
    story.append(tot)
    story.append(Spacer(1, 4))
    story.append(Paragraph(f"<b>Total In Words:</b> {amount_in_words(total, 'INR')}", normal))

    _build_bank_terms(story, normal, bold, small)
    doc.build(story)
    buf.seek(0)
    return buf



def generate_multi_month_usd_invoice(invoice_number, invoice_date, due_date, terms,
                                      partner_info, month_data, description):
    """Generate USD invoice with separate line per month.
    month_data: list of dicts {month, amount}
    """
    total_amount = sum(m["amount"] for m in month_data)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             leftMargin=15*mm, rightMargin=15*mm,
                             topMargin=15*mm, bottomMargin=15*mm)
    normal, bold, small, title, sub, right = _base_styles()
    story = []
    _build_header(story, normal, bold, small, title, sub,
                  invoice_number, invoice_date, due_date, terms, partner_info, "USD",
                  balance_due_str=f"Balance Due: <b>${total_amount:,.2f}</b>")

    headers = ["Sr.", "Item & Description", "Amount ($)", "Total ($)"]
    rows = [headers]
    for i, m in enumerate(month_data, 1):
        rows.append([
            str(i),
            Paragraph(
                f"Impressions of Online Media Advertising (oRTB)<br/>"
                f"<i>FOR {str(m['month']).upper()} (AFTER OFFSET)</i>", normal),
            f"${m['amount']:,.2f}",
            f"${m['amount']:,.2f}"
        ])
    col_w = [10*mm, 116*mm, 27*mm, 27*mm]
    t = Table(rows, colWidths=col_w, repeatRows=1)
    t.setStyle(_invoice_table_style())
    story.append(t)
    story.append(Spacer(1, 4))
    total_data = [["", "", "Total", f"${total_amount:,.2f}"]]
    tot = Table(total_data, colWidths=col_w)
    tot.setStyle(TableStyle([
        ("FONTNAME", (2,0), (-1,0), "Helvetica-Bold"),
        ("ALIGN",    (2,0), (-1,0), "RIGHT"),
        ("LINEABOVE",(0,0), (-1,0), 0.5, colors.grey),
    ]))
    story.append(tot)
    story.append(Spacer(1, 4))
    story.append(Paragraph(f"<b>Total In Words:</b> {amount_in_words(total_amount, 'USD')}", normal))
    _build_bank_terms(story, normal, bold, small)
    doc.build(story)
    buf.seek(0)
    return buf


def generate_multi_month_inr_cgst_sgst_invoice(invoice_number, invoice_date, due_date, terms,
                                                 partner_info, month_data, description):
    """Generate INR CGST+SGST invoice with separate line per month.
    month_data: list of dicts {month, usd_amount, fx_rate}
    """
    grand_taxable = 0.0
    grand_cgst = 0.0
    grand_sgst = 0.0
    grand_total = 0.0
    for m in month_data:
        _t = round(m["usd_amount"] * m["fx_rate"], 2)
        _c = round(_t * 0.09, 2)
        _s = round(_t * 0.09, 2)
        grand_taxable += _t
        grand_cgst += _c
        grand_sgst += _s
        grand_total += _t + _c + _s
    grand_taxable = round(grand_taxable, 2)
    grand_cgst    = round(grand_cgst, 2)
    grand_sgst    = round(grand_sgst, 2)
    grand_total   = round(grand_total, 2)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             leftMargin=15*mm, rightMargin=15*mm,
                             topMargin=15*mm, bottomMargin=15*mm)
    normal, bold, small, title, sub, right = _base_styles()
    story = []
    _build_header(story, normal, bold, small, title, sub,
                  invoice_number, invoice_date, due_date, terms, partner_info, "INR_CGST_SGST",
                  balance_due_str=f"Balance Due: <b>Rs.{grand_total:,.2f}</b>")
    if partner_info.get("address"):
        story.append(Paragraph("Place Of Supply: Maharashtra (27)", normal))
        story.append(Spacer(1, 4))

    headers = ["Sr.", "Item & Description", "HSN/SAC", "Taxable Amt", "CGST", "SGST", "Total (Rs.)"]
    rows = [headers]
    for i, m in enumerate(month_data, 1):
        _t = round(m["usd_amount"] * m["fx_rate"], 2)
        _c = round(_t * 0.09, 2)
        _s = round(_t * 0.09, 2)
        _row_total = round(_t + _c + _s, 2)
        desc = (f"Impressions of Online Media Advertising (oRTB) INR<br/>"
                f"<i>FOR {str(m['month']).upper()} (AFTER OFFSET)</i><br/>"
                f"${m['usd_amount']:,.2f} × INR {m['fx_rate']:.4f}")
        rows.append([str(i), Paragraph(desc, normal), "998361",
                     f"Rs.{_t:,.2f}", f"Rs.{_c:,.2f}\n9%", f"Rs.{_s:,.2f}\n9%", f"Rs.{_row_total:,.2f}"])
    col_w = [8*mm, 67*mm, 18*mm, 26*mm, 18*mm, 18*mm, 25*mm]
    t = Table(rows, colWidths=col_w, repeatRows=1)
    t.setStyle(_invoice_table_style())
    story.append(t)
    story.append(Spacer(1, 4))
    total_data = [["", "", "", f"Rs.{grand_taxable:,.2f}", f"Rs.{grand_cgst:,.2f}", f"Rs.{grand_sgst:,.2f}", f"Rs.{grand_total:,.2f}"]]
    tot = Table(total_data, colWidths=col_w)
    tot.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("ALIGN",    (3,0), (-1,0), "RIGHT"),
        ("LINEABOVE",(0,0), (-1,0), 0.5, colors.grey),
    ]))
    story.append(tot)
    story.append(Spacer(1, 4))
    story.append(Paragraph(f"<b>Total In Words:</b> {amount_in_words(grand_total, 'INR')}", normal))
    _build_bank_terms(story, normal, bold, small)
    doc.build(story)
    buf.seek(0)
    return buf


def generate_multi_month_inr_igst_invoice(invoice_number, invoice_date, due_date, terms,
                                            partner_info, month_data, description):
    """Generate INR IGST invoice with separate line per month.
    month_data: list of dicts {month, usd_amount, fx_rate}
    """
    grand_taxable = 0.0
    grand_igst    = 0.0
    grand_total   = 0.0
    for m in month_data:
        _t = round(m["usd_amount"] * m["fx_rate"], 2)
        _i = round(_t * 0.18, 2)
        grand_taxable += _t
        grand_igst    += _i
        grand_total   += _t + _i
    grand_taxable = round(grand_taxable, 2)
    grand_igst    = round(grand_igst, 2)
    grand_total   = round(grand_total, 2)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             leftMargin=15*mm, rightMargin=15*mm,
                             topMargin=15*mm, bottomMargin=15*mm)
    normal, bold, small, title, sub, right = _base_styles()
    story = []
    _build_header(story, normal, bold, small, title, sub,
                  invoice_number, invoice_date, due_date, terms, partner_info, "INR_IGST",
                  balance_due_str=f"Balance Due: <b>Rs.{grand_total:,.2f}</b>")
    story.append(Paragraph("Place Of Supply: As per billing address", normal))
    story.append(Spacer(1, 4))

    headers = ["Sr.", "Item & Description", "HSN/SAC", "Taxable Amt", "IGST", "Total (Rs.)"]
    rows = [headers]
    for i, m in enumerate(month_data, 1):
        _t = round(m["usd_amount"] * m["fx_rate"], 2)
        _i = round(_t * 0.18, 2)
        _row_total = round(_t + _i, 2)
        desc = (f"Impressions of Online Media Advertising (oRTB) INR<br/>"
                f"<i>FOR {str(m['month']).upper()} (AFTER OFFSET)</i><br/>"
                f"${m['usd_amount']:,.2f} × INR {m['fx_rate']:.4f}")
        rows.append([str(i), Paragraph(desc, normal), "998361",
                     f"Rs.{_t:,.2f}", f"Rs.{_i:,.2f}\n18%", f"Rs.{_row_total:,.2f}"])
    col_w = [8*mm, 73*mm, 18*mm, 30*mm, 24*mm, 27*mm]
    t = Table(rows, colWidths=col_w, repeatRows=1)
    t.setStyle(_invoice_table_style())
    story.append(t)
    story.append(Spacer(1, 4))
    total_data = [["", "", "", f"Rs.{grand_taxable:,.2f}", f"Rs.{grand_igst:,.2f}", f"Rs.{grand_total:,.2f}"]]
    tot = Table(total_data, colWidths=col_w)
    tot.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("ALIGN",    (3,0), (-1,0), "RIGHT"),
        ("LINEABOVE",(0,0), (-1,0), 0.5, colors.grey),
    ]))
    story.append(tot)
    story.append(Spacer(1, 4))
    story.append(Paragraph(f"<b>Total In Words:</b> {amount_in_words(grand_total, 'INR')}", normal))
    _build_bank_terms(story, normal, bold, small)
    doc.build(story)
    buf.seek(0)
    return buf


def generate_credit_note_pdf(cn_number, cn_date, original_inv_number, reason,
                              partner_info, month_label, amount, currency,
                              taxable=0.0, cgst=0.0, sgst=0.0, igst=0.0, fx_rate=1.0):
    """Generate a Credit Note PDF mirroring the invoice layout."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             leftMargin=15*mm, rightMargin=15*mm,
                             topMargin=15*mm, bottomMargin=15*mm)
    normal, bold, small, title, sub, right = _base_styles()
    story = []

    # Balance line
    if currency == "INR":
        bal_str = f"Credit Amount: <b>Rs.{amount:,.2f}</b>"
    else:
        bal_str = f"Credit Amount: <b>${amount:,.2f}</b>"
    # ── Zone 1: Logo left | CREDIT NOTE centred on page ──────────────────
    logo_cell = ""
    if os.path.exists(LOGO_PATH):
        try:
            logo_cell = RLImage(LOGO_PATH, width=40*mm, height=14*mm)
        except Exception:
            logo_cell = Paragraph(COMPANY["name"], title)
    else:
        logo_cell = Paragraph(COMPANY["name"], title)

    title_c = ParagraphStyle("CN_title_c", fontSize=22, textColor=BRAND_BLUE,
                              fontName="Helvetica-Bold", alignment=TA_CENTER)
    # [45 | 90 | 45] = 180 mm — title perfectly centred on page
    logo_row = [[logo_cell, Paragraph("CREDIT NOTE", title_c), ""]]
    logo_tbl = Table(logo_row, colWidths=[45*mm, 90*mm, 45*mm])
    logo_tbl.setStyle(TableStyle([
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING",   (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 0),
    ]))
    story.append(logo_tbl)

    # ── Zone 2: Credit Amount ───────────────────────────────────────────────
    story.append(Spacer(1, 6))
    cn_bal_right = ParagraphStyle("CN_bal_r", fontSize=9, fontName="Helvetica-Bold",
                                   alignment=TA_RIGHT, textColor=TEXT_DARK, leading=13)
    story.append(Paragraph(bal_str, cn_bal_right))
    story.append(Spacer(1, 8))

    # ── Zone 3: Company (left 85 mm) | CN meta + Bill To (right 95 mm) ─────

    company_text = (
        f"<b>{COMPANY['name']}</b><br/>"
        f"{COMPANY['address'].replace(chr(10), '<br/>')}<br/>"
        f"{COMPANY['email']}<br/>{COMPANY['website']}"
    )
    if currency == "INR":
        company_text += f"<br/><b>GSTIN</b> {COMPANY['gstin']}"

    bill_to_lines = [f"<b>{partner_info.get('legal_name','')}</b>"]
    addr = partner_info.get("address", "")
    if addr:
        bill_to_lines.append(addr.replace("\n", "<br/>"))
    if partner_info.get("country"):
        bill_to_lines.append(partner_info["country"])
    if currency == "INR" and partner_info.get("gstin"):
        bill_to_lines.append(f"GSTIN {partner_info['gstin']}")
    bill_to_text = "<br/>".join(bill_to_lines)

    cn_right_para = Paragraph(
        f"<b>CN# {cn_number}</b><br/>"
        f"<b>CN Date :</b> {cn_date}<br/>"
        f"<b>Original Invoice :</b> {original_inv_number}<br/>"
        f"<b>Reason :</b> {reason}<br/><br/>"
        f"<b>Bill To</b><br/>{bill_to_text}",
        normal
    )
    info_data = [[Paragraph(company_text, normal), cn_right_para]]
    info_tbl = Table(info_data, colWidths=[85*mm, 95*mm])
    info_tbl.setStyle(TableStyle([
        ("VALIGN",       (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING",   (0, 0), (-1, -1), 0),
    ]))
    story.append(info_tbl)
    story.append(Spacer(1, 10))

    # Line items
    if currency == "INR" and (cgst > 0 or igst > 0):
        if cgst > 0:
            headers = ["Sr.", "Item & Description", "HSN/SAC", "Taxable Amount", "CGST", "SGST", "Total (Rs.)"]
            rows = [headers, [
                "1",
                Paragraph(f"Credit Note against Invoice {original_inv_number}<br/>"
                          f"<i>FOR {month_label.upper()}</i>", normal),
                "998361",
                f"Rs.{taxable:,.2f}",
                f"Rs.{cgst:,.2f}\n9%",
                f"Rs.{sgst:,.2f}\n9%",
                f"Rs.{amount:,.2f}"
            ]]
            col_w = [8*mm, 70*mm, 18*mm, 26*mm, 18*mm, 18*mm, 22*mm]
            total_data = [["", "", "", "", "", "Total", f"Rs.{amount:,.2f}"]]
            tot = Table(total_data, colWidths=col_w)
            tot.setStyle(TableStyle([("FONTNAME",(5,0),(-1,0),"Helvetica-Bold"),("ALIGN",(5,0),(-1,0),"RIGHT"),("LINEABOVE",(0,0),(-1,0),0.5,colors.grey)]))
        else:
            headers = ["Sr.", "Item & Description", "HSN/SAC", "Taxable Amount", "IGST", "Total (Rs.)"]
            rows = [headers, [
                "1",
                Paragraph(f"Credit Note against Invoice {original_inv_number}<br/>"
                          f"<i>FOR {month_label.upper()}</i>", normal),
                "998361",
                f"Rs.{taxable:,.2f}",
                f"Rs.{igst:,.2f}\n18%",
                f"Rs.{amount:,.2f}"
            ]]
            col_w = [8*mm, 77*mm, 18*mm, 30*mm, 22*mm, 25*mm]
            total_data = [["", "", "", "", "Total", f"Rs.{amount:,.2f}"]]
            tot = Table(total_data, colWidths=col_w)
            tot.setStyle(TableStyle([("FONTNAME",(4,0),(-1,0),"Helvetica-Bold"),("ALIGN",(4,0),(-1,0),"RIGHT"),("LINEABOVE",(0,0),(-1,0),0.5,colors.grey)]))
    else:
        headers = ["Sr.", "Item & Description", "Amount ($)", "Total ($)"]
        rows = [headers, [
            "1",
            Paragraph(f"Credit Note against Invoice {original_inv_number}<br/>"
                      f"<i>FOR {month_label.upper()}</i>", normal),
            f"${amount:,.2f}",
            f"${amount:,.2f}"
        ]]
        col_w = [10*mm, 116*mm, 27*mm, 27*mm]
        total_data = [["", "", "Total", f"${amount:,.2f}"]]
        tot = Table(total_data, colWidths=col_w)
        tot.setStyle(TableStyle([("FONTNAME",(2,0),(-1,0),"Helvetica-Bold"),("ALIGN",(2,0),(-1,0),"RIGHT"),("LINEABOVE",(0,0),(-1,0),0.5,colors.grey)]))

    t = Table(rows, colWidths=col_w, repeatRows=1)
    t.setStyle(_invoice_table_style())
    story.append(t)
    story.append(Spacer(1, 4))
    story.append(tot)
    story.append(Spacer(1, 4))

    currency_label = "INR" if currency == "INR" else "USD"
    story.append(Paragraph(f"<b>Total In Words:</b> {amount_in_words(amount, currency_label)}", normal))
    _build_bank_terms(story, normal, bold, small)
    doc.build(story)
    buf.seek(0)
    return buf


def generate_dsp_statement(dsp_name, dsp_df):
    """Generate DSP account statement PDF."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             leftMargin=15*mm, rightMargin=15*mm,
                             topMargin=15*mm, bottomMargin=15*mm)
    normal, bold, small, title, sub, right = _base_styles()
    story = []

    # ── Logo + Title ──────────────────────────────────
    logo_cell = ""
    if os.path.exists(LOGO_PATH):
        try:
            logo_cell = RLImage(LOGO_PATH, width=40*mm, height=14*mm)
        except Exception:
            logo_cell = Paragraph(COMPANY["name"], bold)
    else:
        logo_cell = Paragraph(COMPANY["name"], bold)

    hdr_row = [[logo_cell, Paragraph("DSP Account Statement", title)]]
    hdr_tbl = Table(hdr_row, colWidths=[44*mm, 146*mm])
    hdr_tbl.setStyle(TableStyle([
        ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING", (0,0), (-1,-1), 0),
        ("RIGHTPADDING",(0,0), (-1,-1), 0),
        ("TOPPADDING",  (0,0), (-1,-1), 0),
        ("BOTTOMPADDING",(0,0),(-1,-1), 0),
    ]))
    story.append(hdr_tbl)
    story.append(Spacer(1, 4))
    story.append(Paragraph(f"<b>{COMPANY['name']}</b>", bold))
    story.append(Spacer(1, 4))
    story.append(Paragraph(f"<b>Partner:</b> {dsp_name}&nbsp;&nbsp;&nbsp;"
                            f"<b>Generated:</b> {datetime.today().strftime('%d-%b-%Y')}", normal))
    story.append(Spacer(1, 8))

    df = dsp_df[dsp_df["DSP Name"] == dsp_name].copy()
    df = df.sort_values("Month")

    # Load invoice history to get Invoice No & Invoice Date per month
    try:
        _conn = get_db()
        inv_hist = pd.read_sql(
            "SELECT dsp_name, month, invoice_number, invoice_date FROM invoice_details "
            "WHERE is_credit_note=0 AND is_deleted=0",
            _conn
        )
        _conn.close()
        # Build map with both the raw month string AND a normalised Mon-YYYY key
        inv_hist_map = {}
        for _, _ir in inv_hist.iterrows():
            _raw_key  = str(_ir["dsp_name"]).strip() + "|" + str(_ir["month"]).strip()
            # Also store under normalised Mon-YYYY so lookup succeeds regardless of source format
            try:
                _norm_month = pd.to_datetime(str(_ir["month"]), format="%b-%Y", errors="coerce")
                if pd.isna(_norm_month):
                    _norm_month = pd.to_datetime(str(_ir["month"]), errors="coerce")
                _norm_key = str(_ir["dsp_name"]).strip() + "|" + _norm_month.strftime("%b-%Y")
            except Exception:
                _norm_key = _raw_key
            _val = {"invoice_number": _ir["invoice_number"], "invoice_date": _ir["invoice_date"]}
            inv_hist_map[_raw_key]  = _val
            inv_hist_map[_norm_key] = _val
    except Exception:
        inv_hist_map = {}

    headers = ["Month", "Invoice Date", "Invoice No.", "Receivable $", "Due Date",
               "Received Date", "Received $", "Received In", "Shortage $", "Status"]
    rows = [headers]
    for _, row in df.iterrows():
        receivable = float(row.get("Receivable $", 0) or 0)
        received   = float(row.get("Received Amount $", 0) or 0)
        shortage   = receivable - received
        status     = "Paid" if received >= receivable and receivable > 0 else ("Partial" if received > 0 else "Unpaid")

        month_str = str(row.get("Month", ""))
        try:
            month_fmt = pd.to_datetime(month_str).strftime("%b-%Y")
        except Exception:
            month_fmt = month_str

        dsp_key = str(row.get("DSP Name", "")).strip() + "|" + month_fmt
        inv_info = inv_hist_map.get(dsp_key, {})
        inv_no   = inv_info.get("invoice_number", "")
        inv_date = ""
        if inv_info.get("invoice_date"):
            try:
                inv_date = pd.to_datetime(inv_info["invoice_date"]).strftime("%d-%b-%Y")
            except Exception:
                inv_date = str(inv_info.get("invoice_date", ""))

        # Format due date as DD-MMM-YYYY
        due_raw = row.get("Due Date", "")
        due_fmt = ""
        try:
            _due_parsed = pd.to_datetime(due_raw, errors="coerce")
            if pd.notna(_due_parsed):
                due_fmt = _due_parsed.strftime("%d-%b-%Y")
        except Exception:
            pass

        # Clean Received Date — hide NaT/nan
        _recv_date_raw = row.get("Received Date", "")
        try:
            _recv_date_parsed = pd.to_datetime(_recv_date_raw, errors="coerce")
            recv_date_fmt = "" if pd.isna(_recv_date_parsed) else _recv_date_parsed.strftime("%d-%b-%Y")
        except Exception:
            recv_date_fmt = str(_recv_date_raw or "")
        if recv_date_fmt.lower() in ("nat", "nan", "none"):
            recv_date_fmt = ""

        # Clean Received In — hide "Select" placeholder
        _recv_in_raw = str(row.get("Received In", "") or "").strip()
        recv_in_fmt  = "" if _recv_in_raw.lower() in ("select", "none", "nan") else _recv_in_raw

        rows.append([
            month_fmt,
            inv_date,
            inv_no,
            f"${receivable:,.2f}",
            due_fmt,
            recv_date_fmt,
            f"${received:,.2f}",
            recv_in_fmt,
            f"${shortage:,.2f}",
            status
        ])

    # Totals row
    total_recv = df["Receivable $"].apply(pd.to_numeric, errors="coerce").sum()
    total_paid = df["Received Amount $"].apply(pd.to_numeric, errors="coerce").sum()
    rows.append(["TOTAL", "", "", f"${total_recv:,.2f}", "", "", f"${total_paid:,.2f}", "",
                 f"${total_recv-total_paid:,.2f}", ""])

    col_w = [16*mm, 20*mm, 28*mm, 20*mm, 20*mm, 20*mm, 18*mm, 18*mm, 18*mm, 14*mm]
    t = Table(rows, colWidths=col_w, repeatRows=1)
    t.setStyle(_invoice_table_style())
    t.setStyle(TableStyle([
        ("BACKGROUND",  (0, len(rows)-1), (-1, len(rows)-1), BRAND_BLUE),
        ("TEXTCOLOR",   (0, len(rows)-1), (-1, len(rows)-1), colors.white),
        ("FONTNAME",    (0, len(rows)-1), (-1, len(rows)-1), "Helvetica-Bold"),
    ]))
    story.append(t)

    doc.build(story)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────
# EMAIL SENDER
# ─────────────────────────────────────────────
BCC_EMAIL = "vaishali.peakmyads@gmail.com"

def _smtp_send(smtp_host, smtp_port, smtp_user, smtp_pass,
               all_recipients, msg_string):
    """
    Sends email via GoDaddy SMTP (smtpout.secureserver.net : 465 SSL/TLS).
    Tries three SSL context modes so transient handshake errors are handled.
    """
    import ssl as _ssl, smtplib as _smtplib

    errors = []

    # ── Attempt 1: Standard SSL context with ehlo ──────────────────────────
    try:
        _ctx = _ssl.create_default_context()
        with _smtplib.SMTP_SSL(smtp_host, int(smtp_port),
                                context=_ctx, timeout=20) as _srv:
            _srv.ehlo()
            _srv.login(smtp_user, smtp_pass)
            _srv.sendmail(smtp_user, all_recipients, msg_string)
        return  # success
    except Exception as _e:
        errors.append(f"Attempt 1 (SSL+ehlo): {_e}")

    # ── Attempt 2: Relaxed SSL context (disables cert verification) ────────
    # GoDaddy sometimes has cert chain issues in certain environments
    try:
        _ctx2 = _ssl.SSLContext(_ssl.PROTOCOL_TLS_CLIENT)
        _ctx2.check_hostname = False
        _ctx2.verify_mode    = _ssl.CERT_NONE
        with _smtplib.SMTP_SSL(smtp_host, int(smtp_port),
                                context=_ctx2, timeout=20) as _srv2:
            _srv2.ehlo()
            _srv2.login(smtp_user, smtp_pass)
            _srv2.sendmail(smtp_user, all_recipients, msg_string)
        return  # success
    except Exception as _e2:
        errors.append(f"Attempt 2 (SSL relaxed): {_e2}")

    # ── Attempt 3: STARTTLS on port 587 as last resort ─────────────────────
    try:
        _ctx3 = _ssl.create_default_context()
        with _smtplib.SMTP(smtp_host, 587, timeout=20) as _srv3:
            _srv3.ehlo()
            _srv3.starttls(context=_ctx3)
            _srv3.ehlo()
            _srv3.login(smtp_user, smtp_pass)
            _srv3.sendmail(smtp_user, all_recipients, msg_string)
        return  # success
    except Exception as _e3:
        errors.append(f"Attempt 3 (STARTTLS 587): {_e3}")

    # All failed — raise with full detail so user sees the real reason
    raise ConnectionError(
        "All SMTP connection attempts failed:\n" + "\n".join(errors)
    )


def send_invoice_email(to_emails, subject, body, pdf_buffer, pdf_filename,
                       smtp_host, smtp_port, smtp_user, smtp_pass):
    msg = MIMEMultipart()
    msg["From"]    = smtp_user
    msg["To"]      = ", ".join(to_emails)
    msg["Bcc"]     = BCC_EMAIL
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "html"))

    part = MIMEBase("application", "octet-stream")
    part.set_payload(pdf_buffer.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{pdf_filename}"')
    msg.attach(part)

    all_recipients = to_emails + [BCC_EMAIL]
    _smtp_send(smtp_host, smtp_port, smtp_user, smtp_pass,
               all_recipients, msg.as_string())


# ─────────────────────────────────────────────
# SAVE INVOICE TO DB
# ─────────────────────────────────────────────
def save_invoice(invoice_number, invoice_date, dsp_name, month, invoice_type,
                 amount, tax_amount, total_amount, currency, due_date, status="Draft",
                 igst_amount=0.0, cgst_amount=0.0, sgst_amount=0.0, taxable_amount=0.0,
                 is_credit_note=0, credit_note_ref="", credit_note_reason="", fx_rate=0.0,
                 month_breakdown=""):
    import json as _json
    conn = get_db()
    try:
        conn.execute("""
            INSERT OR REPLACE INTO invoice_details
            (invoice_number, invoice_date, dsp_name, month, invoice_type,
             amount, tax_amount, total_amount, currency, due_date, status, created_at,
             igst_amount, cgst_amount, sgst_amount, taxable_amount,
             is_credit_note, credit_note_ref, credit_note_reason, fx_rate, month_breakdown)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (invoice_number, invoice_date, dsp_name, month, invoice_type,
              amount, tax_amount, total_amount, currency, due_date, status,
              datetime.today().isoformat(),
              igst_amount, cgst_amount, sgst_amount, taxable_amount,
              is_credit_note, credit_note_ref, credit_note_reason, fx_rate, month_breakdown))
        conn.commit()
    finally:
        conn.close()


# ─────────────────────────────────────────────
# MAIN STREAMLIT RENDER FUNCTION
# ─────────────────────────────────────────────

def _sync_invoice_status(dsp_df):
    """
    Sync invoice status in invoice_details based on DSP payment data.

    Logic (same for single-month and multi-month):
      SUM(Received Amount $) across all component months of this invoice
      SUM(Receivable $)      across all component months of this invoice

      total_received >= total_receivable  (and receivable > 0)  →  Paid
      0 < total_received < total_receivable                      →  Partially Paid
      total_received == 0                                        →  Raised (revert if was Paid/Partial)

    Works for INR and USD invoices — comparison is always in USD (DSP tab is always USD).
    """
    try:
        _conn = get_db()
        _rows = _conn.execute(
            "SELECT invoice_number, dsp_name, month, status "
            "FROM invoice_details WHERE is_deleted=0 AND is_credit_note=0"
        ).fetchall()

        # Normalise any month string → "Mon-YYYY"
        def _nm(m):
            try:
                _p = pd.to_datetime(str(m).strip(), format="%b-%Y", errors="coerce")
                if pd.isna(_p):
                    _p = pd.to_datetime(str(m).strip(), errors="coerce")
                return _p.strftime("%b-%Y") if pd.notna(_p) else str(m).strip()
            except Exception:
                return str(m).strip()

        # Strip $ / commas and parse as float
        def _usd(v):
            try:
                s = str(v or 0).replace("$","").replace(",","").strip()
                return float(s) if s else 0.0
            except Exception:
                return 0.0

        # Build DSP lookup: (dsp_name, month_str) → {received, receivable}
        _lk = {}
        for _, _dr in dsp_df.iterrows():
            _key = (str(_dr.get("DSP Name","")).strip(), _nm(str(_dr.get("Month",""))))
            _lk[_key] = {
                "received":   _usd(_dr.get("Received Amount $", 0)),
                "receivable": _usd(_dr.get("Receivable $",      0)),
            }

        for _inv_no, _dsp, _month_raw, _cur_status in _rows:
            _dsp_k  = str(_dsp).strip()
            _parts  = [_nm(m) for m in str(_month_raw).split(",") if m.strip()]

            # Sum received and receivable across ALL component months
            _total_received   = 0.0
            _total_receivable = 0.0
            _found_any        = False

            for _mp in _parts:
                _dk = (_dsp_k, _mp)
                if _dk not in _lk:
                    continue   # month not in current DSP data — skip
                _found_any        = True
                _total_received   += _lk[_dk]["received"]
                _total_receivable += _lk[_dk]["receivable"]

            if not _found_any:
                continue   # no DSP data for this invoice's months — leave status unchanged

            # Determine new status from totals
            if _total_receivable > 0 and _total_received >= _total_receivable:
                _new_status = "Paid"
            elif _total_received > 0:
                _new_status = "Partially Paid"
            else:
                # No payment received — revert to Raised if it was previously marked Paid/Partial
                _new_status = "Raised" if _cur_status in ("Paid", "Partially Paid") else _cur_status

            if _new_status != _cur_status:
                _conn.execute(
                    "UPDATE invoice_details SET status=? WHERE invoice_number=?",
                    (_new_status, _inv_no)
                )

        _conn.commit()
        _conn.close()
    except Exception:
        pass

def render_invoice_module(dsp_df: pd.DataFrame):
    """
    Call this from app.py inside the DSP tab.
    dsp_df: the full DSP dataframe from session state.
    """
    
    st.markdown('''
            <div style="background:linear-gradient(135deg,#003366 0%,#005599 100%);
            border-radius:10px;padding:18px 24px;margin-bottom:20px;
            box-shadow:0 4px 16px rgba(0,51,102,.2);
            display:flex;align-items:center;height:55px;gap:14px;">
            <div style="font-size:32px;">🧾</div>
            <div>
                <div style="color:white;font-size:20px;font-weight:800;">
                    Invoice Manager</div>
                <div style="color:#90caf9;font-size:12px;margin-top:3px;">
                    Create Invoice &nbsp;|&nbsp;
                    Invoice History &nbsp;|&nbsp;
                    Send Reminder &nbsp;|&nbsp;
                    DSP Statement &nbsp;|&nbsp;
                    GST Reports</div>
            </div></div>
        ''', unsafe_allow_html=True)
    
    # ── FY / Month / Quarter filter bar ─────────────────────────────────────
    _inv_hdr_c2, _inv_hdr_c3, _inv_hdr_c4 = st.columns([2, 2, 2])

    # Build FY list from actual data in invoice_details (month column)
    _fy_today = datetime.today()
    _fy_s     = _fy_today.year if _fy_today.month >= 4 else _fy_today.year - 1
    _cur_fy   = f"FY {_fy_s}-{str(_fy_s+1)[2:]}"

    def _month_to_fy_label(month_str):
        """Convert 'May-2025' → 'FY 2025-26', 'Apr-2026' → 'FY 2026-27'."""
        try:
            _dt = pd.to_datetime(month_str.strip(), format="%b-%Y", errors="coerce")
            if pd.isna(_dt):
                return None
            _yr = _dt.year if _dt.month >= 4 else _dt.year - 1
            return f"FY {_yr}-{str(_yr+1)[2:]}"
        except Exception:
            return None

    try:
        _fy_conn = get_db()
        _fy_months_raw = pd.read_sql(
            "SELECT DISTINCT month FROM invoice_details WHERE is_deleted=0 AND (month IS NOT NULL AND month != '')",
            _fy_conn
        )["month"].dropna().tolist()
        _fy_conn.close()
        # Each invoice may have comma-separated months (multi-month): split and process all
        _fy_set = set()
        for _mval in _fy_months_raw:
            for _m in str(_mval).split(","):
                _lbl = _month_to_fy_label(_m.strip())
                if _lbl:
                    _fy_set.add(_lbl)
    except Exception:
        _fy_set = set()

    # Always include current FY; sort descending
    _fy_set.add(_cur_fy)
    _fy_opts = sorted(_fy_set, reverse=True)
    # Ensure current FY is first
    if _cur_fy in _fy_opts:
        _fy_opts = [_cur_fy] + [f for f in _fy_opts if f != _cur_fy]

    with _inv_hdr_c2:
        _inv_filter_fy = st.selectbox("Financial Year", _fy_opts, index=0, key="inv_filter_fy")

    _quarter_opts = ["Q1 (Apr–Jun)", "Q2 (Jul–Sep)", "Q3 (Oct–Dec)", "Q4 (Jan–Mar)"]

    # Build FY-ordered month labels with year, e.g. "Apr-2026" … "Mar-2027"
    _fy_yr = int(_inv_filter_fy.split()[1].split("-")[0])   # e.g. 2026 from "FY 2026-27"
    _fy_month_sequence = [
        (4,  _fy_yr),  (5,  _fy_yr),  (6,  _fy_yr),
        (7,  _fy_yr),  (8,  _fy_yr),  (9,  _fy_yr),
        (10, _fy_yr),  (11, _fy_yr),  (12, _fy_yr),
        (1,  _fy_yr+1),(2,  _fy_yr+1),(3,  _fy_yr+1),
    ]
    _month_label_list = [
        datetime(yr, mn, 1).strftime("%b-%Y")
        for mn, yr in _fy_month_sequence
    ]

    # Derive which month / quarter is already chosen
    _inv_filter_month   = st.session_state.get("inv_filter_month_sel",   "All")
    _inv_filter_quarter = st.session_state.get("inv_filter_quarter_sel", "All")

    with _inv_hdr_c3:
        _month_disabled = (_inv_filter_quarter not in ("", "All"))
        _inv_filter_month = st.selectbox(
            "Month",
            ["All"] + _month_label_list,
            key="inv_filter_month_sel",
            disabled=_month_disabled,
        )

    with _inv_hdr_c4:
        _inv_filter_quarter = st.selectbox(
            "Quarter",
            ["All"] + _quarter_opts,
            key="inv_filter_quarter_sel",
            disabled=(_inv_filter_month not in ("", "All")),
        )

    _q_month_ranges = {
        "Q1 (Apr–Jun)": (4, 6),
        "Q2 (Jul–Sep)": (7, 9),
        "Q3 (Oct–Dec)": (10, 12),
        "Q4 (Jan–Mar)": (1, 3),
    }

    def _inv_filter_dates(fy_yr, sel_month, sel_quarter):
        """Return (date_from, date_to) for the chosen FY / Month / Quarter filter.
        Month labels are like 'Apr-2026' so we parse them directly — no ambiguity."""
        from datetime import date as _d
        import calendar
        if sel_month not in ("", "All"):
            # Parse "Apr-2026" → month=4, year=2026
            _pm = pd.to_datetime(sel_month, format="%b-%Y", errors="coerce")
            if pd.isna(_pm):
                _pm = pd.to_datetime(sel_month, errors="coerce")
            mn, yr = _pm.month, _pm.year
            last_day = calendar.monthrange(yr, mn)[1]
            return _d(yr, mn, 1), _d(yr, mn, last_day)
        elif sel_quarter not in ("", "All"):
            m1, m2 = _q_month_ranges[sel_quarter]
            # Q4 (Jan–Mar) belongs to calendar year fy_yr+1
            yr1 = fy_yr + 1 if m1 <= 3 else fy_yr
            yr2 = fy_yr + 1 if m2 <= 3 else fy_yr
            last_day2 = calendar.monthrange(yr2, m2)[1]
            return _d(yr1, m1, 1), _d(yr2, m2, last_day2)
        else:
            # Full FY: 1 Apr fy_yr → 31 Mar fy_yr+1
            return _d(fy_yr, 4, 1), _d(fy_yr + 1, 3, 31)

    _inv_date_from, _inv_date_to = _inv_filter_dates(
        _fy_yr, _inv_filter_month, _inv_filter_quarter
    )

    
    render_invoice_subnav()                          # ← ADD THIS
    tab_create, tab_history, tab_reminder, tab_statement, tab_gst = st.tabs([
        "📄 Create Invoice",
        "📋 Invoice History",
        "🔔 Send Reminder",
        "📊 DSP Statement",
        "📑 GST Report"
    ])

    # ════════════════════════════════════════
    # TAB 1 — CREATE INVOICE
    # ════════════════════════════════════════
    with tab_create:
        
        st.markdown('''
            <div style="background:linear-gradient(135deg,#003366 0%,#005599 100%);
            border-radius:10px;padding:18px 24px;margin-bottom:20px;
            box-shadow:0 4px 16px rgba(0,51,102,.2);
            display:flex;align-items:center;height:4px;gap:14px;">
            <div>
                <div style="color:white;font-size:20px;font-weight:800;">
                    New Invoice</div>
            </div></div>
        ''', unsafe_allow_html=True)

        col1, col2, col3 = st.columns([2, 2, 2])

        dsp_names = sorted(dsp_df["DSP Name"].dropna().unique().tolist())

        with col1:
            sel_dsp = st.selectbox(
                "DSP Name",
                options=["-- Select --"] + dsp_names,
                index=0,
                key="inv_dsp_name"
            )
 
        if sel_dsp == "-- Select --":
            st.info("👆 Please select a DSP to continue.")
        else:
            # ── Auto-sync invoice status from DSP payment data ──
            _sync_invoice_status(dsp_df)

            # Filter months for selected DSP — exclude Paid and Deleted
            _excl_conn = get_db()
            _excl_rows = _excl_conn.execute(
                "SELECT month FROM invoice_details WHERE dsp_name=? AND status IN ('Paid','Sent','Raised') AND is_deleted=0 AND is_credit_note=0",
                (sel_dsp,)
            ).fetchall()
            _excl_conn.close()
            # Split comma-joined multi-month strings into individual month tokens
            _excluded_months = set()
            for r in _excl_rows:
                for _m in str(r[0]).split(","):
                    _excluded_months.add(_m.strip())

            months_for_dsp = dsp_df[dsp_df["DSP Name"] == sel_dsp]["Month"].dropna().unique().tolist()
            months_for_dsp = sorted(months_for_dsp, key=lambda m: pd.to_datetime(m, format="%b-%Y", errors="coerce"))
            months_for_dsp = [m for m in months_for_dsp if str(m) not in _excluded_months]

            if not months_for_dsp:
                st.success("✅ All invoices for this DSP have been paid. No months available.")
                st.stop()

            with col2:
                _multi_month = st.checkbox("Multi-month invoice", value=False, key="inv_multi_month")

            if _multi_month:
                # Multi-month: multiselect for months
                sel_months_multi = st.multiselect(
                    "Select Months (multi)",
                    options=months_for_dsp,
                    default=[months_for_dsp[0]] if months_for_dsp else [],
                    key="inv_months_multi"
                )
                sel_month = sel_months_multi[0] if sel_months_multi else months_for_dsp[0]

                if sel_months_multi:
                    # Per-month table — FX rate shown for ALL invoice types (INR and USD)
                    # so per-month FX is always stored in month_breakdown for GST reporting
                    _type_map_early = {
                        "USD":                              "USD",
                        "INR + CGST/SGST (Maharashtra)":   "INR_CGST_SGST",
                        "INR + IGST (Outside Maharashtra)":"INR_IGST",
                    }
                    _inv_type_label_early = st.session_state.get("inv_type_sel", "USD")
                    _inv_type_key_early   = _type_map_early.get(_inv_type_label_early, "USD")
                    _inv_is_inr = _inv_type_key_early != "USD"

                    # Column header row
                    _hdr = st.columns([2, 2, 2, 2])
                    _hdr[0].caption("Month")
                    _hdr[1].caption("Amount (USD)")
                    _hdr[2].caption("FX Rate (INR/USD)")
                    _hdr[3].caption("INR Equivalent")

                    _multi_per_month = []
                    _total_multi = 0.0
                    _grand_inr   = 0.0

                    for _mm in sel_months_multi:
                        _rd  = dsp_df[(dsp_df["DSP Name"] == sel_dsp) & (dsp_df["Month"].astype(str) == str(_mm))]
                        _amt = float(_rd.iloc[0].get("Receivable $", 0) or 0) if not _rd.empty else 0.0
                        _total_multi += _amt

                        # Auto-fetch month-end FX rate for this month (always, for both INR and USD)
                        _fx_key = f"_mm_fx_{sel_dsp}_{_mm}"
                        if _fx_key not in st.session_state:
                            _fx_auto_mm = 87.50
                            try:
                                import requests as _req
                                _dt_mm = pd.to_datetime(str(_mm), format="%b-%Y")
                                _ld_mm = (_dt_mm + pd.offsets.MonthEnd(0)).strftime("%Y-%m-%d")
                                _r_mm  = _req.get(
                                    f"https://api.frankfurter.app/{_ld_mm}?from=USD&to=INR", timeout=6
                                )
                                if _r_mm.status_code == 200:
                                    _fx_auto_mm = round(_r_mm.json()["rates"]["INR"], 4)
                            except Exception:
                                pass
                            st.session_state[_fx_key] = _fx_auto_mm

                        _m_cols = st.columns([2, 2, 2, 2])
                        with _m_cols[0]:
                            st.markdown(f"**{_mm}**")
                        with _m_cols[1]:
                            st.markdown(f"${_amt:,.2f}")
                        with _m_cols[2]:
                            _fx_val = st.number_input(
                                f"FX ({_mm})",
                                min_value=1.0, step=0.01, format="%.4f",
                                key=_fx_key,
                                label_visibility="collapsed"
                            )
                        with _m_cols[3]:
                            _inr_val = round(_amt * _fx_val, 2)
                            st.markdown(f"₹{_inr_val:,.2f}")
                        _grand_inr += _inr_val
                        _multi_per_month.append({"month": _mm, "usd_amount": _amt, "fx_rate": _fx_val})

                    # Totals row
                    _t_cols = st.columns([2, 2, 2, 2])
                    with _t_cols[0]:
                        st.markdown("**TOTAL**")
                    with _t_cols[1]:
                        st.markdown(f"**${_total_multi:,.2f}**")
                    with _t_cols[2]:
                        st.caption("↑ per month")
                    with _t_cols[3]:
                        st.markdown(f"**₹{_grand_inr:,.2f}**")

                    auto_amount = _total_multi
                    _month_label = ", ".join(str(m) for m in sel_months_multi)
                    st.session_state["_inv_multi_per_month"] = _multi_per_month
                else:
                    auto_amount = 0.0
                    _month_label = ""
            else:
                with st.container():
                    _mc1, _mc2 = st.columns([3, 1])
                    with _mc1:
                        sel_month = st.selectbox("Month", months_for_dsp, key="inv_month")
                    sel_months_multi = [sel_month]
                    _month_label = str(sel_month)

                # Auto-fetch amount for the selected DSP + Month
                row_data = dsp_df[
                    (dsp_df["DSP Name"] == sel_dsp) &
                    (dsp_df["Month"].astype(str) == str(sel_month))
                ]
                auto_amount = 0.0
                if not row_data.empty:
                    r = row_data.iloc[0]
                    auto_amount = float(r.get("Receivable $", 0) or 0)

            auto_terms  = "Net 60"

            # Get partner info
            partner = get_partner_info(sel_dsp)
            if partner.get("payment_terms"):
                auto_terms = partner["payment_terms"]

            # ── Invalidate stale widget values when DSP or Month changes ──
            _combo_key = f"{sel_dsp}|{_month_label}"
            if st.session_state.get("_inv_last_combo") != _combo_key:
                st.session_state["_inv_last_combo"] = _combo_key
                # Clear all dependent widget keys so they re-render with fresh values
                for _k in ("inv_amount", "inv_terms", "inv_fx_rate", "inv_due_date",
                           "_inv_last_month", "_inv_prev_date_terms"):
                    if _k in st.session_state:
                        del st.session_state[_k]

            # Pre-seed amount and terms if not yet in session state
            if "inv_amount" not in st.session_state:
                st.session_state["inv_amount"] = float(auto_amount)
            if "inv_terms" not in st.session_state:
                st.session_state["inv_terms"] = str(auto_terms)

            # Detect invoice type
            detected_type = determine_invoice_type(partner)
            type_options  = ["USD", "INR + CGST/SGST (Maharashtra)", "INR + IGST (Outside Maharashtra)"]
            type_map      = {
                "USD":          "USD",
                "INR_CGST_SGST":"INR + CGST/SGST (Maharashtra)",
                "INR_IGST":     "INR + IGST (Outside Maharashtra)"
            }
            default_idx = list(type_map.values()).index(type_map.get(detected_type, "USD"))

            with col3:
                inv_type_label = st.selectbox(
                    "Invoice Type",
                    type_options,
                    index=default_idx,
                    key="inv_type_sel"
                )

            inv_type_key = {v: k for k, v in type_map.items()}[inv_type_label]

            # ── Form fields ──────────────────────────
            col4, col5, col6 = st.columns([2, 2, 2])
            with col4:
                inv_amount = st.number_input(
                    "Amount (USD)",
                    min_value=0.0, step=0.01, key="inv_amount"
                )

            # ── FX Rate — auto-fetch on month change (same pattern as Add Cost) ──
            fx_rate = 1.0

            month_str_for_fx = str(sel_month)

            # Detect month change → invalidate cached FX & Due Date
            if st.session_state.get("_inv_last_month") != month_str_for_fx:
                st.session_state["_inv_last_month"] = month_str_for_fx
                # Force FX widget to re-render with new value
                if "inv_fx_rate" in st.session_state:
                    del st.session_state["inv_fx_rate"]

            # Detect invoice-date or terms change → invalidate Due Date widget
            _date_key  = str(st.session_state.get("inv_date", date.today()))
            _terms_key = str(st.session_state.get("inv_terms", auto_terms))
            _prev_date_terms = st.session_state.get("_inv_prev_date_terms", "")
            _curr_date_terms = _date_key + "|" + _terms_key
            if _prev_date_terms != _curr_date_terms:
                st.session_state["_inv_prev_date_terms"] = _curr_date_terms
                if "inv_due_date" in st.session_state:
                    del st.session_state["inv_due_date"]

            if inv_type_key != "USD":
                # Fetch FX rate for this month
                fx_auto = 87.50
                try:
                    import requests
                    dt = pd.to_datetime(month_str_for_fx, format="%b-%Y")
                    last_day = (dt + pd.offsets.MonthEnd(0)).strftime("%Y-%m-%d")
                    resp = requests.get(
                        f"https://api.frankfurter.app/{last_day}?from=USD&to=INR",
                        timeout=8
                    )
                    if resp.status_code == 200:
                        fx_auto = round(resp.json()["rates"]["INR"], 4)
                except Exception:
                    pass

                # If widget key was just deleted (month changed), pre-seed the value
                if "inv_fx_rate" not in st.session_state:
                    st.session_state["inv_fx_rate"] = fx_auto

                with col5:
                    fx_rate = st.number_input(
                        "FX Rate (INR per USD)",
                        min_value=1.0, step=0.01, format="%.4f",
                        key="inv_fx_rate"
                    )

            with col6:
                inv_date = st.date_input("Invoice Date", value=date.today(), key="inv_date")

            col7, col8 = st.columns([2, 2])
            with col7:
                inv_terms = st.text_input("Payment Terms", key="inv_terms")

            # ── Auto-compute Due Date: Invoice Date + Term Days - 1 ──
            # Always recompute from current inv_date + inv_terms
            _actual_inv_date = st.session_state.get("inv_date", date.today())
            _actual_terms    = st.session_state.get("inv_terms", auto_terms)
            auto_due_computed = compute_due_date(_actual_inv_date, _actual_terms)

            # If widget key was just deleted, pre-seed with freshly computed value
            if "inv_due_date" not in st.session_state:
                st.session_state["inv_due_date"] = auto_due_computed

            with col8:
                inv_due = st.text_input(
                    "Due Date (auto-computed)",
                    key="inv_due_date",
                    help="Format: DD-MMM-YYYY  |  Logic: Invoice Date + Term Days - 1"
                )

            # Partner info preview
            with st.expander("📋 Partner Details (from Partner List)", expanded=False):
                if partner:
                    pc1, pc2 = st.columns(2)
                    with pc1:
                        st.write(f"**Legal Name:** {partner.get('legal_name','—')}")
                        st.write(f"**Address:** {partner.get('address','—')}")
                        st.write(f"**Country:** {partner.get('country','—')}")
                    with pc2:
                        st.write(f"**GSTIN:** {partner.get('gstin','—')}")
                        st.write(f"**Email:** {partner.get('email1','—')}")
                        st.write(f"**Finance Email:** {partner.get('finance_email','—')}")
                else:
                    st.warning(f"No partner record found for '{sel_dsp}' in Partner List.")

            # ── Preview & Generate ───────────────────
            inv_date_str = inv_date.strftime("%d/%m/%Y")

            st.markdown("---")
            gc1, gc2, gc3 = st.columns([2, 2, 2])

            with gc1:
                if st.button("👁️ Preview & Download PDF", key="inv_preview_btn", type="primary"):
                    conn = get_db()
                    inv_number = generate_invoice_number(conn)
                    conn.close()

                    try:
                        _is_multi = len(sel_months_multi) > 1

                        if inv_type_key == "USD":
                            if _is_multi:
                                # Build per-month data (USD has no FX rate)
                                _month_data_usd = []
                                for _mm in sel_months_multi:
                                    _rd = dsp_df[(dsp_df["DSP Name"] == sel_dsp) & (dsp_df["Month"].astype(str) == str(_mm))]
                                    _amt_m = float(_rd.iloc[0].get("Receivable $", 0) or 0) if not _rd.empty else 0.0
                                    _month_data_usd.append({"month": _mm, "amount": _amt_m})
                                # Use the user-entered total amount (may override per-month)
                                _total_from_data = sum(m["amount"] for m in _month_data_usd)
                                if _total_from_data == 0 or abs(inv_amount - _total_from_data) > 0.01:
                                    # proportionally scale if user changed total
                                    for _md in _month_data_usd:
                                        _md["amount"] = round(inv_amount / len(sel_months_multi), 2)
                                pdf_buf = generate_multi_month_usd_invoice(
                                    inv_number, inv_date_str, inv_due, inv_terms,
                                    partner, _month_data_usd, ""
                                )
                            else:
                                pdf_buf = generate_usd_invoice(
                                    inv_number, inv_date_str, inv_due, inv_terms,
                                    partner, str(sel_month), inv_amount, ""
                                )
                            fname    = f"{inv_number.replace('/','_')}.pdf"
                            currency = "USD"
                            tax_amt  = 0
                            total    = inv_amount
                            taxable  = inv_amount

                        elif inv_type_key == "INR_CGST_SGST":
                            if _is_multi:
                                # Use per-month FX rates stored in session state
                                _month_data_inr = st.session_state.get("_inv_multi_per_month", [])
                                if not _month_data_inr:
                                    # Fallback: use single fx_rate for all months
                                    for _mm in sel_months_multi:
                                        _rd = dsp_df[(dsp_df["DSP Name"] == sel_dsp) & (dsp_df["Month"].astype(str) == str(_mm))]
                                        _amt_m = float(_rd.iloc[0].get("Receivable $", 0) or 0) if not _rd.empty else 0.0
                                        _month_data_inr.append({"month": _mm, "usd_amount": _amt_m, "fx_rate": fx_rate})
                                pdf_buf = generate_multi_month_inr_cgst_sgst_invoice(
                                    inv_number, inv_date_str, inv_due, inv_terms,
                                    partner, _month_data_inr, ""
                                )
                                taxable  = round(sum(m["usd_amount"] * m["fx_rate"] for m in _month_data_inr), 2)
                                cgst_amt = round(taxable * 0.09, 2)
                                sgst_amt = round(taxable * 0.09, 2)
                            else:
                                taxable  = round(inv_amount * fx_rate, 2)
                                cgst_amt = round(taxable * 0.09, 2)
                                sgst_amt = round(taxable * 0.09, 2)
                                pdf_buf = generate_inr_cgst_sgst_invoice(
                                    inv_number, inv_date_str, inv_due, inv_terms,
                                    partner, str(sel_month), inv_amount, fx_rate, ""
                                )
                            tax_amt = cgst_amt + sgst_amt
                            total   = taxable + tax_amt
                            fname    = f"{inv_number.replace('/','_')}_INR.pdf"
                            currency = "INR"

                        else:
                            if _is_multi:
                                # Use per-month FX rates stored in session state
                                _month_data_inr = st.session_state.get("_inv_multi_per_month", [])
                                if not _month_data_inr:
                                    for _mm in sel_months_multi:
                                        _rd = dsp_df[(dsp_df["DSP Name"] == sel_dsp) & (dsp_df["Month"].astype(str) == str(_mm))]
                                        _amt_m = float(_rd.iloc[0].get("Receivable $", 0) or 0) if not _rd.empty else 0.0
                                        _month_data_inr.append({"month": _mm, "usd_amount": _amt_m, "fx_rate": fx_rate})
                                pdf_buf = generate_multi_month_inr_igst_invoice(
                                    inv_number, inv_date_str, inv_due, inv_terms,
                                    partner, _month_data_inr, ""
                                )
                                taxable  = round(sum(m["usd_amount"] * m["fx_rate"] for m in _month_data_inr), 2)
                                igst_amt = round(taxable * 0.18, 2)
                            else:
                                taxable  = round(inv_amount * fx_rate, 2)
                                igst_amt = round(taxable * 0.18, 2)
                                pdf_buf = generate_inr_igst_invoice(
                                    inv_number, inv_date_str, inv_due, inv_terms,
                                    partner, str(sel_month), inv_amount, fx_rate, ""
                                )
                            tax_amt = igst_amt
                            total   = taxable + tax_amt
                            fname    = f"{inv_number.replace('/','_')}_INR_IGST.pdf"
                            currency = "INR"

                        st.session_state["_inv_buf"]        = pdf_buf
                        st.session_state["_inv_fname"]      = fname
                        st.session_state["_inv_number"]     = inv_number
                        st.session_state["_inv_type"]       = inv_type_key
                        st.session_state["_inv_amount"]     = inv_amount
                        st.session_state["_inv_tax"]        = tax_amt
                        st.session_state["_inv_total"]      = total
                        st.session_state["_inv_currency"]   = currency
                        st.session_state["_inv_dsp"]        = sel_dsp
                        st.session_state["_inv_month"]       = _month_label
                        st.session_state["_inv_months_list"] = sel_months_multi
                        st.session_state["_inv_due"]         = inv_due
                        st.session_state["_inv_date"]       = inv_date_str
                        # Store INR breakdown for email
                        if inv_type_key == "INR_CGST_SGST":
                            st.session_state["_inv_taxable"]  = taxable
                            st.session_state["_inv_cgst"]     = cgst_amt
                            st.session_state["_inv_sgst"]     = sgst_amt
                            st.session_state["_inv_igst"]     = 0.0
                        elif inv_type_key == "INR_IGST":
                            st.session_state["_inv_taxable"]  = taxable
                            st.session_state["_inv_igst"]     = igst_amt
                            st.session_state["_inv_cgst"]     = 0.0
                            st.session_state["_inv_sgst"]     = 0.0
                        else:
                            st.session_state["_inv_taxable"]  = 0.0
                            st.session_state["_inv_igst"]     = 0.0
                            st.session_state["_inv_cgst"]     = 0.0
                            st.session_state["_inv_sgst"]     = 0.0
                        # Store fx_rate — for USD single-month, auto-fetch month-end rate
                        if inv_type_key == "USD" and not _is_multi:
                            _usd_fx_auto = 0.0
                            try:
                                import requests as _rfx
                                _dt_fx = pd.to_datetime(str(sel_month), format="%b-%Y")
                                _ld_fx = (_dt_fx + pd.offsets.MonthEnd(0)).strftime("%Y-%m-%d")
                                _rr = _rfx.get(
                                    f"https://api.frankfurter.app/{_ld_fx}?from=USD&to=INR", timeout=6
                                )
                                if _rr.status_code == 200:
                                    _usd_fx_auto = round(_rr.json()["rates"]["INR"], 4)
                            except Exception:
                                pass
                            st.session_state["_inv_fx_rate"] = _usd_fx_auto
                        else:
                            st.session_state["_inv_fx_rate"] = fx_rate
                        st.success(f"✅ Invoice {inv_number} generated!")

                    except Exception as e:
                        st.error(f"PDF generation failed: {e}")

            # Download button (appears after generation)
            if st.session_state.get("_inv_buf"):
                buf_copy = io.BytesIO(st.session_state["_inv_buf"].getvalue())
                with gc2:
                    st.download_button(
                        "📥 Download PDF",
                        data=buf_copy,
                        file_name=st.session_state["_inv_fname"],
                        mime="application/pdf",
                        key="inv_download_btn"
                    )

                with gc3:
                    if st.button("💾 Save to Invoice History", key="inv_save_btn"):
                        import json as _json_sv
                        # Multi-month: save ONE record with totals (months joined in month field)
                        _save_months  = st.session_state.get("_inv_months_list", [st.session_state.get("_inv_month","")])
                        _base_inv_no  = st.session_state["_inv_number"]
                        _n_months     = len(_save_months)
                        _multi_pm     = st.session_state.get("_inv_multi_per_month", [])
                        # Month label: "Oct-2025, Jan-2026" for multi, single string for single
                        _month_str    = st.session_state.get("_inv_month", str(_save_months[0]) if _save_months else "")
                        # For FX rate: use average of per-month rates (or single rate)
                        if _multi_pm:
                            _avg_fx = round(sum(m.get("fx_rate", 0) for m in _multi_pm) / len(_multi_pm), 4)
                        else:
                            _avg_fx = st.session_state.get("_inv_fx_rate", 0.0)
                        # Build month_breakdown JSON: [{month, usd_amount, fx_rate, inr_amount}]
                        if _multi_pm and _n_months > 1:
                            _breakdown = []
                            for _bm in _multi_pm:
                                _bfx   = float(_bm.get("fx_rate", _avg_fx) or _avg_fx)
                                _busd  = float(_bm.get("usd_amount", 0) or 0)
                                _binr  = round(_busd * _bfx, 2)
                                _breakdown.append({
                                    "month":      str(_bm.get("month","")),
                                    "usd_amount": _busd,
                                    "fx_rate":    _bfx,
                                    "inr_amount": _binr,
                                })
                            _breakdown_json = _json_sv.dumps(_breakdown)
                        else:
                            _breakdown_json = ""
                        # Save ONE row with full totals
                        save_invoice(
                            _base_inv_no,
                            st.session_state["_inv_date"],
                            st.session_state["_inv_dsp"],
                            _month_str,
                            st.session_state["_inv_type"],
                            st.session_state["_inv_amount"],
                            st.session_state["_inv_tax"],
                            st.session_state["_inv_total"],
                            st.session_state["_inv_currency"],
                            st.session_state["_inv_due"],
                            status="Raised",
                            igst_amount=st.session_state.get("_inv_igst",    0.0),
                            cgst_amount=st.session_state.get("_inv_cgst",    0.0),
                            sgst_amount=st.session_state.get("_inv_sgst",    0.0),
                            taxable_amount=st.session_state.get("_inv_taxable", 0.0),
                            fx_rate=_avg_fx,
                            month_breakdown=_breakdown_json,
                        )
                        st.success("✅ Saved to Invoice History!")

            # ── Send by Email ─────────────────────────
            if st.session_state.get("_inv_buf"):
                st.markdown("#### 📧 Send Invoice via Email")
                partner_emails = [
                    e for e in [
                        partner.get("email1",""),
                        partner.get("email2",""),
                        partner.get("finance_email","")
                    ] if e
                ]
                email_default = ", ".join(partner_emails)

                em1, em2 = st.columns([3, 1])
                with em1:
                    to_email_str = st.text_input(
                        "To (comma separated)", value=email_default, key="inv_to_email"
                    )
                with em2:
                    st.markdown(f"**BCC:** {BCC_EMAIL}", unsafe_allow_html=False)

                # ── Build subject & body ──────────────────
                _inv_no      = st.session_state.get("_inv_number", "")
                _inv_amt     = st.session_state.get("_inv_amount", 0.0)
                _inv_total   = st.session_state.get("_inv_total", 0.0)
                _inv_due_d   = st.session_state.get("_inv_due", "")
                _inv_dsp_n   = st.session_state.get("_inv_dsp", "")
                _inv_cur     = st.session_state.get("_inv_currency", "USD")
                _inv_type_k  = st.session_state.get("_inv_type", "USD")
                _legal       = partner.get("legal_name", _inv_dsp_n)

                # GST breakdown values
                _inv_taxable = st.session_state.get("_inv_taxable", 0.0)
                _inv_cgst    = st.session_state.get("_inv_cgst", 0.0)
                _inv_sgst    = st.session_state.get("_inv_sgst", 0.0)
                _inv_igst    = st.session_state.get("_inv_igst", 0.0)

                # Format due date as DD/MM/YYYY for subject line
                try:
                    _due_for_subj = pd.to_datetime(_inv_due_d, dayfirst=True).strftime("%d/%m/%Y")
                except Exception:
                    _due_for_subj = _inv_due_d

                # For INR invoices, show INR total; for USD show USD
                if _inv_type_k in ("INR_CGST_SGST", "INR_IGST"):
                    _amount_label = f"₹{_inv_total:,.2f}"
                    _due_amount_display = f"₹{_inv_total:,.2f}"
                else:
                    _amount_label = f"${_inv_amt:,.2f}"
                    _due_amount_display = f"${_inv_amt:,.2f}"

                # ── Format month label for subject (single vs multi-month) ──
                _inv_month_raw = st.session_state.get("_inv_month", "")
                _inv_months_list = st.session_state.get("_inv_months_list", [])
                # Build a nice "Aug-2025, Sep-2025 and Oct-2025" style string
                def _fmt_month_label(months_raw, months_list):
                    """Convert month data to human-readable label for email."""
                    if months_list and len(months_list) > 1:
                        _ml = [str(m) for m in months_list if str(m).strip()]
                        if len(_ml) == 1:
                            return _ml[0]
                        return ", ".join(_ml[:-1]) + " and " + _ml[-1]
                    # Single month from comma-split of raw string
                    _parts = [m.strip() for m in str(months_raw).split(",") if m.strip()]
                    if len(_parts) > 1:
                        return ", ".join(_parts[:-1]) + " and " + _parts[-1]
                    return str(months_raw)

                _month_label_for_email = _fmt_month_label(_inv_month_raw, _inv_months_list)

                email_subject = (
                    f"Invoice for {_month_label_for_email} from {COMPANY['name']} "
                    f"(Invoice No.: {_inv_no} - {_amount_label}) "
                    f"| Due On {_due_for_subj}"
                )

                # Build GST breakdown rows for INR invoices
                if _inv_type_k == "INR_CGST_SGST":
                    _gst_rows = f"""
      <tr><td><b>Taxable Amount</b></td><td>&nbsp;-&nbsp;</td><td>₹{_inv_taxable:,.2f}</td></tr>
      <tr><td><b>CGST (9%)</b></td><td>&nbsp;-&nbsp;</td><td>₹{_inv_cgst:,.2f}</td></tr>
      <tr><td><b>SGST (9%)</b></td><td>&nbsp;-&nbsp;</td><td>₹{_inv_sgst:,.2f}</td></tr>"""
                elif _inv_type_k == "INR_IGST":
                    _gst_rows = f"""
      <tr><td><b>Taxable Amount</b></td><td>&nbsp;-&nbsp;</td><td>₹{_inv_taxable:,.2f}</td></tr>
      <tr><td><b>IGST (18%)</b></td><td>&nbsp;-&nbsp;</td><td>₹{_inv_igst:,.2f}</td></tr>"""
                else:
                    _gst_rows = ""

                email_body = f"""
    <p>Dear {_legal},</p>
    <p>It was a pleasure doing business with you.</p>
    <p>Please find the attached invoice, kindly pay the same on or before due date.</p>
    <br/>
    <table>
      <tr><td><b>Customer Name</b></td><td>&nbsp;-&nbsp;</td><td>{_legal}</td></tr>
      <tr><td><b>Invoice No.</b></td><td>&nbsp;-&nbsp;</td><td>{_inv_no}</td></tr>
      <tr><td><b>Invoice Month</b></td><td>&nbsp;-&nbsp;</td><td>{_month_label_for_email}</td></tr>{_gst_rows}
      <tr><td><b>Due Amount</b></td><td>&nbsp;-&nbsp;</td><td>{_due_amount_display}</td></tr>
      <tr><td><b>Due Date</b></td><td>&nbsp;-&nbsp;</td><td>{_due_for_subj}</td></tr>
    </table>
    <br/>
    <p>You can email us at <a href="mailto:{COMPANY['email']}">{COMPANY['email']}</a> for any clarifications.</p>
    <br/>
    <p>Regards,<br/><b>Finance Team - {COMPANY['name']}</b></p>
    """

                # ── Email Preview ────────────────────────
                with st.expander("👁️ Email Preview (Subject & Body)", expanded=True):
                    st.markdown(f"**Subject:** {email_subject}")
                    st.markdown("---")
                    st.markdown(email_body, unsafe_allow_html=True)

                # ── SMTP Settings ─────────────────────────
                if _smtp_configured_via_secrets():
                    st.caption("✅ **Email credentials loaded from secrets** — no password needed.")
                    _inv_smtp_host, _inv_smtp_port, _inv_smtp_user, _inv_smtp_pass = _get_smtp_creds()
                else:
                    with st.expander("⚙️ Email / SMTP Settings", expanded=False):
                        _sc1, _sc2 = st.columns(2)
                        with _sc1:
                            st.text_input("SMTP Host",   value="smtpout.secureserver.net", key="smtp_host")
                            st.text_input("From Email",  value="finance@peakmyads.com",    key="smtp_user")
                        with _sc2:
                            st.number_input("SMTP Port", value=465, key="smtp_port")
                            st.text_input("Password / App Password", type="password",      key="smtp_pass")
                        st.caption("SMTP: smtpout.secureserver.net : 465 (SSL/TLS)")
                    _inv_smtp_host, _inv_smtp_port, _inv_smtp_user, _inv_smtp_pass = _get_smtp_creds()

                send_btn = st.button("📤 Send Email", key="inv_send_btn", type="primary")

                if send_btn:
                    to_list = [e.strip() for e in to_email_str.split(",") if e.strip()]
                    if not to_list:
                        st.error("Please enter at least one email address.")
                    elif not _inv_smtp_pass:
                        st.error("SMTP password not set. Configure secrets.toml or enter password in SMTP Settings.")
                    else:
                        # Store everything in session state and open preview dialog
                        st.session_state["_inv_preview_open"]    = True
                        st.session_state["_inv_preview_to"]      = to_list
                        st.session_state["_inv_preview_subject"] = email_subject
                        st.session_state["_inv_preview_body"]    = email_body
                        st.session_state["_inv_preview_host"]    = _inv_smtp_host
                        st.session_state["_inv_preview_port"]    = _inv_smtp_port
                        st.session_state["_inv_preview_user"]    = _inv_smtp_user
                        st.session_state["_inv_preview_pass"]    = _inv_smtp_pass
                        st.rerun()

                @st.dialog("📧 Final Email Preview — Please Review Before Sending", width="large")
                def _inv_email_preview_dialog():
                    _to_list = st.session_state.get("_inv_preview_to", [])
                    _subject = st.session_state.get("_inv_preview_subject", "")
                    _body    = st.session_state.get("_inv_preview_body", "")
                    _host    = st.session_state.get("_inv_preview_host", "")
                    _port    = st.session_state.get("_inv_preview_port", 465)
                    _user    = st.session_state.get("_inv_preview_user", "")
                    _pass    = st.session_state.get("_inv_preview_pass", "")
                    _fname   = st.session_state.get("_inv_fname", "invoice.pdf")

                    st.markdown(f"**📬 To:** {', '.join(_to_list)}")
                    st.markdown(f"**📌 Subject:** {_subject}")
                    st.markdown(f"**📎 Attachment:** `{_fname}`")
                    st.divider()
                    st.markdown("**📝 Email Body:**")
                    st.markdown(_body, unsafe_allow_html=True)
                    st.divider()

                    _dc1, _dc2 = st.columns(2)
                    with _dc1:
                        if st.button("✅ Confirm & Send", type="primary",
                                     use_container_width=True, key="inv_confirm_send"):
                            try:
                                buf_send = io.BytesIO(st.session_state["_inv_buf"].getvalue())
                                send_invoice_email(
                                    _to_list, _subject, _body,
                                    buf_send, _fname,
                                    _host, int(_port), _user, _pass
                                )
                                import json as _json_cs
                                _save_months  = st.session_state.get("_inv_months_list", [st.session_state.get("_inv_month", "")])
                                _base_inv_no  = st.session_state["_inv_number"]
                                _n_months     = len(_save_months)
                                _multi_pm     = st.session_state.get("_inv_multi_per_month", [])
                                _month_str    = st.session_state.get("_inv_month", str(_save_months[0]) if _save_months else "")
                                if _multi_pm:
                                    _avg_fx = round(sum(m.get("fx_rate", 0) for m in _multi_pm) / len(_multi_pm), 4)
                                else:
                                    _avg_fx = st.session_state.get("_inv_fx_rate", 0.0)
                                if _multi_pm and _n_months > 1:
                                    _breakdown = []
                                    for _bm in _multi_pm:
                                        _bfx  = float(_bm.get("fx_rate", _avg_fx) or _avg_fx)
                                        _busd = float(_bm.get("usd_amount", 0) or 0)
                                        _binr = round(_busd * _bfx, 2)
                                        _breakdown.append({
                                            "month":      str(_bm.get("month", "")),
                                            "usd_amount": _busd,
                                            "fx_rate":    _bfx,
                                            "inr_amount": _binr,
                                        })
                                    _breakdown_json = _json_cs.dumps(_breakdown)
                                else:
                                    _breakdown_json = ""
                                save_invoice(
                                    _base_inv_no,
                                    st.session_state["_inv_date"],
                                    st.session_state["_inv_dsp"],
                                    _month_str,
                                    st.session_state["_inv_type"],
                                    st.session_state["_inv_amount"],
                                    st.session_state["_inv_tax"],
                                    st.session_state["_inv_total"],
                                    st.session_state["_inv_currency"],
                                    st.session_state["_inv_due"],
                                    status="Sent",
                                    igst_amount=st.session_state.get("_inv_igst", 0.0),
                                    cgst_amount=st.session_state.get("_inv_cgst", 0.0),
                                    sgst_amount=st.session_state.get("_inv_sgst", 0.0),
                                    taxable_amount=st.session_state.get("_inv_taxable", 0.0),
                                    fx_rate=_avg_fx,
                                    month_breakdown=_breakdown_json,
                                )
                                # Close preview, open success popup
                                st.session_state["_inv_preview_open"] = False
                                st.session_state["_inv_success_open"] = True
                                st.session_state["_inv_success_msg"]  = (
                                    f"✅ **Invoice sent to {', '.join(_to_list)}!**  \n"
                                    f"BCC: {BCC_EMAIL}"
                                )
                                st.rerun()
                            except Exception as e:
                                st.error(f"Email failed: {e}")
                    with _dc2:
                        if st.button("✖ Cancel", use_container_width=True, key="inv_cancel_send"):
                            st.session_state["_inv_preview_open"] = False
                            st.rerun()

                if st.session_state.get("_inv_preview_open"):
                    _inv_email_preview_dialog()

                @st.dialog("📨 Email Sent", width="small")
                def _inv_success_dialog():
                    st.markdown(st.session_state.get("_inv_success_msg", ""))
                    st.markdown("")
                    if st.button("Close", use_container_width=True, key="inv_success_close", type="primary"):
                        st.session_state["_inv_success_open"] = False
                        # Reset DSP dropdown explicitly to blank selection
                        st.session_state["inv_dsp_name"] = "-- Select --"
                        # Clear all generated invoice state
                        for _k in ["inv_month", "inv_fx_rate",
                                   "_inv_buf", "_inv_fname", "_inv_number",
                                   "_inv_date", "_inv_dsp", "_inv_month",
                                   "_inv_type", "_inv_amount", "_inv_tax",
                                   "_inv_total", "_inv_currency", "_inv_due",
                                   "_inv_igst", "_inv_cgst", "_inv_sgst",
                                   "_inv_taxable", "_inv_fx_rate",
                                   "_inv_months_list", "_inv_preview_open",
                                   "_inv_preview_to", "_inv_preview_subject",
                                   "_inv_preview_body", "_inv_success_msg"]:
                            st.session_state.pop(_k, None)
                        st.rerun()

                if st.session_state.get("_inv_success_open"):
                    _inv_success_dialog()

    # ════════════════════════════════════════
    # TAB 2 — INVOICE HISTORY
    # ════════════════════════════════════════
    with tab_history:
        
        st.markdown('''
            <div style="background:linear-gradient(135deg,#003366 0%,#005599 100%);
            border-radius:10px;padding:18px 24px;margin-bottom:20px;
            box-shadow:0 4px 16px rgba(0,51,102,.2);
            display:flex;align-items:center;height:4px;gap:14px;">
            <div>
                <div style="color:white;font-size:20px;font-weight:800;">
                    Invoice History</div>
            </div></div>
        ''', unsafe_allow_html=True)
        
        # ── Helper: build reminder email content from a row ──────────
        def _build_reminder_content(row):
            dsp      = row.get("dsp_name", "")
            inv_no   = row.get("invoice_number", "")
            month    = row.get("month", "")
            due_raw  = row.get("due_date", "")
            cur      = str(row.get("currency", "USD")).upper()
            total    = float(row.get("total_amount", 0) or 0)
            amt      = float(row.get("amount", 0) or 0)
            igst     = float(row.get("igst_amount", 0) or 0)
            cgst     = float(row.get("cgst_amount", 0) or 0)
            sgst     = float(row.get("sgst_amount", 0) or 0)
            taxable  = float(row.get("taxable_amount", 0) or 0)

            due_amt_str = f"₹{total:,.2f}" if cur == "INR" else f"${amt:,.2f}"
            try:
                due_fmt = pd.to_datetime(due_raw, dayfirst=True).strftime("%d/%m/%Y")
            except Exception:
                due_fmt = due_raw

            partner  = get_partner_info(dsp)
            legal    = partner.get("legal_name", dsp)
            email    = ", ".join(filter(None, [partner.get("email1",""), partner.get("finance_email","")]))

            if igst > 0:
                _tax_row = f"\n  <tr><td><b>Taxable Amount</b></td><td>&nbsp;-&nbsp;</td><td>₹{taxable:,.2f}</td></tr>" if taxable > 0 else ""
                _gst_rows = _tax_row + f"\n  <tr><td><b>IGST (18%)</b></td><td>&nbsp;-&nbsp;</td><td>₹{igst:,.2f}</td></tr>"
            elif cgst > 0 or sgst > 0:
                _tax_row = f"\n  <tr><td><b>Taxable Amount</b></td><td>&nbsp;-&nbsp;</td><td>₹{taxable:,.2f}</td></tr>" if taxable > 0 else ""
                _gst_rows = (
                    _tax_row +
                    f"\n  <tr><td><b>CGST (9%)</b></td><td>&nbsp;-&nbsp;</td><td>₹{cgst:,.2f}</td></tr>"
                    f"\n  <tr><td><b>SGST (9%)</b></td><td>&nbsp;-&nbsp;</td><td>₹{sgst:,.2f}</td></tr>"
                )
            else:
                _gst_rows = ""

            subject = (
                f"Payment Reminder — {COMPANY['name']} "
                f"(Invoice No.: {inv_no} - {due_amt_str}) | Due On {due_fmt}"
            )
            body = f"""
<p>Dear {legal},</p>
<p>This is a gentle reminder that the following invoice is outstanding and payment is due.</p>
<br/>
<table>
  <tr><td><b>Customer Name</b></td><td>&nbsp;-&nbsp;</td><td>{legal}</td></tr>
  <tr><td><b>Invoice No.</b></td><td>&nbsp;-&nbsp;</td><td>{inv_no}</td></tr>
  <tr><td><b>Month</b></td><td>&nbsp;-&nbsp;</td><td>{month}</td></tr>{_gst_rows}
  <tr><td><b>Due Amount</b></td><td>&nbsp;-&nbsp;</td><td>{due_amt_str}</td></tr>
  <tr><td><b>Due Date</b></td><td>&nbsp;-&nbsp;</td><td>{due_fmt}</td></tr>
</table>
<br/>
<p>Please arrange payment at your earliest convenience.</p>
<p>You can email us at <a href="mailto:{COMPANY['email']}">{COMPANY['email']}</a> for any clarifications.</p>
<br/>
<p>Regards,<br/><b>Finance Team - {COMPANY['name']}</b></p>
"""
            return subject, body, email, legal, inv_no

        conn = get_db()
        show_deleted = st.checkbox("Show Deleted Invoices", value=False, key="hist_show_deleted")
        show_cn      = st.checkbox("Show Credit Notes", value=True, key="hist_show_cn")
        hist_df = pd.read_sql("SELECT * FROM invoice_details ORDER BY created_at DESC", conn)
        # Auto-sync invoice status from DSP payment data (refreshes on every History tab open)
        _sync_invoice_status(dsp_df)
        # Reload hist_df after sync so grid reflects updated statuses
        hist_df = pd.read_sql("SELECT * FROM invoice_details ORDER BY created_at DESC", conn)
        if not show_deleted:
            hist_df = hist_df[hist_df.get("is_deleted", pd.Series(0, index=hist_df.index)).fillna(0).astype(int) == 0]
        if not show_cn:
            hist_df = hist_df[hist_df.get("is_credit_note", pd.Series(0, index=hist_df.index)).fillna(0).astype(int) == 0]
        conn.close()

        # ── Apply FY / Month / Quarter filter ───────────────────────────
        if not hist_df.empty:
            _hdt = pd.to_datetime(hist_df["invoice_date"], dayfirst=True, errors="coerce")
            _mask = (_hdt >= pd.Timestamp(_inv_date_from)) & (_hdt <= pd.Timestamp(_inv_date_to))
            hist_df = hist_df[_mask]
            if hist_df.empty:
                st.info(f"No invoices found for {_inv_filter_fy}"
                        + (f" — {_inv_filter_month}" if _inv_filter_month != 'All' else "")
                        + (f" — {_inv_filter_quarter}" if _inv_filter_quarter != 'All' else ""))

        # ── Active period label ──────────────────────────────────────────
        _period_label = _inv_filter_fy
        if _inv_filter_month not in ("", "All"):
            _period_label += f" | {_inv_filter_month}"
        elif _inv_filter_quarter not in ("", "All"):
            _period_label += f" | {_inv_filter_quarter}"
        st.caption(f"Showing: **{_period_label}** &nbsp;({_inv_date_from.strftime('%d %b %Y')} – {_inv_date_to.strftime('%d %b %Y')})")
        if hist_df.empty:
            st.info("No invoices saved yet.")
        else:
            from st_aggrid import AgGrid, GridOptionsBuilder, JsCode

            # ── Ensure GST columns exist (may be 0 for old records) ──────
            for _gc in ("igst_amount", "cgst_amount", "sgst_amount", "taxable_amount"):
                if _gc not in hist_df.columns:
                    hist_df[_gc] = 0.0

            # ── Column formatting ──────────────────────────────────────
            currency_fmt = JsCode("""
            function(params) {
                if (params.value == null || params.value === '') return '';
                var cur = (params.data && params.data.currency)
                          ? params.data.currency.toString().trim().toUpperCase()
                          : 'USD';
                var symbol = (cur === 'INR') ? '\u20b9' : '$';
                return symbol + parseFloat(params.value).toLocaleString(undefined, {minimumFractionDigits: 2});
            }
            """)

            inr_fmt = JsCode("""
            function(params) {
                if (params.value == null || params.value === '' || parseFloat(params.value) === 0) return '-';
                return '\u20b9' + parseFloat(params.value).toLocaleString(undefined, {minimumFractionDigits: 2});
            }
            """)

            status_style = JsCode("""
            function(params) {
                if (!params.value) return {};
                const v = params.value.toLowerCase();
                if (v === 'sent')            return { color: '#1565C0', fontWeight: 'bold' };
                if (v === 'raised')          return { color: '#0277BD', fontWeight: 'bold' };
                if (v === 'draft')           return { color: '#E65100', fontWeight: 'bold' };
                if (v === 'paid')            return { color: '#2E7D32', fontWeight: 'bold' };
                if (v === 'partially paid')  return { color: '#7B3F00', fontWeight: 'bold' };
                if (v === 'overdue')         return { color: '#B71C1C', fontWeight: 'bold' };
                if (v === 'deleted')         return { color: '#9E9E9E', fontWeight: 'bold', textDecoration: 'line-through' };
                if (v === 'credit note')     return { color: '#6A1B9A', fontWeight: 'bold' };
                return {};
            }
            """)

            gb_hist = GridOptionsBuilder.from_dataframe(hist_df)

            # Hide internal/technical columns
            for _hc in ["id", "sent_at", "fx_rate", "is_deleted", "is_credit_note",
                         "credit_note_ref", "credit_note_reason"]:
                if _hc in hist_df.columns:
                    gb_hist.configure_column(_hc, hide=True)

            usd_fmt = JsCode("""
            function(params) {
                if (params.value == null || params.value === '') return '';
                return '$' + parseFloat(params.value).toLocaleString(undefined, {minimumFractionDigits: 2});
            }
            """)

            if "amount" in hist_df.columns:
                gb_hist.configure_column("amount", type=["numericColumn"], valueFormatter=usd_fmt, headerName="Amount (USD)")
            for cur_col in ["tax_amount", "total_amount"]:
                if cur_col in hist_df.columns:
                    gb_hist.configure_column(cur_col, type=["numericColumn"], valueFormatter=currency_fmt)

            gb_hist.configure_column("taxable_amount", headerName="Taxable (INR)", type=["numericColumn"], valueFormatter=inr_fmt)
            gb_hist.configure_column("igst_amount",    headerName="IGST (INR)",    type=["numericColumn"], valueFormatter=inr_fmt)
            gb_hist.configure_column("cgst_amount",    headerName="CGST (INR)",    type=["numericColumn"], valueFormatter=inr_fmt)
            gb_hist.configure_column("sgst_amount",    headerName="SGST (INR)",    type=["numericColumn"], valueFormatter=inr_fmt)

            if "status" in hist_df.columns:
                gb_hist.configure_column("status", cellStyle=status_style)

            pinned_style = JsCode("""
            function(params) {
                if (params.column.pinned) {
                    return { fontWeight: 'bold', borderRight: '2px solid #003366' };
                }
            }
            """)
            for pin_col in ["invoice_number", "dsp_name"]:
                if pin_col in hist_df.columns:
                    gb_hist.configure_column(pin_col, pinned="left", cellStyle=pinned_style)

            gb_hist.configure_default_column(resizable=True, sortable=True, filter=False)
            gb_hist.configure_selection(selection_mode="single", use_checkbox=False)

            total_row = {col: "" for col in hist_df.columns}
            total_row["invoice_number"] = "TOTAL"
            for tot_col in ["amount", "tax_amount", "total_amount", "taxable_amount", "igst_amount", "cgst_amount", "sgst_amount"]:
                if tot_col in hist_df.columns:
                    total_row[tot_col] = pd.to_numeric(hist_df[tot_col], errors="coerce").sum()

            hist_grid_opts = gb_hist.build()
            hist_grid_opts["pinnedBottomRowData"] = [total_row]
            hist_grid_opts["getRowStyle"] = JsCode("""
            function(params) {
                if (params.node.rowPinned) {
                    return { backgroundColor: '#003366', color: 'white', fontWeight: 'bold', fontSize: '14px' };
                }
            }
            """)

            hist_custom_css = {
                ".ag-header": {
                    "background-color": "#003366 !important",
                    "color": "white !important",
                    "font-weight": "bold !important",
                    "font-size": "14px !important"
                },
                ".ag-header-cell-label": {
                    "color": "white !important",
                    "font-weight": "bold !important"
                },
                ".ag-row-selected": {
                    "background-color": "#d0e4f7 !important",
                    "border-left": "4px solid #003366 !important"
                }
            }

            st.caption("👆 Click any row to select it, then use **🔔 Reminder** or **✏️ Edit** buttons below.")

            grid_response = AgGrid(
                hist_df,
                gridOptions=hist_grid_opts,
                allow_unsafe_jscode=True,
                fit_columns_on_grid_load=True,
                height=400,
                custom_css=hist_custom_css,
                update_mode="SELECTION_CHANGED",
            )

            # ── Detect selected row ───────────────────────────────────
            sel_rows = grid_response.get("selected_rows", None)
            # Handle both list and DataFrame
            if sel_rows is None:
                sel_row_data = None
            elif isinstance(sel_rows, pd.DataFrame):
                sel_row_data = sel_rows.iloc[0].to_dict() if not sel_rows.empty else None
            elif isinstance(sel_rows, list) and len(sel_rows) > 0:
                sel_row_data = sel_rows[0]
            else:
                sel_row_data = None

            # ── Action buttons below grid ─────────────────────────────
            if sel_row_data:
                _sel_inv_no  = sel_row_data.get("invoice_number", "")
                _sel_dsp     = sel_row_data.get("dsp_name", "")
                _sel_cur     = str(sel_row_data.get("currency", "USD")).upper()
                _sel_total   = float(sel_row_data.get("total_amount", 0) or 0)
                _sel_amt     = float(sel_row_data.get("amount", 0) or 0)
                _amt_disp    = f"₹{_sel_total:,.2f}" if _sel_cur == "INR" else f"${_sel_amt:,.2f}"

                st.markdown(
                    f"""<div style='background:#e8f4fd;border:1.5px solid #003366;border-radius:8px;
                    padding:10px 18px;margin:10px 0 4px 0;display:flex;align-items:center;gap:12px'>
                    <span style='font-size:15px;font-weight:700;color:#003366'>
                    Selected: {_sel_inv_no} &nbsp;|&nbsp; {_sel_dsp} &nbsp;|&nbsp; {_amt_disp}
                    </span></div>""",
                    unsafe_allow_html=True
                )
                
                act_c1, act_c2, act_c3, act_c4, act_c5, act_c6 = st.columns([1,1,1,1,1,1])
                with act_c1:
                    open_reminder = st.button("🔔 Reminder", key="hist_open_rem_btn", type="primary", use_container_width=True)
                with act_c2:
                    open_edit = st.button("✏️ Edit", key="hist_open_edit_btn", use_container_width=True)
                with act_c3:
                    _is_already_deleted = int(sel_row_data.get("is_deleted", 0) or 0) == 1
                    open_delete = st.button(
                        "♻️ Restore" if _is_already_deleted else "🗑️ Delete",
                        key="hist_open_del_btn", use_container_width=True
                    )
                with act_c4:
                    _is_cn_row = int(sel_row_data.get("is_credit_note", 0) or 0) == 1
                    open_cn = st.button(
                        "📝 Credit Note", key="hist_open_cn_btn",
                        use_container_width=True,
                        disabled=_is_cn_row or _is_already_deleted,
                        help="Cannot issue CN against a CN or deleted invoice" if (_is_cn_row or _is_already_deleted) else ""
                    )
                with act_c5:
                    open_download = st.button("⬇️ Download", key="hist_open_dl_btn", use_container_width=True)
                with act_c6:
                    open_resend = st.button("📤 Resend Email", key="hist_open_resend_btn", use_container_width=True)

                # Track which panel is open via session state
                if open_reminder:
                    st.session_state["_hist_panel"] = "reminder"
                    st.session_state["_hist_sel_inv"] = _sel_inv_no
                if open_edit:
                    st.session_state["_hist_panel"] = "edit"
                    st.session_state["_hist_sel_inv"] = _sel_inv_no
                if open_cn:
                    st.session_state["_hist_panel"] = "credit_note"
                    st.session_state["_hist_sel_inv"] = _sel_inv_no
                if open_download:
                    st.session_state["_hist_panel"] = "download"
                    st.session_state["_hist_sel_inv"] = _sel_inv_no
                if open_resend:
                    st.session_state["_hist_panel"] = "resend"
                    st.session_state["_hist_sel_inv"] = _sel_inv_no
                if open_delete:
                    # Toggle deleted flag immediately
                    _new_del_flag = 0 if _is_already_deleted else 1
                    _new_status   = sel_row_data.get("status", "Sent") if _is_already_deleted else "Deleted"
                    try:
                        _dc = get_db()
                        _dc.execute(
                            "UPDATE invoice_details SET is_deleted=?, status=? WHERE invoice_number=?",
                            (_new_del_flag, _new_status, _sel_inv_no)
                        )
                        _dc.commit()
                        _dc.close()
                        if _new_del_flag:
                            st.warning(f"🗑️ Invoice {_sel_inv_no} marked as Deleted (record preserved).")
                        else:
                            st.success(f"♻️ Invoice {_sel_inv_no} restored.")
                        st.rerun()
                    except Exception as _de:
                        st.error(f"Failed: {_de}")

                _active_panel = st.session_state.get("_hist_panel", None)
                _active_inv   = st.session_state.get("_hist_sel_inv", None)

                # Only show panel if it matches the currently selected invoice
                if _active_panel and _active_inv == _sel_inv_no:

                    # ══════════════════════════════════════════════
                    # REMINDER PANEL  (modal-style container)
                    # ══════════════════════════════════════════════
                    if _active_panel == "reminder":
                        st.markdown("""
                        <div style='border:2px solid #1565C0;border-radius:10px;
                        background:#f0f6ff;padding:4px 0 0 0;margin-top:8px'>
                        <div style='background:#1565C0;color:white;padding:8px 18px;
                        border-radius:8px 8px 0 0;font-size:16px;font-weight:700'>
                        🔔 Send Payment Reminder</div></div>""", unsafe_allow_html=True)

                        with st.container():
                            rem_subj, rem_body, rem_email_def, rem_legal, rem_inv_n = _build_reminder_content(sel_row_data)

                            st.divider()
                            
                            rp1, rp2 = st.columns([3, 1])
                            with rp1:
                                rem_to_inp = st.text_input("Reminder To (comma separated)", value=rem_email_def, key="panel_rem_to")
                            with rp2:
                                st.markdown(f"**BCC:** {BCC_EMAIL}")
                                
                            st.divider()

                            with st.expander("👁️ Email Preview (Subject & Body)", expanded=True):
                                st.markdown(f"**Subject:** {rem_subj}")
                                st.markdown("---")
                                st.markdown(rem_body, unsafe_allow_html=True)

                            if _smtp_configured_via_secrets():
                                st.caption("✅ **Email credentials loaded from secrets** — no password needed.")
                                _pr_host, _pr_port, _pr_user, _pr_pass = _get_smtp_creds()
                            else:
                                with st.expander("⚙️ SMTP Settings", expanded=False):
                                    rsc1, rsc2 = st.columns(2)
                                    with rsc1:
                                        st.text_input("SMTP Host",  value="smtpout.secureserver.net", key="panel_rem_host")
                                        st.text_input("From Email", value="finance@peakmyads.com",    key="panel_rem_user")
                                    with rsc2:
                                        st.number_input("SMTP Port", value=465, key="panel_rem_port")
                                        st.text_input("App Password", type="password", key="panel_rem_pass")
                                    st.caption("SMTP: smtpout.secureserver.net : 465 (SSL/TLS)")
                                _pr_host = st.session_state.get("panel_rem_host", "smtpout.secureserver.net")
                                _pr_port = int(st.session_state.get("panel_rem_port", 465))
                                _pr_user = st.session_state.get("panel_rem_user", "finance@peakmyads.com")
                                _pr_pass = st.session_state.get("panel_rem_pass", "")

                            st.divider()

                            rb1, rb2 = st.columns([1, 5])
                            with rb1:
                                if st.button("📤 Send Reminder", key="panel_rem_send", type="primary"):
                                    to_list_p = [e.strip() for e in rem_to_inp.split(",") if e.strip()]
                                    if not to_list_p:
                                        st.error("Please enter at least one recipient.")
                                    elif not _pr_pass:
                                        st.error("SMTP password not set. Configure secrets.toml or enter in SMTP Settings.")
                                    else:
                                        try:
                                            msg = MIMEMultipart()
                                            msg["From"]    = _pr_user
                                            msg["To"]      = ", ".join(to_list_p)
                                            msg["Bcc"]     = BCC_EMAIL
                                            msg["Subject"] = rem_subj
                                            msg.attach(MIMEText(rem_body, "html"))
                                            _smtp_send(_pr_host, int(_pr_port), _pr_user, _pr_pass,
                                                       to_list_p + [BCC_EMAIL], msg.as_string())
                                            st.success(f"✅ Reminder sent to {', '.join(to_list_p)}!")
                                            st.session_state["_hist_panel"] = None
                                        except Exception as e:
                                            st.error(f"Failed: {e}")
                            with rb2:
                                if st.button("✖ Close", key="panel_rem_close"):
                                    st.session_state["_hist_panel"] = None
                                    st.rerun()

                    # ══════════════════════════════════════════════
                    # EDIT PANEL  (modal-style container)
                    # ══════════════════════════════════════════════
                    elif _active_panel == "edit":
                        st.markdown("""
                        <div style='border:2px solid #E65100;border-radius:10px;
                        background:#fff8f4;padding:4px 0 0 0;margin-top:8px'>
                        <div style='background:#E65100;color:white;padding:8px 18px;
                        border-radius:8px 8px 0 0;font-size:16px;font-weight:700'>
                        ✏️ Edit Invoice</div></div>""", unsafe_allow_html=True)
                        
                        st.divider()

                        with st.container():
                            _e = sel_row_data   # shorthand

                            # Parse existing values
                            _e_inv_no   = _e.get("invoice_number", "")
                            _e_dsp      = _e.get("dsp_name", "")
                            _e_month    = _e.get("month", "")
                            _e_inv_type = _e.get("invoice_type", "USD")
                            _e_amount   = float(_e.get("amount", 0) or 0)
                            _e_tax      = float(_e.get("tax_amount", 0) or 0)
                            _e_total    = float(_e.get("total_amount", 0) or 0)
                            _e_cur      = _e.get("currency", "USD")
                            _e_due      = _e.get("due_date", "")
                            _e_status   = _e.get("status", "Draft")
                            _e_taxable  = float(_e.get("taxable_amount", 0) or 0)
                            _e_igst     = float(_e.get("igst_amount", 0) or 0)
                            _e_cgst     = float(_e.get("cgst_amount", 0) or 0)
                            _e_sgst     = float(_e.get("sgst_amount", 0) or 0)

                            try:
                                _e_inv_date = datetime.strptime(_e.get("invoice_date",""), "%d/%m/%Y").date()
                            except Exception:
                                _e_inv_date = date.today()

                            st.markdown(f"**Invoice:** `{_e_inv_no}` &nbsp;|&nbsp; **DSP:** {_e_dsp} &nbsp;|&nbsp; **Month:** {_e_month}")
                            st.markdown("---")

                            ec1, ec2, ec3 = st.columns(3)
                            with ec1:
                                e_inv_date = st.date_input("Invoice Date", value=_e_inv_date, key="edit_inv_date")
                            with ec2:
                                e_due = st.text_input("Due Date (DD-MMM-YYYY)", value=_e_due, key="edit_due_date")
                            with ec3:
                                _status_opts = ["Draft", "Raised", "Sent", "Partially Paid", "Paid", "Overdue", "Deleted"]
                                e_status = st.selectbox(
                                    "Status",
                                    _status_opts,
                                    index=_status_opts.index(_e_status) if _e_status in _status_opts else 0,
                                    key="edit_status"
                                )

                            ec4, ec5 = st.columns(2)
                            with ec4:
                                e_amount = st.number_input("Amount (USD)", value=_e_amount, min_value=0.0, step=0.01, key="edit_amount")
                            with ec5:
                                e_inv_type_label = st.selectbox(
                                    "Invoice Type",
                                    ["USD", "INR + CGST/SGST (Maharashtra)", "INR + IGST (Outside Maharashtra)"],
                                    index={"USD":0,"INR_CGST_SGST":1,"INR_IGST":2}.get(_e_inv_type, 0),
                                    key="edit_inv_type"
                                )

                            e_inv_type_key = {"USD":"USD","INR + CGST/SGST (Maharashtra)":"INR_CGST_SGST","INR + IGST (Outside Maharashtra)":"INR_IGST"}[e_inv_type_label]

                            # INR fields — editable for INR invoices
                            if e_inv_type_key != "USD":
                                st.markdown("**INR Amounts** *(edit if needed)*")
                                ei1, ei2, ei3, ei4 = st.columns(4)
                                with ei1:
                                    e_taxable = st.number_input("Taxable (₹)", value=_e_taxable, min_value=0.0, step=0.01, key="edit_taxable")
                                if e_inv_type_key == "INR_CGST_SGST":
                                    with ei2:
                                        e_cgst = st.number_input("CGST (₹)", value=_e_cgst, min_value=0.0, step=0.01, key="edit_cgst")
                                    with ei3:
                                        e_sgst = st.number_input("SGST (₹)", value=_e_sgst, min_value=0.0, step=0.01, key="edit_sgst")
                                    with ei4:
                                        e_igst = 0.0
                                        e_tax_amt = e_cgst + e_sgst
                                        e_total   = e_taxable + e_tax_amt
                                        st.markdown("**Total (₹)**")
                                        st.markdown(f"<span style='font-size:16px;font-weight:700;color:#003366'>₹{e_total:,.2f}</span>", unsafe_allow_html=True)
                                else:
                                    with ei2:
                                        e_igst = st.number_input("IGST (₹)", value=_e_igst, min_value=0.0, step=0.01, key="edit_igst")
                                    with ei3:
                                        e_cgst = 0.0
                                        e_sgst = 0.0
                                    with ei4:
                                        e_tax_amt = e_igst
                                        e_total   = e_taxable + e_tax_amt
                                        st.markdown("**Total (₹)**")
                                        st.markdown(f"<span style='font-size:16px;font-weight:700;color:#003366'>₹{e_total:,.2f}</span>", unsafe_allow_html=True)
                                e_currency = "INR"
                            else:
                                e_taxable = 0.0; e_igst = 0.0; e_cgst = 0.0; e_sgst = 0.0
                                e_tax_amt = 0.0; e_total = e_amount; e_currency = "USD"

                            st.markdown("---")
                            eb1, eb2 = st.columns([1, 5])
                            with eb1:
                                if st.button("💾 Save Changes", key="edit_save_btn", type="primary"):
                                    try:
                                        conn = get_db()
                                        conn.execute("""
                                            UPDATE invoice_details SET
                                                invoice_date   = ?,
                                                invoice_type   = ?,
                                                amount         = ?,
                                                tax_amount     = ?,
                                                total_amount   = ?,
                                                currency       = ?,
                                                due_date       = ?,
                                                status         = ?,
                                                taxable_amount = ?,
                                                igst_amount    = ?,
                                                cgst_amount    = ?,
                                                sgst_amount    = ?
                                            WHERE invoice_number = ?
                                        """, (
                                            e_inv_date.strftime("%d/%m/%Y"),
                                            e_inv_type_key,
                                            e_amount,
                                            e_tax_amt,
                                            e_total,
                                            e_currency,
                                            e_due,
                                            e_status,
                                            e_taxable,
                                            e_igst,
                                            e_cgst,
                                            e_sgst,
                                            _e_inv_no
                                        ))
                                        conn.commit()
                                        conn.close()
                                        st.success(f"✅ Invoice {_e_inv_no} updated successfully!")
                                        st.session_state["_hist_panel"] = None
                                        st.rerun()
                                    except Exception as e:
                                        st.error(f"Save failed: {e}")
                            with eb2:
                                if st.button("✖ Close", key="edit_close_btn"):
                                    st.session_state["_hist_panel"] = None
                                    st.rerun()

                    # ══════════════════════════════════════════════
                    # CREDIT NOTE PANEL
                    # ══════════════════════════════════════════════
                    elif _active_panel == "credit_note":
                        # ── Hard guard: CN cannot be issued against another CN or deleted invoice ──
                        _guard_is_cn  = int(sel_row_data.get("is_credit_note", 0) or 0) == 1
                        _guard_is_del = int(sel_row_data.get("is_deleted",     0) or 0) == 1
                        if _guard_is_cn:
                            st.error("⛔ Credit notes cannot be issued against another credit note.")
                            if st.button("✖ Close", key="cn_guard_cn_close"):
                                st.session_state["_hist_panel"] = None
                                st.rerun()
                        elif _guard_is_del:
                            st.error("⛔ Credit notes cannot be issued against a deleted invoice.")
                            if st.button("✖ Close", key="cn_guard_del_close"):
                                st.session_state["_hist_panel"] = None
                                st.rerun()
                        else:
                          st.markdown("""
                          <div style='border:2px solid #6A1B9A;border-radius:10px;
                          background:#fdf4ff;padding:4px 0 0 0;margin-top:8px'>
                          <div style='background:#6A1B9A;color:white;padding:8px 18px;
                          border-radius:8px 8px 0 0;font-size:16px;font-weight:700'>
                          📝 Issue Credit Note</div></div>""", unsafe_allow_html=True)

                          with st.container():
                            st.divider()
                            _orig = sel_row_data
                            _orig_inv_no  = _orig.get("invoice_number", "")
                            _orig_dsp     = _orig.get("dsp_name", "")
                            _orig_month   = _orig.get("month", "")
                            _orig_type    = _orig.get("invoice_type", "USD")
                            _orig_cur     = str(_orig.get("currency", "USD")).upper()
                            _orig_amt     = float(_orig.get("amount", 0) or 0)
                            _orig_total   = float(_orig.get("total_amount", 0) or 0)
                            _orig_taxable = float(_orig.get("taxable_amount", 0) or 0)
                            _orig_igst    = float(_orig.get("igst_amount", 0) or 0)
                            _orig_cgst    = float(_orig.get("cgst_amount", 0) or 0)
                            _orig_sgst    = float(_orig.get("sgst_amount", 0) or 0)

                            # ── If taxable_amount not stored (old records), back-calc from total ──
                            if _orig_taxable == 0 and _orig_cur == "INR":
                                if _orig_type == "INR_CGST_SGST":
                                    # total = taxable * 1.18 → taxable = total / 1.18
                                    _orig_taxable = round(_orig_total / 1.18, 2)
                                    _orig_cgst    = round(_orig_taxable * 0.09, 2)
                                    _orig_sgst    = round(_orig_taxable * 0.09, 2)
                                    _orig_igst    = 0.0
                                elif _orig_type == "INR_IGST":
                                    _orig_taxable = round(_orig_total / 1.18, 2)
                                    _orig_igst    = round(_orig_taxable * 0.18, 2)
                                    _orig_cgst    = 0.0
                                    _orig_sgst    = 0.0

                            st.markdown(f"**Issuing credit note against:** `{_orig_inv_no}` | {_orig_dsp} | {_orig_month}")
                            st.markdown("---")

                            # ── Calculate already-issued CN total against this invoice ──
                            try:
                                _prev_cn_conn = get_db()
                                _prev_cn_total = _prev_cn_conn.execute(
                                    "SELECT COALESCE(SUM(total_amount),0) FROM invoice_details "
                                    "WHERE credit_note_ref=? AND is_credit_note=1 AND is_deleted=0",
                                    (_orig_inv_no,)
                                ).fetchone()[0]
                                _prev_cn_conn.close()
                            except Exception:
                                _prev_cn_total = 0.0

                            if _orig_cur == "INR":
                                _orig_full = _orig_total
                                _balance   = max(0.0, round(_orig_full - float(_prev_cn_total), 2))
                                _bal_label = f"₹{_balance:,.2f}"
                            else:
                                _orig_full = _orig_amt
                                _balance   = max(0.0, round(_orig_full - float(_prev_cn_total), 2))
                                _bal_label = f"${_balance:,.2f}"

                            if _prev_cn_total > 0:
                                st.info(
                                    f"ℹ️ Previously credited: "
                                    f"{'₹' if _orig_cur=='INR' else '$'}{float(_prev_cn_total):,.2f} &nbsp;|&nbsp; "
                                    f"**Remaining balance available for credit: {_bal_label}**"
                                )
                            if _balance <= 0:
                                st.error("⚠️ This invoice has been fully credited. No further credit note can be issued.")
                                if st.button("✖ Close", key="cn_bal_close"):
                                    st.session_state["_hist_panel"] = None
                                    st.rerun()
                                st.stop()

                            cn_c1, cn_c2 = st.columns(2)
                            with cn_c1:
                                cn_date = st.date_input("Credit Note Date", value=date.today(), key="cn_date")
                            with cn_c2:
                                cn_reason = st.text_input("Reason for Credit Note", value="Billing Adjustment", key="cn_reason")

                            # Amount to credit — capped at remaining balance
                            if _orig_cur == "INR":
                                cn_full_amt = _balance   # remaining, not full original
                                cn_full_label = f"₹{cn_full_amt:,.2f}"
                            else:
                                cn_full_amt = _balance
                                cn_full_label = f"${cn_full_amt:,.2f}"

                            # If INR invoice but taxable_amount is 0 (old record),
                            # back-calculate it from total and tax rates
                            if _orig_cur == "INR" and _orig_taxable == 0.0 and cn_full_amt > 0:
                                if _orig_type == "INR_CGST_SGST":
                                    # total = taxable * 1.18  →  taxable = total / 1.18
                                    _orig_taxable = round(cn_full_amt / 1.18, 2)
                                    _orig_cgst    = round(_orig_taxable * 0.09, 2)
                                    _orig_sgst    = round(_orig_taxable * 0.09, 2)
                                    _orig_igst    = 0.0
                                elif _orig_type == "INR_IGST":
                                    _orig_taxable = round(cn_full_amt / 1.18, 2)
                                    _orig_igst    = round(_orig_taxable * 0.18, 2)
                                    _orig_cgst    = 0.0
                                    _orig_sgst    = 0.0

                            cn_c3, cn_c4 = st.columns(2)
                            with cn_c3:
                                cn_partial = st.checkbox("Partial Credit Note", value=False, key="cn_partial")
                            with cn_c4:
                                if cn_partial:
                                    if _orig_cur == "INR":
                                        cn_amt = st.number_input("Credit Amount (₹)", value=cn_full_amt, min_value=0.01, max_value=cn_full_amt, step=0.01, key="cn_amt")
                                    else:
                                        cn_amt = st.number_input("Credit Amount ($)", value=cn_full_amt, min_value=0.01, max_value=cn_full_amt, step=0.01, key="cn_amt")
                                else:
                                    cn_amt = cn_full_amt
                                    st.markdown(f"**Full Credit:** {cn_full_label}")

                            # ─────────────────────────────────────────────────────────────
                            # FX RATE SECTION  (for BOTH USD and INR credit notes)
                            # ─────────────────────────────────────────────────────────────
                            import json as _cn_json, math as _cn_math

                            def _fetch_cn_fx(month_str, fallback=87.50):
                                """Return month-end INR/USD rate from Frankfurter API."""
                                try:
                                    import requests as _rq
                                    _dt = pd.to_datetime(month_str.strip(), format="%b-%Y", errors="coerce")
                                    if pd.isna(_dt):
                                        _dt = pd.to_datetime(month_str.strip(), errors="coerce")
                                    if pd.isna(_dt):
                                        return fallback
                                    _ld = (_dt + pd.offsets.MonthEnd(0)).strftime("%Y-%m-%d")
                                    _rsp = _rq.get(
                                        f"https://api.frankfurter.app/{_ld}?from=USD&to=INR", timeout=6
                                    )
                                    if _rsp.status_code == 200:
                                        _d = _rsp.json()
                                        if "rates" in _d and "INR" in _d["rates"]:
                                            return round(_d["rates"]["INR"], 4)
                                except Exception:
                                    pass
                                return fallback

                            def _parse_breakdown(val):
                                """Safely parse month_breakdown — handles NaN, None, empty."""
                                if val is None:
                                    return []
                                if isinstance(val, float) and _cn_math.isnan(val):
                                    return []
                                s = str(val).strip()
                                if not s or not s.startswith("["):
                                    return []
                                try:
                                    return _cn_json.loads(s)
                                except Exception:
                                    return []

                            def _bd_fx_for_month(breakdown, month_str):
                                """Find fx_rate in breakdown for a month string — exact then fuzzy match."""
                                try:
                                    _tgt = pd.to_datetime(month_str.strip(), format="%b-%Y", errors="coerce")
                                except Exception:
                                    _tgt = None
                                for _b in breakdown:
                                    _bfx = float(_b.get("fx_rate", 0) or 0)
                                    if _bfx <= 1:
                                        continue
                                    _bm = str(_b.get("month","")).strip()
                                    if _bm == month_str.strip():          # exact
                                        return _bfx
                                    try:                                   # fuzzy month+year
                                        _bdt = pd.to_datetime(_bm, format="%b-%Y", errors="coerce")
                                        if (pd.notna(_bdt) and _tgt is not None and pd.notna(_tgt)
                                                and _bdt.month == _tgt.month and _bdt.year == _tgt.year):
                                            return _bfx
                                    except Exception:
                                        pass
                                return 0.0

                            # ── Stale session-state guard: reset per-month keys when invoice changes ──
                            _cn_tracker_key = "_cn_panel_inv_tracker"
                            if st.session_state.get(_cn_tracker_key) != _orig_inv_no:
                                _stale = [k for k in list(st.session_state.keys())
                                          if k.startswith(("_cn_mm_fx_", "_cn_mm_usd_", "_cn_usd_amount_"))]
                                for _sk in _stale:
                                    del st.session_state[_sk]
                                st.session_state[_cn_tracker_key] = _orig_inv_no

                            # Parse original invoice breakdown (NaN-safe)
                            _orig_breakdown  = _parse_breakdown(_orig.get("month_breakdown"))
                            _orig_fx_stored  = float(_orig.get("fx_rate", 0) or 0)
                            _orig_total_usd  = float(_orig.get("amount",   0) or 0)
                            _bd_usd_lookup   = {
                                str(b.get("month","")).strip(): float(b.get("usd_amount", 0) or 0)
                                for b in _orig_breakdown
                            }

                            # Split orig_month → individual month strings
                            _cn_raw_months = [m.strip() for m in str(_orig_month).split(",") if m.strip()]
                            _cn_is_multi   = len(_cn_raw_months) > 1
                            _cn_per_month_data = []

                            # ── Multi-month CN: per-month FX table ──────────────────────────
                            if _cn_is_multi:
                                st.markdown("---")
                                st.markdown("**💱 Per-Month Amount & FX Rate**")
                                st.caption(
                                    "FX rates pre-filled from the original invoice's stored month-end rates. "
                                    "Edit if needed before generating."
                                )

                                _cn_mm_hdr = st.columns([2, 2, 2, 2])
                                _cn_mm_hdr[0].caption("Month")
                                _cn_mm_hdr[1].caption("Amount (USD)")
                                _cn_mm_hdr[2].caption("FX Rate (INR/USD)")

                                # Decide column-3 header based on currency
                                _is_inr_cn = (_orig_cur == "INR")
                                _has_gst   = _orig_type in ("INR_CGST_SGST", "INR_IGST")

                                _cn_mm_hdr[3].caption("INR Total (incl. GST)" if (_is_inr_cn and _has_gst) else "INR Taxable" if _is_inr_cn else "INR Equiv.")

                                _cn_total_usd = 0.0
                                _cn_total_inr_grand = 0.0   # sum of per-month INR totals (incl. GST)
                                for _cn_m in _cn_raw_months:
                                    _cn_mm_usd_key = f"_cn_mm_usd_{_orig_inv_no}_{_cn_m}"
                                    _cn_mm_fx_key  = f"_cn_mm_fx_{_orig_inv_no}_{_cn_m}"

                                    # Default USD: from original breakdown, else split evenly
                                    _cn_m_usd_default = _bd_usd_lookup.get(
                                        _cn_m,
                                        round(_orig_total_usd / len(_cn_raw_months), 4)
                                    )

                                    # Default FX: original breakdown → stored avg FX → API → 87.5
                                    if _cn_mm_fx_key not in st.session_state:
                                        _bd_fx = _bd_fx_for_month(_orig_breakdown, _cn_m)
                                        if _bd_fx > 1:
                                            st.session_state[_cn_mm_fx_key] = _bd_fx
                                        elif _orig_fx_stored > 1:
                                            st.session_state[_cn_mm_fx_key] = _orig_fx_stored
                                        else:
                                            st.session_state[_cn_mm_fx_key] = _fetch_cn_fx(_cn_m, 87.50)

                                    _cn_mm_cols = st.columns([2, 2, 2, 2])
                                    with _cn_mm_cols[0]:
                                        st.markdown(f"**{_cn_m}**")
                                    with _cn_mm_cols[1]:
                                        _cn_m_usd = st.number_input(
                                            f"USD ({_cn_m})", min_value=0.0,
                                            value=float(_cn_m_usd_default),
                                            step=0.01, format="%.4f",
                                            key=_cn_mm_usd_key,
                                            label_visibility="collapsed"
                                        )
                                    with _cn_mm_cols[2]:
                                        _cn_m_fx = st.number_input(
                                            f"FX ({_cn_m})", min_value=1.0,
                                            step=0.01, format="%.4f",
                                            key=_cn_mm_fx_key,
                                            label_visibility="collapsed"
                                        )

                                    # For INR invoices: taxable = USD × FX; total incl. GST
                                    _cn_m_taxable_inr = round(_cn_m_usd * _cn_m_fx, 2)
                                    if _is_inr_cn and _has_gst:
                                        _cn_m_total_inr = round(_cn_m_taxable_inr * 1.18, 2)
                                    elif _is_inr_cn:
                                        _cn_m_total_inr = _cn_m_taxable_inr
                                    else:
                                        _cn_m_total_inr = _cn_m_taxable_inr  # USD CN: just equiv

                                    with _cn_mm_cols[3]:
                                        if _is_inr_cn:
                                            if _has_gst:
                                                st.markdown(
                                                    f"₹{_cn_m_total_inr:,.2f}"
                                                    f"<br><small style='color:gray'>taxable ₹{_cn_m_taxable_inr:,.2f}</small>",
                                                    unsafe_allow_html=True
                                                )
                                            else:
                                                st.markdown(f"₹{_cn_m_taxable_inr:,.2f}")
                                        else:
                                            st.markdown(f"₹{_cn_m_taxable_inr:,.2f}")

                                    _cn_total_usd         += _cn_m_usd
                                    _cn_total_inr_grand   += _cn_m_total_inr
                                    _cn_per_month_data.append({
                                        "month":      _cn_m,
                                        "usd_amount": round(_cn_m_usd,          4),
                                        "fx_rate":    round(_cn_m_fx,            4),
                                        "inr_amount": round(_cn_m_taxable_inr,   2),  # taxable INR (no GST)
                                    })

                                # ── Summary ──
                                _cn_total_taxable_inr = round(sum(d["usd_amount"] * d["fx_rate"] for d in _cn_per_month_data), 2)
                                if _is_inr_cn and _has_gst:
                                    st.markdown(
                                        f"**Total USD:** ${_cn_total_usd:,.4f} &nbsp;|&nbsp; "
                                        f"**Total Taxable (₹):** ₹{_cn_total_taxable_inr:,.2f} &nbsp;|&nbsp; "
                                        f"**Total Inv. Value (₹):** ₹{_cn_total_inr_grand:,.2f}"
                                    )
                                else:
                                    st.markdown(
                                        f"**Total USD:** ${_cn_total_usd:,.4f} &nbsp;|&nbsp; "
                                        f"**INR Equiv:** ₹{_cn_total_taxable_inr:,.2f}"
                                    )

                                # Average FX for flat db field
                                _cn_fx_rate = round(
                                    sum(d["fx_rate"] for d in _cn_per_month_data) / len(_cn_per_month_data), 4
                                ) if _cn_per_month_data else 87.50

                                # ─── KEY FIX: cn_amt must be in INR for INR invoices ───
                                if _is_inr_cn:
                                    cn_amt = _cn_total_inr_grand   # INR total incl. GST
                                else:
                                    cn_amt = _cn_total_usd         # USD total (no GST)

                            else:
                                # ── Single-month CN: one FX rate ────────────────────────────
                                st.markdown("---")
                                _cn_single_month = _cn_raw_months[0] if _cn_raw_months else str(_orig_month)
                                _cn_fx_key = f"_cn_fx_{_orig_inv_no}_{_cn_single_month}"

                                if _cn_fx_key not in st.session_state:
                                    _bd_fx = _bd_fx_for_month(_orig_breakdown, _cn_single_month)
                                    if _bd_fx > 1:
                                        st.session_state[_cn_fx_key] = _bd_fx
                                    elif _orig_fx_stored > 1:
                                        st.session_state[_cn_fx_key] = _orig_fx_stored
                                    else:
                                        st.session_state[_cn_fx_key] = _fetch_cn_fx(_cn_single_month, 87.50)

                                if _orig_cur == "INR":
                                    st.markdown("**💱 USD Amount & FX Rate** *(Credit Note is raised in USD then converted to INR)*")
                                    _orig_taxable_stored = float(_orig.get("taxable_amount", 0) or 0)
                                    if _orig_fx_stored > 1 and _orig_taxable_stored > 0:
                                        _cn_usd_default = round(_orig_taxable_stored / _orig_fx_stored, 4)
                                    else:
                                        _cur_fx_init = st.session_state.get(_cn_fx_key, 87.5)
                                        _cn_usd_default = round(cn_full_amt / 1.18 / max(_cur_fx_init, 1), 4)

                                    _cn_usd_c1, _cn_usd_c2, _cn_usd_c3 = st.columns(3)
                                    with _cn_usd_c1:
                                        cn_usd_amount = st.number_input(
                                            "Credit Amount (USD)", min_value=0.0001,
                                            value=max(0.0001, _cn_usd_default),
                                            step=0.01, format="%.4f",
                                            key=f"_cn_usd_amount_{_orig_inv_no}"
                                        )
                                    with _cn_usd_c2:
                                        _cn_fx_rate = st.number_input(
                                            "FX Rate (INR/USD)", min_value=1.0, step=0.01, format="%.4f",
                                            key=_cn_fx_key,
                                            help="Month-end rate from original invoice; edit if needed"
                                        )
                                    cn_taxable_inr = round(cn_usd_amount * _cn_fx_rate, 2)
                                    with _cn_usd_c3:
                                        st.markdown(f"**Taxable (INR):** ₹{cn_taxable_inr:,.2f}")
                                    cn_amt = round(cn_taxable_inr * 1.18, 2) if _orig_type in ("INR_CGST_SGST", "INR_IGST") else cn_taxable_inr
                                    _cn_per_month_data = [{"month": _cn_single_month, "usd_amount": cn_usd_amount, "fx_rate": _cn_fx_rate, "inr_amount": cn_taxable_inr}]
                                else:
                                    # USD single-month: show FX for GST report
                                    st.markdown("**💱 FX Rate for GST Report** *(month-end INR/USD rate)*")
                                    _cn_usd_c1, _cn_usd_c2 = st.columns(2)
                                    with _cn_usd_c1:
                                        st.markdown(f"**Credit Amount (USD):** ${cn_amt:,.4f}")
                                    with _cn_usd_c2:
                                        _cn_fx_rate = st.number_input(
                                            "FX Rate (INR/USD)", min_value=1.0, step=0.01, format="%.4f",
                                            key=_cn_fx_key,
                                            help="Month-end rate from original invoice; stored in GST Report"
                                        )
                                    _cn_inr_eq = round(cn_amt * _cn_fx_rate, 2)
                                    st.caption(f"INR equiv ≈ ₹{_cn_inr_eq:,.2f}")
                                    _cn_per_month_data = [{"month": _cn_single_month, "usd_amount": cn_amt, "fx_rate": _cn_fx_rate, "inr_amount": _cn_inr_eq}]

                            # ── Compute taxable / GST amounts from final cn_amt ──────────
                            if _orig_cur == "INR":
                                if _orig_type == "INR_CGST_SGST":
                                    cn_taxable = round(cn_amt / 1.18, 2)
                                    cn_cgst    = round(cn_taxable * 0.09, 2)
                                    cn_sgst    = round(cn_taxable * 0.09, 2)
                                    cn_igst    = 0.0
                                    cn_tax_amt = round(cn_cgst + cn_sgst, 2)
                                elif _orig_type == "INR_IGST":
                                    cn_taxable = round(cn_amt / 1.18, 2)
                                    cn_igst    = round(cn_taxable * 0.18, 2)
                                    cn_cgst    = 0.0
                                    cn_sgst    = 0.0
                                    cn_tax_amt = cn_igst
                                else:
                                    cn_taxable = cn_amt
                                    cn_cgst = cn_sgst = cn_igst = 0.0
                                    cn_tax_amt = 0.0
                            else:
                                # USD: no GST
                                cn_taxable = cn_amt
                                cn_cgst = cn_sgst = cn_igst = 0.0
                                cn_tax_amt = 0.0

                            if _orig_cur == "INR":
                                st.markdown(
                                    f"**Credit Note Amount:** ₹{cn_amt:,.2f} &nbsp;|&nbsp; "
                                    f"Taxable: ₹{cn_taxable:,.2f} &nbsp;|&nbsp; Tax: ₹{cn_tax_amt:,.2f}"
                                )
                            else:
                                _cn_total_inr_disp = round(
                                    sum(d["usd_amount"] * d["fx_rate"] for d in _cn_per_month_data), 2
                                ) if _cn_per_month_data else 0.0
                                st.markdown(
                                    f"**Credit Note Amount:** ${cn_amt:,.2f} &nbsp;|&nbsp; "
                                    f"INR equiv ≈ ₹{_cn_total_inr_disp:,.2f}"
                                )

                            cn_b1, cn_b2 = st.columns([1, 5])
                            with cn_b1:
                                if st.button("📥 Generate & Save Credit Note", key="cn_generate_btn", type="primary"):
                                    try:
                                        _cn_conn = get_db()
                                        _cn_number = generate_credit_note_number(_cn_conn)
                                        _cn_conn.close()
                                        _cn_partner  = get_partner_info(_orig_dsp)
                                        _cn_date_str = cn_date.strftime("%d/%m/%Y")

                                        cn_buf = generate_credit_note_pdf(
                                            _cn_number, _cn_date_str, _orig_inv_no, cn_reason,
                                            _cn_partner, _orig_month, cn_amt, _orig_cur,
                                            taxable=cn_taxable, cgst=cn_cgst, sgst=cn_sgst, igst=cn_igst,
                                            fx_rate=_cn_fx_rate
                                        )
                                        # Always save month_breakdown so GST Report gets per-month FX
                                        _cn_bd_json = _cn_json.dumps(_cn_per_month_data) if _cn_per_month_data else ""

                                        save_invoice(
                                            _cn_number, _cn_date_str, _orig_dsp, _orig_month,
                                            _orig_type, cn_amt, cn_tax_amt, cn_amt, _orig_cur,
                                            "", status="Credit Note",
                                            igst_amount=cn_igst, cgst_amount=cn_cgst,
                                            sgst_amount=cn_sgst, taxable_amount=cn_taxable,
                                            is_credit_note=1,
                                            credit_note_ref=_orig_inv_no,
                                            credit_note_reason=cn_reason,
                                            fx_rate=_cn_fx_rate,
                                            month_breakdown=_cn_bd_json,
                                        )
                                        st.session_state["_cn_buf"]    = cn_buf
                                        st.session_state["_cn_fname"]  = f"{_cn_number.replace('/','_')}.pdf"
                                        st.session_state["_cn_number"] = _cn_number
                                        st.success(f"✅ Credit Note {_cn_number} generated and saved!")
                                    except Exception as _cne:
                                        st.error(f"Failed: {_cne}")
                            with cn_b2:
                                if st.button("✖ Close", key="cn_close_btn"):
                                    st.session_state["_hist_panel"] = None
                                    st.rerun()

                            if st.session_state.get("_cn_buf"):
                                _cn_dl_copy = io.BytesIO(st.session_state["_cn_buf"].getvalue())
                                st.download_button(
                                    f"📥 Download {st.session_state.get('_cn_number','Credit Note')} PDF",
                                    data=_cn_dl_copy,
                                    file_name=st.session_state["_cn_fname"],
                                    mime="application/pdf",
                                    key="cn_download_btn"
                                )

                    # ══════════════════════════════════════════════
                    # DOWNLOAD PANEL
                    # ══════════════════════════════════════════════
                    elif _active_panel == "download":
                        st.markdown("""
                        <div style='border:2px solid #2E7D32;border-radius:10px;
                        background:#f1f8f1;padding:4px 0 0 0;margin-top:8px'>
                        <div style='background:#2E7D32;color:white;padding:8px 18px;
                        border-radius:8px 8px 0 0;font-size:16px;font-weight:700'>
                        ⬇️ Download Invoice / Credit Note</div></div>""",
                        unsafe_allow_html=True)

                        with st.container():
                            st.markdown(f"**Regenerating PDF for:** `{_sel_inv_no}` — {_sel_dsp}")
                            try:
                                _dl_buf = _regenerate_invoice_pdf(sel_row_data)
                                _dl_fname = f"{_sel_inv_no.replace('/','_')}.pdf"
                                dl_c1, dl_c2 = st.columns([1, 5])
                                with dl_c1:
                                    st.download_button(
                                        f"📥 Download PDF",
                                        data=_dl_buf.getvalue(),
                                        file_name=_dl_fname,
                                        mime="application/pdf",
                                        key="hist_dl_pdf_btn",
                                        type="primary"
                                    )
                                with dl_c2:
                                    if st.button("✖ Close", key="hist_dl_close"):
                                        st.session_state["_hist_panel"] = None
                                        st.rerun()
                            except Exception as _dl_e:
                                st.error(f"Could not regenerate PDF: {_dl_e}")
                                if st.button("✖ Close", key="hist_dl_close_err"):
                                    st.session_state["_hist_panel"] = None
                                    st.rerun()

                    # ══════════════════════════════════════════════
                    # RESEND EMAIL PANEL
                    # ══════════════════════════════════════════════
                    elif _active_panel == "resend":
                        st.markdown("""
                        <div style='border:2px solid #E65100;border-radius:10px;
                        background:#fff8f4;padding:4px 0 0 0;margin-top:8px'>
                        <div style='background:#E65100;color:white;padding:8px 18px;
                        border-radius:8px 8px 0 0;font-size:16px;font-weight:700'>
                        📤 Resend Invoice / Credit Note via Email</div></div>""",
                        unsafe_allow_html=True)

                        with st.container():
                            # Regenerate PDF
                            try:
                                _rs_buf   = _regenerate_invoice_pdf(sel_row_data)
                                _rs_fname = f"{_sel_inv_no.replace('/','_')}.pdf"
                            except Exception as _rs_pdf_e:
                                st.error(f"Could not regenerate PDF: {_rs_pdf_e}")
                                if st.button("✖ Close", key="rs_close_err"):
                                    st.session_state["_hist_panel"] = None
                                    st.rerun()
                                st.stop()

                            # Pre-fill recipient from partner info
                            _rs_partner = get_partner_info(_sel_dsp)
                            _rs_emails  = ", ".join(filter(None, [
                                _rs_partner.get("email1",""),
                                _rs_partner.get("email2",""),
                                _rs_partner.get("finance_email","")
                            ]))
                            _rs_inv_is_cn = int(sel_row_data.get("is_credit_note",0) or 0) == 1
                            _rs_month     = str(sel_row_data.get("month",""))
                            _rs_cur       = str(sel_row_data.get("currency","USD")).upper()
                            _rs_total     = float(sel_row_data.get("total_amount",0) or 0)
                            _rs_amt       = float(sel_row_data.get("amount",0) or 0)
                            _rs_taxable   = float(sel_row_data.get("taxable_amount",0) or 0)
                            _rs_cgst      = float(sel_row_data.get("cgst_amount",0) or 0)
                            _rs_sgst      = float(sel_row_data.get("sgst_amount",0) or 0)
                            _rs_igst      = float(sel_row_data.get("igst_amount",0) or 0)
                            _rs_inv_type  = str(sel_row_data.get("invoice_type","USD"))
                            _rs_due_raw   = str(sel_row_data.get("due_date","") or "")
                            _rs_legal     = _rs_partner.get("legal_name", _sel_dsp)

                            # Amount label (same logic as create tab)
                            if _rs_cur == "INR":
                                _rs_amount_label    = f"₹{_rs_total:,.2f}"
                                _rs_due_amount_disp = f"₹{_rs_total:,.2f}"
                            else:
                                _rs_amount_label    = f"${_rs_amt:,.2f}"
                                _rs_due_amount_disp = f"${_rs_amt:,.2f}"

                            # Due date for subject
                            try:
                                _rs_due_subj = pd.to_datetime(_rs_due_raw, dayfirst=True).strftime("%d/%m/%Y") if _rs_due_raw else "N/A"
                            except Exception:
                                _rs_due_subj = _rs_due_raw or "N/A"

                            # Month label (single: "Aug-2025", multi: "Aug-2025, Sep-2025 and Oct-2025")
                            _rs_month_parts = [m.strip() for m in _rs_month.split(",") if m.strip()]
                            if len(_rs_month_parts) > 1:
                                _rs_month_label = ", ".join(_rs_month_parts[:-1]) + " and " + _rs_month_parts[-1]
                            else:
                                _rs_month_label = _rs_month

                            # ── Subject — same format as create tab ──────────────
                            if _rs_inv_is_cn:
                                _rs_subject = (
                                    f"Credit Note for {_rs_month_label} from {COMPANY['name']} "
                                    f"(CN No.: {_sel_inv_no} - {_rs_amount_label})"
                                )
                            else:
                                _rs_subject = (
                                    f"Invoice for {_rs_month_label} from {COMPANY['name']} "
                                    f"(Invoice No.: {_sel_inv_no} - {_rs_amount_label}) "
                                    f"| Due On {_rs_due_subj}"
                                )

                            # ── GST breakdown rows for INR invoices ──────────────
                            if _rs_inv_type == "INR_CGST_SGST":
                                _rs_gst_rows = f"""
      <tr><td><b>Taxable Amount</b></td><td>&nbsp;-&nbsp;</td><td>₹{_rs_taxable:,.2f}</td></tr>
      <tr><td><b>CGST (9%)</b></td><td>&nbsp;-&nbsp;</td><td>₹{_rs_cgst:,.2f}</td></tr>
      <tr><td><b>SGST (9%)</b></td><td>&nbsp;-&nbsp;</td><td>₹{_rs_sgst:,.2f}</td></tr>"""
                            elif _rs_inv_type == "INR_IGST":
                                _rs_gst_rows = f"""
      <tr><td><b>Taxable Amount</b></td><td>&nbsp;-&nbsp;</td><td>₹{_rs_taxable:,.2f}</td></tr>
      <tr><td><b>IGST (18%)</b></td><td>&nbsp;-&nbsp;</td><td>₹{_rs_igst:,.2f}</td></tr>"""
                            else:
                                _rs_gst_rows = ""

                            # ── Body — same format as create tab ─────────────────
                            if _rs_inv_is_cn:
                                _rs_body = f"""
    <p>Dear {_rs_legal},</p>
    <p>Please find attached the Credit Note <b>{_sel_inv_no}</b> issued against your account.</p>
    <br/>
    <table>
      <tr><td><b>Customer Name</b></td><td>&nbsp;-&nbsp;</td><td>{_rs_legal}</td></tr>
      <tr><td><b>Credit Note No.</b></td><td>&nbsp;-&nbsp;</td><td>{_sel_inv_no}</td></tr>
      <tr><td><b>Month(s)</b></td><td>&nbsp;-&nbsp;</td><td>{_rs_month_label}</td></tr>
      <tr><td><b>Credit Amount</b></td><td>&nbsp;-&nbsp;</td><td>{_rs_amount_label}</td></tr>
    </table>
    <br/>
    <p>You can email us at <a href="mailto:{COMPANY['email']}">{COMPANY['email']}</a> for any clarifications.</p>
    <br/>
    <p>Regards,<br/><b>Finance Team - {COMPANY['name']}</b></p>"""
                            else:
                                _rs_body = f"""
    <p>Dear {_rs_legal},</p>
    <p>It was a pleasure doing business with you.</p>
    <p>Please find the attached invoice, kindly pay the same on or before due date.</p>
    <br/>
    <table>
      <tr><td><b>Customer Name</b></td><td>&nbsp;-&nbsp;</td><td>{_rs_legal}</td></tr>
      <tr><td><b>Invoice No.</b></td><td>&nbsp;-&nbsp;</td><td>{_sel_inv_no}</td></tr>
      <tr><td><b>Invoice Month</b></td><td>&nbsp;-&nbsp;</td><td>{_rs_month_label}</td></tr>{_rs_gst_rows}
      <tr><td><b>Due Amount</b></td><td>&nbsp;-&nbsp;</td><td>{_rs_due_amount_disp}</td></tr>
      <tr><td><b>Due Date</b></td><td>&nbsp;-&nbsp;</td><td>{_rs_due_subj}</td></tr>
    </table>
    <br/>
    <p>You can email us at <a href="mailto:{COMPANY['email']}">{COMPANY['email']}</a> for any clarifications.</p>
    <br/>
    <p>Regards,<br/><b>Finance Team - {COMPANY['name']}</b></p>"""

                            rs_c1, rs_c2 = st.columns(2)
                            with rs_c1:
                                _rs_to = st.text_input(
                                    "To (comma separated)", value=_rs_emails, key="rs_to_inp"
                                )
                            with rs_c2:
                                _rs_subj_inp = st.text_input(
                                    "Subject", value=_rs_subject, key="rs_subj_inp"
                                )

                            # Preview (same as create tab)
                            with st.expander("👁️ Email Preview", expanded=True):
                                st.markdown(f"**Subject:** {_rs_subject}")
                                st.markdown("---")
                                st.markdown(_rs_body, unsafe_allow_html=True)

                            with st.expander("📝 Edit Email Body", expanded=False):
                                _rs_body_inp = st.text_area(
                                    "Body (HTML)", value=_rs_body, height=200, key="rs_body_inp"
                                )
                            _rs_body_send = st.session_state.get("rs_body_inp", _rs_body)

                            # SMTP credentials
                            if _smtp_configured_via_secrets():
                                st.caption("✅ **Email credentials loaded from secrets** — no password needed.")
                                _rs_host, _rs_port, _rs_user, _rs_pass = _get_smtp_creds()
                            else:
                                with st.expander("⚙️ SMTP Settings", expanded=False):
                                    _rsc1, _rsc2 = st.columns(2)
                                    with _rsc1:
                                        st.text_input("SMTP Host", value="smtpout.secureserver.net", key="rs_smtp_host")
                                        st.text_input("From Email", value="finance@peakmyads.com",   key="rs_smtp_user")
                                    with _rsc2:
                                        st.number_input("SMTP Port", value=465, key="rs_smtp_port")
                                        st.text_input("Password", type="password", key="rs_smtp_pass")
                                _rs_host = st.session_state.get("rs_smtp_host","smtpout.secureserver.net")
                                _rs_port = int(st.session_state.get("rs_smtp_port", 465))
                                _rs_user = st.session_state.get("rs_smtp_user","finance@peakmyads.com")
                                _rs_pass = st.session_state.get("rs_smtp_pass","")

                            rs_btn1, rs_btn2, rs_btn3 = st.columns([1, 1, 4])
                            with rs_btn1:
                                if st.button("📤 Send", key="rs_send_btn", type="primary", use_container_width=True):
                                    _rs_to_list = [e.strip() for e in _rs_to.split(",") if e.strip()]
                                    if not _rs_to_list:
                                        st.error("Enter at least one recipient.")
                                    elif not _rs_pass:
                                        st.error("SMTP password not set.")
                                    else:
                                        try:
                                            send_invoice_email(
                                                _rs_to_list,
                                                st.session_state.get("rs_subj_inp", _rs_subject),
                                                _rs_body_send,
                                                io.BytesIO(_rs_buf.getvalue()),
                                                _rs_fname,
                                                _rs_host, int(_rs_port),
                                                _rs_user, _rs_pass
                                            )
                                            st.success(f"✅ {_rs_label} resent to {', '.join(_rs_to_list)}!")
                                            st.session_state["_hist_panel"] = None
                                        except Exception as _rs_e:
                                            st.error(f"Email failed: {_rs_e}")
                            with rs_btn2:
                                if st.button("✖ Close", key="rs_close_btn", use_container_width=True):
                                    st.session_state["_hist_panel"] = None
                                    st.rerun()
                            with rs_btn3:
                                # Also offer direct download alongside resend
                                st.download_button(
                                    "📥 Download PDF instead",
                                    data=_rs_buf.getvalue(),
                                    file_name=_rs_fname,
                                    mime="application/pdf",
                                    key="rs_dl_btn"
                                )

            else:
                st.info("👆 Click a row above to select it, then choose an action: "
                        "**🔔 Reminder** | **✏️ Edit** | **📝 Credit Note** | "
                        "**⬇️ Download** | **📤 Resend Email**")

    # ════════════════════════════════════════
    # TAB 3 — SEND REMINDER
    # ════════════════════════════════════════
    with tab_reminder:
        
        st.markdown('''
            <div style="background:linear-gradient(135deg,#003366 0%,#005599 100%);
            border-radius:10px;padding:18px 24px;margin-bottom:20px;
            box-shadow:0 4px 16px rgba(0,51,102,.2);
            display:flex;align-items:center;height:4px;gap:14px;">
            <div>
                <div style="color:white;font-size:20px;font-weight:800;">
                    🔔 Payment Reminders</div>
            </div></div>
        ''', unsafe_allow_html=True)
        
        st.caption("Shows DSP partners with zero Received Amount (unpaid).")

        # Exclude months already paid in invoice_details
        try:
            _paid_conn = get_db()
            _paid_rows = _paid_conn.execute(
                "SELECT dsp_name, month FROM invoice_details WHERE status='Paid' AND is_deleted=0 AND is_credit_note=0"
            ).fetchall()
            _paid_conn.close()
            _paid_keys = {(r[0], r[1]) for r in _paid_rows}
        except Exception:
            _paid_keys = set()

        _unpaid_mask = pd.to_numeric(dsp_df["Received Amount $"], errors="coerce").fillna(0) == 0
        unpaid = dsp_df[_unpaid_mask][["Month", "DSP Name", "Receivable $", "Due Date"]].copy()
        unpaid["Receivable $"] = pd.to_numeric(unpaid["Receivable $"], errors="coerce")
        unpaid = unpaid[~unpaid.apply(lambda r: (r["DSP Name"], str(r["Month"])) in _paid_keys, axis=1)]
        unpaid["Due Date"] = pd.to_datetime(unpaid["Due Date"], errors="coerce").dt.strftime("%d-%b-%Y")

        if unpaid.empty:
            st.success("✅ No outstanding DSP invoices!")
        else:
            st.warning(f"{len(unpaid)} outstanding row(s) found.")

            from st_aggrid import AgGrid, GridOptionsBuilder, JsCode

            currency_fmt_rem = JsCode("""
            function(params) {
                if (params.value == null || params.value === '') return '';
                return '$' + parseFloat(params.value).toLocaleString(undefined, {minimumFractionDigits: 2});
            }
            """)

            gb_rem = GridOptionsBuilder.from_dataframe(unpaid)

            if "Receivable $" in unpaid.columns:
                gb_rem.configure_column("Receivable $", type=["numericColumn"], valueFormatter=currency_fmt_rem)

            pinned_style_rem = JsCode("""
            function(params) {
                if (params.column.pinned) {
                    return { fontWeight: 'bold', borderRight: '2px solid #003366' };
                }
            }
            """)
            for pin_col in ["DSP Name", "Month"]:
                if pin_col in unpaid.columns:
                    gb_rem.configure_column(pin_col, pinned="left", cellStyle=pinned_style_rem)

            gb_rem.configure_default_column(resizable=True, sortable=True, filter=False)

            # Grand total row
            rem_total_row = {col: "" for col in unpaid.columns}
            rem_total_row["DSP Name"] = "TOTAL"
            rem_total_row["Receivable $"] = pd.to_numeric(unpaid["Receivable $"], errors="coerce").sum()

            rem_grid_opts = gb_rem.build()
            rem_grid_opts["pinnedBottomRowData"] = [rem_total_row]
            rem_grid_opts["getRowStyle"] = JsCode("""
            function(params) {
                if (params.node.rowPinned) {
                    return {
                        backgroundColor: '#003366',
                        color: 'white',
                        fontWeight: 'bold',
                        fontSize: '14px'
                    };
                }
            }
            """)

            rem_custom_css = {
                ".ag-header": {
                    "background-color": "#003366 !important",
                    "color": "white !important",
                    "font-weight": "bold !important",
                    "font-size": "14px !important"
                },
                ".ag-header-cell-label": {
                    "color": "white !important",
                    "font-weight": "bold !important"
                }
            }

            AgGrid(
                unpaid,
                gridOptions=rem_grid_opts,
                allow_unsafe_jscode=True,
                fit_columns_on_grid_load=True,
                height=300,
                custom_css=rem_custom_css
            )

            rem_dsp = st.selectbox(
                "Select DSP to remind",
                sorted(unpaid["DSP Name"].unique().tolist()),
                key="rem_dsp_sel"
            )
            rem_partner = get_partner_info(rem_dsp)
            rem_email_default = ", ".join(filter(None, [
                rem_partner.get("email1",""),
                rem_partner.get("finance_email","")
            ]))

            re1, re2 = st.columns([3, 1])
            with re1:
                rem_to = st.text_input("Reminder To", value=rem_email_default, key="rem_to_email")
            with re2:
                st.markdown(f"**BCC:** {BCC_EMAIL}", unsafe_allow_html=False)

            # ── Build outstanding table for this DSP ──────────────────
            this_unpaid = unpaid[unpaid["DSP Name"] == rem_dsp].copy()
            total_outstanding = this_unpaid["Receivable $"].sum()

            _table_rows_html = ""
            for _, _ur in this_unpaid.iterrows():
                _table_rows_html += (
                    "<tr>"
                    "<td style='padding:6px 12px;border:1px solid #ddd'>" + str(_ur['Month']) + "</td>"
                    "<td style='padding:6px 12px;border:1px solid #ddd;text-align:right'>$" + f"{float(_ur['Receivable $'] or 0):,.2f}" + "</td>"
                    "<td style='padding:6px 12px;border:1px solid #ddd'>" + str(_ur['Due Date']) + "</td>"
                    "</tr>"
                )
            _table_rows_html += (
                "<tr style='background:#003366;color:white;font-weight:bold'>"
                "<td style='padding:6px 12px;border:1px solid #003366'>Total Outstanding</td>"
                "<td style='padding:6px 12px;border:1px solid #003366;text-align:right'>$" + f"{total_outstanding:,.2f}" + "</td>"
                "<td style='padding:6px 12px;border:1px solid #003366'></td>"
                "</tr>"
            )
            _outstanding_table = (
                "<table style='border-collapse:collapse;width:100%;font-size:14px'>"
                "<thead><tr style='background:#003366;color:white'>"
                "<th style='padding:8px 12px;border:1px solid #003366;text-align:left'>Month</th>"
                "<th style='padding:8px 12px;border:1px solid #003366;text-align:right'>Amount</th>"
                "<th style='padding:8px 12px;border:1px solid #003366;text-align:left'>Due Date</th>"
                "</tr></thead>"
                "<tbody>" + _table_rows_html + "</tbody></table>"
            )

            _legal_name = rem_partner.get("legal_name", rem_dsp)
            rem_subject = f"Payment Reminder — {COMPANY['name']} | Outstanding: ${total_outstanding:,.2f}"
            rem_body = (
                f"<p>Dear {_legal_name},</p>"
                "<p>This is a gentle reminder that the following invoices are outstanding and payment is due:</p>"
                "<br/>"
                + _outstanding_table +
                "<br/>"
                "<p>Please arrange payment at your earliest convenience.</p>"
                f"<p>You can email us at <a href='mailto:{COMPANY['email']}'>{COMPANY['email']}</a> for any clarifications.</p>"
                "<br/>"
                f"<p>Regards,<br/><b>Finance Team - {COMPANY['name']}</b></p>"
            )

            # Email preview for reminder
            with st.expander("👁️ Reminder Email Preview", expanded=False):
                st.markdown(f"**Subject:** {rem_subject}")
                st.markdown("---")
                st.markdown(rem_body, unsafe_allow_html=True)

            if _smtp_configured_via_secrets():
                st.caption("✅ **Email credentials loaded from secrets** — no password needed.")
                _rem_host, _rem_port, _rem_user, _rem_pass = _get_smtp_creds()
            else:
                with st.expander("⚙️ SMTP Settings", expanded=False):
                    rsc1, rsc2 = st.columns(2)
                    with rsc1:
                        st.text_input("SMTP Host",  value="smtpout.secureserver.net", key="rem_smtp_host")
                        st.text_input("From Email", value="finance@peakmyads.com",    key="rem_smtp_user")
                    with rsc2:
                        st.number_input("SMTP Port", value=465, key="rem_smtp_port")
                        st.text_input("App Password", type="password", key="rem_smtp_pass")
                    st.caption("SMTP: smtpout.secureserver.net : 465 (SSL/TLS)")
                _rem_host, _rem_port, _rem_user, _rem_pass = _get_smtp_creds()

            rem_send = st.button("📤 Send Reminder", key="rem_send_btn", type="primary")

            if rem_send:
                to_list = [e.strip() for e in rem_to.split(",") if e.strip()]
                if not to_list:
                    st.error("No email address entered.")
                elif not _rem_pass:
                    st.error("SMTP password not set. Configure secrets.toml or enter in SMTP Settings.")
                else:
                    try:
                        msg = MIMEMultipart()
                        msg["From"]    = _rem_user
                        msg["To"]      = ", ".join(to_list)
                        msg["Bcc"]     = BCC_EMAIL
                        msg["Subject"] = rem_subject
                        msg.attach(MIMEText(rem_body, "html"))
                        all_rem_recipients = to_list + [BCC_EMAIL]
                        _smtp_send(_rem_host, int(_rem_port), _rem_user, _rem_pass,
                                   all_rem_recipients, msg.as_string())
                        st.success(f"✅ Reminder sent to {', '.join(to_list)}! (BCC: {BCC_EMAIL})")
                    except Exception as e:
                        st.error(f"Reminder failed: {e}")

    # ════════════════════════════════════════
    # TAB 4 — DSP STATEMENT
    # ════════════════════════════════════════
    with tab_statement:
        st.markdown('''
            <div style="background:linear-gradient(135deg,#003366 0%,#005599 100%);
            border-radius:10px;padding:18px 24px;margin-bottom:20px;
            box-shadow:0 4px 16px rgba(0,51,102,.2);
            display:flex;align-items:center;height:4px;gap:14px;">
            <div>
                <div style="color:white;font-size:20px;font-weight:800;">
                    📊 DSP Account Statement</div>
            </div></div>
        ''', unsafe_allow_html=True)
        
        dsp_names_all = sorted(dsp_df["DSP Name"].dropna().unique().tolist())

        stmt_dsp = st.selectbox("Select DSP", dsp_names_all, key="stmt_dsp_sel")

        # Preview table — join Invoice Date & No. from invoice_details
        stmt_df = dsp_df[dsp_df["DSP Name"] == stmt_dsp][
            ["Month", "Receivable $", "Due Date", "Received Date",
             "Received Amount $", "Received In", "Reason"]
        ].copy()

        # Normalise Month for display & join key
        def _fmt_month(m):
            try:
                return pd.to_datetime(str(m)).strftime("%b-%Y")
            except Exception:
                return str(m)
        stmt_df["Month"] = stmt_df["Month"].apply(_fmt_month)

        # Fetch Invoice No. & Invoice Date from invoice_details
        try:
            _si_conn = get_db()
            _si_hist = pd.read_sql(
                "SELECT dsp_name, month, invoice_number, invoice_date FROM invoice_details "
                "WHERE is_credit_note=0 AND is_deleted=0", _si_conn
            )
            _si_conn.close()
            _si_map = {}
            for _, _sr in _si_hist.iterrows():
                _sk = str(_sr["dsp_name"]).strip() + "|" + str(_sr["month"]).strip()
                try:
                    _nm = pd.to_datetime(str(_sr["month"]), format="%b-%Y", errors="coerce")
                    if pd.isna(_nm):
                        _nm = pd.to_datetime(str(_sr["month"]), errors="coerce")
                    _nk = str(_sr["dsp_name"]).strip() + "|" + _nm.strftime("%b-%Y")
                except Exception:
                    _nk = _sk
                _v = {"invoice_number": _sr["invoice_number"], "invoice_date": _sr["invoice_date"]}
                _si_map[_sk] = _v
                _si_map[_nk] = _v
        except Exception:
            _si_map = {}

        def _get_inv_no(row):
            k = stmt_dsp + "|" + str(row["Month"])
            return _si_map.get(k, {}).get("invoice_number", "")

        def _get_inv_date(row):
            k = stmt_dsp + "|" + str(row["Month"])
            _d = _si_map.get(k, {}).get("invoice_date", "")
            if not _d:
                return ""
            try:
                return pd.to_datetime(str(_d)).strftime("%d-%b-%Y")
            except Exception:
                return str(_d)

        stmt_df.insert(1, "Invoice Date", stmt_df.apply(_get_inv_date, axis=1))
        stmt_df.insert(2, "Invoice No.",  stmt_df.apply(_get_inv_no,   axis=1))

        # Fix Due Date — NaT → ""
        stmt_df["Due Date"] = stmt_df["Due Date"].apply(
            lambda v: "" if pd.isna(pd.to_datetime(v, errors="coerce"))
                      else pd.to_datetime(v, errors="coerce").strftime("%d-%b-%Y")
        )
        # Fix Received Date — NaT → ""
        stmt_df["Received Date"] = stmt_df["Received Date"].apply(
            lambda v: "" if pd.isna(pd.to_datetime(v, errors="coerce"))
                      else pd.to_datetime(v, errors="coerce").strftime("%d-%b-%Y")
        )
        # Fix Received In — hide "Select" placeholder
        stmt_df["Received In"] = stmt_df["Received In"].apply(
            lambda v: "" if str(v or "").strip().lower() in ("select", "none", "nan", "") else str(v)
        )

        from st_aggrid import AgGrid, GridOptionsBuilder, JsCode

        currency_fmt_stmt = JsCode("""
        function(params) {
            if (params.value == null || params.value === \'\') return \'\';
            return \'$\' + parseFloat(params.value).toLocaleString(undefined, {minimumFractionDigits: 2});
        }
        """)

        gb_stmt = GridOptionsBuilder.from_dataframe(stmt_df)

        for cur_col in ["Receivable $", "Received Amount $"]:
            if cur_col in stmt_df.columns:
                gb_stmt.configure_column(cur_col, type=["numericColumn"], valueFormatter=currency_fmt_stmt)

        shortage_style = JsCode("""
        function(params) {
            let recv = parseFloat(params.data["Receivable $"]) || 0;
            let paid = parseFloat(params.data["Received Amount $"]) || 0;
            if (recv > 0 && paid >= recv) return { color: \'#2E7D32\', fontWeight: \'bold\' };
            if (recv > 0 && paid < recv)  return { color: \'#C62828\', fontWeight: \'bold\' };
            return {};
        }
        """)
        if "Received Amount $" in stmt_df.columns:
            gb_stmt.configure_column("Received Amount $",
                type=["numericColumn"],
                valueFormatter=currency_fmt_stmt,
                cellStyle=shortage_style
            )

        pinned_style_stmt = JsCode("""
        function(params) {
            if (params.column.pinned) {
                return { fontWeight: \'bold\', borderRight: \'2px solid #003366\' };
            }
        }
        """)
        if "Month" in stmt_df.columns:
            gb_stmt.configure_column("Month", pinned="left", cellStyle=pinned_style_stmt)

        gb_stmt.configure_default_column(resizable=True, sortable=True, filter=False)

        stmt_total_row = {col: "" for col in stmt_df.columns}
        stmt_total_row["Month"] = "TOTAL"
        for tot_col in ["Receivable $", "Received Amount $"]:
            if tot_col in stmt_df.columns:
                stmt_total_row[tot_col] = pd.to_numeric(stmt_df[tot_col], errors="coerce").sum()

        stmt_grid_opts = gb_stmt.build()
        stmt_grid_opts["pinnedBottomRowData"] = [stmt_total_row]
        stmt_grid_opts["getRowStyle"] = JsCode("""
        function(params) {
            if (params.node.rowPinned) {
                return {
                    backgroundColor: \'#003366\',
                    color: \'white\',
                    fontWeight: \'bold\',
                    fontSize: \'14px\'
                };
            }
        }
        """)

        stmt_custom_css = {
            ".ag-header": {
                "background-color": "#003366 !important",
                "color": "white !important",
                "font-weight": "bold !important",
                "font-size": "14px !important"
            },
            ".ag-header-cell-label": {
                "color": "white !important",
                "font-weight": "bold !important"
            }
        }

        AgGrid(
            stmt_df,
            gridOptions=stmt_grid_opts,
            allow_unsafe_jscode=True,
            fit_columns_on_grid_load=True,
            height=400,
            custom_css=stmt_custom_css
        )

        # ── Statement Email ──────────────────────
        stmt_partner = get_partner_info(stmt_dsp)
        stmt_emails_default = ", ".join(filter(None, [
            stmt_partner.get("email1", ""),
            stmt_partner.get("finance_email", "")
        ]))

        se1, se2 = st.columns([3, 1])
        with se1:
            stmt_to_email = st.text_input("Send Statement To", value=stmt_emails_default, key="stmt_to_email")
        with se2:
            st.markdown(f"**BCC:** {BCC_EMAIL}")

        # Statement email subject & body
        stmt_subject = f"Account Statement from {COMPANY['name']} | {stmt_dsp}"
        stmt_body = f"""
<p>Dear {stmt_partner.get('legal_name', stmt_dsp)},</p>
<p>Please find attached your account statement from <b>{COMPANY['name']}</b>.</p>
<p>Kindly review and revert for any clarifications.</p>
<br/>
<p>You can email us at <a href="mailto:{COMPANY['email']}">{COMPANY['email']}</a> for any queries.</p>
<br/>
<p>Regards,<br/><b>Finance Team - {COMPANY['name']}</b></p>
"""
        with st.expander("👁️ Statement Email Preview", expanded=False):
            st.markdown(f"**Subject:** {stmt_subject}")
            st.markdown("---")
            st.markdown(stmt_body, unsafe_allow_html=True)

        if _smtp_configured_via_secrets():
            st.caption("✅ **Email credentials loaded from secrets** — no password needed.")
            _stmt_host, _stmt_port, _stmt_user, _stmt_pass = _get_smtp_creds()
        else:
            with st.expander("⚙️ SMTP Settings (Statement)", expanded=False):
                ssc1, ssc2 = st.columns(2)
                with ssc1:
                    st.text_input("SMTP Host",  value="smtpout.secureserver.net", key="stmt_smtp_host")
                    st.text_input("From Email", value="finance@peakmyads.com",    key="stmt_smtp_user")
                with ssc2:
                    st.number_input("SMTP Port", value=465, key="stmt_smtp_port")
                    st.text_input("Password", type="password", key="stmt_smtp_pass")
                st.caption("SMTP: smtpout.secureserver.net : 465 (SSL/TLS)")
            _stmt_host, _stmt_port, _stmt_user, _stmt_pass = _get_smtp_creds()

        btn_col1, btn_col2 = st.columns([1, 1])
        with btn_col1:
            if st.button("📥 Generate & Download Statement PDF", key="stmt_download_btn", type="primary"):
                try:
                    stmt_buf = generate_dsp_statement(stmt_dsp, dsp_df)
                    st.session_state["_stmt_buf"]  = stmt_buf
                    st.session_state["_stmt_dsp"]  = stmt_dsp
                    st.session_state["_stmt_fname"] = f"Statement_{stmt_dsp}_{datetime.today().strftime('%Y%m%d')}.pdf"
                    st.success("✅ Statement PDF generated!")
                except Exception as e:
                    st.error(f"Statement generation failed: {e}")

        if st.session_state.get("_stmt_buf"):
            stmt_buf_copy = io.BytesIO(st.session_state["_stmt_buf"].getvalue())
            with btn_col2:
                st.download_button(
                    "⬇️ Download Statement PDF",
                    data=stmt_buf_copy,
                    file_name=st.session_state["_stmt_fname"],
                    mime="application/pdf",
                    key="stmt_dl_actual"
                )

            if st.button("📤 Send Statement via Email", key="stmt_send_btn"):
                stmt_to_list = [e.strip() for e in stmt_to_email.split(",") if e.strip()]
                if not stmt_to_list:
                    st.error("Please enter at least one email address.")
                elif not _stmt_pass:
                    st.error("SMTP password not set. Configure secrets.toml or enter in SMTP Settings.")
                else:
                    try:
                        buf_stmt_send = io.BytesIO(st.session_state["_stmt_buf"].getvalue())
                        send_invoice_email(
                            stmt_to_list, stmt_subject, stmt_body,
                            buf_stmt_send, st.session_state["_stmt_fname"],
                            _stmt_host, int(_stmt_port), _stmt_user, _stmt_pass
                        )
                        st.success(f"✅ Statement sent to {', '.join(stmt_to_list)}! (BCC: {BCC_EMAIL})")
                    except Exception as e:
                        st.error(f"Email failed: {e}")

    # ════════════════════════════════════════
    # TAB 5 — GST REPORT
    # ════════════════════════════════════════
    with tab_gst:
        
        st.markdown('''
            <div style="background:linear-gradient(135deg,#003366 0%,#005599 100%);
            border-radius:10px;padding:18px 24px;margin-bottom:20px;
            box-shadow:0 4px 16px rgba(0,51,102,.2);
            display:flex;align-items:center;height:4px;gap:14px;">
            <div>
                <div style="color:white;font-size:20px;font-weight:800;">
                    📑 GST Report</div>
            </div></div>
        ''', unsafe_allow_html=True)
        
        st.caption("INR invoices (GST filing) + USD invoices (Export). GSTIN from List of Partners.")

        gf1, gf2, gf3, gf4 = st.columns(4)
        with gf1:
            gst_incl_cn  = st.checkbox("Include Credit Notes", value=False, key="gst_incl_cn")
        with gf2:
            gst_incl_del = st.checkbox("Include Deleted",       value=False, key="gst_incl_del")
        with gf3:
            gst_all_dsp  = st.checkbox("All DSPs",              value=True,  key="gst_all_dsp")
        with gf4:
            gst_all_per  = st.checkbox("All Periods",           value=True,  key="gst_all_per")

        # ── CN Data Repair Tool (fix CNs saved with wrong INR amounts) ──────
        with st.expander("🔧 Repair Credit Notes with Incorrect Amounts", expanded=False):
            st.caption(
                "Fixes INR credit notes that were saved with a raw USD amount instead of the "
                "correct INR value (USD × FX × 1.18). Only affects CNs where total_amount ≤ "
                "the USD amount stored — safe to run multiple times."
            )
            _fix_gc = get_db()
            import json as _fix_json, math as _fix_math

            def _parse_bd_fix(v):
                if v is None or (isinstance(v, float) and _fix_math.isnan(v)):
                    return []
                s = str(v).strip()
                return _fix_json.loads(s) if s.startswith("[") else []

            _cn_rows_fix = pd.read_sql(
                "SELECT * FROM invoice_details WHERE is_credit_note=1 AND currency='INR' AND is_deleted=0",
                _fix_gc
            )
            _fix_gc.close()

            _cn_broken = []
            for _, _fcr in _cn_rows_fix.iterrows():
                _f_amt    = float(_fcr.get("total_amount", 0) or 0)
                _f_usd    = float(_fcr.get("amount", 0) or 0)
                _f_fx     = float(_fcr.get("fx_rate", 0) or 0)
                _f_bd     = _parse_bd_fix(_fcr.get("month_breakdown"))
                _f_type   = str(_fcr.get("invoice_type",""))
                # A CN is "broken" if total_amount ≈ USD amount (i.e., stored as dollars not rupees)
                if _f_amt > 0 and _f_usd > 0 and abs(_f_amt - _f_usd) < 0.01 and _f_fx > 1:
                    # Compute correct INR total from breakdown or flat fx
                    if _f_bd:
                        _correct_taxable = round(sum(float(b.get("usd_amount",0)) * float(b.get("fx_rate",_f_fx)) for b in _f_bd), 2)
                    else:
                        _correct_taxable = round(_f_usd * _f_fx, 2)
                    if _f_type in ("INR_CGST_SGST", "INR_IGST"):
                        _correct_total = round(_correct_taxable * 1.18, 2)
                        _correct_igst  = round(_correct_taxable * 0.18, 2) if _f_type == "INR_IGST" else 0.0
                        _correct_cgst  = round(_correct_taxable * 0.09, 2) if _f_type == "INR_CGST_SGST" else 0.0
                        _correct_sgst  = _correct_cgst
                    else:
                        _correct_total   = _correct_taxable
                        _correct_igst    = _correct_cgst = _correct_sgst = 0.0
                    _cn_broken.append({
                        "id":              _fcr.get("id"),
                        "invoice_number":  _fcr.get("invoice_number",""),
                        "dsp_name":        _fcr.get("dsp_name",""),
                        "month":           _fcr.get("month",""),
                        "wrong_total":     _f_amt,
                        "correct_taxable": _correct_taxable,
                        "correct_total":   _correct_total,
                        "correct_igst":    _correct_igst,
                        "correct_cgst":    _correct_cgst,
                        "correct_sgst":    _correct_sgst,
                    })

            if not _cn_broken:
                st.success("✅ No broken credit notes found — all CNs look correct.")
            else:
                st.warning(f"Found **{len(_cn_broken)}** credit note(s) with incorrect amounts:")
                _fix_preview_df = pd.DataFrame(_cn_broken)[
                    ["invoice_number","dsp_name","month","wrong_total","correct_taxable","correct_total"]
                ]
                _fix_preview_df.columns = ["CN No.","DSP","Month","Stored Total (₹ — WRONG)","Correct Taxable (₹)","Correct Total (₹)"]
                st.dataframe(_fix_preview_df, use_container_width=True, hide_index=True)

                if st.button("⚡ Fix All Listed Credit Notes", key="gst_fix_cn_btn", type="primary"):
                    _fix_conn2 = get_db()
                    _fixed_count = 0
                    for _fb in _cn_broken:
                        _fix_conn2.execute(
                            """UPDATE invoice_details
                               SET taxable_amount=?, igst_amount=?, cgst_amount=?, sgst_amount=?,
                                   total_amount=?, amount=?
                               WHERE id=?""",
                            (
                                _fb["correct_taxable"], _fb["correct_igst"],
                                _fb["correct_cgst"],    _fb["correct_sgst"],
                                _fb["correct_total"],   _fb["correct_total"],
                                _fb["id"]
                            )
                        )
                        _fixed_count += 1
                    _fix_conn2.commit()
                    _fix_conn2.close()
                    st.success(f"✅ Fixed {_fixed_count} credit note(s). Refresh the page to see updated GST report.")
                    st.rerun()

        if not gst_all_dsp:
            _gst_dsp_opts = sorted(dsp_df["DSP Name"].dropna().unique().tolist())
            gst_dsp_sel   = st.multiselect("Filter DSP(s)", _gst_dsp_opts, key="gst_dsp_sel")
        else:
            gst_dsp_sel = []

        if not gst_all_per:
            _pd1, _pd2 = st.columns(2)
            with _pd1:
                gst_dt_from = st.date_input("From Date",
                    value=date(date.today().year if date.today().month >= 4 else date.today().year - 1, 4, 1),
                    key="gst_dt_from")
            with _pd2:
                gst_dt_to = st.date_input("To Date", value=date.today(), key="gst_dt_to")
        else:
            gst_dt_from = None
            gst_dt_to   = None

        # Load ALL invoices
        _gc = get_db()
        _base_filter = "SELECT * FROM invoice_details WHERE 1=1"
        if not gst_incl_cn:  _base_filter += " AND is_credit_note=0"
        if not gst_incl_del: _base_filter += " AND is_deleted=0"
        _base_filter += " ORDER BY invoice_date ASC"
        _all_df = pd.read_sql(_base_filter, _gc)
        _gc.close()

        if gst_dsp_sel:
            _all_df = _all_df[_all_df["dsp_name"].isin(gst_dsp_sel)]

        if gst_dt_from and gst_dt_to and not _all_df.empty:
            _gdt = pd.to_datetime(_all_df["invoice_date"], dayfirst=True, errors="coerce")
            _all_df = _all_df[(_gdt >= pd.Timestamp(gst_dt_from)) & (_gdt <= pd.Timestamp(gst_dt_to))]

        gst_inv_df = _all_df[_all_df["currency"] == "INR"].copy() if not _all_df.empty else pd.DataFrame()
        usd_inv_df = _all_df[_all_df["currency"] == "USD"].copy() if not _all_df.empty else pd.DataFrame()

        # Single metrics bar
        _cnt_inr = int((_all_df["currency"]=="INR").sum()) if not _all_df.empty else 0
        _cnt_usd = int((_all_df["currency"]=="USD").sum()) if not _all_df.empty else 0
        _cnt_cn  = int((_all_df.get("is_credit_note", pd.Series(dtype=int))==1).sum()) if not _all_df.empty else 0
        _cnt_del = int((_all_df.get("is_deleted", pd.Series(dtype=int))==1).sum()) if not _all_df.empty else 0
        _tot_inr_val = float(gst_inv_df["total_amount"].sum()) if not gst_inv_df.empty else 0.0
        _tot_usd_val = float(usd_inv_df["amount"].sum()) if not usd_inv_df.empty else 0.0

        m1, m2, m3, m4, m5, m6 = st.columns(6)
        m1.metric("INR Invoices",     _cnt_inr)
        m2.metric("Export (USD)",     _cnt_usd)
        m3.metric("Credit Notes",     _cnt_cn)
        m4.metric("Deleted",          _cnt_del)
        m5.metric("Total INR (₹)",    f"{_tot_inr_val:,.0f}")
        m6.metric("Total USD ($)",    f"{_tot_usd_val:,.2f}")
        st.markdown("---")

        if _all_df.empty:
            st.info("No invoices found for the selected filters.")
        else:
            # Partner cache
            try:
                _pc = get_db()
                _praw = pd.read_sql("SELECT * FROM partner_list", _pc)
                _pc.close()
                _short_col = next(
                    (c for c in _praw.columns if "short" in c.lower() or "bidscube" in c.lower()), None
                )
            except Exception:
                _praw = pd.DataFrame()
                _short_col = None

            _ST = {
                "maharashtra":"Maharashtra","pune":"Maharashtra","mumbai":"Maharashtra",
                "navi mumbai":"Maharashtra","thane":"Maharashtra",
                "delhi":"Delhi","new delhi":"Delhi","gurugram":"Haryana","gurgaon":"Haryana",
                "karnataka":"Karnataka","bangalore":"Karnataka","bengaluru":"Karnataka",
                "telangana":"Telangana","hyderabad":"Telangana",
                "tamil nadu":"Tamil Nadu","chennai":"Tamil Nadu",
                "gujarat":"Gujarat","ahmedabad":"Gujarat","surat":"Gujarat",
                "haryana":"Haryana","uttar pradesh":"Uttar Pradesh","noida":"Uttar Pradesh",
                "west bengal":"West Bengal","kolkata":"West Bengal",
                "rajasthan":"Rajasthan","kerala":"Kerala","madhya pradesh":"Madhya Pradesh",
                "andhra pradesh":"Andhra Pradesh","odisha":"Odisha","bihar":"Bihar",
                "jharkhand":"Jharkhand","punjab":"Punjab","goa":"Goa","assam":"Assam",
            }

            def _fetch_p(dsp):
                if _short_col and not _praw.empty:
                    m = _praw[_praw[_short_col] == dsp]
                    if not m.empty:
                        r = m.iloc[0]
                        return {
                            "gstin":   str(r.get("gstin") or r.get("GSTIN") or "").strip(),
                            "address": str(r.get("address") or r.get("Registered Address") or "").strip(),
                            "country": str(r.get("country") or r.get("Country") or "").strip(),
                        }
                return get_partner_info(dsp)

            def _pos(p):
                addr = str(p.get("address","")).lower()
                ctry = str(p.get("country","")).lower()
                for kw, sname in _ST.items():
                    if kw in addr:
                        return sname
                if ctry and "india" not in ctry:
                    return str(p.get("country","")).strip() or "Outside India"
                segs = [s.strip() for s in addr.replace("\n",",").split(",") if s.strip()]
                return segs[-1].title() if segs else "India"

            import json as _inr_json_top, math as _inr_math_top

            def _inr_parse_bd(v):
                """NaN-safe breakdown parser."""
                if v is None or (isinstance(v, float) and _inr_math_top.isnan(v)):
                    return []
                s = str(v).strip()
                return _inr_json_top.loads(s) if s.startswith("[") else []

            def _inr_gst_from_taxable(taxable, inv_type):
                """Return (igst, cgst, sgst, total) from taxable INR amount."""
                if inv_type == "INR_IGST":
                    igst  = round(taxable * 0.18, 2)
                    return igst, 0.0, 0.0, round(taxable + igst, 2)
                elif inv_type == "INR_CGST_SGST":
                    cgst = sgst = round(taxable * 0.09, 2)
                    return 0.0, cgst, sgst, round(taxable + cgst + sgst, 2)
                return 0.0, 0.0, 0.0, taxable

            _pcache = {}

            # ── Build combined row list ────────────────────────────
            all_rows = []

            # INR rows
            for _, inv in gst_inv_df.iterrows():
                _dsp = inv.get("dsp_name","")
                if _dsp not in _pcache:
                    _pcache[_dsp] = _fetch_p(_dsp)
                _p       = _pcache[_dsp]
                _gstin   = _p.get("gstin","") or ""
                _supply  = _pos(_p)
                _inv_no  = inv.get("invoice_number","")
                _inv_date= inv.get("invoice_date","")
                try:
                    _inv_date = pd.to_datetime(str(_inv_date), dayfirst=True, errors="coerce").strftime("%d-%b-%Y")
                except Exception:
                    pass
                _inv_type = inv.get("invoice_type","")
                _taxable  = float(inv.get("taxable_amount", 0) or 0)
                _igst     = float(inv.get("igst_amount",    0) or 0)
                _cgst     = float(inv.get("cgst_amount",    0) or 0)
                _sgst     = float(inv.get("sgst_amount",    0) or 0)
                _total    = float(inv.get("total_amount",   0) or 0)
                _is_cn    = int(inv.get("is_credit_note",   0) or 0)
                _is_del   = int(inv.get("is_deleted",       0) or 0)
                # Fallback: recompute tax from total if stored values are zero
                if _taxable == 0 and _total > 0:
                    _taxable = round(_total / 1.18, 2)
                    if _igst == 0 and _cgst == 0 and _sgst == 0:
                        if _inv_type == "INR_CGST_SGST":
                            _cgst = _sgst = round(_taxable * 0.09, 2)
                        elif _inv_type == "INR_IGST":
                            _igst = round(_taxable * 0.18, 2)
                _tax_rate = "18% (CGST+SGST)" if _inv_type=="INR_CGST_SGST" else \
                            "18% (IGST)"       if _inv_type=="INR_IGST"       else "GST"
                _cat = "Credit Note" if _is_cn else ("Deleted" if _is_del else "Regular")
                _stored_fx_inr = float(inv.get("fx_rate", 0) or 0)
                _usd_amt_inr   = float(inv.get("amount", 0) or 0)
                _bd_inr        = _inr_parse_bd(inv.get("month_breakdown"))

                if _bd_inr:
                    # ── Multi-month INR invoice / CN ──────────────────────────────
                    # For CNs: ALWAYS recompute INR from usd_amount × fx_rate in breakdown
                    # (stored taxable/total may be in USD if saved by old buggy code)
                    # For regular invoices: use stored taxable split proportionally
                    #   unless the breakdown itself has valid usd+fx (safe to recompute)
                    _bd_has_fx = any(
                        float(b.get("fx_rate", 0) or 0) > 1 and float(b.get("usd_amount", 0) or 0) > 0
                        for b in _bd_inr
                    )
                    _n_bd = len(_bd_inr)
                    # Total USD across all breakdown rows (for proportional fallback)
                    _bd_total_usd = sum(float(b.get("usd_amount", 0) or 0) for b in _bd_inr)

                    for _brow_inr in _bd_inr:
                        _b_month_inr = str(_brow_inr.get("month",""))
                        _b_usd_inr   = float(_brow_inr.get("usd_amount", 0) or 0)
                        _b_fx_inr    = float(_brow_inr.get("fx_rate", _stored_fx_inr) or _stored_fx_inr)
                        _b_inr_stored= float(_brow_inr.get("inr_amount", 0) or 0)

                        if _is_cn and _bd_has_fx and _b_usd_inr > 0 and _b_fx_inr > 1:
                            # ★ CN: recompute everything from USD × FX — ignores wrong stored values
                            _b_taxable_inr = round(_b_usd_inr * _b_fx_inr, 2)
                            _b_igst, _b_cgst, _b_sgst, _b_total_inr = _inr_gst_from_taxable(_b_taxable_inr, _inv_type)
                            _b_inr_equiv   = _b_taxable_inr   # for "Amount (INR equiv.)" col
                        elif _bd_has_fx and _b_usd_inr > 0 and _b_fx_inr > 1:
                            # Regular multi-month INR invoice with FX breakdown: also recompute
                            _b_taxable_inr = round(_b_usd_inr * _b_fx_inr, 2)
                            _b_igst, _b_cgst, _b_sgst, _b_total_inr = _inr_gst_from_taxable(_b_taxable_inr, _inv_type)
                            _b_inr_equiv   = _b_inr_stored if _b_inr_stored > 0 else _b_taxable_inr
                        else:
                            # Legacy: split stored taxable/tax proportionally
                            _r = (_b_usd_inr / _bd_total_usd) if _bd_total_usd > 0 else (1 / _n_bd)
                            _b_taxable_inr = round(_taxable * _r, 2)
                            _b_igst        = round(_igst    * _r, 2)
                            _b_cgst        = round(_cgst    * _r, 2)
                            _b_sgst        = round(_sgst    * _r, 2)
                            _b_total_inr   = round(_total   * _r, 2)
                            _b_inr_equiv   = _b_inr_stored if _b_inr_stored > 0 else round(_b_usd_inr * _b_fx_inr, 2)

                        all_rows.append({
                            "Category":            _cat,
                            "GSTIN":               _gstin,
                            "DSP Name":            _dsp,
                            "Invoice No.":         f"{_inv_no} [{_b_month_inr}]",
                            "Month":               _b_month_inr,
                            "Invoice Date":        _inv_date,
                            "Place of Supply":     _supply,
                            "Tax Rate %":          _tax_rate,
                            "Taxable Value (₹)":   round(_b_taxable_inr, 2),
                            "IGST (₹)":            round(_b_igst,        2),
                            "CGST (₹)":            round(_b_cgst,        2),
                            "SGST (₹)":            round(_b_sgst,        2),
                            "Total Invoice Value (₹)": round(_b_total_inr, 2),
                            "Amount (USD)":        round(_b_usd_inr, 2),
                            "FX Rate":             round(_b_fx_inr,  4) if _b_fx_inr > 0 else "",
                            "Amount (INR equiv.)": round(_b_inr_equiv, 2),
                        })
                else:
                    # ── Single-month INR invoice / CN ─────────────────────────────
                    # For CNs with valid fx_rate: recompute INR from stored USD × FX
                    if _is_cn and _stored_fx_inr > 1 and _usd_amt_inr > 0:
                        _s_taxable = round(_usd_amt_inr * _stored_fx_inr, 2)
                        _s_igst, _s_cgst, _s_sgst, _s_total = _inr_gst_from_taxable(_s_taxable, _inv_type)
                    else:
                        _s_taxable = round(_taxable, 2)
                        _s_igst    = round(_igst,    2)
                        _s_cgst    = round(_cgst,    2)
                        _s_sgst    = round(_sgst,    2)
                        _s_total   = round(_total,   2)
                    all_rows.append({
                        "Category":            _cat,
                        "GSTIN":               _gstin,
                        "DSP Name":            _dsp,
                        "Invoice No.":         ("CN: " if _is_cn else "") + _inv_no,
                        "Month":               inv.get("month",""),
                        "Invoice Date":        _inv_date,
                        "Place of Supply":     _supply,
                        "Tax Rate %":          _tax_rate,
                        "Taxable Value (₹)":   _s_taxable,
                        "IGST (₹)":            _s_igst,
                        "CGST (₹)":            _s_cgst,
                        "SGST (₹)":            _s_sgst,
                        "Total Invoice Value (₹)": _s_total,
                        "Amount (USD)":        _usd_amt_inr,
                        "FX Rate":             round(_stored_fx_inr, 4) if _stored_fx_inr > 0 else "",
                        "Amount (INR equiv.)": round(_s_taxable, 2),
                    })

            import json as _gst_json
            # USD rows — expand multi-month invoices into per-month rows for GST filing
            for _, inv in usd_inv_df.iterrows():
                _dsp = inv.get("dsp_name","")
                if _dsp not in _pcache:
                    _pcache[_dsp] = _fetch_p(_dsp)
                _p       = _pcache[_dsp]
                _gstin   = _p.get("gstin","") or ""
                _supply  = _pos(_p)
                _inv_no  = inv.get("invoice_number","")
                _inv_date= inv.get("invoice_date","")
                try:
                    _inv_date = pd.to_datetime(str(_inv_date), dayfirst=True, errors="coerce").strftime("%d-%b-%Y")
                except Exception:
                    pass
                _amt       = float(inv.get("amount",       0) or 0)
                _stored_fx = float(inv.get("fx_rate",      0) or 0)
                _is_cn     = int(inv.get("is_credit_note", 0) or 0)
                _is_del    = int(inv.get("is_deleted",     0) or 0)
                _cat       = "Credit Note" if _is_cn else ("Deleted" if _is_del else "Export")
                _breakdown_raw = str(inv.get("month_breakdown","") or "")

                if _breakdown_raw and _breakdown_raw.startswith("["):
                    # Multi-month: expand into one row per breakdown entry
                    try:
                        _bd = _gst_json.loads(_breakdown_raw)
                        for _bi, _brow in enumerate(_bd):
                            _b_month  = str(_brow.get("month",""))
                            _b_usd    = float(_brow.get("usd_amount", 0) or 0)
                            _b_fx     = float(_brow.get("fx_rate",    0) or 0)
                            _b_inr    = float(_brow.get("inr_amount", 0) or 0)
                            if _b_inr == 0 and _b_usd > 0 and _b_fx > 0:
                                _b_inr = round(_b_usd * _b_fx, 2)
                            _row_label = f"{_inv_no} [{_b_month}]"
                            all_rows.append({
                                "Category":            _cat,
                                "GSTIN":               _gstin,
                                "DSP Name":            _dsp,
                                "Invoice No.":         _row_label,
                                "Month":               _b_month,
                                "Invoice Date":        _inv_date,
                                "Place of Supply":     _supply,
                                "Tax Rate %":          "0% (Export)",
                                "Taxable Value (₹)":   0.0,
                                "IGST (₹)":            0.0,
                                "CGST (₹)":            0.0,
                                "SGST (₹)":            0.0,
                                "Total Invoice Value (₹)": 0.0,
                                "Amount (USD)":        round(_b_usd, 2),
                                "FX Rate":             _b_fx if _b_fx > 0 else "",
                                "Amount (INR equiv.)": _b_inr,
                            })
                    except Exception:
                        # Fallback to single row if JSON parse fails
                        _inr_equiv = round(_amt * _stored_fx, 2) if _stored_fx > 0 else 0.0
                        all_rows.append({
                            "Category":            _cat,
                            "GSTIN":               _gstin,
                            "DSP Name":            _dsp,
                            "Invoice No.":         ("CN: " if _is_cn else "") + _inv_no,
                            "Month":               inv.get("month",""),
                            "Invoice Date":        _inv_date,
                            "Place of Supply":     _supply,
                            "Tax Rate %":          "0% (Export)",
                            "Taxable Value (₹)":   0.0,
                            "IGST (₹)":            0.0,
                            "CGST (₹)":            0.0,
                            "SGST (₹)":            0.0,
                            "Total Invoice Value (₹)": 0.0,
                            "Amount (USD)":        round(_amt, 2),
                            "FX Rate":             _stored_fx if _stored_fx > 0 else "",
                            "Amount (INR equiv.)": _inr_equiv,
                        })
                else:
                    # Single-month invoice
                    _inr_equiv = round(_amt * _stored_fx, 2) if _stored_fx > 0 else 0.0
                    all_rows.append({
                        "Category":            _cat,
                        "GSTIN":               _gstin,
                        "DSP Name":            _dsp,
                        "Invoice No.":         ("CN: " if _is_cn else "") + _inv_no,
                        "Month":               inv.get("month",""),
                        "Invoice Date":        _inv_date,
                        "Place of Supply":     _supply,
                        "Tax Rate %":          "0% (Export)",
                        "Taxable Value (₹)":   0.0,
                        "IGST (₹)":            0.0,
                        "CGST (₹)":            0.0,
                        "SGST (₹)":            0.0,
                        "Total Invoice Value (₹)": 0.0,
                        "Amount (USD)":        round(_amt, 2),
                        "FX Rate":             _stored_fx if _stored_fx > 0 else "",
                        "Amount (INR equiv.)": _inr_equiv,
                    })

            combined_df = pd.DataFrame(all_rows)

            # Tabs for INR / USD / All
            _gtab1, _gtab2, _gtab3 = st.tabs(["🇮🇳 INR Invoices", "🌐 Export (USD)", "📋 All Combined"])

            from st_aggrid import AgGrid, GridOptionsBuilder, JsCode

            def _make_gst_grid(df, height=380):
                if df.empty:
                    st.info("No records.")
                    return
                # Ensure FX Rate column is numeric (replace "" with 0)
                if "FX Rate" in df.columns:
                    df = df.copy()
                    df["FX Rate"] = pd.to_numeric(df["FX Rate"], errors="coerce").fillna(0)
                _inr_fmt = JsCode("""function(p){
                    if(p.value==null||p.value===''||isNaN(parseFloat(p.value))||parseFloat(p.value)===0)return'-';
                    return '₹'+parseFloat(p.value).toLocaleString(undefined,{minimumFractionDigits:2});
                }""")
                _usd_fmt = JsCode("""function(p){
                    if(p.value==null||p.value===''||isNaN(parseFloat(p.value))||parseFloat(p.value)===0)return'-';
                    return '$'+parseFloat(p.value).toLocaleString(undefined,{minimumFractionDigits:2});
                }""")
                _cat_style = JsCode("""function(p){
                    if(!p.value)return{};
                    const v=p.value.toLowerCase();
                    if(v==='export') return{color:'#0277BD',fontWeight:'bold'};
                    if(v==='regular')return{color:'#2E7D32',fontWeight:'bold'};
                    if(v==='credit note')return{color:'#6A1B9A',fontWeight:'bold'};
                    if(v==='deleted')return{color:'#9E9E9E',fontWeight:'bold'};
                    return{};
                }""")
                gb = GridOptionsBuilder.from_dataframe(df)
                for _c in ["Taxable Value (₹)","IGST (₹)","CGST (₹)","SGST (₹)","Total Invoice Value (₹)","Amount (INR equiv.)"]:
                    if _c in df.columns:
                        gb.configure_column(_c, type=["numericColumn"], valueFormatter=_inr_fmt, minWidth=120)
                for _c in ["Amount (USD)"]:
                    if _c in df.columns:
                        gb.configure_column(_c, type=["numericColumn"], valueFormatter=_usd_fmt, minWidth=115)
                if "Category" in df.columns:
                    gb.configure_column("Category", cellStyle=_cat_style, minWidth=110)
                if "DSP Name" in df.columns:
                    gb.configure_column("DSP Name", minWidth=130, pinned="left")
                if "Invoice No." in df.columns:
                    gb.configure_column("Invoice No.", minWidth=130, pinned="left")
                if "GSTIN" in df.columns:
                    gb.configure_column("GSTIN", minWidth=150)
                if "Month" in df.columns:
                    gb.configure_column("Month", minWidth=100)
                if "Invoice Date" in df.columns:
                    gb.configure_column("Invoice Date", minWidth=110)
                if "Place of Supply" in df.columns:
                    gb.configure_column("Place of Supply", minWidth=140)
                if "Tax Rate %" in df.columns:
                    gb.configure_column("Tax Rate %", minWidth=130)
                if "FX Rate" in df.columns:
                    gb.configure_column("FX Rate", minWidth=90)
                gb.configure_default_column(resizable=True, sortable=True, filter=True)

                # Numeric sum columns — FX Rate is a rate not an amount, exclude from sum
                _num_cols = [c for c in ["Taxable Value (₹)","IGST (₹)","CGST (₹)","SGST (₹)",
                                         "Total Invoice Value (₹)","Amount (USD)","Amount (INR equiv.)"]
                             if c in df.columns]
                _tot = {col: "" for col in df.columns}
                _tot["DSP Name"]    = "TOTAL"
                _tot["Invoice No."] = f"{len(df)} records"
                if "FX Rate" in df.columns:
                    _tot["FX Rate"] = ""   # Don't sum FX rates
                for _tc in _num_cols:
                    _tot[_tc] = round(df[_tc].sum(), 2)
                opts = gb.build()
                opts["pinnedBottomRowData"] = [_tot]
                opts["getRowStyle"] = JsCode("""function(p){if(p.node.rowPinned){return{backgroundColor:'#003366',color:'white',fontWeight:'bold'}}}""")
                _css = {
                    ".ag-header":{"background-color":"#003366 !important","color":"white !important","font-weight":"bold !important"},
                    ".ag-header-cell-label":{"color":"white !important","font-weight":"bold !important"}
                }
                AgGrid(df, gridOptions=opts, allow_unsafe_jscode=True,
                       fit_columns_on_grid_load=False, height=height, custom_css=_css)

            with _gtab1:
                # INR tab: show FX Rate (used for this invoice) but not Amount (INR equiv.) — Total (₹) is the INR value
                _inr_disp = combined_df[combined_df["Tax Rate %"] != "0% (Export)"].drop(
                    columns=["Amount (USD)","Amount (INR equiv.)"], errors="ignore"
                ).copy()
                _make_gst_grid(_inr_disp)

            with _gtab2:
                _usd_disp = combined_df[combined_df["Tax Rate %"] == "0% (Export)"].drop(
                    columns=["Taxable Value (₹)","IGST (₹)","CGST (₹)","SGST (₹)","Total Invoice Value (₹)"],
                    errors="ignore"
                ).copy()
                # Per-row FX: use stored month-end rate; fallback only for old invoices (fx=0 or 1)
                if not _usd_disp.empty:
                    _fx_usd_fallback = st.number_input(
                        "Fallback FX Rate (INR/USD) — for old invoices with no stored rate",
                        min_value=1.0, value=87.50, step=0.01, format="%.4f",
                        key="gst_usd_fx",
                        help="Applied only to old invoices without a stored month-end rate. New invoices use their own month-specific rate automatically."
                    )
                    def _apply_fx(row):
                        _fx = float(row.get("FX Rate") or 0)
                        # fx <= 1: not set (0) or old USD default (1) → use fallback
                        return round(_fx if _fx > 1 else _fx_usd_fallback, 4)
                    _usd_disp = _usd_disp.copy()
                    _usd_disp["FX Rate"] = _usd_disp.apply(_apply_fx, axis=1)
                    _usd_disp["Amount (INR equiv.)"] = (
                        _usd_disp["Amount (USD)"] * _usd_disp["FX Rate"]
                    ).round(2)
                    st.caption("💡 Each row uses its month-end FX rate (stored at invoice creation). Rows using the fallback above are old invoices — re-raise or Edit to update.")
                    _make_gst_grid(_usd_disp)

            with _gtab3:
                # Combined: apply FX rate for all rows
                _fx_fallback_all = float(st.session_state.get("gst_usd_fx", 87.50) or 87.50)
                _combined_all = combined_df.copy()
                for _idx, _row in _combined_all.iterrows():
                    _sfx = float(_row.get("FX Rate") or 0)
                    if _row["Tax Rate %"] == "0% (Export)":
                        # USD row: use stored FX if > 1, else fallback
                        _sfx = _sfx if _sfx > 1 else _fx_fallback_all
                        _combined_all.at[_idx, "FX Rate"] = round(_sfx, 4)
                        _combined_all.at[_idx, "Amount (INR equiv.)"] = round(
                            float(_row.get("Amount (USD)", 0) or 0) * _sfx, 2
                        )
                    else:
                        # INR row: FX rate stored in DB (from when invoice was raised)
                        # Amount (INR equiv.) = total_amount (already INR)
                        # Show the stored FX rate if available
                        _combined_all.at[_idx, "FX Rate"] = round(_sfx, 4) if _sfx > 0 else ""
                        # INR equiv already set to total_amount in row build; keep it
                _make_gst_grid(_combined_all, height=440)

            st.markdown("---")
            st.markdown("##### 📥 Export GST Report")
            _per_lbl = (
                f"{gst_dt_from.strftime('%d%b%Y')}_to_{gst_dt_to.strftime('%d%b%Y')}"
                if (gst_dt_from and gst_dt_to) else datetime.today().strftime("%b%Y")
            )

            _inr_export = combined_df[combined_df["Tax Rate %"] != "0% (Export)"].drop(
                columns=["Amount (USD)","FX Rate","Amount (INR equiv.)"], errors="ignore"
            ).copy()
            _usd_export = combined_df[combined_df["Tax Rate %"] == "0% (Export)"].drop(
                columns=["Taxable Value (₹)","IGST (₹)","CGST (₹)","SGST (₹)","Total Invoice Value (₹)"],
                errors="ignore"
            ).copy()

            # ── Fix Export Detail FX: use per-row stored rate, fallback only when not set ──
            _fx_e = float(st.session_state.get("gst_usd_fx", 87.50) or 87.50)
            def _fix_usd_fx(row):
                _rfx = float(row.get("FX Rate") or 0)
                return round(_rfx if _rfx > 1 else _fx_e, 4)
            if not _usd_export.empty:
                _usd_export["FX Rate"] = _usd_export.apply(_fix_usd_fx, axis=1)
                _usd_export["Amount (INR equiv.)"] = (
                    _usd_export["Amount (USD)"] * _usd_export["FX Rate"]
                ).round(2)

            # ── GSTR Summary — INR types + Export + Credit Notes ─────────────────
            _gstr_rows = []

            # INR: Intra/Inter-State (only non-CN rows)
            _inr_reg = gst_inv_df[gst_inv_df.get("is_credit_note", pd.Series(0, index=gst_inv_df.index)).fillna(0).astype(int) == 0] if not gst_inv_df.empty else pd.DataFrame()
            for _itype, _label in [("INR_CGST_SGST","Intra-State (CGST+SGST)"),
                                    ("INR_IGST",     "Inter-State (IGST)")]:
                _sub = _inr_reg[_inr_reg["invoice_type"] == _itype] if not _inr_reg.empty else pd.DataFrame()
                if _sub.empty:
                    continue
                _stax = float(_sub["taxable_amount"].fillna(0).sum())
                _stot = float(_sub["total_amount"].fillna(0).sum())
                if _stax == 0 and _stot > 0:
                    _stax = round(_stot / 1.18, 2)
                _gstr_rows.append({
                    "GST Type":            _label,
                    "No. of Invoices":     len(_sub),
                    "No. of Credit Notes": 0,
                    "Taxable Value (INR)": round(_stax, 2),
                    "IGST (INR)":          round(float(_sub["igst_amount"].fillna(0).sum()), 2),
                    "CGST (INR)":          round(float(_sub["cgst_amount"].fillna(0).sum()), 2),
                    "SGST (INR)":          round(float(_sub["sgst_amount"].fillna(0).sum()), 2),
                    "Total Invoice (INR)": round(_stot, 2),
                    "Amount (USD)":        "",
                    "Amount (INR equiv.)": "",
                })

            # INR Credit Notes
            _inr_cn = gst_inv_df[gst_inv_df.get("is_credit_note", pd.Series(0, index=gst_inv_df.index)).fillna(0).astype(int) == 1] if not gst_inv_df.empty else pd.DataFrame()
            for _itype, _label in [("INR_CGST_SGST","Credit Notes – Intra-State (CGST+SGST)"),
                                    ("INR_IGST",     "Credit Notes – Inter-State (IGST)")]:
                _sub = _inr_cn[_inr_cn["invoice_type"] == _itype] if not _inr_cn.empty else pd.DataFrame()
                if _sub.empty:
                    continue
                # Recompute from breakdown (same logic as GST display renderer)
                _cn_taxable_sum = 0.0; _cn_igst_sum = 0.0; _cn_cgst_sum = 0.0
                _cn_sgst_sum = 0.0; _cn_total_sum = 0.0
                for _, _cr in _sub.iterrows():
                    _cr_bd   = _inr_parse_bd(_cr.get("month_breakdown"))
                    _cr_fx   = float(_cr.get("fx_rate", 0) or 0)
                    _cr_usd  = float(_cr.get("amount", 0) or 0)
                    _cr_type = str(_cr.get("invoice_type",""))
                    _cr_bd_has_fx = any(float(b.get("fx_rate",0) or 0) > 1 and float(b.get("usd_amount",0) or 0) > 0 for b in _cr_bd)
                    if _cr_bd and _cr_bd_has_fx:
                        for _b in _cr_bd:
                            _bt = round(float(_b.get("usd_amount",0) or 0) * float(_b.get("fx_rate",_cr_fx) or _cr_fx), 2)
                            _bi, _bc, _bs, _btot = _inr_gst_from_taxable(_bt, _cr_type)
                            _cn_taxable_sum += _bt; _cn_igst_sum += _bi
                            _cn_cgst_sum += _bc; _cn_sgst_sum += _bs; _cn_total_sum += _btot
                    elif _cr_fx > 1 and _cr_usd > 0:
                        _bt = round(_cr_usd * _cr_fx, 2)
                        _bi, _bc, _bs, _btot = _inr_gst_from_taxable(_bt, _cr_type)
                        _cn_taxable_sum += _bt; _cn_igst_sum += _bi
                        _cn_cgst_sum += _bc; _cn_sgst_sum += _bs; _cn_total_sum += _btot
                    else:
                        _cn_taxable_sum += float(_cr.get("taxable_amount",0) or 0)
                        _cn_igst_sum    += float(_cr.get("igst_amount",    0) or 0)
                        _cn_cgst_sum    += float(_cr.get("cgst_amount",    0) or 0)
                        _cn_sgst_sum    += float(_cr.get("sgst_amount",    0) or 0)
                        _cn_total_sum   += float(_cr.get("total_amount",   0) or 0)
                _gstr_rows.append({
                    "GST Type":            _label,
                    "No. of Invoices":     0,
                    "No. of Credit Notes": len(_sub),
                    "Taxable Value (INR)": round(_cn_taxable_sum, 2),
                    "IGST (INR)":          round(_cn_igst_sum, 2),
                    "CGST (INR)":          round(_cn_cgst_sum, 2),
                    "SGST (INR)":          round(_cn_sgst_sum, 2),
                    "Total Invoice (INR)": round(_cn_total_sum, 2),
                    "Amount (USD)":        "",
                    "Amount (INR equiv.)": "",
                })

            # ── Helper: strip [Mon-YYYY] suffix to get base invoice number ──
            import re as _re_ai
            def _base_inv(inv_str):
                """Strip ' [Mon-YYYY]' suffix → base invoice number."""
                return _re_ai.sub(r"\s*\[.*?\]$", "", str(inv_str)).replace("CN: ","").strip()

            # Export (USD) invoices & CNs
            if not _usd_export.empty:
                _exp_reg = _usd_export[_usd_export["Category"] != "Credit Note"] if "Category" in _usd_export.columns else _usd_export
                _exp_cn  = _usd_export[_usd_export["Category"] == "Credit Note"] if "Category" in _usd_export.columns else pd.DataFrame()
                _e_usd   = float(_exp_reg["Amount (USD)"].fillna(0).sum())
                _e_inr   = float(_exp_reg["Amount (INR equiv.)"].fillna(0).sum())
                if not _exp_reg.empty:
                    # Count unique base invoice numbers (not split rows)
                    _exp_reg_unique = _exp_reg["Invoice No."].apply(_base_inv).nunique()
                    _gstr_rows.append({
                        "GST Type":            "Export – 0% (USD Invoices)",
                        "No. of Invoices":     _exp_reg_unique,
                        "No. of Credit Notes": 0,
                        "Taxable Value (INR)": "",
                        "IGST (INR)":          "",
                        "CGST (INR)":          "",
                        "SGST (INR)":          "",
                        "Total Invoice (INR)": "",
                        "Amount (USD)":        round(_e_usd, 2),
                        "Amount (INR equiv.)": round(_e_inr, 2),
                    })
                if not _exp_cn.empty:
                    _ec_usd = float(_exp_cn["Amount (USD)"].fillna(0).sum())
                    _ec_inr = float(_exp_cn["Amount (INR equiv.)"].fillna(0).sum())
                    _exp_cn_unique = _exp_cn["Invoice No."].apply(_base_inv).nunique()
                    _gstr_rows.append({
                        "GST Type":            "Credit Notes – Export (USD)",
                        "No. of Invoices":     0,
                        "No. of Credit Notes": _exp_cn_unique,
                        "Taxable Value (INR)": "",
                        "IGST (INR)":          "",
                        "CGST (INR)":          "",
                        "SGST (INR)":          "",
                        "Total Invoice (INR)": "",
                        "Amount (USD)":        round(_ec_usd, 2),
                        "Amount (INR equiv.)": round(_ec_inr, 2),
                    })

            # ── "All Invoice" tab: one row per actual invoice/CN (no month splits) ──
            # Group combined_df rows by the base invoice number (strip [Month] suffix),
            # sum all financial columns, compute weighted FX = INR equiv / USD.

            _ai_cols = [
                "Category","GSTIN","DSP Name","Invoice No.","Month","Invoice Date",
                "Place of Supply","Tax Rate %",
                "Taxable Value (₹)","IGST (₹)","CGST (₹)","SGST (₹)","Total Invoice Value (₹)",
                "Amount (USD)","FX Rate","Amount (INR equiv.)"
            ]
            # Use the fully FX-resolved combined_df (_combined_all from display, or rebuild here)
            _ai_src = combined_df.copy()
            # Apply FX resolution same as All Combined tab
            for _idx2, _row2 in _ai_src.iterrows():
                _sfx2 = float(_row2.get("FX Rate") or 0)
                if _row2["Tax Rate %"] == "0% (Export)":
                    _sfx2 = _sfx2 if _sfx2 > 1 else _fx_e
                    _ai_src.at[_idx2, "FX Rate"] = round(_sfx2, 4)
                    _ai_src.at[_idx2, "Amount (INR equiv.)"] = round(
                        float(_row2.get("Amount (USD)", 0) or 0) * _sfx2, 2
                    )

            _ai_src["_base_inv"] = _ai_src["Invoice No."].apply(_base_inv)

            _num_cols  = ["Taxable Value (₹)","IGST (₹)","CGST (₹)","SGST (₹)",
                          "Total Invoice Value (₹)","Amount (USD)","Amount (INR equiv.)"]

            _all_inv_rows = []
            for _base, _grp in _ai_src.groupby("_base_inv", sort=False):
                _first = _grp.iloc[0]
                # Months: join unique months (original, not expanded)
                _months_ai = ", ".join(_grp["Month"].dropna().unique().tolist())
                # Sum numeric columns
                _sums = {c: round(float(_grp[c].fillna(0).sum()), 2)
                         if c in _grp.columns else 0.0 for c in _num_cols}
                # Weighted FX = INR equiv / USD (if USD > 0)
                _ai_usd = _sums["Amount (USD)"]
                _ai_inr = _sums["Amount (INR equiv.)"]
                _ai_fx  = round(_ai_inr / _ai_usd, 4) if _ai_usd > 0 else ""
                _all_inv_rows.append({
                    "Category":                str(_first.get("Category","")),
                    "GSTIN":                   str(_first.get("GSTIN","")),
                    "DSP Name":                str(_first.get("DSP Name","")),
                    "Invoice No.":             _base,
                    "Month":                   _months_ai,
                    "Invoice Date":            str(_first.get("Invoice Date","")),
                    "Place of Supply":         str(_first.get("Place of Supply","")),
                    "Tax Rate %":              str(_first.get("Tax Rate %","")),
                    "Taxable Value (₹)":       _sums["Taxable Value (₹)"],
                    "IGST (₹)":                _sums["IGST (₹)"],
                    "CGST (₹)":                _sums["CGST (₹)"],
                    "SGST (₹)":                _sums["SGST (₹)"],
                    "Total Invoice Value (₹)": _sums["Total Invoice Value (₹)"],
                    "Amount (USD)":            _ai_usd if _ai_usd > 0 else "",
                    "FX Rate":                 _ai_fx,
                    "Amount (INR equiv.)":     _ai_inr if _ai_inr > 0 else "",
                })
            _all_inv_df = pd.DataFrame(_all_inv_rows)

            exp1, exp2, exp3 = st.columns(3)

            with exp1:
                _xbuf = io.BytesIO()
                with pd.ExcelWriter(_xbuf, engine="openpyxl") as _ew:
                    if not _inr_export.empty:
                        _inr_export.to_excel(_ew, index=False, sheet_name="INR Invoices")
                    if not _usd_export.empty:
                        _usd_export.to_excel(_ew, index=False, sheet_name="Export (USD)")
                    combined_df.to_excel(_ew, index=False, sheet_name="All Combined")
                    try:
                        from openpyxl.styles import Font, PatternFill, Alignment
                        _hfill = PatternFill("solid", fgColor="003366")
                        _hfont = Font(bold=True, color="FFFFFF")
                        for _sn in _ew.sheets:
                            _ws = _ew.sheets[_sn]
                            for _cell in _ws[1]:
                                _cell.font = _hfont; _cell.fill = _hfill
                                _cell.alignment = Alignment(horizontal="center")
                            for _cc in _ws.columns:
                                _ws.column_dimensions[_cc[0].column_letter].width = min(
                                    max(len(str(_c.value or "")) for _c in _cc) + 4, 45)
                    except Exception:
                        pass
                _xbuf.seek(0)
                st.download_button("📊 Download Excel (.xlsx)", data=_xbuf,
                    file_name=f"GST_Report_{_per_lbl}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="gst_excel_dl")

            with exp2:
                _csv_data = _inr_export.to_csv(index=False).encode("utf-8-sig") if not _inr_export.empty else b""
                if _csv_data:
                    st.download_button("📄 Download INR CSV", data=_csv_data,
                        file_name=f"GST_INR_{_per_lbl}.csv", mime="text/csv", key="gst_csv_dl")
                else:
                    st.caption("No INR data.")

            with exp3:
                if _gstr_rows or not _all_inv_df.empty:
                    _gbuf = io.BytesIO()
                    with pd.ExcelWriter(_gbuf, engine="openpyxl") as _ew2:
                        # Sheet 1: GSTR Summary
                        if _gstr_rows:
                            pd.DataFrame(_gstr_rows).to_excel(_ew2, index=False, sheet_name="GSTR Summary")
                        # Sheet 2: INR Detail
                        if not _inr_export.empty:
                            _inr_export.to_excel(_ew2, index=False, sheet_name="INR Detail")
                        # Sheet 3: Export Detail — per-row FX (not flat fallback)
                        if not _usd_export.empty:
                            _usd_export.to_excel(_ew2, index=False, sheet_name="Export Detail")
                        # Sheet 4: All Invoice — one row per actual invoice/CN
                        if not _all_inv_df.empty:
                            _all_inv_df.to_excel(_ew2, index=False, sheet_name="All Invoice")

                        # Style all sheets
                        try:
                            from openpyxl.styles import Font, PatternFill, Alignment, numbers
                            _hfill2 = PatternFill("solid", fgColor="003366")
                            _hfont2 = Font(bold=True, color="FFFFFF")
                            _cn_fill = PatternFill("solid", fgColor="FFF3CD")   # light amber for CN rows
                            for _sn2 in _ew2.sheets:
                                _ws2 = _ew2.sheets[_sn2]
                                # Header row
                                for _cell2 in _ws2[1]:
                                    _cell2.font  = _hfont2
                                    _cell2.fill  = _hfill2
                                    _cell2.alignment = Alignment(horizontal="center", wrap_text=True)
                                # Data rows
                                for _rw2 in _ws2.iter_rows(min_row=2):
                                    for _c2 in _rw2:
                                        _c2.alignment = Alignment(horizontal="right" if isinstance(_c2.value, (int, float)) else "left")
                                    # Highlight Credit Note rows amber
                                    if _rw2[0].value and "Credit Note" in str(_rw2[0].value):
                                        for _c2 in _rw2:
                                            _c2.fill = _cn_fill
                                # Auto-width
                                for _cc2 in _ws2.columns:
                                    _ws2.column_dimensions[_cc2[0].column_letter].width = min(
                                        max((len(str(_c2.value or "")) for _c2 in _cc2), default=8) + 3, 40
                                    )
                        except Exception:
                            pass
                    _gbuf.seek(0)
                    _gst_dl_col, _gst_email_col = st.columns([1, 1])
                    with _gst_dl_col:
                        st.download_button("📑 Download GSTR Summary", data=_gbuf,
                            file_name=f"GSTR_Summary_{_per_lbl}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="gst_gstr_dl")
                    with _gst_email_col:
                        if st.button("📤 Send GSTR via Email", key="gst_send_email_btn",
                                     type="primary", use_container_width=True):
                            st.session_state["_gst_email_panel_open"] = True
                            st.session_state["_gst_email_buf"]   = _gbuf.getvalue()
                            st.session_state["_gst_email_fname"] = f"GSTR_Summary_{_per_lbl}.xlsx"
                            st.session_state["_gst_email_per"]   = _per_lbl
                            st.rerun()

                    # ── Modal dialog ──────────────────────────────────────────
                    @st.dialog("📤 Send GSTR Summary via Email", width="large")
                    def _gst_email_dialog():
                        _per   = st.session_state.get("_gst_email_per", _per_lbl)
                        _subj_default = f"GSTR Summary — {_per} | {COMPANY['name']}"
                        _body_default = (
                            f"<p>Dear Team,</p>"
                            f"<p>Please find attached the <b>GSTR Summary Report</b> "
                            f"for <b>{_per}</b>.</p>"
                            f"<p>The report includes INR Invoices, Export (USD) Invoices, "
                            f"Credit Notes, and a consolidated All Invoice view.</p>"
                            f"<br/><p>Regards,<br/>"
                            f"<b>Finance Team — {COMPANY['name']}</b></p>"
                        )

                        d_c1, d_c2 = st.columns([1, 1])
                        with d_c1:
                            _d_to = st.text_input(
                                "📧 To (comma separated)",
                                value=COMPANY.get("email",""), key="d_gst_to"
                            )
                        with d_c2:
                            _d_subj = st.text_input(
                                "📌 Subject", value=_subj_default, key="d_gst_subj"
                            )

                        with st.expander("📝 Email Body Preview / Edit", expanded=False):
                            st.markdown(_body_default, unsafe_allow_html=True)
                            st.divider()
                            st.text_area("Edit HTML body (optional)",
                                         value=_body_default, height=130, key="d_gst_body")
                        _d_body_send = st.session_state.get("d_gst_body", _body_default)

                        if _smtp_configured_via_secrets():
                            st.success("✅ Email credentials loaded from secrets — ready to send.", icon="🔐")
                            _dh, _dp, _du, _dpass = _get_smtp_creds()
                        else:
                            with st.expander("⚙️ SMTP Settings", expanded=False):
                                _sc1, _sc2 = st.columns(2)
                                with _sc1:
                                    st.text_input("SMTP Host",  value="smtpout.secureserver.net", key="d_ge_host")
                                    st.text_input("From Email", value="finance@peakmyads.com",    key="d_ge_user")
                                with _sc2:
                                    st.number_input("SMTP Port", value=465, key="d_ge_port")
                                    st.text_input("Password", type="password", key="d_ge_pass")
                            _dh    = st.session_state.get("d_ge_host","smtpout.secureserver.net")
                            _dp    = int(st.session_state.get("d_ge_port", 465))
                            _du    = st.session_state.get("d_ge_user","finance@peakmyads.com")
                            _dpass = st.session_state.get("d_ge_pass","")

                        st.divider()
                        db1, db2 = st.columns([1, 1])
                        with db1:
                            if st.button("📤 Send Now", type="primary",
                                         use_container_width=True, key="d_gst_send"):
                                _d_to_list = [e.strip() for e in _d_to.split(",") if e.strip()]
                                if not _d_to_list:
                                    st.error("Enter at least one recipient.")
                                elif not _dpass:
                                    st.error("SMTP password not set.")
                                else:
                                    try:
                                        _dm = MIMEMultipart()
                                        _dm["From"]    = _du
                                        _dm["To"]      = ", ".join(_d_to_list)
                                        _dm["Bcc"]     = BCC_EMAIL
                                        _dm["Subject"] = st.session_state.get("d_gst_subj", _subj_default)
                                        _dm.attach(MIMEText(_d_body_send, "html"))
                                        _da = MIMEBase("application",
                                            "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                                        _da.set_payload(st.session_state["_gst_email_buf"])
                                        encoders.encode_base64(_da)
                                        _da.add_header("Content-Disposition",
                                            f'attachment; filename="{st.session_state["_gst_email_fname"]}"')
                                        _dm.attach(_da)
                                        _smtp_send(_dh, int(_dp), _du, _dpass,
                                                   _d_to_list + [BCC_EMAIL], _dm.as_string())
                                        st.success(f"✅ GSTR Summary sent to {', '.join(_d_to_list)}!")
                                        st.session_state["_gst_email_panel_open"] = False
                                    except Exception as _de:
                                        st.error(f"Email failed: {_de}")
                        with db2:
                            if st.button("✖ Cancel", use_container_width=True, key="d_gst_cancel"):
                                st.session_state["_gst_email_panel_open"] = False
                                st.rerun()

                    if st.session_state.get("_gst_email_panel_open"):
                        _gst_email_dialog()
                else:
                    st.caption("No GSTR summary available.")
