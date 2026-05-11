"""
Revenue Tracker Software
Author: Sumit
Description:
Revenue, Cost, Billing & Collection Tracker
"""

from login import login_screen, get_allowed_tabs, admin_change_password
from invoice_module import render_invoice_module
from navbar import render_navbar
from dashboard_module import render_dashboard_tab
from ageing_module import render_ageing_tab
from bc_report_module import render_bc_report_tab
import streamlit as st
import pandas as pd
import os
from datetime import datetime
import traceback
import io

from st_aggrid import AgGrid, GridOptionsBuilder, JsCode
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Table
from reportlab.lib import utils
from st_aggrid import JsCode

import base64

def get_image_base64(path):
    with open(path, "rb") as img_file:
        return base64.b64encode(img_file.read()).decode()


import dropbox

def get_dropbox_client():
    import dropbox

    APP_KEY = "xn40ddpn2ow57xg"
    APP_SECRET = "9n8shqjaywwigor"
    REFRESH_TOKEN = "z8O-84xJC_4AAAAAAAAAAdqrxTIhS0ruI0K2BlFDkUUWu0SZzRwZaz3g34uw-DUf"

    return dropbox.Dropbox(
        oauth2_refresh_token=REFRESH_TOKEN,
        app_key=APP_KEY,
        app_secret=APP_SECRET
    )
    
    return dbx
# -------------------------------
# GOOGLE SHEETS CONNECTION (GLOBAL)
# -------------------------------
import sqlite3
import pandas as pd
import streamlit as st
import sys, os

# ==========================================
# PORTABLE DB PATH (works as .py and .exe)
# ==========================================
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

DB_PATH = os.path.join(BASE_DIR, "tracker.db")

def get_db_connection():
    return sqlite3.connect(DB_PATH, check_same_thread=False)

# ==========================================
# AUTO-CREATE DB TABLES ON FIRST RUN
# ==========================================
def init_db():
    conn = get_db_connection()
    cur = conn.cursor()

    cur.executescript("""
        CREATE TABLE IF NOT EXISTS master_data (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            month         TEXT,
            partner_name  TEXT,
            dsp_amount    REAL DEFAULT 0,
            ssp_amount    REAL DEFAULT 0,
            c_dsp         REAL DEFAULT 0,
            c_ssp         REAL DEFAULT 0
        );
        
        -- ✅ ADD THIS BLOCK
        CREATE TABLE IF NOT EXISTS login_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT,
            login_time TEXT
        );

        CREATE TABLE IF NOT EXISTS dsp_data (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            month           TEXT,
            dsp_name        TEXT,
            receivable      REAL DEFAULT 0,
            due_date        TEXT,
            received_date   TEXT,
            received_amount REAL DEFAULT 0,
            received_in     TEXT,
            reason          TEXT,
            UNIQUE(month, dsp_name)
        );

        CREATE TABLE IF NOT EXISTS ssp_data (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            month        TEXT,
            ssp_name     TEXT,
            payable      REAL DEFAULT 0,
            due_date     TEXT,
            payment_date TEXT,
            paid_amount  REAL DEFAULT 0,
            paid_by      TEXT,
            reason       TEXT,
            inv_status   TEXT DEFAULT ''
        );

        CREATE UNIQUE INDEX IF NOT EXISTS idx_ssp_month_name
            ON ssp_data(month, ssp_name);

        CREATE TABLE IF NOT EXISTS partner_list (
            id                          INTEGER PRIMARY KEY AUTOINCREMENT,
            agreement_date              TEXT,
            legal_name                  TEXT,
            short_name                  TEXT,
            address                     TEXT,
            country                     TEXT,
            entity_type                 TEXT,
            gstin                       TEXT,
            payment_terms               TEXT,
            contact_person              TEXT,
            designation                 TEXT,
            contact_no                  TEXT,
            email1                      TEXT,
            email2                      TEXT,
            email3                      TEXT,
            finance_contact             TEXT,
            finance_email               TEXT
        );

        CREATE TABLE IF NOT EXISTS cost_centre (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            category       TEXT,
            cost_name      TEXT,
            sub_cost       TEXT,
            financial_year TEXT,
            month          TEXT,
            currency       TEXT,
            amount_usd     REAL DEFAULT 0,
            fx_rate        REAL DEFAULT 0,
            amount_inr     REAL DEFAULT 0
        );
    """)

    conn.commit()
    conn.close()

# Run on every startup — safe because of IF NOT EXISTS
init_db()

# ==========================================
# MIGRATION: add inv_status column if missing
# ==========================================
def migrate_add_inv_status():
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("ALTER TABLE ssp_data ADD COLUMN inv_status TEXT DEFAULT ''")
        conn.commit()
    except Exception:
        pass  # column already exists
    conn.close()

migrate_add_inv_status()

# ==========================================
# SSP INVOICE: UPDATE STATUS IN DB
# ==========================================
def update_ssp_inv_status(ssp_name: str, month: str, status: str):
    """Set inv_status for a specific (month, ssp_name) row."""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "UPDATE ssp_data SET inv_status = ? WHERE ssp_name = ? AND month = ?",
        (status, ssp_name, month)
    )
    if cur.rowcount == 0:
        # Row doesn't exist yet — insert a minimal placeholder
        cur.execute(
            "INSERT OR IGNORE INTO ssp_data (month, ssp_name, inv_status) VALUES (?, ?, ?)",
            (month, ssp_name, status)
        )
    conn.commit()
    conn.close()

# ==========================================
# SSP INVOICE: UPLOAD TO DROPBOX
# ==========================================
def upload_invoice_to_dropbox(file_bytes: bytes, ssp_name: str, month: str, file_ext: str) -> str:
    """
    Upload invoice to Dropbox under /SSP Invoice/<ssp_name>/
    Filename: <ssp_name>_<month>_<upload_date>.<ext>
    Returns the Dropbox path.
    """
    import dropbox as _dropbox

    dbx = get_dropbox_client()
    safe_name  = ssp_name.strip().replace("/", "-").replace("\\", "-")
    folder_path = f"/SSP Invoice/{safe_name}"

    # Create folder if it doesn't exist
    try:
        dbx.files_get_metadata(folder_path)
    except _dropbox.exceptions.ApiError:
        try:
            dbx.files_create_folder_v2(folder_path)
        except Exception:
            pass  # race condition – folder created between check & create

    upload_date  = datetime.now().strftime("%d%b%Y")
    safe_month   = month.replace("-", "")          # e.g. Mar2025
    filename     = f"{safe_name}_{safe_month}_{upload_date}.{file_ext}"
    dropbox_path = f"{folder_path}/{filename}"

    dbx.files_upload(
        file_bytes,
        dropbox_path,
        mode=_dropbox.files.WriteMode.overwrite
    )
    return dropbox_path

# ---------------- SSP UPSERT ---------------- #

def upsert_ssp_data(df: pd.DataFrame):
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Ensure UNIQUE constraint exists on (month, ssp_name)
    try:
        cur.execute("""
            CREATE UNIQUE INDEX IF NOT EXISTS idx_ssp_month_name
            ON ssp_data(month, ssp_name)
        """)
        conn.commit()
    except Exception:
        pass  # index may already exist

    for _, row in df.iterrows():
        due_date     = pd.to_datetime(row["Due Date"],     errors="coerce")
        payment_date = pd.to_datetime(row["Payment Date"], errors="coerce")

        cur.execute("""
        INSERT INTO ssp_data (
            month, ssp_name, payable,
            due_date, payment_date, paid_amount,
            paid_from, reason
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(month, ssp_name)
        DO UPDATE SET
            payment_date=excluded.payment_date,
            paid_amount=excluded.paid_amount,
            paid_from=excluded.paid_from,
            reason=excluded.reason
        """, (
            row["Month"],
            row["SSP Name"],
            float(row["Payable $"]),
            due_date.strftime("%Y-%m-%d")     if pd.notna(due_date)     else None,
            payment_date.strftime("%Y-%m-%d") if pd.notna(payment_date) else None,
            float(row["Paid Amount $"]),
            row["Paid From"],
            row["Reason"]
        ))

    conn.commit()
    conn.close()

# ---------------- DSP UPSERT ---------------- #

def upsert_dsp_data(df: pd.DataFrame):
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Ensure UNIQUE constraint exists on (month, dsp_name)
    try:
        cur.execute("""
            CREATE UNIQUE INDEX IF NOT EXISTS idx_dsp_month_name
            ON dsp_data(month, dsp_name)
        """)
        conn.commit()
    except Exception:
        pass  # index may already exist

    for _, row in df.iterrows():
        due_date     = pd.to_datetime(row["Due Date"],     errors="coerce")
        received_date = pd.to_datetime(row["Received Date"], errors="coerce")

        cur.execute("""
        INSERT INTO dsp_data (
            month, dsp_name, receivable,
            due_date, received_date, received_amount,
            received_in, reason
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(month, dsp_name)
        DO UPDATE SET
            received_date=excluded.received_date,
            received_amount=excluded.received_amount,
            received_in=excluded.received_in,
            reason=excluded.reason
        """, (
            row["Month"],
            row["DSP Name"],
            float(row["Receivable $"]),
            due_date.strftime("%Y-%m-%d")     if pd.notna(due_date)     else None,
            received_date.strftime("%Y-%m-%d") if pd.notna(received_date) else None,
            float(row["Received Amount $"]),
            row["Received In"],
            row["Reason"]
        ))

    conn.commit()
    conn.close()
# ==========================================
# IMPORT EXCEL TO SQLITE (GENERIC)
# ==========================================

def get_table_columns(table_name):
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute(f"PRAGMA table_info({table_name})")
    cols = [row[1] for row in cursor.fetchall()]

    conn.close()
    return cols


def insert_dataframe_to_db(df, table_name):
    conn = get_db_connection()
    cursor = conn.cursor()

    df = df.copy()

    # ==========================================
    # 🔥 FIX: CONVERT TIMESTAMP → STRING
    # ==========================================
    for col in df.columns:
        if "month" in col.lower():
            df[col] = pd.to_datetime(df[col], errors="coerce")
            df[col] = df[col].dt.strftime("%b-%Y")   # ✅ Feb-2026 format

        elif "date" in col.lower():
            df[col] = pd.to_datetime(df[col], errors="coerce")
            df[col] = df[col].dt.strftime("%Y-%m-%d")  # ISO format — unambiguous

    # ==========================================
    # 🔥 ROUND NUMERIC COLUMNS TO 2 DECIMAL
    # ==========================================
    for col in df.columns:
        if df[col].dtype in ["float64", "int64"]:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
            df[col] = df[col].round(2)
    
    # Replace NaT with empty
    for col in df.columns:
        if "month" not in col.lower():
            df[col] = df[col].fillna("")

    cols = list(df.columns)
    cols_quoted = [f'"{col}"' for col in cols]

    placeholders = ",".join(["?"] * len(cols))

    query = f"""
        INSERT INTO {table_name} ({",".join(cols_quoted)})
        VALUES ({placeholders})
    """

    data = [tuple(row) for row in df.values]

    cursor.executemany(query, data)

    conn.commit()
    conn.close()

    return len(data)

@st.cache_data(ttl=60)
def calculate_kpis(df_master):

    df = df_master.copy()

    df["DSP $ (BC)"] = pd.to_numeric(df["DSP $ (BC)"], errors="coerce").fillna(0)
    df["SSP $ (BC)"] = pd.to_numeric(df["SSP $ (BC)"], errors="coerce").fillna(0)
    df["C DSP $"] = pd.to_numeric(df.get("C DSP $", 0), errors="coerce").fillna(0)
    df["C SSP $"] = pd.to_numeric(df.get("C SSP $", 0), errors="coerce").fillna(0)

    df["Net $ (BC)"] = df["DSP $ (BC)"] - df["SSP $ (BC)"]
    df["C Net $"] = df["C DSP $"] - df["C SSP $"]

    total_dsp = df["DSP $ (BC)"].sum()
    total_ssp = df["SSP $ (BC)"].sum()
    total_net = df["Net $ (BC)"].sum()

    total_c_dsp = df["C DSP $"].sum()
    total_c_ssp = df["C SSP $"].sum()
    total_c_net = df["C Net $"].sum()

    ivt = total_net - total_c_net
    ivt_percent = (ivt / total_dsp * 100) if total_dsp != 0 else 0
    c_profit_percent = (total_c_net / total_c_dsp * 100) if total_c_dsp != 0 else 0

    return (
        df,
        total_dsp,
        total_ssp,
        total_net,
        total_c_dsp,
        total_c_ssp,
        total_c_net,
        ivt,
        ivt_percent,
        c_profit_percent
    )

import sqlite3
import pandas as pd

@st.cache_data(ttl=300)
def load_master_data():
    conn = get_db_connection()
    df = pd.read_sql("SELECT * FROM master_data", conn)
    conn.close()

    # 🔥 Always define expected columns
    expected_cols = [
        "id",
        "Month",
        "Partner Name",
        "DSP $ (BC)",
        "SSP $ (BC)",
        "C DSP $",
        "C SSP $"
    ]

    if df.empty:
        return pd.DataFrame(columns=expected_cols)

    # Rename DB → App columns
    df.rename(columns={
        "month": "Month",
        "partner_name": "Partner Name",
        "dsp_amount": "DSP $ (BC)",
        "ssp_amount": "SSP $ (BC)",
        "c_dsp": "C DSP $",
        "c_ssp": "C SSP $"
    }, inplace=True)

    # Ensure all expected columns exist
    for col in expected_cols:
        if col not in df.columns:
            df[col] = 0

    # Convert Month safely (handles Jul-25, Jul-2025, etc.)
    df["Month"] = df["Month"].astype(str).str.strip()

    # Try 2-digit year first (Jul-25)
    df["Month"] = pd.to_datetime(df["Month"], errors="coerce")

    # Final fallback (generic parser)
    mask = df["Month"].isna()
    if mask.any():
        df.loc[mask, "Month"] = pd.to_datetime(
            df.loc[mask, "Month"],
            errors="coerce"
        )

    # Normalize to first day of month
    df["Month"] = df["Month"].dt.to_period("M").dt.to_timestamp()

    # Convert numeric
    for col in ["DSP $ (BC)", "SSP $ (BC)", "C DSP $", "C SSP $"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # Derived
    df["C Net $"] = df["C DSP $"] - df["C SSP $"]
    df["Net $ (BC)"] = df["DSP $ (BC)"] - df["SSP $ (BC)"]

    return df

# ---------------- SSP FINAL LOADER (mirrors load_dsp_final) ---------------- #

def load_ssp_final(master_df, partner_df):

    # ✅ FIRST-RUN SAFE BLOCK (ADD THIS AT TOP)

    df_master = master_df.copy()

    # If empty → return clean structure
    if df_master.empty:
        return pd.DataFrame()

    # Ensure required columns always exist
    for col in ["C DSP $", "C SSP $"]:
        if col not in df_master.columns:
            df_master[col] = 0

    # Force numeric
    df_master["C DSP $"] = pd.to_numeric(df_master["C DSP $"], errors="coerce").fillna(0)
    df_master["C SSP $"] = pd.to_numeric(df_master["C SSP $"], errors="coerce").fillna(0)

    # Always derive (NEVER trust DB)
    df_master["C Net $"] = df_master["C DSP $"] - df_master["C SSP $"]

    # SSP rows = negative C Net $
    df_ssp = df_master[df_master["C Net $"] < 0].copy()
    df_ssp["SSP Name"] = df_ssp["Partner Name"]

    df_ssp["Month"] = pd.to_datetime(df_ssp["Month"], errors="coerce")
    df_ssp["Month"] = df_ssp["Month"].dt.strftime("%b-%Y")

    # Partner mapping
    if not partner_df.empty:
        for idx, row in df_ssp.iterrows():
            match = partner_df[
                partner_df["Short Name using in Bidscube"] == row["Partner Name"]
            ]
            if match.empty:
                continue
            country  = match.iloc[0].get("Country", "")
            net_term = match.iloc[0].get("Payment Terms", "")
            df_ssp.loc[idx, "USD/INR"]  = "INR" if country == "India (IN)" else "USD"
            df_ssp.loc[idx, "NET Term"] = net_term

    # Due Date
    import re
    def calc_due_ssp(month_str, net_term):
        try:
            month_dt  = pd.to_datetime(month_str, format="%b-%Y")
            last_date = (month_dt + pd.offsets.MonthEnd(0))
            days = 0
            if isinstance(net_term, str):
                m = re.search(r"\d+", net_term)
                if m:
                    days = int(m.group())
            return last_date + pd.Timedelta(days=days + 1)
        except:
            return ""

    df_ssp["Payable $"] = df_ssp["C Net $"].abs()

    df_ssp["Due Date"] = df_ssp.apply(
        lambda x: calc_due_ssp(x["Month"], x.get("NET Term", "")), axis=1
    )

    # Defaults
    df_ssp["Payment Date"]  = pd.NaT
    df_ssp["Paid Amount $"] = 0.00
    df_ssp["Paid From"]       = ""
    df_ssp["Reason"]        = ""

    # Load saved SSP data from DB
    conn = get_db_connection()
    try:
        df_db = pd.read_sql("SELECT * FROM ssp_data", conn)
    except:
        df_db = pd.DataFrame()
    conn.close()

    if not df_db.empty:
        df_db.rename(columns={
            "ssp_name":    "SSP Name",
            "month":       "Month",
            "payable":     "Payable $",
            "due_date":    "Due Date",
            "payment_date":"Payment Date",
            "paid_amount": "Paid Amount $",
            "paid_from":     "Paid From",
            "reason":      "Reason",
            "inv_status":  "Invoice Status"
        }, inplace=True)

        df_db["Month"] = pd.to_datetime(df_db["Month"], errors="coerce")
        df_db["Month"] = df_db["Month"].dt.strftime("%b-%Y")

        df_final = pd.merge(
            df_ssp, df_db,
            on=["Month", "SSP Name"],
            how="left",
            suffixes=("", "_db")
        )

        for col in ["Payment Date", "Paid Amount $", "Paid From", "Reason", "Invoice Status"]:
            if f"{col}_db" in df_final.columns:
                df_final[col] = df_final[f"{col}_db"].combine_first(df_final[col])

        df_final.drop(columns=[c for c in df_final.columns if c.endswith("_db")], inplace=True)
        df_final["Payment Date"] = pd.to_datetime(df_final["Payment Date"], errors="coerce")
        df_final["Due Date"]     = pd.to_datetime(df_final["Due Date"],     errors="coerce")
    else:
        df_final = df_ssp.copy()
        df_final["Invoice Status"] = ""

    today = pd.Timestamp.today()
    df_final["Shortage"] = df_final["Payable $"] - df_final["Paid Amount $"]
    df_final["Outstanding $"] = df_final["Shortage"]   # alias used by Dashboard/Summary
    df_final["Age"] = (
        (today - pd.to_datetime(df_final["Due Date"], errors="coerce"))
        .dt.days.fillna(0)
    )

    # Ensure Invoice Status column exists
    if "Invoice Status" not in df_final.columns:
        df_final["Invoice Status"] = ""
    df_final["Invoice Status"] = df_final["Invoice Status"].fillna("")

    # Auto-escalate: if fully paid, bump to INV Paid
    mask_paid = (
        (df_final["Paid Amount $"] >= df_final["Payable $"]) &
        (df_final["Payable $"] > 0) &
        (df_final["Invoice Status"] == "INV Received")
    )
    df_final.loc[mask_paid, "Invoice Status"] = "INV Paid"

    df_final = df_final[[
        "Month", "SSP Name", "Payable $", "Due Date",
        "Payment Date", "Paid Amount $", "Paid From",
        "Shortage", "Reason", "Age", "Invoice Status"
    ]]

    return df_final

# ==========================================
# NEW DSP HYBRID LOADER
# ==========================================
def load_dsp_final(master_df, partner_df):

    # ✅ FIRST-RUN SAFE BLOCK (ADD THIS AT TOP)

    df_master = master_df.copy()

    # If empty → return clean structure
    if df_master.empty:
        return pd.DataFrame()

    # Ensure required columns always exist
    for col in ["C DSP $", "C SSP $"]:
        if col not in df_master.columns:
            df_master[col] = 0

    # Force numeric
    df_master["C DSP $"] = pd.to_numeric(df_master["C DSP $"], errors="coerce").fillna(0)
    df_master["C SSP $"] = pd.to_numeric(df_master["C SSP $"], errors="coerce").fillna(0)

    # Always derive (NEVER trust DB)
    df_master["C Net $"] = df_master["C DSP $"] - df_master["C SSP $"]

    # DSP rows = negative C Net $
    df_dsp = df_master[df_master["C Net $"] > 0].copy()
    df_dsp["DSP Name"] = df_dsp["Partner Name"]

    df_dsp["Month"] = pd.to_datetime(df_dsp["Month"], errors="coerce")
    df_dsp["Month"] = df_dsp["Month"].dt.strftime("%b-%Y")

    # Partner mapping
    if not partner_df.empty:
        for idx, row in df_dsp.iterrows():
            match = partner_df[
                partner_df["Short Name using in Bidscube"] == row["Partner Name"]
            ]
            if match.empty:
                continue
            country  = match.iloc[0].get("Country", "")
            net_term = match.iloc[0].get("Payment Terms", "")
            df_dsp.loc[idx, "USD/INR"]  = "INR" if country == "India (IN)" else "USD"
            df_dsp.loc[idx, "NET Term"] = net_term

    # Due Date
    import re
    def calc_due_dsp(month_str, net_term):
        try:
            month_dt  = pd.to_datetime(month_str, format="%b-%Y")
            last_date = (month_dt + pd.offsets.MonthEnd(0))
            days = 0
            if isinstance(net_term, str):
                m = re.search(r"\d+", net_term)
                if m:
                    days = int(m.group())
            return last_date + pd.Timedelta(days=days + 1)
        except:
            return ""

    df_dsp["Receivable $"] = df_dsp["C Net $"].abs()

    df_dsp["Due Date"] = df_dsp.apply(
        lambda x: calc_due_dsp(x["Month"], x.get("NET Term", "")), axis=1
    )

    # Defaults
    df_dsp["Received Date"]  = pd.NaT
    df_dsp["Received Amount $"] = 0.00
    df_dsp["Received In"]       = ""
    df_dsp["Reason"]        = ""

    # Load saved DSP data from DB
    conn = get_db_connection()
    try:
        df_db = pd.read_sql("SELECT * FROM dsp_data", conn)
    except:
        df_db = pd.DataFrame()
    conn.close()

    if not df_db.empty:
        df_db.rename(columns={
            "dsp_name":    "DSP Name",
            "month":       "Month",
            "receivable":     "Receivable $",
            "due_date":    "Due Date",
            "received_date":"Received Date",
            "received_amount": "Received Amount $",
            "received_in":     "Received In",
            "reason":      "Reason"
        }, inplace=True)

        df_db["Month"] = pd.to_datetime(df_db["Month"], errors="coerce")
        df_db["Month"] = df_db["Month"].dt.strftime("%b-%Y")

        df_final = pd.merge(
            df_dsp, df_db,
            on=["Month", "DSP Name"],
            how="left",
            suffixes=("", "_db")
        )

        for col in ["Received Date", "Received Amount $", "Received In", "Reason"]:
            if f"{col}_db" in df_final.columns:
                df_final[col] = df_final[f"{col}_db"].combine_first(df_final[col])

        df_final.drop(columns=[c for c in df_final.columns if c.endswith("_db")], inplace=True)
        df_final["Received Date"] = pd.to_datetime(df_final["Received Date"], errors="coerce")
        df_final["Due Date"]     = pd.to_datetime(df_final["Due Date"],     errors="coerce")
    else:
        df_final = df_dsp.copy()

    today = pd.Timestamp.today()
    df_final["Shortage"] = df_final["Receivable $"] - df_final["Received Amount $"]
    df_final["Outstanding $"] = df_final["Shortage"]   # alias used by Dashboard/Summary
    df_final["Age"] = (
        (today - pd.to_datetime(df_final["Due Date"], errors="coerce"))
        .dt.days.fillna(0)
    )

    df_final = df_final[[
        "Month", "DSP Name", "Receivable $", "Due Date",
        "Received Date", "Received Amount $", "Received In",
        "Shortage", "Reason", "Age"
    ]]

    return df_final

@st.cache_data(ttl=60)
def load_partner_list():
    conn = get_db_connection()
    df = pd.read_sql("SELECT * FROM partner_list", conn)
    conn.close()

    expected_cols = [
        "Agreement Start Date",
        "Legal Entity Name",
        "Short Name using in Bidscube",
        "Registered Address",
        "Country",
        "Foreign / Indian Entity",
        "GSTIN",
        "Payment Terms",
        "Contact Person",
        "Designation",
        "Contact No.",
        "Email 1",
        "Email 2",
        "Email 3",
        "Finance Contact",
        "Finance Email"
    ]

    if df.empty:
        return pd.DataFrame(columns=expected_cols)

    # old DB support
    df.rename(columns={
        "agreement_date": "Agreement Start Date",
        "legal_name": "Legal Entity Name",
        "short_name": "Short Name using in Bidscube",
        "address": "Registered Address",
        "entity_type": "Foreign / Indian Entity",
        "payment_terms": "Payment Terms",
        "contact_person": "Contact Person",
        "contact_no": "Contact No.",
        "email1": "Email 1",
        "email2": "Email 2",
        "email3": "Email 3",
        "finance_contact": "Finance Contact",
        "finance_email": "Finance Email"
    }, inplace=True)

    for col in expected_cols:
        if col not in df.columns:
            df[col] = ""

    # Try ISO YYYY-MM-DD first, fall back to M/D/YYYY (legacy UI saves),
    # then D/M/YYYY (common for manually-entered dates by Indian/UK users).
    _raw = df["Agreement Start Date"].astype(str).str.strip()
    _parsed = pd.to_datetime(_raw, format="%Y-%m-%d", errors="coerce")       # ISO
    _mask   = _parsed.isna()
    if _mask.any():
        _parsed[_mask] = pd.to_datetime(_raw[_mask], format="%m/%d/%Y", errors="coerce")  # M/D/YYYY
    _mask = _parsed.isna()
    if _mask.any():
        _parsed[_mask] = pd.to_datetime(_raw[_mask], format="%-m/%-d/%Y", errors="coerce")  # M/D/YYYY no-pad
    _mask = _parsed.isna()
    if _mask.any():
        _parsed[_mask] = pd.to_datetime(_raw[_mask], dayfirst=True, errors="coerce")      # D/M/YYYY fallback
    df["Agreement Start Date"] = _parsed

    return df

# ADD THIS HERE
@st.cache_data(ttl=120)
def load_cost_centre():
    conn = get_db_connection()
    try:
        df = pd.read_sql("SELECT * FROM cost_centre", conn)

        expected_cols = [
            "id",
            "Category",
            "Cost Name",
            "Sub Cost",
            "Financial Year",
            "Month",
            "Currency",
            "Amount USD",
            "FX Rate",
            "Amount INR"
        ]

        if df.empty:
            return pd.DataFrame(columns=expected_cols)

        # DB → App rename
        df.rename(columns={
            "category": "Category",
            "cost_name": "Cost Name",
            "sub_cost": "Sub Cost",
            "financial_year": "Financial Year",
            "month": "Month",
            "currency": "Currency",
            "amount_usd": "Amount USD",
            "fx_rate": "FX Rate",
            "amount_inr": "Amount INR"
        }, inplace=True)

        # ensure columns
        for col in expected_cols:
            if col not in df.columns:
                if col in ["Amount USD", "FX Rate", "Amount INR"]:
                    df[col] = 0.0
                else:
                    df[col] = ""

        # Month cleanup
        # -----------------------------
        # FIX MONTH FORMAT (Jul-25 / Jul-2025 / 2025-07-01)
        # -----------------------------
        df["Month"] = df["Month"].astype(str).str.strip()

        # ✅ ONLY THIS LINE (no multiple parsing)
        df["Month"] = pd.to_datetime(df["Month"], errors="coerce")

        # 🚨 IMPORTANT: DO NOT overwrite again later

        # numeric cleanup
        for col in ["Amount USD", "FX Rate", "Amount INR"]:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

        return df[expected_cols]

    except Exception as e:
        st.error(f"Cost Centre load error: {e}")
        return pd.DataFrame(columns=[
            "id",
            "Category",
            "Cost Name",
            "Sub Cost",
            "Financial Year",
            "Month",
            "Currency",
            "Amount USD",
            "FX Rate",
            "Amount INR"
        ])
    finally:
        conn.close()

import numpy as np
import re
from datetime import datetime

# ==========================================
# FX RATE ENGINE (USD → INR)
# ==========================================

import requests
import pandas as pd
import streamlit as st

@st.cache_data(ttl=86400)
def get_fx_rate(month_str):

    import pandas as pd
    import requests

    try:
        dt = pd.to_datetime(month_str, format="%b-%Y")
        last_day = (dt + pd.offsets.MonthEnd(0)).strftime("%Y-%m-%d")

        url = f"https://api.frankfurter.app/{last_day}?from=USD&to=INR"

        r = requests.get(url, timeout=10)

        if r.status_code != 200:
            return 0.0

        data = r.json()

        return round(data["rates"]["INR"], 4)

    except:
        return 0.0

def prepare_dataframe_for_gsheet(df: pd.DataFrame):
    clean_df = df.copy()
    date_columns = []

    for col in clean_df.columns:

        # 🔹 Try detect date columns by name
        if "date" in col.lower():
            parsed_dates = pd.to_datetime(clean_df[col], errors="coerce", dayfirst=True)

            if parsed_dates.notna().mean() > 0.5:
                clean_df[col] = parsed_dates
                date_columns.append(col)
                continue

        # 🔹 Try numeric detection
        numeric_series = pd.to_numeric(clean_df[col], errors="coerce")
        numeric_ratio = numeric_series.notna().mean()

        if numeric_ratio > 0.7:
            clean_df[col] = numeric_series
            clean_df[col].replace([np.inf, -np.inf], 0, inplace=True)
            clean_df[col].fillna(0, inplace=True)
            clean_df[col] = clean_df[col].round(2)
        else:
            clean_df[col] = clean_df[col].astype(str)
            clean_df[col].replace("nan", "", inplace=True)

    return clean_df, date_columns

# -------------------------------
# Utility Functions
# -------------------------------

def load_sheet(sheet_name: str) -> pd.DataFrame:
    if not os.path.exists(FILE_PATH):
        return pd.DataFrame()

    try:
        return pd.read_excel(FILE_PATH, sheet_name=sheet_name)
    except:
        return pd.DataFrame()


def save_sheet(df: pd.DataFrame, sheet_name: str):
    with pd.ExcelWriter(FILE_PATH, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)


def format_usd(value):
    try:
        return f"${value:,.2f}"
    except:
        return "$0.00"

def clean_for_gsheet(df: pd.DataFrame) -> pd.DataFrame:
    import numpy as np
    df = df.copy()

    # Replace infinities
    df = df.replace([np.inf, -np.inf], 0)

    # Handle numeric columns
    numeric_cols = df.select_dtypes(include=["number"]).columns
    df[numeric_cols] = df[numeric_cols].fillna(0)

    # Handle non-numeric columns
    non_numeric_cols = df.select_dtypes(exclude=["number"]).columns
    df[non_numeric_cols] = df[non_numeric_cols].fillna("")

    return df

@st.cache_data(ttl=60)
def load_dsp_sheet():
    conn = get_db_connection()
    df = pd.read_sql("SELECT * FROM dsp_data", conn)
    conn.close()

    expected_cols = [
        "id",
        "Month",
        "DSP Name",
        "Receivable $",
        "Received Amount $",
        "Due Date",
        "Received Date"
    ]

    if df.empty:
        return pd.DataFrame(columns=expected_cols)

    df.rename(columns={
        "id": "id",
        "month": "Month",
        "dsp_name": "DSP Name",
        "receivable": "Receivable $",
        "received": "Received Amount $",
        "due_date": "Due Date",
        "received_date": "Received Date"
    }, inplace=True)

    for col in expected_cols:
        if col not in df.columns:
            df[col] = 0 if "$" in col else ""

    # -----------------------------
    # FIX MONTH FORMAT (Jul-25 etc.)
    # -----------------------------
    df["Month"] = df["Month"].astype(str).str.strip()

    df["Month"] = pd.to_datetime(
        df["Month"],
        format="%b-%y",
        errors="coerce"
    )

    mask = df["Month"].isna()
    if mask.any():
        df.loc[mask, "Month"] = pd.to_datetime(
            df.loc[mask, "Month"],
            format="%b-%Y",
            errors="coerce"
        )

    mask = df["Month"].isna()
    if mask.any():
        df.loc[mask, "Month"] = pd.to_datetime(
            df.loc[mask, "Month"],
            errors="coerce"
        )

    df["Month"] = df["Month"].dt.to_period("M").dt.to_timestamp()

    # -----------------------------
    # DATE FIELDS
    # -----------------------------
    df["Due Date"] = pd.to_datetime(df["Due Date"], errors="coerce", dayfirst=True)
    df["Received Date"] = pd.to_datetime(df["Received Date"], errors="coerce", dayfirst=True)

    # -----------------------------
    # NUMERIC FIELDS
    # -----------------------------
    df["Receivable $"] = pd.to_numeric(df["Receivable $"], errors="coerce").fillna(0)
    df["Received Amount $"] = pd.to_numeric(df["Received Amount $"], errors="coerce").fillna(0)
    
    return df


@st.cache_data(ttl=60)
def load_ssp_sheet():
    conn = get_db_connection()
    df = pd.read_sql("SELECT * FROM ssp_data", conn)
    conn.close()

    expected_cols = [
        "Month", "SSP Name", "Payable $", "Paid Amount $",
        "Due Date", "Payment Date", "Outstanding $"
    ]

    if df.empty:
        return pd.DataFrame(columns=expected_cols)

    df.rename(columns={
        "month": "Month",
        "ssp_name": "SSP Name",
        "payable": "Payable $",
        "paid": "Paid Amount $",
        "due_date": "Due Date",
        "payment_date": "Payment Date"
    }, inplace=True)

    for col in expected_cols:
        if col not in df.columns:
            df[col] = 0 if "$" in col else ""

    df["Month"] = df["Month"].astype(str).str.strip()

    df["Month"] = pd.to_datetime(
        df["Month"],
        format="%b-%y",
        errors="coerce"
    )

    mask = df["Month"].isna()
    if mask.any():
        df.loc[mask, "Month"] = pd.to_datetime(
            df.loc[mask, "Month"],
            format="%b-%Y",
            errors="coerce"
        )

    mask = df["Month"].isna()
    if mask.any():
        df.loc[mask, "Month"] = pd.to_datetime(
            df.loc[mask, "Month"],
            errors="coerce"
        )

    df["Month"] = df["Month"].dt.to_period("M").dt.to_timestamp()
    df["Due Date"] = pd.to_datetime(df["Due Date"], errors="coerce", dayfirst=True)
    df["Payment Date"] = pd.to_datetime(df["Payment Date"], errors="coerce", dayfirst=True)

    df["Payable $"] = pd.to_numeric(df["Payable $"], errors="coerce").fillna(0)
    df["Paid Amount $"] = pd.to_numeric(df["Paid Amount $"], errors="coerce").fillna(0)
    df["Outstanding $"] = df["Payable $"] - df["Paid Amount $"]

    return df

# ==========================================
# CENTRAL DATA STORE (LOAD ONCE ONLY)
# ==========================================

def initialize_session_data():

    if "data_initialized" not in st.session_state:

        try:
            st.session_state.master_df = load_master_data()
        except:
            st.session_state.master_df = pd.DataFrame()

        try:
            st.session_state.partner_df = load_partner_list()
        except:
            st.session_state.partner_df = pd.DataFrame()

        try:
            st.session_state.dsp_df = load_dsp_final(
                st.session_state.master_df,
                st.session_state.partner_df
            )
        except:
            st.session_state.dsp_df = pd.DataFrame()

        try:
            st.session_state.ssp_df = load_ssp_final(
                st.session_state.master_df,
                st.session_state.partner_df
            )
        except:
            st.session_state.ssp_df = pd.DataFrame()

        try:
            st.session_state.cost_df = load_cost_centre()
        except:
            st.session_state.cost_df = pd.DataFrame()

        st.session_state.data_initialized = True

# -------------------------------
# MODAL STATE CONTROL
# -------------------------------
if "show_add_cost_modal" not in st.session_state:
    st.session_state.show_add_cost_modal = False
    
# -------------------------------
# Streamlit App Config
# -------------------------------

st.set_page_config(
    page_title="PEAKADS LLP - Revenue Tracker",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ── Responsive design system ─────────────────────────────────────────────────
from responsive import setup_responsive
sc = setup_responsive()   # detects screen, injects CSS, returns ScreenConfig
# Use sc.grid_height(base) wherever AgGrid height= appears below
# ─────────────────────────────────────────────────────────────────────────────

st.markdown("""
<style>

/* REMOVE ALL STREAMLIT DEFAULT SPACING */
html, body {
    margin: 0 !important;
    padding: 0 !important;
}

/* Remove top & bottom padding completely */
.block-container {
    padding-top: 0 !important;
    padding-bottom: 0 !important;

    /* Controlled left & right spacing */
    padding-left: 15px !important;
    padding-right: 15px !important;

    margin: 0 !important;
}

/* Remove Streamlit header + footer */
header { display: none !important; }
footer { display: none !important; }

/* Remove app container spacing */
[data-testid="stAppViewContainer"] {
    margin: 0 !important;
    padding: 0 !important;
}

/* Remove vertical block spacing */
[data-testid="stVerticalBlock"] {
    gap: 0rem !important;
}

</style>
""", unsafe_allow_html=True)

from datetime import date

def generate_financial_years():
    from datetime import date

    today = date.today()
    current_year = today.year
    current_month = today.month

    # Current FY start year
    current_fy_start = current_year if current_month >= 4 else current_year - 1

    # Show previous FY + current FY
    fy_list = [
        f"{current_fy_start - 1}-{str(current_fy_start)[-2:]}",
        f"{current_fy_start}-{str(current_fy_start + 1)[-2:]}"
    ]

    return fy_list



def _tab_fy_list():
    """FY labels in Invoice Manager format: 'FY 2025-26', 'FY 2026-27'."""
    today = datetime.today()
    s = today.year if today.month >= 4 else today.year - 1
    return [f"FY {s-1}-{str(s)[-2:]}", f"FY {s}-{str(s+1)[-2:]}"]

def _fy_yr(fy_label):
    """'FY 2026-27' → 2026 (the start calendar year)."""
    return int(fy_label.replace("FY ","").split("-")[0])

def _fy_months(fy_label):
    """Return ordered month labels Apr→Mar for the given FY label.
    Falls back to current FY when fy_label is 'All'."""
    if not fy_label or fy_label == "All":
        today = datetime.today()
        yr = today.year if today.month >= 4 else today.year - 1
    else:
        yr = _fy_yr(fy_label)
    seq = [(4,yr),(5,yr),(6,yr),(7,yr),(8,yr),(9,yr),
           (10,yr),(11,yr),(12,yr),(1,yr+1),(2,yr+1),(3,yr+1)]
    return [datetime(y,m,1).strftime("%b-%Y") for m,y in seq]

def _fy_date_range(fy_label):
    """'FY 2026-27' → (Timestamp('2026-04-01'), Timestamp('2027-03-31'))."""
    yr = _fy_yr(fy_label)
    return pd.Timestamp(f"{yr}-04-01"), pd.Timestamp(f"{yr+1}-03-31")

_QUARTER_OPTS   = ["Q1 (Apr–Jun)", "Q2 (Jul–Sep)", "Q3 (Oct–Dec)", "Q4 (Jan–Mar)"]
_QUARTER_MONTHS = {
    "Q1 (Apr–Jun)": (4, 6),
    "Q2 (Jul–Sep)": (7, 9),
    "Q3 (Oct–Dec)": (10, 12),
    "Q4 (Jan–Mar)": (1,  3),
}

def _quarter_date_range(fy_label, q_label):
    """'FY 2026-27', 'Q1 (Apr–Jun)' → (Timestamp, Timestamp)."""
    import calendar
    if not fy_label or fy_label == "All":
        today = datetime.today()
        yr = today.year if today.month >= 4 else today.year - 1
    else:
        yr = _fy_yr(fy_label)
    m1, m2 = _QUARTER_MONTHS[q_label]
    yr1  = yr+1 if m1 <= 3 else yr
    yr2  = yr+1 if m2 <= 3 else yr
    last = calendar.monthrange(yr2, m2)[1]
    return pd.Timestamp(f"{yr1}-{m1:02d}-01"), pd.Timestamp(f"{yr2}-{m2:02d}-{last}")


def get_fy_date_range(fy_string):
    start_year = int(fy_string.split("-")[0])
    start_date = pd.to_datetime(f"{start_year}-04-01")
    end_date = pd.to_datetime(f"{start_year+1}-03-31")
    return start_date, end_date


def get_quarter_range(fy_string, quarter):
    start_year = int(fy_string.split("-")[0])

    mapping = {
        "Q1": (4, 6),
        "Q2": (7, 9),
        "Q3": (10, 12),
        "Q4": (1, 3)
    }

    start_month, end_month = mapping[quarter]

    if quarter == "Q4":
        start = pd.Timestamp(start_year + 1, start_month, 1)
        end = pd.Timestamp(start_year + 1, end_month, 1) + pd.offsets.MonthEnd(0)
    else:
        start = pd.Timestamp(start_year, start_month, 1)
        end = pd.Timestamp(start_year, end_month, 1) + pd.offsets.MonthEnd(0)

    return start, end

st.markdown("""
<style>

/* KPI Background Container */
.kpi-container {
    background: linear-gradient(135deg, #87CEFA, #e8eef7);
    padding: 25px;
    border-radius: 12px;
    margin-bottom: 20px;
    box-shadow: 0px 4px 10px rgba(0,0,0,0.08);
}

/* Make metrics inside look clean */
div[data-testid="stMetric"] {
    background-color: #00FFFF;
    padding: 15px;
    border-radius: 10px;
    box-shadow: 0px 2px 6px rgba(0,0,0,0.05);
}

</style>
""", unsafe_allow_html=True)

st.markdown("""
<style>

/* ═══════════════════════════════════════════════════════════
   MODERN COMPACT TABS  —  Main tabs + all sub-tabs
   ═══════════════════════════════════════════════════════════ */

/* ── Tab list row: dark pill-bar background ─────────────── */
div[data-baseweb="tab-list"] {
    background    : linear-gradient(90deg, #001a3d 0%, #003070 50%, #001a3d 100%) !important;
    border-radius : 10px !important;
    padding       : 5px 8px !important;
    gap           : 4px !important;
    overflow-x    : auto !important;
    scrollbar-width: none !important;
    flex-wrap     : nowrap !important;
    margin-bottom : 6px !important;
}
div[data-baseweb="tab-list"]::-webkit-scrollbar { display: none !important; }

/* ── Kill Streamlit's default red underline & bottom border ─ */
div[data-baseweb="tab-highlight"],
div[data-baseweb="tab-border"] {
    display: none !important;
    height : 0   !important;
}

/* ── Every tab button: compact pill ─────────────────────── */
div[data-testid="stTabs"] button[role="tab"] {
    background    : transparent !important;
    color         : rgba(255,255,255,0.65) !important;
    font-size     : 12.5px !important;
    font-weight   : 600    !important;
    padding       : 6px 13px !important;
    border-radius : 7px    !important;
    border        : 1.5px solid transparent !important;
    white-space   : nowrap !important;
    transition    : background 0.18s ease, color 0.18s ease,
                    border-color 0.18s ease, transform 0.12s ease !important;
    min-width     : 0      !important;
    letter-spacing: 0.2px  !important;
    font-family   : -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif !important;
    margin        : 0      !important;
    position      : relative !important;
}

div[data-testid="stTabs"] button[role="tab"] p {
    font-size  : 13.5px !important;
    font-weight: 600    !important;
    color      : white !important;
    margin     : 0      !important;
}

/* ── Hover ───────────────────────────────────────────────── */
div[data-testid="stTabs"] button[role="tab"]:hover {
    background  : #9a009a !important;
    color       : #fff !important;
    border-color: rgba(255,255,255,0.20) !important;
    transform   : translateY(-1px) !important;
}

/* ── Active / selected ───────────────────────────────────── */
div[data-testid="stTabs"] button[role="tab"][aria-selected="true"] {
    background   : linear-gradient(135deg, #FF5E0E 0%, #FF8F00 100%) !important;
    color        : #fff !important;
    border-color : rgba(255,140,0,0.5) !important;
    box-shadow   : 0 2px 10px rgba(255,94,14,0.40),
                   0 1px  3px rgba(0,0,0,0.20) !important;
    transform    : translateY(-1px) !important;
    font-weight  : 700 !important;
}

/* ── Tab content area: remove default top border/padding ─── */
div[data-testid="stTabsContent"] {
    padding-top   : 8px !important;
    border-top    : none !important;
}

/* ── Scrollable on small screens ────────────────────────── */
@media (max-width: 900px) {
    div[data-baseweb="tab-list"] {
        padding: 4px 6px !important;
    }
    div[data-testid="stTabs"] button[role="tab"] {
        font-size: 11.5px !important;
        padding  : 5px 10px !important;
    }
}

</style>
""", unsafe_allow_html=True)



st.markdown("""
<style>

/* ═══════════════════════════════════════════════════════════════
   GLOBAL BUTTON DESIGN SYSTEM
   ═══════════════════════════════════════════════════════════════
   Covers: default, primary, danger-labelled, logout, full-width
   ═══════════════════════════════════════════════════════════════ */

/* ── Base reset: applies to EVERY st.button ─────────────────── */
div.stButton > button,
div.stDownloadButton > button {
    font-family   : -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif !important;
    font-size     : 12.5px   !important;
    font-weight   : 600      !important;
    letter-spacing: 0.2px    !important;
    padding       : 6px 15px !important;
    border-radius : 8px      !important;
    border        : 1.5px solid transparent !important;
    cursor        : pointer   !important;
    transition    : background 0.18s ease, border-color 0.18s ease,
                    box-shadow 0.18s ease, transform 0.12s ease !important;
    white-space   : nowrap    !important;
    line-height   : 1.4       !important;
}

/* ── DEFAULT button (secondary) — Navy blue ──────────────────── */
div.stButton > button[kind="secondary"],
div.stDownloadButton > button {
    background : linear-gradient(135deg, #1a4b8c 0%, #0d2d5e 100%) !important;
    color      : #e8f0ff !important;
    border-color: rgba(100,160,255,0.25) !important;
    box-shadow : 0 2px 6px rgba(0,0,0,0.25),
                 inset 0 1px 0 rgba(255,255,255,0.08) !important;
}
div.stButton > button[kind="secondary"]:hover,
div.stDownloadButton > button:hover {
    background  : linear-gradient(135deg, #1f5aa8 0%, #103570 100%) !important;
    border-color: rgba(100,160,255,0.45) !important;
    box-shadow  : 0 4px 12px rgba(0,0,0,0.30),
                  inset 0 1px 0 rgba(255,255,255,0.12) !important;
    transform   : translateY(-1px) !important;
    color       : #fff !important;
}
div.stButton > button[kind="secondary"]:active,
div.stDownloadButton > button:active {
    transform  : translateY(0px) !important;
    box-shadow : 0 1px 4px rgba(0,0,0,0.25) !important;
}

/* ── PRIMARY button — Brand orange ──────────────────────────── */
div.stButton > button[kind="primary"] {
    background : linear-gradient(135deg, #FF5E0E 0%, #FF8F00 100%) !important;
    color      : #fff !important;
    border-color: rgba(255,140,0,0.35) !important;
    box-shadow : 0 3px 10px rgba(255,94,14,0.35),
                 inset 0 1px 0 rgba(255,255,255,0.18) !important;
}
div.stButton > button[kind="primary"]:hover {
    background  : linear-gradient(135deg, #ff6f20 0%, #ffA010 100%) !important;
    border-color: rgba(255,160,0,0.5) !important;
    box-shadow  : 0 5px 16px rgba(255,94,14,0.45),
                  inset 0 1px 0 rgba(255,255,255,0.22) !important;
    transform   : translateY(-1px) !important;
}
div.stButton > button[kind="primary"]:active {
    transform  : translateY(0px) !important;
    box-shadow : 0 2px 6px rgba(255,94,14,0.30) !important;
}

/* ── LOGOUT button — Red, fixed hover ───────────────────────── */
button[data-testid="stBaseButton-secondary"]#logout_button,
div.stButton > button[key="logout_button"],
div[data-testid="stHorizontalBlock"] div.stButton:last-child > button {
    background  : linear-gradient(135deg, #c0392b 0%, #96281b 100%) !important;
    color       : #fff !important;
    border-color: rgba(255,100,100,0.30) !important;
    box-shadow  : 0 3px 10px rgba(150,40,27,0.40) !important;
    font-weight : 700 !important;
}
div[data-testid="stHorizontalBlock"] div.stButton:last-child > button:hover {
    background  : linear-gradient(135deg, #008000 0%, #008000 100%) !important;
    box-shadow  : 0 5px 14px rgba(192,57,43,0.50) !important;
    transform   : translateY(-1px) !important;
    color       : #fff !important;
}

/* ── Full-width buttons keep pill shape ──────────────────────── */
div.stButton > button[use_container_width="true"],
div.stButton > button:has(+ *[class*="full"]) {
    width: 100% !important;
}

/* ── p tag inside buttons: inherit properly ──────────────────── */
div.stButton > button p,
div.stDownloadButton > button p {
    font-size  : inherit !important;
    font-weight: inherit !important;
    color      : inherit !important;
    margin     : 0       !important;
}

/* ── DISABLED state — elegant locked look (global) ───────────── */
div.stButton > button:disabled,
div.stDownloadButton > button:disabled {
    background   : rgba(80,80,100,0.08) !important;
    color        : rgba(120,120,140,0.50) !important;
    border       : 1.5px dashed rgba(120,120,140,0.28) !important;
    box-shadow   : none !important;
    cursor       : not-allowed !important;
    transform    : none !important;
    font-style   : italic !important;
}

</style>
""", unsafe_allow_html=True)

# ---------------------------
# SESSION INITIALIZATION
# ---------------------------

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

if "user" not in st.session_state:
    st.session_state.user = ""

if "role" not in st.session_state:
    st.session_state.role = ""

if "backup_running" not in st.session_state:
    st.session_state.backup_running = False

from login import login_screen, get_allowed_tabs

# ---------------------------
# LOGIN CHECK
# ---------------------------

if not st.session_state.logged_in:
    login_screen()
    st.stop()

# ---------------------------
# LOGIN CHECK
# ---------------------------

if not st.session_state.logged_in:
    login_screen()
    st.stop()


initialize_session_data()

allowed_tabs = get_allowed_tabs()

# Ensure "Ageing" tab is always present right after "Dashboard"
if "Ageing" not in allowed_tabs:
    idx = allowed_tabs.index("Dashboard") + 1 if "Dashboard" in allowed_tabs else 1
    allowed_tabs.insert(idx, "Ageing")

tab_map = {
    "Dashboard": "📊 Dashboard",
    "Ageing": "📈 AR/AP Ageing",
    "Master Data": "📁 Master Data",
    "DSP (Customers)": "🧑 DSP (Customers)",
    "Invoice Manager": "🧾 Invoice Manager",
    "SSP (Vendors)": "📤 SSP (Vendors)",
    "List of Partners": "🤝 List of Partners",
    "Costs Centre": "💰 Costs Centre",
    "P&L": "📉 P&L",
    "Admin Control": "⚙️ Admin Control Panel",
    "Edit Database": "🛠️ Edit Database",
    "BC Report":     "📋 BC Report",
}

tab_titles = [tab_map[t] for t in allowed_tabs]

# Tabs row
render_navbar()          # ← ADD THIS
tabs_list = st.tabs(tab_titles)
tabs = dict(zip(allowed_tabs, tabs_list))

if "Dashboard" in tabs:
    with tabs["Dashboard"]:
        render_dashboard_tab(
            master_df = st.session_state.get("master_df", pd.DataFrame()),
            dsp_df    = st.session_state.get("dsp_df"),
            ssp_df    = st.session_state.get("ssp_df"),
            partner_df = st.session_state.get("partner_df"),   # ← add this
        )

# ====================================================
# BC REPORT TAB
# ====================================================

if "BC Report" in tabs:
    with tabs["BC Report"]:
        render_bc_report_tab()

# ====================================================
# LIST OF PARTNERS TAB
# ====================================================

if "List of Partners" in tabs:
    with tabs["List of Partners"]:
        
        st.markdown('''
                <div style="background:linear-gradient(135deg,#003366 0%,#005599 100%);
                border-radius:10px;padding:18px 24px;margin-bottom:20px;
                box-shadow:0 4px 16px rgba(0,51,102,.2);
                display:flex;align-items:center;height:55px;gap:14px;">
                <div style="font-size:32px;">🤝</div>
                <div>
                    <div style="color:white;font-size:20px;font-weight:800;">
                        List of Partners</div>
                </div></div>
            ''', unsafe_allow_html=True)

        col2, col3, col4 = st.columns([1, 1, 3.5])
        
        
        # -----------------------------
        # LOAD PARTNER DATA
        # -----------------------------
        df_partner = load_partner_list().copy()
        st.session_state.partner_df = df_partner.copy()

        if df_partner.empty:
            st.warning("No Partner Data Found")
        else:

            # -----------------------------
            # REQUIRED COLUMNS
            # -----------------------------
            required_columns = [
                "Agreement Start Date",
                "Short Name using in Bidscube",
                "Country",
                "Foreign / Indian Entity",
                "GSTIN",
                "Payment Terms",
                "Contact Person",
                "Email 1",
                "Finance Contact",
                "Finance Email"
            ]

            for col in required_columns:
                if col not in df_partner.columns:
                    df_partner[col] = ""

            df_partner = df_partner[required_columns].copy()

            # -----------------------------
            # FORMAT DATE
            # -----------------------------
            df_partner["Agreement Start Date"] = pd.to_datetime(
                df_partner["Agreement Start Date"],
                errors="coerce",
                dayfirst=True
            )

            df_partner["Agreement Start Date"] = df_partner["Agreement Start Date"].dt.strftime("%d-%b-%Y")

            # -----------------------------
            # DISPLAY COLUMN NAMES
            # -----------------------------
            df_display = df_partner.rename(columns={
                "Short Name using in Bidscube": "Partner",
                "Agreement Start Date": "Start Date",
                "Foreign / Indian Entity": "Foreign / Indian"
            })

        # -----------------------------
        # SEARCH + REFRESH
        # -----------------------------
        

        with col2:
            search_text = st.text_input(
                "🔍 Search Partner",
                placeholder="Global search...",
                key="partner_search"
            )

        with col3:
            st.markdown("<br>", unsafe_allow_html=True)
            refresh_clicked = st.button("🔄 Refresh", key="partner_list_refresh")

            if refresh_clicked:
                # Force fresh DB read
                load_partner_list.clear()

                fresh_df = load_partner_list()

                st.session_state.partner_df = fresh_df.copy()

                st.success("Partner list refreshed successfully")
                st.rerun()
                
        # ====================================================
        # ADD PARTNER POPUP — defined BEFORE the button that calls it
        # ====================================================
        @st.dialog("Add New Partner", width="large")
        def add_form_popup():

            st.markdown("## 📝 Partner Onboarding Form")

            # -----------------------------
            # ROW 1 (3 columns)
            # -----------------------------
            col1, col2, col3 = st.columns(3)

            with col1:
                agreement_date = st.date_input("Agreement Start Date *")
                legal_name = st.text_input("Legal Entity Name *")

            with col2:
                short_name = st.text_input("Short Name *")
                payment_terms = st.selectbox(
                    "Payment Terms *",
                    ["Net 30", "Net 45", "Net 60", "Net 90"]
                )

            with col3:
                contact_no = st.text_input("Contact No.")
                email1 = st.text_input("Email 1")

            # -----------------------------
            # ROW 2
            # -----------------------------
            col4, col5, col6 = st.columns(3)

            with col4:
                country = st.selectbox(
                    "Country",
                    ["India (IN)", "US", "UK", "Singapore", "UAE"]
                )

            with col5:
                contact_person = st.text_input("Contact Person")
                designation = st.text_input("Designation")

            with col6:
                email2 = st.text_input("Email 2")
                email3 = st.text_input("Email 3")

            # -----------------------------
            # ROW 3
            # -----------------------------
            entity_type = "Indian" if country == "India (IN)" else "Foreign"

            gstin = st.text_input(
                "GSTIN *" if country == "India (IN)" else "GSTIN",
                disabled=(entity_type != "Indian")
            )

            # -----------------------------
            # ADDRESS
            # -----------------------------
            st.markdown("### 📍 Registered Address")
            address = st.text_area("", height=80)

            # -----------------------------
            # FINANCE
            # -----------------------------
            st.markdown("### 💰 Finance Details")

            col7, col8 = st.columns(2)

            with col7:
                finance_contact = st.text_input("Finance Contact")

            with col8:
                finance_email = st.text_input("Finance Email *")

            # -----------------------------
            # BUTTON (FIXED WIDTH)
            # -----------------------------
            st.markdown("<br>", unsafe_allow_html=True)

            colA, colB, colC = st.columns([1, 2, 1])

            with colB:
                save_clicked = st.button(
                    "💾 Save Partner",
                    use_container_width=True
                )

            # -----------------------------
            # SAVE LOGIC (same as yours)
            # -----------------------------
            if save_clicked:
                # ── Mandatory field validation ──────────────────────────────
                _errors = []
                if not short_name.strip():
                    _errors.append("• Short Name is required")
                if not legal_name.strip():
                    _errors.append("• Legal Entity Name is required")
                if not payment_terms:
                    _errors.append("• Payment Terms is required")
                if not finance_email.strip():
                    _errors.append("• Finance Email is required")
                if country == "India (IN)" and not gstin.strip():
                    _errors.append("• GSTIN is mandatory for Indian partners")

                if _errors:
                    st.error("Please fix the following before saving:\n\n" + "\n".join(_errors))
                    st.stop()
                # ── End validation ──────────────────────────────────────────

                conn = get_db_connection()
                cursor = conn.cursor()

                # Detect actual column name via PRAGMA (works with old and new DB)
                pragma = cursor.execute("PRAGMA table_info(partner_list)").fetchall()
                col_names = [row[1] for row in pragma]
                _sn_col = "short_name" if "short_name" in col_names else "Short Name using in Bidscube"
                cursor.execute(f'SELECT "{_sn_col}" FROM partner_list')
                existing = [str(r[0]).strip().lower() for r in cursor.fetchall() if r[0]]

                if short_name.strip().lower() in existing:
                    st.error("Partner already exists")
                    conn.close()
                    return

                # Map from new snake_case names -> old display names (for old DBs)
                pragma = cursor.execute("PRAGMA table_info(partner_list)").fetchall()
                col_names = [row[1] for row in pragma]

                def _col(new_name, old_name):
                    return new_name if new_name in col_names else old_name

                c_agreement   = _col("agreement_date",  "Agreement Start Date")
                c_legal       = _col("legal_name",       "Legal Entity Name")
                c_short       = _col("short_name",       "Short Name using in Bidscube")
                c_address     = _col("address",          "Registered Address")
                c_country     = _col("country",          "Country")
                c_entity      = _col("entity_type",      "Foreign / Indian Entity")
                c_gstin       = _col("gstin",            "GSTIN")
                c_payment     = _col("payment_terms",    "Payment Terms")
                c_contact_p   = _col("contact_person",   "Contact Person")
                c_desig       = _col("designation",      "Designation")
                c_contact_no  = _col("contact_no",       "Contact No.")
                c_email1      = _col("email1",           "Email 1")
                c_email2      = _col("email2",           "Email 2")
                c_email3      = _col("email3",           "Email 3")
                c_fin_contact = _col("finance_contact",  "Finance Contact")
                c_fin_email   = _col("finance_email",    "Finance Email")

                cursor.execute(f"""
                INSERT INTO partner_list (
                    "{c_agreement}", "{c_legal}", "{c_short}", "{c_address}",
                    "{c_country}", "{c_entity}", "{c_gstin}", "{c_payment}",
                    "{c_contact_p}", "{c_desig}", "{c_contact_no}",
                    "{c_email1}", "{c_email2}", "{c_email3}",
                    "{c_fin_contact}", "{c_fin_email}"
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    agreement_date.strftime("%Y-%m-%d"),  # ISO format — unambiguous
                    legal_name,
                    short_name,
                    address,
                    country,
                    entity_type,
                    gstin,
                    payment_terms,
                    contact_person,
                    designation,
                    contact_no,
                    email1,
                    email2,
                    email3,
                    finance_contact,
                    finance_email
                ))

                conn.commit()
                conn.close()

                st.success("✅ Partner Added Successfully")
                load_partner_list.clear()
                st.rerun()

        with col4:
            st.markdown("<br>", unsafe_allow_html=True)

            if st.button("+ Add Partner", key="add_form_btn"):
                st.session_state.open_partner_popup = True
                st.rerun()
                
            if st.session_state.get("open_partner_popup", False):

                st.session_state.open_partner_popup = False

                add_form_popup()


        st.divider()

        # -----------------------------
        # AGGRID
        # -----------------------------
        if not df_partner.empty:    
            from st_aggrid import AgGrid, GridOptionsBuilder, JsCode

            gb = GridOptionsBuilder.from_dataframe(df_display)

            date_comparator = JsCode("""
            function(date1, date2) {
                if (!date1) return -1;
                if (!date2) return 1;

                function parseDate(str) {
                    const parts = str.split("-");
                    if (parts.length !== 3) return new Date(0);
                    return new Date(parts[1] + " " + parts[0] + ", " + parts[2]);
                }

                const d1 = parseDate(date1);
                const d2 = parseDate(date2);

                return d1 - d2;
            }
            """)

            gb.configure_column("Start Date", comparator=date_comparator)
            gb.configure_column("Email 1", flex=2)
            gb.configure_column("Finance Email", flex=2)

            gb.configure_default_column(
                resizable=True,
                sortable=True,
                filter=True,
                flex=1
            )

            gb.configure_grid_options(
                domLayout="normal",
                suppressHorizontalScroll=False
            )

            gridOptions = gb.build()

            if search_text:
                gridOptions["quickFilterText"] = search_text

            custom_css = {
                ".ag-root-wrapper": {
                    "overflow": "auto"
                },
                ".ag-body-horizontal-scroll": {
                    "height": "8px"
                },
                ".ag-header": {
                    "background-color": "#003366 !important",
                    "color": "white !important",
                    "font-weight": "bold !important",
                    "font-size": "14px !important"
                },
                ".ag-header-cell-label": {
                    "color": "white !important",
                    "font-weight": "bold !important"
                }
            }

            AgGrid(
                df_display,
                gridOptions=gridOptions,
                allow_unsafe_jscode=True,
                height=sc.grid_height(700),
                fit_columns_on_grid_load=True,
                custom_css=custom_css
            )

# ====================================================
# 2️⃣ MASTER DATA TAB
# ====================================================

if "Master Data" in tabs:
    with tabs["Master Data"]:
        
        st.markdown('''
                <div style="background:linear-gradient(135deg,#003366 0%,#005599 100%);
                border-radius:10px;padding:18px 24px;margin-bottom:20px;
                box-shadow:0 4px 16px rgba(0,51,102,.2);
                display:flex;align-items:center;height:55px;gap:14px;">
                <div style="font-size:32px;">📁</div>
                <div>
                    <div style="color:white;font-size:20px;font-weight:800;">
                        Master Data</div>
                </div></div>
            ''', unsafe_allow_html=True)

        # ── FY / Month / Quarter filters (Invoice Manager style) ──────────────
        _mst_fy_list = _tab_fy_list()
        col2, col3, col4, col5, col6 = st.columns([0.3, 0.3, 0.3, 0.5, 0.5])

        with col2:
            selected_fy = st.selectbox(
                "Financial Year",
                options=["All"] + _mst_fy_list,
                index=0,
                key="master_fy"
            )

        _mst_q = st.session_state.get("master_quarter", "All")
        _mst_m = st.session_state.get("master_month",   "All")

        with col3:
            selected_month = st.selectbox(
                "Month",
                options=["All"] + _fy_months(selected_fy),
                index=0,
                key="master_month",
                disabled=(_mst_q not in ("", "All")),
            )

        with col4:
            selected_quarter = st.selectbox(
                "Quarter",
                options=["All"] + _QUARTER_OPTS,
                index=0,
                key="master_quarter",
                disabled=(_mst_m not in ("", "All")),
            )

        with col5:
            search_text = st.text_input(
                "🔍 Search",
                placeholder="Global Search...",
                key="master_search"
            )

        with col6:
            st.markdown("<br>", unsafe_allow_html=True)
            refresh_clicked = st.button(
                "🔄 Refresh",
                key="master_refresh_button"
            )

        if refresh_clicked:
            load_master_data.clear()
            st.session_state.master_df = load_master_data()
            st.rerun()

        if "master_df" not in st.session_state:
            st.session_state.master_df = load_master_data()

        df_master = st.session_state.master_df
        if df_master.empty:
            st.info("📭 No data yet. Please upload or add Master Data.")
            st.stop()

        df_filtered = df_master.copy()
        df_filtered["Month"] = pd.to_datetime(df_filtered["Month"], errors="coerce")

        if selected_fy not in ("", "All"):
            _mst_fy_start, _mst_fy_end = _fy_date_range(selected_fy)
            df_filtered = df_filtered[
                (df_filtered["Month"] >= _mst_fy_start) &
                (df_filtered["Month"] <= _mst_fy_end)
            ]
            if selected_quarter not in ("", "All"):
                q_start, q_end = _quarter_date_range(selected_fy, selected_quarter)
                df_filtered = df_filtered[
                    (df_filtered["Month"] >= q_start) &
                    (df_filtered["Month"] <= q_end)
                ]
            elif selected_month not in ("", "All"):
                df_filtered = df_filtered[
                    df_filtered["Month"].dt.strftime("%b-%Y") == selected_month
                ]
        
        df_partner = st.session_state.get("partner_df", pd.DataFrame()).copy()

        if df_filtered.empty:
            st.warning("No Master Data Found")
        else:

            # Ensure Month is proper datetime
            df_filtered["Month"] = pd.to_datetime(df_filtered["Month"], errors="coerce")

            # Sort filtered data only
            df_master = df_filtered.sort_values("Month").copy()

            # Display format only
            df_master["Month"] = df_master["Month"].dt.strftime("%b-%Y")
        
        if not df_master.empty:

            df_master["Net $ (BC)"] = df_master["DSP $ (BC)"] - df_master["SSP $ (BC)"]

            df_master["Month"] = pd.to_datetime(df_master["Month"], format="%b-%Y", errors="coerce")
            df_master["Month"] = df_master["Month"].dt.strftime("%b-%Y")

            for _str_col in ["GSTIN", "NET Term", "I/F", "USD/INR"]:
                if _str_col not in df_master.columns:
                    df_master[_str_col] = ""
                else:
                    df_master[_str_col] = df_master[_str_col].astype(object)

            for index, row in df_master.iterrows():

                if df_partner.empty:
                    continue

                if "Short Name using in Bidscube" not in df_partner.columns:
                    continue

                partner_match = df_partner[
                    df_partner["Short Name using in Bidscube"] == row["Partner Name"]
                ]

                if partner_match.empty:
                    continue

                country  = str(partner_match.iloc[0].get("Country", "") or "")
                gstin    = str(partner_match.iloc[0].get("GSTIN", "") or "")
                net_term = str(partner_match.iloc[0].get("Payment Terms", "") or "")

                if country == "India (IN)":
                    df_master.loc[index, "I/F"] = "Indian"
                    df_master.loc[index, "USD/INR"] = "INR"
                else:
                    df_master.loc[index, "I/F"] = "Foreign"
                    df_master.loc[index, "USD/INR"] = "USD"

                df_master.loc[index, "GSTIN"] = gstin
                df_master.loc[index, "NET Term"] = net_term

            # Ensure required columns exist
            for col in ["C DSP $", "C SSP $", "C Net $", "Category (DSP/SSP)"]:
                if col not in df_master.columns:
                    df_master[col] = 0.0

            st.divider()
            
            # ===== AGGRID MASTER DATA TABLE =====

            from st_aggrid import AgGrid, GridOptionsBuilder, JsCode

            # Ensure numeric columns
            numeric_cols = [
                "DSP $ (BC)",
                "SSP $ (BC)",
                "Net $ (BC)",
                "C DSP $",
                "C SSP $",
                "C Net $",
            ]

            df_master["C DSP $"] = pd.to_numeric(df_master["C DSP $"], errors="coerce").fillna(0)
            df_master["C SSP $"] = pd.to_numeric(df_master["C SSP $"], errors="coerce").fillna(0)

            for _sc_col in ["GSTIN", "NET Term", "I/F", "USD/INR", "Category (DSP/SSP)"]:
                if _sc_col in df_master.columns:
                    df_master[_sc_col] = df_master[_sc_col].fillna("").astype(str).replace("nan", "").replace("None", "")

            month_comparator = JsCode("""
            function(date1, date2) {
                function parseMonth(str) {
                    if (!str) return new Date(0);
                    const [mon, year] = str.split("-");
                    return new Date(mon + " 1, " + year);
                }
                const d1 = parseMonth(date1);
                const d2 = parseMonth(date2);
                return d1 - d2;
            }
            """)
            
            # -------- GRID BUILDER --------
            from st_aggrid import GridOptionsBuilder, JsCode
            
            # Currency formatter
            currency_formatter = JsCode("""
            function(params) {
                if (params.value == null || params.value === '') return '';
                return '$' + parseFloat(params.value).toLocaleString(undefined, {minimumFractionDigits: 2});
            }
            """)
            
            bg_style_js = JsCode("""
            function(params) {

                if (params.node.rowPinned) {
                    return {};   // skip footer
                }

                let col = params.colDef.field;

                if (col === "C DSP $" || col === "C SSP $") {
                    return { backgroundColor: "#FFF4E5" };   // very light orange
                }

                if (col === "C Net $") {
                    return { backgroundColor: "#E8F5E9" };   // light green
                }

                if (col === "Net $ (BC)") {
                    return { backgroundColor: "#F3E5F5" };   // light purple
                    fontWeight: "bold"
                }

                return {};
            }
            """)
            
            gb = GridOptionsBuilder.from_dataframe(df_master)
            if "id" in df_master.columns:
                gb.configure_column("id", hide=True)
            
            gb.configure_column(
                "Month",
                comparator=month_comparator,
                type=["textColumn"],
                editable=False
            )
            
            negative_style = JsCode("""
            function(params) {
                if (params.value < 0) {
                    return {
                        color: 'red',
                        fontWeight: 'bold'
                    };
                }
            }
            """)
            
            # Editable Table
            numeric_cols = [
                "DSP $ (BC)",
                "SSP $ (BC)",
                "Net $ (BC)",
                "C DSP $",
                "C SSP $",
                "C Net $",
            ]
            
            editable_cols = ["C DSP $", "C SSP $"]
            
            # Ensure numeric conversion before grid
            for col in numeric_cols:
                if col in df_master.columns:
                    df_master[col] = pd.to_numeric(df_master[col], errors="coerce").fillna(0)

            # Apply uniform currency formatting to ALL numeric columns
            for col in numeric_cols:
                if col in df_master.columns:
                    gb.configure_column(
                        col,
                        type=["numericColumn"],
                        editable=(col in editable_cols),
                        valueFormatter=currency_formatter,
                        cellStyle=bg_style_js
                    )

                                    
            from st_aggrid import JsCode

            # C Net = C DSP - C SSP (Excel style)
            net_value_getter = JsCode("""
            function(params) {
                let dsp = parseFloat(params.data["C DSP $"]) || 0;
                let ssp = parseFloat(params.data["C SSP $"]) || 0;
                return dsp - ssp;
            }
            """)

            # Category depends on C Net
            category_value_getter = JsCode("""
            function(params) {
            
                // 🚫 Do NOT apply to footer row
                if (params.node.rowPinned) {
                    return "";
                }
                
                let dsp = parseFloat(params.data["C DSP $"]) || 0;
                let ssp = parseFloat(params.data["C SSP $"]) || 0;
                let net = dsp - ssp;
                return net >= 0 ? "DSP" : "SSP";
            }
            """)

            gb.configure_column(
                "C Net $",
                editable=False,
                type=["numericColumn"],
                valueGetter=net_value_getter,
                valueFormatter=currency_formatter,
                cellStyle=JsCode("""
                    function(params) {

                        if (params.node.rowPinned) return {};

                        let style = {
                            backgroundColor: "#E8F5E9",   // light green
                            fontWeight: "bold"
                        };

                        if (params.value < 0) {
                            style.color = "red";
                        }

                        return style;
                    }
                """)
            )
            
            gb.configure_column(
                "Net $ (BC)",
                type=["numericColumn"],
                valueFormatter=currency_formatter,
                cellStyle=JsCode("""
                    function(params) {

                        if (params.node.rowPinned) return {};

                        let style = {
                            backgroundColor: "#F3E5F5",   // light purple
                            fontWeight: "bold"           // <-- BOLD ALWAYS
                        };

                        if (params.value < 0) {
                            style.color = "red";
                        }

                        return style;
                    }
                """)
            )

            gb.configure_column(
                "Category (DSP/SSP)",
                editable=False,
                valueGetter=category_value_getter
            )
                    
            # Freeze first 2 columns
            first_two_cols = df_master.columns[:2]

            for col in first_two_cols:
                gb.configure_column(col, pinned="left")
                
            from st_aggrid import JsCode

            pinned_style = JsCode("""
            function(params) {
                if (params.column.pinned) {
                    return {
                        fontWeight: 'bold',
                        borderRight: '2px solid #003366'
                    };
                }
            }
            """)

            for col in first_two_cols:
                gb.configure_column(
                    col,
                    pinned="left",
                    flex=1.4,
                    minWidth=100,
                    cellStyle=pinned_style
                )
            
            gb.configure_default_column(
                resizable=True,
                sortable=True,
                filter=False,
                flex=1,
                minWidth=90
            )

            # Red negative styling
            negative_style = JsCode("""
            function(params) {
                if (params.value < 0) {
                    return {color: 'red', fontWeight: 'bold'};
                }
            }
            """)

            # -------- REAL-TIME GRAND TOTAL (JS BASED) --------

            grand_total_js = JsCode("""
            function(api) {
                let totals = {};
                let numericCols = ["DSP $ (BC)", "SSP $ (BC)", "Net $ (BC)", "C DSP $", "C SSP $", "C Net $"];

                numericCols.forEach(col => totals[col] = 0);

                api.forEachNodeAfterFilter(function(node) {
                    numericCols.forEach(function(col) {
                        let val = parseFloat(node.data[col]);
                        if (!isNaN(val)) {
                            totals[col] += val;
                        }
                    });
                });

                let totalRow = { Month: "Grand Total" };
                numericCols.forEach(col => totalRow[col] = totals[col]);

                api.setPinnedBottomRowData([totalRow]);
            }
            """)
            
            gridOptions = gb.build()
            gridOptions["onFirstDataRendered"] = grand_total_js
            gridOptions["onFilterChanged"] = grand_total_js
            gridOptions["onModelUpdated"] = grand_total_js
            gridOptions["onCellValueChanged"] = grand_total_js
            gridOptions["onGridReady"] = JsCode("""
                function(params) { params.api.sizeColumnsToFit(); }
            """)
            
            if search_text:
                gridOptions["quickFilterText"] = search_text

            # Footer styling
            gridOptions["getRowStyle"] = JsCode("""
            function(params) {
                if (params.node.rowPinned) {
                    return {
                        backgroundColor: '#003366',
                        color: 'white',
                        fontWeight: 'bold',
                        fontSize: '14px'
                    };
                }
            }
            """)
            
            custom_css = {
                ".ag-header": {
                    "background-color": "#003366 !important",
                    "color": "white !important",
                    "font-weight": "bold !important",
                    "font-size": "14px !important"
                },
                ".ag-header-cell-label": {
                    "color": "white !important",
                    "font-weight": "bold !important"
                }
            }
            
            st.markdown("""
            <script>
            document.addEventListener("DOMContentLoaded", function() {
                const counters = document.querySelectorAll(".kpi-value");
                counters.forEach(counter => {
                    const updateCount = () => {
                        const target = +counter.getAttribute("data-value");
                        const duration = 800;
                        const stepTime = 20;
                        const steps = duration / stepTime;
                        let count = 0;
                        const increment = target / steps;

                        const timer = setInterval(() => {
                            count += increment;
                            if (Math.abs(count) >= Math.abs(target)) {
                                counter.innerText = counter.innerText.includes('%')
                                    ? target.toFixed(2) + '%'
                                    : '$' + target.toLocaleString(undefined, {minimumFractionDigits: 2});
                                clearInterval(timer);
                            } else {
                                counter.innerText = counter.innerText.includes('%')
                                    ? count.toFixed(2) + '%'
                                    : '$' + count.toLocaleString(undefined, {minimumFractionDigits: 2});
                            }
                        }, stepTime);
                    };
                    updateCount();
                });
            });
            </script>
            """, unsafe_allow_html=True)

            # -------- GRAND TOTAL (Search + Month Reactive) --------

            filtered_df = df_master.copy()

            if search_text:
                search_lower = search_text.lower()

                mask = filtered_df.apply(
                    lambda row: row.astype(str).str.lower().str.contains(search_lower).any(),
                    axis=1
                )
                filtered_df = filtered_df[mask]

            grand_total_values = filtered_df.sum(numeric_only=True)

            grand_total_row = {col: "" for col in df_master.columns}
            grand_total_row["Month"] = "Grand Total"

            for col in numeric_cols:
                if col in grand_total_row:
                    grand_total_row[col] = grand_total_values.get(col, 0)

            gridOptions["pinnedBottomRowData"] = [grand_total_row]
            
            from st_aggrid import GridUpdateMode
            
            grid_df = (
                df_master
                .reset_index(drop=True)
                .loc[:, ~df_master.columns.duplicated()]
                .copy()
            )

            grid_response = AgGrid(
                grid_df,
                gridOptions=gridOptions,
                allow_unsafe_jscode=True,
                update_on=["cellValueChanged"],
                data_return_mode="AS_INPUT",
                fit_columns_on_grid_load=True,
                height=sc.grid_height(750),
                custom_css=custom_css
            )
            
                    
            if grid_response["selected_rows"] is not None:
                pass  # ignore selection

            # =========================
            # AUTO SAVE ON EDIT (NO READ VERSION)
            # =========================

            if grid_response and grid_response.get("event") == "cellValueChanged":

                updated_df = pd.DataFrame(grid_response["data"]).reset_index(drop=True)

                editable_cols = ["C DSP $", "C SSP $"]

                for col in editable_cols:
                    updated_df[col] = pd.to_numeric(updated_df[col], errors="coerce").fillna(0)

                previous_df = st.session_state.master_df.reset_index(drop=True)

                if not updated_df[editable_cols].equals(previous_df[editable_cols]):

                    batch_requests = []

                    for idx in range(len(updated_df)):

                        old_row = previous_df.iloc[idx]
                        new_row = updated_df.iloc[idx]

                        if (
                            float(old_row["C DSP $"]) != float(new_row["C DSP $"]) or
                            float(old_row["C SSP $"]) != float(new_row["C SSP $"])
                        ):

                            sheet_row_number = idx + 2

                            batch_requests.append({
                                "range": f"F{sheet_row_number}:H{sheet_row_number}",
                                "values": [[
                                    float(new_row["C DSP $"]),
                                    float(new_row["C SSP $"]),
                                    float(new_row["C DSP $"]) - float(new_row["C SSP $"])
                                ]]
                            })

                    if not updated_df.empty and "id" in updated_df.columns:

                        for idx in range(len(updated_df)):
                            old_row = previous_df.iloc[idx]
                            new_row = updated_df.iloc[idx]

                            if (
                                float(old_row["C DSP $"]) != float(new_row["C DSP $"]) or
                                float(old_row["C SSP $"]) != float(new_row["C SSP $"])
                            ):
                                row_id = new_row["id"]

                                conn = get_db_connection()
                                cursor = conn.cursor()

                                cursor.execute("""
                                UPDATE master_data
                                SET c_dsp = ?, c_ssp = ?
                                WHERE id = ?
                                """, (
                                    float(new_row["C DSP $"]),
                                    float(new_row["C SSP $"]),
                                    int(row_id)
                                ))

                                conn.commit()
                                conn.close()

                        load_master_data.clear()
                        st.session_state.master_df = load_master_data()
                        st.success("Master Data updated successfully")
                        st.rerun()
                                           
            # RED negative styling
            def highlight_negative(val):
                if isinstance(val, (int, float)) and val < 0:
                    return "color: red; font-weight: bold;"
                return ""

                   
        else:
            st.warning("No Master Data Found in Excel")


# ====================================================
# DSP (CUSTOMERS) TAB
# ====================================================

if "DSP (Customers)" in tabs:
    with tabs["DSP (Customers)"]:
        
        st.markdown('''
                <div style="background:linear-gradient(135deg,#003366 0%,#005599 100%);
                border-radius:10px;padding:18px 24px;margin-bottom:20px;
                box-shadow:0 4px 16px rgba(0,51,102,.2);
                display:flex;align-items:center;height:55px;gap:14px;">
                <div style="font-size:32px;">🧑</div>
                <div>
                    <div style="color:white;font-size:20px;font-weight:800;">
                        DSP (Customers)</div>
                </div></div>
            ''', unsafe_allow_html=True)

        # ==============================
        # FILTERS  (Invoice Manager style)
        # ==============================
        _dsp_fy_list = _tab_fy_list()
        col2, col3, col4, col5, col6 = st.columns([0.3, 0.3, 0.3, 0.5, 0.5])

        with col2:
            selected_fy = st.selectbox(
                "Financial Year",
                options=["All"] + _dsp_fy_list,
                index=0,
                key="dsp_fy"
            )

        _dsp_q = st.session_state.get("dsp_quarter", "All")
        _dsp_m = st.session_state.get("dsp_month",   "All")

        with col3:
            selected_month = st.selectbox(
                "Month",
                options=["All"] + _fy_months(selected_fy),
                index=0,
                key="dsp_month",
                disabled=(_dsp_q not in ("", "All")),
            )

        with col4:
            selected_quarter = st.selectbox(
                "Quarter",
                options=["All"] + _QUARTER_OPTS,
                index=0,
                key="dsp_quarter",
                disabled=(_dsp_m not in ("", "All")),
            )

        with col5:
            search_text = st.text_input(
                "🔍 Search DSP",
                placeholder="Global search...",
                key="dsp_search"
            )

        with col6:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("🔄 Refresh", key="dsp_refresh_sel"):
                if "dsp_df" in st.session_state:
                    del st.session_state["dsp_df"]
                if "dsp_edit_df" in st.session_state:
                    del st.session_state["dsp_edit_df"]
                st.rerun()

        st.divider()

        # ==============================
        # LOAD DATA (MASTER + DB MERGE)
        # ==============================
        if "dsp_df" not in st.session_state:
            st.session_state.dsp_df = load_dsp_final(
                st.session_state.master_df,
                st.session_state.partner_df
            )

        df = st.session_state.dsp_df.copy()

        if df.empty:
            st.warning("No DSP Data Found")
            st.stop()

        # ==============================
        # TYPE FIXES
        # ==============================
        df["Month"]        = pd.to_datetime(df["Month"],    errors="coerce")
        df["Due Date"]     = pd.to_datetime(df["Due Date"], errors="coerce")
        df["Receivable $"]    = pd.to_numeric(df["Receivable $"],    errors="coerce").fillna(0)
        df["Received Amount $"]= pd.to_numeric(df["Received Amount $"], errors="coerce").fillna(0)

        # ==============================
        # DERIVED
        # ==============================
        df["Shortage"] = df["Receivable $"] - df["Received Amount $"]
        today = pd.Timestamp.today()
        df["Age"] = (today - df["Due Date"]).dt.days
        
        # Invoice Status — multi-month aware, reads actual status from invoice_details
        try:
            _inv_conn = get_db_connection()
            _inv_df   = pd.read_sql(
                "SELECT dsp_name, month, status FROM invoice_details WHERE is_deleted=0",
                _inv_conn
            )
            _inv_conn.close()
            _inv_keys      = set()   # dsp|month → invoice exists (any status)
            _inv_paid_keys = set()   # dsp|month → Paid
            _inv_part_keys = set()   # dsp|month → Partially Paid
            for _, _ir in _inv_df.iterrows():
                _dsp_k  = str(_ir["dsp_name"]).strip()
                _status = str(_ir.get("status","")).strip()
                for _raw_m in str(_ir["month"]).split(","):
                    _raw_m = _raw_m.strip()
                    try:
                        _nm = pd.to_datetime(_raw_m, format="%b-%Y", errors="coerce")
                        if pd.isna(_nm): _nm = pd.to_datetime(_raw_m, errors="coerce")
                        _fmt_m = _nm.strftime("%b-%Y") if pd.notna(_nm) else _raw_m
                    except Exception:
                        _fmt_m = _raw_m
                    _key = _dsp_k + "|" + _fmt_m
                    _inv_keys.add(_key)
                    if _status == "Paid":            _inv_paid_keys.add(_key)
                    elif _status == "Partially Paid": _inv_part_keys.add(_key)
        except Exception:
            _inv_keys = set(); _inv_paid_keys = set(); _inv_part_keys = set()

        def _inv_status(row):
            _dsp_r = str(row.get("DSP Name","")).strip()
            try:
                _m_r   = pd.to_datetime(row.get("Month",""), errors="coerce")
                _m_str = _m_r.strftime("%b-%Y") if pd.notna(_m_r) else ""
            except Exception:
                _m_str = ""
            k = _dsp_r + "|" + _m_str
            if k in _inv_part_keys:  return "Partially Paid"
            if k in _inv_paid_keys:  return "Paid"
            if k in _inv_keys:       return "Raised"
            return "Pending"

        df["Invoice Status"] = df.apply(_inv_status, axis=1)

        # ==============================
        # APPLY FILTERS (Invoice Manager style)
        # ==============================
        df_filtered = df.copy()

        if selected_fy not in ("", "All"):
            _dsp_fy_start, _dsp_fy_end = _fy_date_range(selected_fy)
            df_filtered = df_filtered[
                (df_filtered["Month"] >= _dsp_fy_start) &
                (df_filtered["Month"] <= _dsp_fy_end)
            ]
            if selected_quarter not in ("", "All"):
                q_start, q_end = _quarter_date_range(selected_fy, selected_quarter)
                df_filtered = df_filtered[
                    (df_filtered["Month"] >= q_start) &
                    (df_filtered["Month"] <= q_end)
                ]
            elif selected_month not in ("", "All"):
                df_filtered = df_filtered[
                    df_filtered["Month"].dt.strftime("%b-%Y") == selected_month
                ]

        # ==============================
        # FORMAT FOR DISPLAY
        # ==============================
        df = df_filtered.copy()
        df.columns = df.columns.str.strip()
        df = df.sort_values(["Month", "DSP Name"])

        if "dsp_edit_df" not in st.session_state or len(st.session_state.dsp_edit_df) != len(df):
            st.session_state.dsp_edit_df = df.copy()

        df["Month"]        = pd.to_datetime(df["Month"],    errors="coerce").dt.strftime("%b-%Y")
        df["Due Date"]     = pd.to_datetime(df["Due Date"], errors="coerce").dt.strftime("%d-%b-%Y")
        df["Received Date"] = pd.to_datetime(df["Received Date"], errors="coerce")
        df["Received Date"] = df["Received Date"].dt.strftime("%Y-%m-%d").where(df["Received Date"].notna(), other="")
        df["Received In"]  = df["Received In"].fillna("Select").replace("", "Select")
        df["Reason"]   = df["Reason"].fillna("")

        # ==============================
        # AGGRID CONFIG
        # ==============================
        from st_aggrid import AgGrid, GridOptionsBuilder, JsCode
        from st_aggrid import GridUpdateMode, DataReturnMode

        dsp_currency_formatter = JsCode("""
        function(params) {
            if (params.value == null || params.value === '') return '';
            return '$' + parseFloat(params.value).toLocaleString(undefined, {minimumFractionDigits: 2});
        }
        """)

        dsp_date_formatter = JsCode("""
        function(params) {
            if (!params.value) return '';
            var d = new Date(params.value);
            if (isNaN(d)) return params.value;
            var day   = String(d.getDate()).padStart(2, '0');
            var month = d.toLocaleString('en-GB', { month: 'short' });
            var year  = d.getFullYear();
            return day + '-' + month + '-' + year;
        }
        """)

        gb_dsp = GridOptionsBuilder.from_dataframe(df)

        # Currency columns (non-calculated)
        for col in ["Receivable $", "Received Amount $"]:
            if col in df.columns:
                gb_dsp.configure_column(col, type=["numericColumn"], valueFormatter=dsp_currency_formatter)

        # Shortage — live JS formula like Excel
        gb_dsp.configure_column(
            "Shortage",
            type=["numericColumn"],
            valueFormatter=dsp_currency_formatter,
            valueGetter=JsCode("""
                function(params) {
                    var receivable = parseFloat(params.data["Receivable $"]) || 0;
                    var received    = parseFloat(params.data["Received Amount $"]) || 0;
                    return Math.round((receivable - received) * 100) / 100;
                }
            """),
            editable=False
        )

        # Editable fields
        gb_dsp.configure_column("Received Amount $", editable=True)
        gb_dsp.configure_column("Reason",        editable=True)

        # Received In dropdown
        gb_dsp.configure_column(
            "Received In",
            editable=True,
            cellEditor="agSelectCellEditor",
            cellEditorParams={
                "values": ["Select", "PayPal", "Payoneer", "Bank Remittance", "Bank Transfer"]
            }
        )

        # Received Date — date picker with clean string output
        gb_dsp.configure_column(
            "Received Date",
            editable=True,
            cellEditor=JsCode("""
                class SspDatePickerEditor {
                    init(params) {
                        this.value = params.value || '';
                        this.input = document.createElement('input');
                        this.input.type = 'date';
                        this.input.style.width = '100%';
                        this.input.style.height = '100%';
                        this.input.style.border = 'none';
                        this.input.style.outline = 'none';
                        this.input.style.fontSize = '14px';
                        this.input.style.padding = '2px';
                        if (this.value) {
                            try {
                                var d = new Date(this.value);
                                if (!isNaN(d)) {
                                    var yyyy = d.getFullYear();
                                    var mm   = String(d.getMonth()+1).padStart(2,'0');
                                    var dd   = String(d.getDate()).padStart(2,'0');
                                    this.input.value = yyyy+'-'+mm+'-'+dd;
                                }
                            } catch(e) {}
                        }
                    }
                    getGui() { return this.input; }
                    afterGuiAttached() {
                        this.input.focus();
                        try { this.input.showPicker(); } catch(e) {}
                    }
                    getValue() { return this.input.value || null; }
                    isPopup() { return false; }
                    destroy() {}
                }
            """),
            valueFormatter=dsp_date_formatter,
            headerTooltip="Click cell to pick a date"
        )

        gb_dsp.configure_column("Month",              pinned="left", flex=1,   minWidth=80)
        gb_dsp.configure_column("DSP Name",           pinned="left", flex=1.4, minWidth=110)
        gb_dsp.configure_column("Receivable $",                      flex=1.2, minWidth=95)
        gb_dsp.configure_column("Due Date",                          flex=1.1, minWidth=90)
        gb_dsp.configure_column("Received Date",                     flex=1.2, minWidth=95)
        gb_dsp.configure_column("Received Amount $",                 flex=1.3, minWidth=105)
        gb_dsp.configure_column("Received In",                       flex=1.2, minWidth=95)
        gb_dsp.configure_column("Shortage",                          flex=1.1, minWidth=90)
        gb_dsp.configure_column("Reason",                            flex=1.1, minWidth=90)
        gb_dsp.configure_column("Age",                               flex=0.6, minWidth=50)
        gb_dsp.configure_column("Invoice Status",
            editable=False,
            flex=1.2,
            minWidth=100,
            cellStyle=JsCode("""
                function(params) {
                    if (params.value === 'Raised')  return {color:'#155724', backgroundColor:'#d4edda', fontWeight:'bold'};
                    if (params.value === 'Pending') return {color:'#856404', backgroundColor:'#fff3cd', fontWeight:'bold'};
                    return {};
                }
            """)
        )
        gb_dsp.configure_default_column(resizable=True, sortable=True)

        # Row colour style
        dsp_row_style = JsCode("""
        function(params) {
            if (params.node.rowPinned) {
                return { backgroundColor: '#003366', color: 'white', fontWeight: 'bold', fontSize: '14px' };
            }
            var status = (params.data["Invoice Status"] || "").toLowerCase().trim();
            if (status === "paid")            return { backgroundColor: "#90ee90" };
            if (status === "partially paid")  return { backgroundColor: "#fff8e1" };
            if (status === "raised")          return { backgroundColor: "#fdecea" };
            if (status === "pending")         return { backgroundColor: "#fdecea" };
            return {};
        }
        """)

        # Grand total footer
        dsp_total_js = JsCode("""
        function(api) {
            var totalPayable = 0; var totalPaid = 0;
            api.forEachNodeAfterFilter(function(node) {
                totalPayable += parseFloat(node.data["Receivable $"])    || 0;
                totalPaid    += parseFloat(node.data["Received Amount $"]) || 0;
            });
            api.setPinnedBottomRowData([{
                "Month": "Grand Total",
                "Receivable $": Number(totalPayable),
                "Received Amount $": Number(totalPaid),
                "Shortage": Number(totalPayable - totalPaid)
            }]);
        }
        """)

        dsp_grid_options = gb_dsp.build()
        dsp_grid_options["stopEditingWhenCellsLoseFocus"] = True
        dsp_grid_options["stopEditingWhenGridLosesFocus"] = True
        dsp_grid_options["getRowStyle"]         = dsp_row_style
        dsp_grid_options["onFirstDataRendered"] = dsp_total_js
        dsp_grid_options["onCellValueChanged"]  = dsp_total_js
        dsp_grid_options["onFilterChanged"]     = dsp_total_js
        dsp_grid_options["onGridReady"]         = JsCode("""
            function(params) { params.api.sizeColumnsToFit(); }
        """)

        # Pinned grand total bottom row
        total_receivable = float(df["Receivable $"].sum())
        total_received    = float(df["Received Amount $"].sum())
        dsp_grid_options["pinnedBottomRowData"] = [{
            "Month": "Grand Total",
            "Receivable $": total_receivable,
            "Received Amount $": total_received,
            "Shortage": float(total_receivable - total_received)
        }]

        if search_text:
            dsp_grid_options["quickFilterText"] = search_text

        dsp_custom_css = {
            ".ag-header": {
                "background-color": "#003366 !important",
                "color": "white !important",
                "font-weight": "bold !important"
            }
        }

        # Ensure JSON-serializable
        df_dsp_grid = df.copy()
        for col in df_dsp_grid.columns:
            if pd.api.types.is_numeric_dtype(df_dsp_grid[col]):
                df_dsp_grid[col] = df_dsp_grid[col].apply(lambda x: float(x) if pd.notnull(x) else None)
        df_dsp_grid = df_dsp_grid.where(pd.notnull(df_dsp_grid), None)

        dsp_grid_response = AgGrid(
            df_dsp_grid,
            gridOptions=dsp_grid_options,
            height=sc.grid_height(650),
            update_mode=GridUpdateMode.MODEL_CHANGED,
            data_return_mode=DataReturnMode.AS_INPUT,
            allow_unsafe_jscode=True,
            custom_css=dsp_custom_css,
            fit_columns_on_grid_load=True,
            enable_enterprise_modules=False,
        )

        # ==============================
        # SAVE BUTTON
        # ==============================
        if dsp_grid_response["data"] is not None:

            dsp_edited_df = pd.DataFrame(dsp_grid_response["data"])
            dsp_edited_df = dsp_edited_df.reset_index(drop=True)

            # Clean Received Date
            dsp_edited_df["Received Date"] = dsp_edited_df["Received Date"].apply(
                lambda x: None if x in [{}, "", None] else x
            )
            dsp_edited_df["Received Date"] = pd.to_datetime(dsp_edited_df["Received Date"], errors="coerce")
            dsp_edited_df["Due Date"]     = pd.to_datetime(dsp_edited_df["Due Date"],     errors="coerce")
            dsp_edited_df["Received Amount $"]= pd.to_numeric(dsp_edited_df["Received Amount $"], errors="coerce").fillna(0)

            # Validation
            dsp_errors = []
            for idx, row in dsp_edited_df.iterrows():
                if (
                    pd.notna(row["Received Date"])
                    or row["Received Amount $"] > 0
                    or row["Received In"] not in ["", "Select", None]
                ):
                    if pd.isna(row["Received Date"]):
                        dsp_errors.append(f"Row {int(idx)+1}: Invalid Received Date")
                    if row["Received In"] in ["", "Select", None]:
                        dsp_errors.append(f"Row {int(idx)+1}: Select Received In")

            save_dsp_clicked = st.button("💾 Save DSP Changes", key="dsp_save_btn")

            if dsp_errors:
                st.error("\n".join(dsp_errors))
            elif save_dsp_clicked:

                def clean_dsp_date(val):
                    if val in [{}, "", None, "None", "NaT"]:
                        return None
                    if isinstance(val, dict):
                        try:
                            y = int(val.get("year", 0))
                            m = int(val.get("month", 0))
                            d = int(val.get("date") or val.get("day", 0))
                            if y and m and d:
                                return pd.Timestamp(year=y, month=m, day=d)
                        except:
                            return None
                    try:
                        result = pd.to_datetime(val, dayfirst=True, errors="coerce")
                        return None if pd.isna(result) else result
                    except:
                        return None

                dsp_edited_df["Received Date"] = dsp_edited_df["Received Date"].apply(clean_dsp_date)
                dsp_edited_df["Received Date"] = dsp_edited_df["Received Date"].apply(
                    lambda x: x.strftime("%Y-%m-%d") if pd.notna(x) else None
                )
                dsp_edited_df["Due Date"] = pd.to_datetime(dsp_edited_df["Due Date"], errors="coerce")
                dsp_edited_df["Received Amount $"] = pd.to_numeric(
                    dsp_edited_df["Received Amount $"], errors="coerce"
                ).fillna(0)
                dsp_edited_df["Month"] = pd.to_datetime(dsp_edited_df["Month"], errors="coerce")
                dsp_edited_df["Month"] = dsp_edited_df["Month"].dt.strftime("%b-%Y")

                # Save to DB
                upsert_dsp_data(dsp_edited_df)

                # ── Sync invoice_details status based on DSP payment totals ──
                try:
                    _sync_conn = get_db_connection()
                    _all_inv   = _sync_conn.execute(
                        "SELECT invoice_number, dsp_name, month, status "
                        "FROM invoice_details WHERE is_deleted=0 AND is_credit_note=0"
                    ).fetchall()

                    def _norm_month_s(m):
                        """Normalise any month format to 'Mon-YYYY' e.g. Aug-2025."""
                        try:
                            s = str(m).strip()
                            # Try Mon-YYYY first
                            _p = pd.to_datetime(s, format="%b-%Y", errors="coerce")
                            if pd.isna(_p):
                                # Try ISO date (2025-08-01) or any other format
                                _p = pd.to_datetime(s, errors="coerce")
                            return _p.strftime("%b-%Y") if pd.notna(_p) else s
                        except Exception:
                            return str(m).strip()

                    def _usd_s(v):
                        try:
                            s = str(v or 0).replace("$","").replace(",","").strip()
                            return float(s) if s else 0.0
                        except Exception:
                            return 0.0

                    # Read ALL dsp_data from DB AFTER upsert — gets fresh committed values
                    # This ensures we use the actually saved amounts, not stale grid data
                    _all_dsp_df = pd.read_sql("SELECT * FROM dsp_data", _sync_conn)
                    _dsp_lk = {}
                    for _, _dr in _all_dsp_df.iterrows():
                        _raw_m = str(_dr.get("Month",""))
                        _key = (str(_dr.get("DSP Name","")).strip(), _norm_month_s(_raw_m))
                        _dsp_lk[_key] = {
                            "received":   _usd_s(_dr.get("Received Amount $", 0)),
                            "receivable": _usd_s(_dr.get("Receivable $",      0)),
                        }

                    for _inv_no, _inv_dsp, _inv_month_raw, _inv_status in _all_inv:
                        _dsp_k = str(_inv_dsp).strip()
                        _parts = [_norm_month_s(m) for m in str(_inv_month_raw).split(",") if m.strip()]
                        _tot_recv = 0.0; _tot_recv_able = 0.0; _found = False
                        for _mp in _parts:
                            _dk = (_dsp_k, _mp)
                            if _dk not in _dsp_lk: continue
                            _found         = True
                            _tot_recv      += _dsp_lk[_dk]["received"]
                            _tot_recv_able += _dsp_lk[_dk]["receivable"]

                        if not _found: continue

                        # Round to 2dp to avoid float precision issues e.g. 5.37 >= 5.37
                        _tot_recv      = round(_tot_recv,      2)
                        _tot_recv_able = round(_tot_recv_able, 2)

                        if _tot_recv_able > 0 and _tot_recv >= _tot_recv_able:
                            _new_st = "Paid"
                        elif _tot_recv > 0:
                            _new_st = "Partially Paid"
                        else:
                            _new_st = "Raised" if _inv_status in ("Paid","Partially Paid") else _inv_status

                        if _new_st != _inv_status:
                            _sync_conn.execute(
                                "UPDATE invoice_details SET status=? WHERE invoice_number=?",
                                (_new_st, _inv_no)
                            )

                    _sync_conn.commit()
                    _sync_conn.close()
                except Exception as _se:
                    st.warning(f"Invoice status sync warning: {_se}")

                st.success("DSP (Customers) saved successfully ✅")

                # Clear session state so fresh data loads on rerun
                if "dsp_df" in st.session_state:
                    del st.session_state["dsp_df"]
                if "dsp_edit_df" in st.session_state:
                    del st.session_state["dsp_edit_df"]

                st.rerun()
                
# ====================================================
# 🧾 INVOICE MODULE
# ====================================================
if "Invoice Manager" in tabs:
    with tabs["Invoice Manager"]:
        
        if "dsp_df" in st.session_state and not st.session_state.dsp_df.empty:
            render_invoice_module(st.session_state.dsp_df)

# ====================================================
# ====================================================
# 📤 UPLOAD INVOICE DIALOG (SSP)
# ====================================================

@st.dialog("📤 Upload SSP Invoice", width="large")
def upload_invoice_dialog(ssp_df: pd.DataFrame):
    """Modal popup for uploading an SSP invoice to Dropbox (supports multiple months)."""

    if ssp_df is None or ssp_df.empty:
        st.warning("No SSP data available.")
        return

    # --- SSP selector ---
    ssp_names = sorted(ssp_df["SSP Name"].dropna().unique().tolist())
    selected_ssp = st.selectbox(
        "Select SSP Partner",
        options=["Select SSP"] + ssp_names,
        index=0,
        key="inv_dlg_ssp"
    )

    if selected_ssp == "Select SSP":
        st.info("Please select an SSP partner to continue.")
        return

    # --- Multi-month selector: only months without an invoice already ---
    ssp_rows = ssp_df[ssp_df["SSP Name"] == selected_ssp].copy()

    # Exclude months where invoice is already received or paid
    inv_done_statuses = {"INV Received", "INV Paid"}
    if "Invoice Status" in ssp_rows.columns:
        ssp_rows = ssp_rows[
            ~ssp_rows["Invoice Status"].isin(inv_done_statuses)
        ]

    ssp_months_raw = ssp_rows["Month"].dropna().tolist()
    ssp_months_sorted = sorted(
        ssp_months_raw,
        key=lambda m: pd.to_datetime(m, format="%b-%Y", errors="coerce")
    )

    if not ssp_months_sorted:
        st.success("✅ All months for this SSP already have an invoice uploaded.")
        return

    selected_months = st.multiselect(
        "Select Month(s)",
        options=ssp_months_sorted,
        default=[],
        key="inv_dlg_months",
        help="Only months without an existing invoice are shown"
    )

    # --- Amount breakdown for selected months ---
    if selected_months:
        rows_sel = ssp_df[
            (ssp_df["SSP Name"] == selected_ssp) &
            (ssp_df["Month"].isin(selected_months))
        ][["Month", "Payable $", "Paid Amount $"]].copy()

        rows_sel["Payable $"]    = pd.to_numeric(rows_sel["Payable $"],    errors="coerce").fillna(0)
        rows_sel["Paid Amount $"]= pd.to_numeric(rows_sel["Paid Amount $"],errors="coerce").fillna(0)

        total_payable = rows_sel["Payable $"].sum()
        total_paid    = rows_sel["Paid Amount $"].sum()

        # Show per-month breakdown table
        display_rows = rows_sel.copy()
        display_rows["Payable $"]     = display_rows["Payable $"].apply(lambda x: f"${x:,.2f}")
        display_rows["Paid Amount $"] = display_rows["Paid Amount $"].apply(lambda x: f"${x:,.2f}")
        st.dataframe(display_rows.set_index("Month"), use_container_width=True)

        col_pay, col_paid = st.columns(2)
        col_pay.metric("Total Payable $",    f"${total_payable:,.2f}")
        col_paid.metric("Total Paid Amount $", f"${total_paid:,.2f}")
    else:
        total_payable = 0.0
        total_paid    = 0.0
        st.info("Select at least one month to see the amount details.")

    # --- File uploader ---
    uploaded_file = st.file_uploader(
        "Upload Invoice",
        type=["pdf", "jpg", "jpeg", "doc", "docx"],
        key="inv_dlg_file",
        help="One file will be saved to Dropbox and linked to all selected months"
    )

    st.divider()

    col_upload, col_cancel = st.columns([1, 1])

    with col_upload:
        if st.button("📤 Upload to Dropbox", type="primary", use_container_width=True):
            if not selected_months:
                st.error("Please select at least one month.")
            elif uploaded_file is None:
                st.error("Please select a file before uploading.")
            else:
                file_ext   = uploaded_file.name.rsplit(".", 1)[-1].lower()
                file_bytes = uploaded_file.read()

                # Build a compact month label for the filename
                # Single month → "Mar2025", multiple → "Mar2025-May2025"
                sorted_sel = sorted(
                    selected_months,
                    key=lambda m: pd.to_datetime(m, format="%b-%Y", errors="coerce")
                )
                if len(sorted_sel) == 1:
                    month_label = sorted_sel[0].replace("-", "")          # Mar2025
                else:
                    first = sorted_sel[0].replace("-", "")
                    last  = sorted_sel[-1].replace("-", "")
                    month_label = f"{first}-{last}"                        # Mar2025-May2025

                with st.spinner("Uploading to Dropbox…"):
                    try:
                        dropbox_path = upload_invoice_to_dropbox(
                            file_bytes, selected_ssp, month_label, file_ext
                        )

                        # Update status for every selected month individually
                        updated = []
                        for month in sorted_sel:
                            month_rows = ssp_df[
                                (ssp_df["SSP Name"] == selected_ssp) &
                                (ssp_df["Month"] == month)
                            ]
                            m_payable = float(month_rows["Payable $"].values[0])    if not month_rows.empty else 0.0
                            m_paid    = float(month_rows["Paid Amount $"].values[0]) if not month_rows.empty else 0.0
                            m_status  = "INV Paid" if (m_payable > 0 and m_paid >= m_payable) else "INV Received"
                            update_ssp_inv_status(selected_ssp, month, m_status)
                            updated.append(f"**{month}** → {m_status}")

                        # Refresh session state
                        for key in ["ssp_df", "ssp_edit_df"]:
                            if key in st.session_state:
                                del st.session_state[key]

                        months_summary = "\n\n".join(updated)
                        st.success(
                            f"✅ Invoice uploaded for {len(sorted_sel)} month(s)!\n\n"
                            f"**Dropbox path:** `{dropbox_path}`\n\n"
                            f"{months_summary}"
                        )
                        st.rerun()

                    except Exception as exc:
                        st.error(f"❌ Upload failed: {exc}")

    with col_cancel:
        if st.button("Cancel", use_container_width=True):
            st.rerun()


# 5️⃣ SSP (VENDORS) TAB
# ====================================================

if "SSP (Vendors)" in tabs:
    with tabs["SSP (Vendors)"]:
    
        st.markdown('''
                <div style="background:linear-gradient(135deg,#003366 0%,#005599 100%);
                border-radius:10px;padding:18px 24px;margin-bottom:20px;
                box-shadow:0 4px 16px rgba(0,51,102,.2);
                display:flex;align-items:center;height:55px;gap:14px;">
                <div style="font-size:32px;">📤</div>
                <div>
                    <div style="color:white;font-size:20px;font-weight:800;">
                        SSP (Vendors)</div>
                </div></div>
            ''', unsafe_allow_html=True)
        # ==============================
        # FILTERS (Invoice Manager style)
        # ==============================
        _ssp_fy_list = _tab_fy_list()
        col2, col3, col4, col5, col6 = st.columns([0.3, 0.3, 0.3, 0.5, 0.5])

        with col2:
            selected_fy = st.selectbox(
                "Financial Year",
                options=["All"] + _ssp_fy_list,
                index=0,
                key="ssp_fy"
            )

        _ssp_q = st.session_state.get("ssp_quarter", "All")
        _ssp_m = st.session_state.get("ssp_month",   "All")

        with col3:
            selected_month = st.selectbox(
                "Month",
                options=["All"] + _fy_months(selected_fy),
                index=0,
                key="ssp_month",
                disabled=(_ssp_q not in ("", "All")),
            )

        with col4:
            selected_quarter = st.selectbox(
                "Quarter",
                options=["All"] + _QUARTER_OPTS,
                index=0,
                key="ssp_quarter",
                disabled=(_ssp_m not in ("", "All")),
            )

        with col5:
            search_text = st.text_input(
                "🔍 Search SSP",
                placeholder="Global search...",
                key="ssp_search"
            )

        with col6:
            st.markdown("<br>", unsafe_allow_html=True)
            btn_col1, btn_col2 = st.columns(2)
            with btn_col1:
                if st.button("🔄 Refresh", key="ssp_refresh_sel", use_container_width=True):
                    if "ssp_df" in st.session_state:
                        del st.session_state["ssp_df"]
                    if "ssp_edit_df" in st.session_state:
                        del st.session_state["ssp_edit_df"]
                    st.rerun()
            with btn_col2:
                if st.button("📤 Upload INV", key="ssp_upload_inv_btn", use_container_width=True):
                    ssp_src = st.session_state.get("ssp_df", pd.DataFrame())
                    if ssp_src.empty:
                        ssp_src = load_ssp_final(
                            st.session_state.master_df,
                            st.session_state.partner_df
                        )
                    # Format month for display in dialog
                    ssp_src_dlg = ssp_src.copy()
                    ssp_src_dlg["Month"] = pd.to_datetime(
                        ssp_src_dlg["Month"], errors="coerce"
                    ).dt.strftime("%b-%Y")
                    upload_invoice_dialog(ssp_src_dlg)

        st.divider()

        # ==============================
        # LOAD DATA (MASTER + DB MERGE)
        # ==============================
        if "ssp_df" not in st.session_state:
            st.session_state.ssp_df = load_ssp_final(
                st.session_state.master_df,
                st.session_state.partner_df
            )

        df = st.session_state.ssp_df.copy()

        if df.empty:
            st.warning("No SSP Data Found")
            st.stop()

        # ==============================
        # TYPE FIXES
        # ==============================
        df["Month"]        = pd.to_datetime(df["Month"],    errors="coerce")
        df["Due Date"]     = pd.to_datetime(df["Due Date"], errors="coerce")
        df["Payable $"]    = pd.to_numeric(df["Payable $"],    errors="coerce").fillna(0)
        df["Paid Amount $"]= pd.to_numeric(df["Paid Amount $"], errors="coerce").fillna(0)

        # ==============================
        # DERIVED
        # ==============================
        df["Shortage"] = df["Payable $"] - df["Paid Amount $"]
        today = pd.Timestamp.today()
        df["Age"] = (today - df["Due Date"]).dt.days

        # ==============================
        # APPLY FILTERS (Invoice Manager style)
        # ==============================
        df_filtered = df.copy()

        if selected_fy not in ("", "All"):
            _ssp_fy_start, _ssp_fy_end = _fy_date_range(selected_fy)
            df_filtered = df_filtered[
                (df_filtered["Month"] >= _ssp_fy_start) &
                (df_filtered["Month"] <= _ssp_fy_end)
            ]
            if selected_quarter not in ("", "All"):
                q_start, q_end = _quarter_date_range(selected_fy, selected_quarter)
                df_filtered = df_filtered[
                    (df_filtered["Month"] >= q_start) &
                    (df_filtered["Month"] <= q_end)
                ]
            elif selected_month not in ("", "All"):
                df_filtered = df_filtered[
                    df_filtered["Month"].dt.strftime("%b-%Y") == selected_month
                ]

        # ==============================
        # FORMAT FOR DISPLAY
        # ==============================
        df = df_filtered.copy()
        df.columns = df.columns.str.strip()
        df = df.sort_values(["Month", "SSP Name"])

        if "ssp_edit_df" not in st.session_state or len(st.session_state.ssp_edit_df) != len(df):
            st.session_state.ssp_edit_df = df.copy()

        df["Month"]        = pd.to_datetime(df["Month"],    errors="coerce").dt.strftime("%b-%Y")
        df["Due Date"]     = pd.to_datetime(df["Due Date"], errors="coerce").dt.strftime("%d-%b-%Y")
        df["Payment Date"] = pd.to_datetime(df["Payment Date"], errors="coerce")
        df["Payment Date"] = df["Payment Date"].dt.strftime("%Y-%m-%d").where(df["Payment Date"].notna(), other="")
        df["Paid From"]  = df["Paid From"].fillna("Select").replace("", "Select")
        df["Reason"]   = df["Reason"].fillna("")

        # ==============================
        # AGGRID CONFIG
        # ==============================
        from st_aggrid import AgGrid, GridOptionsBuilder, JsCode
        from st_aggrid import GridUpdateMode, DataReturnMode

        ssp_currency_formatter = JsCode("""
        function(params) {
            if (params.value == null || params.value === '') return '';
            return '$' + parseFloat(params.value).toLocaleString(undefined, {minimumFractionDigits: 2});
        }
        """)

        ssp_date_formatter = JsCode("""
        function(params) {
            if (!params.value) return '';
            var d = new Date(params.value);
            if (isNaN(d)) return params.value;
            var day   = String(d.getDate()).padStart(2, '0');
            var month = d.toLocaleString('en-GB', { month: 'short' });
            var year  = d.getFullYear();
            return day + '-' + month + '-' + year;
        }
        """)

        gb_ssp = GridOptionsBuilder.from_dataframe(df)

        # Currency columns (non-calculated)
        for col in ["Payable $", "Paid Amount $"]:
            if col in df.columns:
                gb_ssp.configure_column(col, type=["numericColumn"], valueFormatter=ssp_currency_formatter)

        # Shortage — live JS formula like Excel
        gb_ssp.configure_column(
            "Shortage",
            type=["numericColumn"],
            valueFormatter=ssp_currency_formatter,
            valueGetter=JsCode("""
                function(params) {
                    var payable = parseFloat(params.data["Payable $"]) || 0;
                    var paid    = parseFloat(params.data["Paid Amount $"]) || 0;
                    return Math.round((payable - paid) * 100) / 100;
                }
            """),
            editable=False
        )

        # Editable fields
        gb_ssp.configure_column("Paid Amount $", editable=True)
        gb_ssp.configure_column("Reason",        editable=True)

        # Paid From dropdown
        gb_ssp.configure_column(
            "Paid From",
            editable=True,
            cellEditor="agSelectCellEditor",
            cellEditorParams={
                "values": ["Select", "PayPal", "Payoneer", "Bank Remittance", "Bank Transfer"]
            }
        )

        # Payment Date — date picker with clean string output
        gb_ssp.configure_column(
            "Payment Date",
            editable=True,
            cellEditor=JsCode("""
                class SspDatePickerEditor {
                    init(params) {
                        this.value = params.value || '';
                        this.input = document.createElement('input');
                        this.input.type = 'date';
                        this.input.style.width = '100%';
                        this.input.style.height = '100%';
                        this.input.style.border = 'none';
                        this.input.style.outline = 'none';
                        this.input.style.fontSize = '14px';
                        this.input.style.padding = '2px';
                        if (this.value) {
                            try {
                                var d = new Date(this.value);
                                if (!isNaN(d)) {
                                    var yyyy = d.getFullYear();
                                    var mm   = String(d.getMonth()+1).padStart(2,'0');
                                    var dd   = String(d.getDate()).padStart(2,'0');
                                    this.input.value = yyyy+'-'+mm+'-'+dd;
                                }
                            } catch(e) {}
                        }
                    }
                    getGui() { return this.input; }
                    afterGuiAttached() {
                        this.input.focus();
                        try { this.input.showPicker(); } catch(e) {}
                    }
                    getValue() { return this.input.value || null; }
                    isPopup() { return false; }
                    destroy() {}
                }
            """),
            valueFormatter=ssp_date_formatter,
            headerTooltip="Click cell to pick a date"
        )

        gb_ssp.configure_column("Month",    pinned="left", flex=1,   minWidth=80)
        gb_ssp.configure_column("SSP Name", pinned="left", flex=1.4, minWidth=110)
        gb_ssp.configure_column("Payable $",     flex=1.2, minWidth=95)
        gb_ssp.configure_column("Due Date",      flex=1.1, minWidth=90)
        gb_ssp.configure_column("Payment Date",  flex=1.2, minWidth=95)
        gb_ssp.configure_column("Paid Amount $", flex=1.2, minWidth=100)
        gb_ssp.configure_column("Paid From",     flex=1.2, minWidth=95)
        gb_ssp.configure_column("Shortage",      flex=1.1, minWidth=90)
        gb_ssp.configure_column("Reason",        flex=1.1, minWidth=90)
        gb_ssp.configure_column("Age",           flex=0.6, minWidth=50)

        # Invoice Status — colour-coded badge (read-only)
        if "Invoice Status" not in df.columns:
            df["Invoice Status"] = ""
        df["Invoice Status"] = df["Invoice Status"].fillna("")

        gb_ssp.configure_column(
            "Invoice Status",
            editable=False,
            flex=1.2,
            minWidth=100,
            cellStyle=JsCode("""
                function(params) {
                    var v = params.value || '';
                    if (v === 'INV Paid')     return { backgroundColor: '#d4edda', color: '#155724', fontWeight: 'bold', borderRadius: '4px' };
                    if (v === 'INV Received') return { backgroundColor: '#fff3cd', color: '#856404', fontWeight: 'bold', borderRadius: '4px' };
                    return { color: '#999' };
                }
            """)
        )

        gb_ssp.configure_default_column(resizable=True, sortable=True)

        # Row colour style
        ssp_row_style = JsCode("""
        function(params) {
            if (params.node.rowPinned) {
                return { backgroundColor: '#003366', color: 'white', fontWeight: 'bold', fontSize: '14px' };
            }
            var payable  = parseFloat(params.data["Payable $"])    || 0;
            var paid     = parseFloat(params.data["Paid Amount $"]) || 0;
            var dueDateStr = params.data["Due Date"];
            if (!dueDateStr) return {};
            var parts = dueDateStr.split("-");
            if (parts.length !== 3) return {};
            var months = {Jan:0,Feb:1,Mar:2,Apr:3,May:4,Jun:5,Jul:6,Aug:7,Sep:8,Oct:9,Nov:10,Dec:11};
            var dueDate = new Date(parseInt(parts[2]), months[parts[1]], parseInt(parts[0]));
            var today = new Date(); today.setHours(0,0,0,0);
            if (today > dueDate && paid === 0)              return { backgroundColor: "#fdecea" };
            if (paid >= payable && payable !== 0)           return { backgroundColor: "#90ee90" };
            if (paid > 0 && paid < payable)                 return { backgroundColor: "#fff8e1" };
            return {};
        }
        """)

        # Grand total footer
        ssp_total_js = JsCode("""
        function(api) {
            var totalPayable = 0; var totalPaid = 0;
            api.forEachNodeAfterFilter(function(node) {
                totalPayable += parseFloat(node.data["Payable $"])    || 0;
                totalPaid    += parseFloat(node.data["Paid Amount $"]) || 0;
            });
            api.setPinnedBottomRowData([{
                "Month": "Grand Total",
                "Payable $": Number(totalPayable),
                "Paid Amount $": Number(totalPaid),
                "Shortage": Number(totalPayable - totalPaid)
            }]);
        }
        """)

        ssp_grid_options = gb_ssp.build()
        ssp_grid_options["getRowStyle"]         = ssp_row_style
        ssp_grid_options["onFirstDataRendered"] = ssp_total_js
        ssp_grid_options["onCellValueChanged"]  = ssp_total_js
        ssp_grid_options["onFilterChanged"]     = ssp_total_js
        ssp_grid_options["onGridReady"]         = JsCode("""
            function(params) { params.api.sizeColumnsToFit(); }
        """)

        # Pinned grand total bottom row
        total_payable = float(df["Payable $"].sum())
        total_paid    = float(df["Paid Amount $"].sum())
        ssp_grid_options["pinnedBottomRowData"] = [{
            "Month": "Grand Total",
            "Payable $": total_payable,
            "Paid Amount $": total_paid,
            "Shortage": float(total_payable - total_paid)
        }]

        if search_text:
            ssp_grid_options["quickFilterText"] = search_text

        ssp_custom_css = {
            ".ag-header": {
                "background-color": "#003366 !important",
                "color": "white !important",
                "font-weight": "bold !important"
            }
        }

        # Ensure JSON-serializable
        df_ssp_grid = df.copy()
        for col in df_ssp_grid.columns:
            if pd.api.types.is_numeric_dtype(df_ssp_grid[col]):
                df_ssp_grid[col] = df_ssp_grid[col].apply(lambda x: float(x) if pd.notnull(x) else None)
        df_ssp_grid = df_ssp_grid.where(pd.notnull(df_ssp_grid), None)

        ssp_grid_response = AgGrid(
            df_ssp_grid,
            gridOptions=ssp_grid_options,
            height=sc.grid_height(650),
            update_mode=GridUpdateMode.MODEL_CHANGED,
            data_return_mode=DataReturnMode.AS_INPUT,
            allow_unsafe_jscode=True,
            custom_css=ssp_custom_css,
            fit_columns_on_grid_load=True
        )

        # ==============================
        # SAVE BUTTON
        # ==============================
        if ssp_grid_response["data"] is not None:

            ssp_edited_df = pd.DataFrame(ssp_grid_response["data"])
            ssp_edited_df = ssp_edited_df.reset_index(drop=True)

            # Clean Payment Date
            ssp_edited_df["Payment Date"] = ssp_edited_df["Payment Date"].apply(
                lambda x: None if x in [{}, "", None] else x
            )
            ssp_edited_df["Payment Date"] = pd.to_datetime(ssp_edited_df["Payment Date"], errors="coerce")
            ssp_edited_df["Due Date"]     = pd.to_datetime(ssp_edited_df["Due Date"],     errors="coerce")
            ssp_edited_df["Paid Amount $"]= pd.to_numeric(ssp_edited_df["Paid Amount $"], errors="coerce").fillna(0)

            # Validation
            ssp_errors = []
            for idx, row in ssp_edited_df.iterrows():
                if (
                    pd.notna(row["Payment Date"])
                    or row["Paid Amount $"] > 0
                    or row["Paid From"] not in ["", "Select", None]
                ):
                    if pd.isna(row["Payment Date"]):
                        ssp_errors.append(f"Row {int(idx)+1}: Invalid Payment Date")
                    if row["Paid From"] in ["", "Select", None]:
                        ssp_errors.append(f"Row {int(idx)+1}: Select Paid From")

            save_ssp_clicked = st.button("💾 Save SSP Changes", key="ssp_save_btn")

            if ssp_errors:
                st.error("\n".join(ssp_errors))
            elif save_ssp_clicked:

                def clean_ssp_date(val):
                    if val in [{}, "", None, "None", "NaT"]:
                        return None
                    if isinstance(val, dict):
                        try:
                            y = int(val.get("year", 0))
                            m = int(val.get("month", 0))
                            d = int(val.get("date") or val.get("day", 0))
                            if y and m and d:
                                return pd.Timestamp(year=y, month=m, day=d)
                        except:
                            return None
                    try:
                        result = pd.to_datetime(val, dayfirst=True, errors="coerce")
                        return None if pd.isna(result) else result
                    except:
                        return None

                ssp_edited_df["Payment Date"] = ssp_edited_df["Payment Date"].apply(clean_ssp_date)
                ssp_edited_df["Payment Date"] = ssp_edited_df["Payment Date"].apply(
                    lambda x: x.strftime("%Y-%m-%d") if pd.notna(x) else None
                )
                ssp_edited_df["Due Date"] = pd.to_datetime(ssp_edited_df["Due Date"], errors="coerce")
                ssp_edited_df["Paid Amount $"] = pd.to_numeric(
                    ssp_edited_df["Paid Amount $"], errors="coerce"
                ).fillna(0)
                ssp_edited_df["Month"] = pd.to_datetime(ssp_edited_df["Month"], errors="coerce")
                ssp_edited_df["Month"] = ssp_edited_df["Month"].dt.strftime("%b-%Y")

                # Save to DB
                upsert_ssp_data(ssp_edited_df)

                # Auto-escalate invoice status to INV Paid when fully paid
                for _, row in ssp_edited_df.iterrows():
                    try:
                        payable = float(row.get("Payable $", 0) or 0)
                        paid    = float(row.get("Paid Amount $", 0) or 0)
                        if payable > 0 and paid >= payable:
                            # Only escalate if already at INV Received (don't overwrite empty)
                            conn_chk = get_db_connection()
                            cur_chk  = conn_chk.cursor()
                            cur_chk.execute(
                                "SELECT inv_status FROM ssp_data WHERE ssp_name=? AND month=?",
                                (row["SSP Name"], row["Month"])
                            )
                            res = cur_chk.fetchone()
                            conn_chk.close()
                            if res and res[0] == "INV Received":
                                update_ssp_inv_status(row["SSP Name"], row["Month"], "INV Paid")
                    except Exception:
                        pass

                st.success("SSP (Vendors) saved successfully ✅")

                # Clear session state so fresh data loads on rerun
                if "ssp_df" in st.session_state:
                    del st.session_state["ssp_df"]
                if "ssp_edit_df" in st.session_state:
                    del st.session_state["ssp_edit_df"]

                st.rerun()

# ====================================================
# 💰 COSTS CENTRE TAB (OLD LOGIC + SQLITE VERSION)
# ====================================================

if "Costs Centre" in tabs:
    with tabs["Costs Centre"]:
        
        st.markdown('''
                <div style="background:linear-gradient(135deg,#003366 0%,#005599 100%);
                border-radius:10px;padding:18px 24px;margin-bottom:20px;
                box-shadow:0 4px 16px rgba(0,51,102,.2);
                display:flex;align-items:center;height:55px;gap:14px;">
                <div style="font-size:32px;">💰</div>
                <div>
                    <div style="color:white;font-size:20px;font-weight:800;">
                        Costs Centre</div>
                </div></div>
            ''', unsafe_allow_html=True)

        fy_list = generate_financial_years()

        col2, col3, col4 = st.columns([1, 1, 3.5])

        with col2:
            selected_fy = st.selectbox(
                "Financial Year",
                fy_list,
                index=len(fy_list) - 1   # current FY auto select
            )

        with col3:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("🔄 Refresh", key="costcentre_refresh"):
                load_cost_centre.clear()
                st.session_state.cost_df = load_cost_centre()
                st.rerun()

        with col4:
            st.markdown("<br>", unsafe_allow_html=True)

            if "open_cost_popup" not in st.session_state:
                st.session_state.open_cost_popup = False

            if st.button("+ Add Cost", key="add_cost_btn"):
                st.session_state.open_cost_popup = True

        cc_view = st.radio(
            "Table View",
            ["📅 Monthwise", "📊 Annual/FY Total Only", "📆 Specific Month"],
            horizontal=True,
            key="cc_table_view",
            label_visibility="collapsed"
        )
        cc_sel_month = None
        # Pre-compute month list from FY selection (same logic as data block)
        _cc_fy_start = int(selected_fy.split("-")[0])
        _cc_month_opts = [m.strftime("%b-%Y") for m in pd.date_range(
            start=f"{_cc_fy_start}-04-01", end=f"{_cc_fy_start+1}-03-31", freq="MS")]
        if cc_view == "📆 Specific Month":
            cc_sel_month = st.selectbox(
                "Select Month",
                _cc_month_opts,
                key="cc_month_sel"
            )

        # ====================================================
        # ADD COST POPUP
        # ====================================================

        @st.dialog("Add New Cost")
        def add_cost_popup():

            st.markdown("### Add Cost Details")

            cost_df = load_cost_centre().copy()

            # -------------------------
            # CATEGORY
            # -------------------------
            category = st.selectbox(
                "Category",
                ["Select", "Direct", "Indirect"],
                key="cost_category"
            )

            # -------------------------
            # COST NAME (dynamic)
            # -------------------------
            if not cost_df.empty and category != "Select":
                cost_names = sorted(
                    cost_df[cost_df["Category"] == category]["Cost Name"]
                    .dropna()
                    .astype(str)
                    .unique()
                    .tolist()
                )
            else:
                cost_names = []

            cost_name_selected = st.selectbox(
                "Cost Name",
                ["Select"] + cost_names,
                key="cost_name"
            )

            new_cost_name = st.text_input(
                "Add New Cost Name (optional)",
                key="new_cost_name"
            )

            cost_name = new_cost_name.strip() if new_cost_name.strip() else cost_name_selected

            # -------------------------
            # SUB COST (dynamic)
            # -------------------------
            if not cost_df.empty and category != "Select" and cost_name_selected != "Select":
                sub_cost_list = sorted(
                    cost_df[
                        (cost_df["Category"] == category) &
                        (cost_df["Cost Name"] == cost_name_selected)
                    ]["Sub Cost"]
                    .dropna()
                    .astype(str)
                    .unique()
                    .tolist()
                )
            else:
                sub_cost_list = []

            sub_cost_selected = st.selectbox(
                "Sub Cost",
                ["Select"] + sub_cost_list,
                key="sub_cost"
            )

            new_sub_cost = st.text_input(
                "Add New Sub Cost (optional)",
                key="new_sub_cost"
            )

            sub_cost = new_sub_cost.strip() if new_sub_cost.strip() else sub_cost_selected

            # -------------------------
            # FINANCIAL YEAR
            # -------------------------
            financial_year = st.selectbox(
                "Financial Year",
                fy_list,
                index=0,
                key="cost_fy"
            )

            # -------------------------
            # MONTH LIST
            # -------------------------
            fy_start = int(financial_year.split("-")[0])

            months = pd.date_range(
                start=f"{fy_start}-04-01",
                end=f"{fy_start+1}-03-31",
                freq="MS"
            )

            month_options = ["Select"] + months.strftime("%b-%Y").tolist()

            month = st.selectbox(
                "Month",
                options=month_options,
                key="cost_month"
            )

            # -------------------------
            # AUTO FX FETCH
            # -------------------------
            fx_rate_auto = get_fx_rate(month) if month != "Select" else 0.0

            if "fx_rate" not in st.session_state:
                st.session_state.fx_rate = fx_rate_auto

            if st.session_state.get("last_month") != month:
                st.session_state.fx_rate = fx_rate_auto
                st.session_state.last_month = month

            # -------------------------
            # CURRENCY
            # -------------------------
            currency = st.selectbox(
                "USD / INR",
                ["Select", "USD", "INR"],
                key="cost_currency"
            )

            amount_usd = 0.0
            fx_rate = 0.0
            amount_inr = 0.0

            if currency == "USD":
                amount_usd = st.number_input(
                    "Amount $",
                    min_value=0.0,
                    step=0.01,
                    key="amount_usd"
                )

                fx_rate = st.number_input(
                    "FX Rate",
                    value=float(st.session_state.fx_rate),
                    step=0.01,
                    format="%.4f",
                    key="fx_rate"
                )

                amount_inr = round(amount_usd * fx_rate, 2)

                st.number_input(
                    "Auto Amount ₹",
                    value=float(amount_inr),
                    disabled=True,
                    format="%.2f",
                    key="auto_amount_inr"
                )

            elif currency == "INR":
                amount_inr = st.number_input(
                    "Amount ₹",
                    min_value=0.0,
                    step=0.01,
                    key="amount_inr"
                )

                fx_rate = 0.0
                amount_usd = 0.0

            st.divider()

            c1, c2 = st.columns(2)

            with c1:
                save_cost = st.button("Save Cost", key="save_cost")

            with c2:
                if st.button("Close", key="close_cost_popup"):
                    st.session_state.open_cost_popup = False
                    st.rerun()

            # -------------------------
            # SAVE BUTTON
            # -------------------------
            if save_cost:

                if category == "Select":
                    st.error("Please select Category")
                    return

                if not cost_name or cost_name == "Select":
                    st.error("Please select or enter Cost Name")
                    return

                if not sub_cost or sub_cost == "Select":
                    st.error("Please select or enter Sub Cost")
                    return

                if month == "Select":
                    st.error("Please select Month")
                    return

                if currency == "Select":
                    st.error("Please select Currency")
                    return

                conn = get_db_connection()
                cursor = conn.cursor()

                # Duplicate check
                cursor.execute("""
                    SELECT COUNT(*) FROM cost_centre
                    WHERE category = ?
                      AND cost_name = ?
                      AND sub_cost = ?
                      AND financial_year = ?
                      AND month = ?
                """, (
                    category,
                    cost_name,
                    sub_cost,
                    financial_year,
                    month
                ))

                duplicate_count = cursor.fetchone()[0]

                if duplicate_count > 0:
                    conn.close()
                    st.error("This cost already exists for the selected month.")
                    return

                cursor.execute("""
                    INSERT INTO cost_centre (
                        category,
                        cost_name,
                        sub_cost,
                        financial_year,
                        month,
                        currency,
                        amount_usd,
                        fx_rate,
                        amount_inr
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    category,
                    cost_name,
                    sub_cost,
                    financial_year,
                    str(month),   # ✅ FORCE STRING
                    currency,
                    float(amount_usd),
                    float(fx_rate),
                    float(amount_inr)
                ))

                conn.commit()
                conn.close()

                load_cost_centre.clear()
                st.session_state.cost_df = load_cost_centre()

                st.success("Cost Saved Successfully")
                st.session_state.open_cost_popup = False
                st.rerun()

        # OPEN POPUP
        if st.session_state.get("open_cost_popup", False):

            st.session_state.open_cost_popup = False

            add_cost_popup()

        # ====================================================
        # COST TABLE
        # ====================================================

        st.divider()

        df_cost = load_cost_centre().copy()

        # 🔥 CLEAN ALL TEXT FIELDS (CRITICAL)
        df_cost["Financial Year"] = df_cost["Financial Year"].astype(str).str.strip()
        df_cost["Category"] = df_cost["Category"].astype(str).str.strip()
        df_cost["Currency"] = df_cost["Currency"].astype(str).str.strip()
        df_cost["Month"] = df_cost["Month"].astype(str).str.strip()

        # ✅ UNIVERSAL PARSER (handles ALL formats)
        df_cost["Month"] = pd.to_datetime(df_cost["Month"], errors="coerce")

        # 🚨 DEBUG (must check once)
        # st.write(df_cost[df_cost["Month"].isna()])

        # ✅ FINAL FORMAT FOR PIVOT
        df_cost["Month"] = df_cost["Month"].dt.strftime("%b-%Y")
                
        if df_cost.empty:
            st.info("No Cost Data Found")
        else:

            # --------------------------------
            # FILTER BY FY
            # --------------------------------
            if selected_fy != "All":
                df_cost = df_cost[df_cost["Financial Year"] == selected_fy]

            if df_cost.empty:
                st.info("No data for selected FY")
            else:    

                start_year = int(selected_fy.split("-")[0])

                months = pd.date_range(
                    start=f"{start_year}-04-01",
                    end=f"{start_year+1}-03-31",
                    freq="MS"
                )

                month_cols = [m.strftime("%b-%Y") for m in months]
                
                # --------------------------------
                # CREATE PARTICULARS
                # --------------------------------
                df_cost["Particulars"] = (
                    df_cost["Cost Name"].astype(str) + " - " + df_cost["Sub Cost"].astype(str)
                )

                # =====================================================
                # DIRECT COST
                # =====================================================
                direct_df = df_cost[df_cost["Category"] == "Direct"].copy()

                direct_usd = direct_df[direct_df["Currency"] == "USD"].copy()

                usd_pivot = (
                    direct_usd
                    .pivot_table(
                        index="Particulars",
                        columns="Month",
                        values="Amount USD",
                        aggfunc="sum",
                        fill_value=0
                    )
                )

                usd_pivot = usd_pivot.reindex(columns=month_cols, fill_value=0)
                usd_pivot.insert(0, "Currency", "USD")
                usd_pivot.reset_index(inplace=True)

                total_usd = usd_pivot[month_cols].sum() if not usd_pivot.empty else pd.Series(0, index=month_cols)

                fx_rate = (
                    direct_usd
                    .groupby("Month")["FX Rate"]
                    .mean()
                )

                # 🔥 CRITICAL FIX
                fx_rate = fx_rate.reindex(month_cols)

                # forward fill to avoid zero
                fx_rate = fx_rate.fillna(0) if not fx_rate.empty else pd.Series(0, index=month_cols)

                direct_inr_from_usd = total_usd * fx_rate
                direct_inr_from_usd = direct_inr_from_usd.fillna(0)

                direct_inr = direct_df[direct_df["Currency"] == "INR"].copy()

                inr_pivot = (
                    direct_inr
                    .pivot_table(
                        index="Particulars",
                        columns="Month",
                        values="Amount INR",
                        aggfunc="sum",
                        fill_value=0
                    )
                )

                inr_pivot = inr_pivot.reindex(columns=month_cols, fill_value=0)
                inr_pivot.insert(0, "Currency", "INR")
                inr_pivot.reset_index(inplace=True)

                direct_inr_total = inr_pivot[month_cols].sum() if not inr_pivot.empty else pd.Series(0, index=month_cols)

                total_direct_inr = direct_inr_total + direct_inr_from_usd

                # =====================================================
                # INDIRECT COST
                # =====================================================
                indirect_df = df_cost[df_cost["Category"] == "Indirect"].copy()

                indirect_pivot = (
                    indirect_df
                    .pivot_table(
                        index="Particulars",
                        columns="Month",
                        values="Amount INR",
                        aggfunc="sum",
                        fill_value=0
                    )
                )

                indirect_pivot = indirect_pivot.reindex(columns=month_cols, fill_value=0)
                indirect_pivot.insert(0, "Currency", "INR")
                indirect_pivot.reset_index(inplace=True)

                total_indirect = indirect_pivot[month_cols].sum() if not indirect_pivot.empty else pd.Series(0, index=month_cols)

                # =====================================================
                # BUILD TABLE ROWS
                # =====================================================
                rows = []

                rows.append({"Particulars": "Direct Cost", "Currency": ""})
                rows += usd_pivot.to_dict("records")

                rows.append({
                    "Particulars": "Total USD",
                    "Currency": "USD",
                    **total_usd.to_dict()
                })

                rows.append({
                    "Particulars": "FX Rate",
                    "Currency": "",
                    **fx_rate.to_dict()
                })

                rows.append({
                    "Particulars": "Direct Cost INR",
                    "Currency": "INR",
                    **direct_inr_from_usd.to_dict()
                })

                rows += inr_pivot.to_dict("records")

                rows.append({
                    "Particulars": "Total Direct Cost INR",
                    "Currency": "INR",
                    **total_direct_inr.to_dict()
                })

                rows.append({"Particulars": "Indirect Cost", "Currency": ""})
                rows += indirect_pivot.to_dict("records")

                rows.append({
                    "Particulars": "Total Indirect Cost INR",
                    "Currency": "INR",
                    **total_indirect.to_dict()
                })
                
                # ============================================
                # GRAND TOTAL COST (NEW ROW)
                # ============================================

                grand_total_cost = total_direct_inr + total_indirect

                rows.append({
                    "Particulars": "Grand Total Cost",
                    "Currency": "INR",
                    **grand_total_cost.to_dict()
                })

                df_table = pd.DataFrame(rows)

                df_table["Annual/FY Total"] = df_table[month_cols].sum(axis=1)
                df_table["Annual/FY Total"] = df_table["Annual/FY Total"].astype(object)
                rows_without_total = ["Direct Cost", "Indirect Cost", "FX Rate"]
                df_table.loc[
                    df_table["Particulars"].isin(rows_without_total),
                    "Annual/FY Total"
                ] = np.nan

                if cc_view == "📊 Annual/FY Total Only":
                    df_table = df_table[["Particulars", "Currency", "Annual/FY Total"]]
                elif cc_view == "📆 Specific Month" and cc_sel_month:
                    df_table = df_table[["Particulars", "Currency", cc_sel_month]]
                else:
                    df_table = df_table[["Particulars", "Currency"] + month_cols + ["Annual/FY Total"]]

                # =====================================================
                # GROUPING
                # =====================================================
                df_table["Group"] = ""
                df_table.loc[df_table["Particulars"].str.contains("Direct Cost"), "Group"] = "Direct Cost"
                df_table.loc[df_table["Particulars"].str.contains("Indirect Cost"), "Group"] = "Indirect Cost"
                df_table["Group"] = df_table["Group"].replace("", np.nan).ffill().fillna("")

                # =====================================================
                # AGGRID
                # =====================================================
                from st_aggrid import AgGrid, GridOptionsBuilder, JsCode

                gb = GridOptionsBuilder.from_dataframe(df_table)

                gb.configure_column("Group",        rowGroup=True, hide=True)
                gb.configure_column("Particulars",  pinned="left", minWidth=160)
                gb.configure_column("Currency",     pinned="left", minWidth=70)

                currency_formatter = JsCode("""
                function(params){
                    if(params.value == null || params.value === '') return '';

                    let currency = params.data.Currency;

                    if(currency === "USD"){
                        return '$' + Number(params.value).toLocaleString(undefined,{minimumFractionDigits:2});
                    }

                    if(currency === "INR"){
                        return '₹' + Number(params.value).toLocaleString(undefined,{minimumFractionDigits:2});
                    }

                    return params.value;
                }
                """)

                for col in [c for c in month_cols + ["Annual/FY Total"]
                            if c in df_table.columns]:
                    gb.configure_column(
                        col,
                        type=["numericColumn"],
                        valueFormatter=currency_formatter
                    )

                gb.configure_default_column(resizable=True, sortable=False)

                gridOptions = gb.build()
                gridOptions["stopEditingWhenCellsLoseFocus"] = True
                gridOptions["groupDefaultExpanded"] = 0
                gridOptions["suppressMovableColumns"] = True
                _cc_js = JsCode("""
                function(params){
                    setTimeout(function(){
                        try {
                            if(params.columnApi) params.columnApi.autoSizeAllColumns(false);
                            else if(params.api)  params.api.autoSizeAllColumns(false);
                        } catch(e) {}
                    }, 400);
                }
                """)
                gridOptions["onGridReady"]         = _cc_js
                gridOptions["onFirstDataRendered"] = _cc_js
                gridOptions["suppressHorizontalScroll"] = False
                for _cd in gridOptions.get("columnDefs", []):
                    _cd.pop("flex", None)
                    _cd.pop("width", None)
                    _f = _cd.get("field","") or _cd.get("headerName","")
                    if _f in ("Particulars","Description","Category"): _cd["minWidth"]=160
                    elif _f == "Currency":        _cd["minWidth"]=70
                    elif _f == "Annual/FY Total": _cd["minWidth"]=120
                    else:                         _cd["minWidth"]=95

                _cc_row_style_js = JsCode("""
                function(params){
                    if(!params.data) return;

                    // ✅ GRAND TOTAL ROW STYLE
                    if (params.data.Particulars === "Grand Total Cost") {
                        return {
                            backgroundColor: '#ef871e',
                            color: 'white',
                            fontWeight: 'bold',
                            fontSize: '13px'
                        };
                    }
                    if(
                        params.data.Particulars === "Total USD" ||
                        params.data.Particulars === "Direct Cost INR" ||
                        params.data.Particulars === "Total Direct Cost INR" ||
                        params.data.Particulars === "Total Indirect Cost INR"
                    ){
                        return {
                            backgroundColor:'#003366',
                            color:'white',
                            fontWeight:'bold'
                        };
                    }
                }
                """)
                gridOptions["getRowStyle"] = _cc_row_style_js

                custom_css = {
                    ".ag-header": {
                        "background-color": "#003366 !important",
                        "color": "white !important",
                        "font-weight": "bold !important"
                    }
                }

                AgGrid(
                    df_table,
                    gridOptions=gridOptions,
                    allow_unsafe_jscode=True,
                    height=sc.grid_height(600),
                    fit_columns_on_grid_load=(cc_view != "📅 Monthwise"),
                    custom_css=custom_css,
                    key=f"cost_centre_grid_{cc_view}_{cc_sel_month or ''}"
                )

# ====================================================
# 📊 P&L TAB
# ====================================================

if "P&L" in tabs:
    with tabs["P&L"]:
        
        st.markdown('''
                <div style="background:linear-gradient(135deg,#003366 0%,#005599 100%);
                border-radius:10px;padding:18px 24px;margin-bottom:20px;
                box-shadow:0 4px 16px rgba(0,51,102,.2);
                display:flex;align-items:center;height:55px;gap:14px;">
                <div style="font-size:32px;">📉</div>
                <div>
                    <div style="color:white;font-size:20px;font-weight:800;">
                        Profit & Loss Statement</div>
                </div></div>
            ''', unsafe_allow_html=True)

        fy_list_pnl = generate_financial_years()

        col2, col3, col4, col5 = st.columns([1, 0.6, 0.6, 1.2])

        with col2:
            selected_fy_pnl = st.selectbox(
                "Financial Year",
                fy_list_pnl,
                index=len(fy_list_pnl) - 1,
                key="pnl_fy_sel"
            )

        with col3:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("🔄 Refresh", key="pnl_refresh"):
                load_cost_centre.clear()
                if "master_df" in st.session_state:
                    del st.session_state["master_df"]
                st.rerun()

        with col4:
            st.markdown("<br>", unsafe_allow_html=True)
            export_format_pnl = st.selectbox(
                "Export Format",
                ["For Year (Monthwise)", "For Month"],
                key="pnl_export_format",
                label_visibility="collapsed"
            )
            
        with col5:
            st.markdown("<br>", unsafe_allow_html=True)
            export_clicked_pnl = st.button("📤 Export Visible", key="pnl_export_btn")

        pnl_view = st.radio(
            "Table View",
            ["📅 Monthwise", "📊 Annual/FY Total Only", "📆 Specific Month"],
            horizontal=True,
            key="pnl_table_view",
            label_visibility="collapsed"
        )
        pnl_sel_month = None
        # Pre-compute month list from FY selection
        _pnl_fy_start = int(selected_fy_pnl.split("-")[0])
        _pnl_month_opts = [m.strftime("%b-%Y") for m in pd.date_range(
            start=f"{_pnl_fy_start}-04-01", end=f"{_pnl_fy_start+1}-03-31", freq="MS")]
        if pnl_view == "📆 Specific Month":
            pnl_sel_month = st.selectbox(
                "Select Month",
                _pnl_month_opts,
                key="pnl_month_sel"
            )

        st.divider()

        df_cost_pnl = load_cost_centre().copy()

        # ── Clean fields ──────────────────────────────────
        df_cost_pnl["Financial Year"] = df_cost_pnl["Financial Year"].astype(str).str.strip()
        df_cost_pnl["Category"]       = df_cost_pnl["Category"].astype(str).str.strip()
        df_cost_pnl["Currency"]       = df_cost_pnl["Currency"].astype(str).str.strip()
        df_cost_pnl["Month"]          = pd.to_datetime(
            df_cost_pnl["Month"].astype(str).str.strip(), errors="coerce"
        ).dt.strftime("%b-%Y")

        if df_cost_pnl.empty:
            st.info("No Cost Data Found")
        else:
            df_cost_pnl = df_cost_pnl[df_cost_pnl["Financial Year"] == selected_fy_pnl]

            if df_cost_pnl.empty:
                st.info("No data for selected FY")
            else:
                start_year_pnl = int(selected_fy_pnl.split("-")[0])
                months_pnl = pd.date_range(
                    start=f"{start_year_pnl}-04-01",
                    end=f"{start_year_pnl+1}-03-31",
                    freq="MS"
                )
                month_cols_pnl = [m.strftime("%b-%Y") for m in months_pnl]

                # ── Particulars ───────────────────────────────────
                df_cost_pnl["Particulars"] = (
                    df_cost_pnl["Cost Name"].astype(str) + " - " +
                    df_cost_pnl["Sub Cost"].astype(str)
                )

                # ── Direct Cost ───────────────────────────────────
                direct_pnl = df_cost_pnl[df_cost_pnl["Category"] == "Direct"].copy()

                direct_usd_pnl = direct_pnl[direct_pnl["Currency"] == "USD"].copy()
                usd_pivot_pnl = (
                    direct_usd_pnl
                    .pivot_table(index="Particulars", columns="Month",
                                 values="Amount USD", aggfunc="sum", fill_value=0)
                    .reindex(columns=month_cols_pnl, fill_value=0)
                )
                usd_pivot_pnl.insert(0, "Currency", "USD")
                usd_pivot_pnl.reset_index(inplace=True)
                total_usd_pnl = usd_pivot_pnl[month_cols_pnl].sum()

                fx_rate_pnl = (
                    direct_usd_pnl.groupby("Month")["FX Rate"].mean()
                    .reindex(month_cols_pnl)
                    .fillna(0)
                )
                direct_inr_from_usd_pnl = (total_usd_pnl * fx_rate_pnl).fillna(0)

                direct_inr_pnl = direct_pnl[direct_pnl["Currency"] == "INR"].copy()
                inr_pivot_pnl = (
                    direct_inr_pnl
                    .pivot_table(index="Particulars", columns="Month",
                                 values="Amount INR", aggfunc="sum", fill_value=0)
                    .reindex(columns=month_cols_pnl, fill_value=0)
                )
                inr_pivot_pnl.insert(0, "Currency", "INR")
                inr_pivot_pnl.reset_index(inplace=True)
                direct_inr_total_pnl = inr_pivot_pnl[month_cols_pnl].sum()
                total_direct_inr_pnl = direct_inr_total_pnl + direct_inr_from_usd_pnl

                # ── Indirect Cost ──────────────────────────────────
                indirect_pnl = df_cost_pnl[df_cost_pnl["Category"] == "Indirect"].copy()
                indirect_pivot_pnl = (
                    indirect_pnl
                    .pivot_table(index="Particulars", columns="Month",
                                 values="Amount INR", aggfunc="sum", fill_value=0)
                    .reindex(columns=month_cols_pnl, fill_value=0)
                )
                indirect_pivot_pnl.insert(0, "Currency", "INR")
                indirect_pivot_pnl.reset_index(inplace=True)
                total_indirect_pnl = indirect_pivot_pnl[month_cols_pnl].sum()

                grand_total_cost_pnl = total_direct_inr_pnl + total_indirect_pnl

                # ── Revenue from Master Data (C Net $) ────────────
                df_master_pnl = st.session_state.get("master_df", pd.DataFrame()).copy()
                if df_master_pnl.empty:
                    df_master_pnl = load_master_data()

                revenue_by_month = pd.Series(0.0, index=month_cols_pnl)
                if not df_master_pnl.empty and "C DSP $" in df_master_pnl.columns:
                    df_master_pnl["Month_dt"] = pd.to_datetime(
                        df_master_pnl["Month"], errors="coerce"
                    )
                    df_master_pnl["Month_fmt"] = df_master_pnl["Month_dt"].dt.strftime("%b-%Y")
                    df_master_pnl["C DSP $"] = pd.to_numeric(
                        df_master_pnl["C DSP $"], errors="coerce"
                    ).fillna(0)
                    rev_grouped = (
                        df_master_pnl[df_master_pnl["Month_fmt"].isin(month_cols_pnl)]
                        .groupby("Month_fmt")["C DSP $"]
                        .sum()
                        .reindex(month_cols_pnl, fill_value=0)
                    )
                    revenue_by_month = rev_grouped
                    
                # ── Publisher Costs from Master Data (C Net $) ────────────
                df_master_pnl = st.session_state.get("master_df", pd.DataFrame()).copy()
                if df_master_pnl.empty:
                    df_master_pnl = load_master_data()

                pub_cost_by_month = pd.Series(0.0, index=month_cols_pnl)
                if not df_master_pnl.empty and "C SSP $" in df_master_pnl.columns:
                    df_master_pnl["Month_dt"] = pd.to_datetime(
                        df_master_pnl["Month"], errors="coerce"
                    )
                    df_master_pnl["Month_fmt"] = df_master_pnl["Month_dt"].dt.strftime("%b-%Y")
                    df_master_pnl["C SSP $"] = pd.to_numeric(
                        df_master_pnl["C SSP $"], errors="coerce"
                    ).fillna(0)
                    pub_grouped = (
                        df_master_pnl[df_master_pnl["Month_fmt"].isin(month_cols_pnl)]
                        .groupby("Month_fmt")["C SSP $"]
                        .sum()
                        .reindex(month_cols_pnl, fill_value=0)
                    )
                    pub_cost_by_month = pub_grouped

                # Revenue INR: Revenue USD * per-month FX rate
                # Use FX rates from direct USD cost data; fallback to get_fx_rate()
                fx_for_revenue = pd.Series(index=month_cols_pnl, dtype=float)
                for m in month_cols_pnl:
                    if fx_rate_pnl.get(m, 0) > 0:
                        fx_for_revenue[m] = fx_rate_pnl[m]
                    else:
                        fx_for_revenue[m] = get_fx_rate(m)
                fx_for_revenue = fx_for_revenue.fillna(0)

                revenue_inr_by_month = (revenue_by_month * fx_for_revenue).round(2)

                revenue_fy_total     = round(revenue_by_month.sum(), 2)
                revenue_inr_fy_total = round(revenue_inr_by_month.sum(), 2)
                
                # Pulisher Costs INR: Ccost USD * per-month FX rate
                # Use FX rates from direct USD cost data; fallback to get_fx_rate()
                fx_for_pub = pd.Series(index=month_cols_pnl, dtype=float)
                for m in month_cols_pnl:
                    if fx_rate_pnl.get(m, 0) > 0:
                        fx_for_pub[m] = fx_rate_pnl[m]
                    else:
                        fx_for_pub[m] = get_fx_rate(m)
                fx_for_pub = fx_for_pub.fillna(0)

                pub_cost_inr_by_month = (pub_cost_by_month * fx_for_pub).round(2)

                pub_cost_fy_total     = round(pub_cost_by_month.sum(), 2)
                pub_cost_inr_fy_total = round(pub_cost_inr_by_month.sum(), 2)
                
                # Net Profit = Revenue INR - Grand Total Cost (per month)
                net_profit_by_month = (revenue_inr_by_month - pub_cost_inr_by_month - grand_total_cost_pnl).round(2)
                net_profit_fy_total = round(revenue_inr_fy_total - pub_cost_inr_fy_total - grand_total_cost_pnl.sum(), 2)
                
                # Netted
                netted_by_month = (revenue_inr_by_month - pub_cost_inr_by_month).round(2)
                netted_fy_total = round(revenue_inr_fy_total - pub_cost_inr_fy_total.sum(), 2)
                
                # Netted USD
                netted_usd_by_month = (revenue_by_month - pub_cost_by_month).round(2)
                netted_usd_fy_total = round(revenue_fy_total - pub_cost_fy_total.sum(), 2)
                
                # ── BUILD TABLE ROWS ───────────────────────────────
                rows_pnl = []

                # Row 1: Revenue (USD)
                rows_pnl.append({
                    "Particulars": "Revenue",
                    "Currency": "USD",
                    **revenue_by_month.to_dict()
                })

                # Row 2: Revenue INR
                rows_pnl.append({
                    "Particulars": "Revenue INR",
                    "Currency": "INR",
                    **revenue_inr_by_month.to_dict()
                })
                
                # Direct Cost header
                rows_pnl.append({"Particulars": ""})
                
                # Row 3: Cost (USD)
                rows_pnl.append({
                    "Particulars": "Pubs Cost",
                    "Currency": "USD",
                    **pub_cost_by_month.to_dict()
                })
                
                # Row 4: Revenue INR
                rows_pnl.append({
                    "Particulars": "Pubs Cost INR",
                    "Currency": "INR",
                    **pub_cost_inr_by_month.to_dict()
                })
                
                # Direct Cost header
                rows_pnl.append({"Particulars": ""})
                
                # Netted USD
                rows_pnl.append({
                    "Particulars": "Netted USD",
                    "Currency": "USD",
                    **netted_usd_by_month.to_dict()
                })
                
                # Netted
                rows_pnl.append({
                    "Particulars": "Netted INR",
                    "Currency": "INR",
                    **netted_by_month.to_dict()
                })
                
                # Direct Cost header
                rows_pnl.append({"Particulars": "Direct Cost", "Currency": ""})
                rows_pnl += usd_pivot_pnl.to_dict("records")
                rows_pnl.append({
                    "Particulars": "Total USD",
                    "Currency": "USD",
                    **total_usd_pnl.to_dict()
                })
                rows_pnl.append({
                    "Particulars": "FX Rate",
                    "Currency": "",
                    **fx_rate_pnl.to_dict()
                })
                rows_pnl.append({
                    "Particulars": "Direct Cost INR",
                    "Currency": "INR",
                    **direct_inr_from_usd_pnl.to_dict()
                })
                rows_pnl += inr_pivot_pnl.to_dict("records")
                rows_pnl.append({
                    "Particulars": "Total Direct Cost INR",
                    "Currency": "INR",
                    **total_direct_inr_pnl.to_dict()
                })

                # Indirect Cost
                rows_pnl.append({"Particulars": "Indirect Cost", "Currency": ""})
                rows_pnl += indirect_pivot_pnl.to_dict("records")
                rows_pnl.append({
                    "Particulars": "Total Indirect Cost INR",
                    "Currency": "INR",
                    **total_indirect_pnl.to_dict()
                })

                # Grand Total Cost
                rows_pnl.append({
                    "Particulars": "Grand Total Cost",
                    "Currency": "INR",
                    **grand_total_cost_pnl.to_dict()
                })

                # Net Profit (last row)
                rows_pnl.append({
                    "Particulars": "Net Profit",
                    "Currency": "INR",
                    **net_profit_by_month.to_dict()
                })

                df_pnl = pd.DataFrame(rows_pnl)
                df_pnl = df_pnl.reindex(columns=["Particulars", "Currency"] + month_cols_pnl)

                # ── Round all numeric month columns to 2 decimals & replace NaN/nan with blank ──
                for _mc in month_cols_pnl:
                    df_pnl[_mc] = pd.to_numeric(df_pnl[_mc], errors="coerce")
                    df_pnl[_mc] = df_pnl[_mc].round(2)
                    _header_mask = df_pnl["Particulars"].isin(["Direct Cost", "Indirect Cost", ""])
                    df_pnl.loc[_header_mask, _mc] = np.nan
                    
                # Annual/FY Total
                skip_total_rows = ["Direct Cost", "Indirect Cost", "FX Rate"]
                def pnl_annual(row):
                    if row["Particulars"] in skip_total_rows:
                        return np.nan
                    try:
                        return round(sum(pd.to_numeric(row[c], errors="coerce") or 0 for c in month_cols_pnl), 2)
                    except:
                        return np.nan
                df_pnl["Annual/FY Total"] = df_pnl.apply(pnl_annual, axis=1)

                # Revenue Annual/FY Total override (sum of USD cols)
                df_pnl.loc[df_pnl["Particulars"] == "Revenue", "Annual/FY Total"] = revenue_fy_total
                df_pnl.loc[df_pnl["Particulars"] == "Revenue INR", "Annual/FY Total"] = revenue_inr_fy_total
                df_pnl.loc[df_pnl["Particulars"] == "Net Profit", "Annual/FY Total"] = net_profit_fy_total

                obj_cols = df_pnl.select_dtypes(include="object").columns
                df_pnl[obj_cols] = df_pnl[obj_cols].fillna("").replace("nan", "")

                # -----------------------------
                # DEFINE DIALOG FIRST (IMPORTANT)
                # -----------------------------
                @st.dialog("Export")
                def export_popup():
                
                    # Month picker (full width, shown only when needed)
                    if export_format_pnl == "For Month":
                        pnl_month_options = [c for c in df_pnl.columns
                                             if c not in ("Particulars", "Currency", "Group", "Annual/FY Total")
                                             and not str(c).startswith("::")
                                             and str(c).strip() != ""]
                        sel_month_persistent = st.selectbox(
                            "Select Month to Export",
                            pnl_month_options,
                            key="pnl_export_month_sel"
                        )
                    else:
                        sel_month_persistent = None

                    st.divider()
                    # ── DOWNLOAD BUTTONS (full width) ──────────────────────

                    # ── Helper: build styled openpyxl workbook ──────
                    def _build_pnl_xlsx(df_export, sheet_title="P&L"):
                        from openpyxl import Workbook
                        from openpyxl.styles import (
                            Font, PatternFill, Alignment, Border, Side
                        )
                        import math

                        wb = Workbook()
                        ws = wb.active
                        ws.title = sheet_title[:31]

                        header_fill   = PatternFill("solid", fgColor="003366")
                        header_font   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
                        revenue_fill  = PatternFill("solid", fgColor="9370DB")
                        rev_inr_fill  = PatternFill("solid", fgColor="800080")
                        pubs_fill     = PatternFill("solid", fgColor="B30000")
                        pubs_inr_fill = PatternFill("solid", fgColor="670000")
                        net_usd_fill  = PatternFill("solid", fgColor="008000")
                        net_inr_fill  = PatternFill("solid", fgColor="1B5E20")
                        direct_fill   = PatternFill("solid", fgColor="FFD0D7")
                        grand_fill    = PatternFill("solid", fgColor="EF871E")
                        total_fill    = PatternFill("solid", fgColor="003366")
                        profit_fill   = PatternFill("solid", fgColor="1B5E20")
                        white_font    = Font(bold=True, color="FFFFFF", name="Arial", size=10)
                        black_font    = Font(bold=True, color="000000", name="Arial", size=10)
                        normal_font   = Font(name="Arial", size=10)
                        thin_side     = Side(style="thin", color="CCCCCC")
                        thin_border   = Border(bottom=thin_side)

                        cols = [c for c in df_export.columns if not str(c).startswith("::")]

                        # Header row
                        for ci, col_name in enumerate(cols, start=1):
                            cell = ws.cell(row=1, column=ci, value=col_name)
                            cell.fill      = header_fill
                            cell.font      = header_font
                            cell.alignment = Alignment(horizontal="center", vertical="center")

                        row_style_map = {
                            "Revenue":                  (revenue_fill,  white_font),
                            "Revenue INR":              (rev_inr_fill,  white_font),
                            "Pubs Cost":                (pubs_fill,     white_font),
                            "Pubs Cost INR":            (pubs_inr_fill, white_font),
                            "Netted USD":               (net_usd_fill,  white_font),
                            "Netted INR":               (net_inr_fill,  white_font),
                            "Direct Cost":              (direct_fill,   black_font),
                            "Indirect Cost":            (direct_fill,   black_font),
                            "Grand Total Cost":         (grand_fill,    white_font),
                            "Net Profit":               (profit_fill,   white_font),
                            "Total USD":                (total_fill,    white_font),
                            "Direct Cost INR":          (total_fill,    white_font),
                            "Total Direct Cost INR":    (total_fill,    white_font),
                            "Total Indirect Cost INR":  (total_fill,    white_font),
                        }

                        currency_cols = [c for c in cols if c not in ("Particulars", "Currency", "Group")]
                        usd_fmt = u'_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
                        inr_fmt = u'_(\u20b9* #,##0.00_);_(\u20b9* (#,##0.00);_(\u20b9* "-"??_);_(@_)'
                        fx_fmt  = '#,##0.00'

                        def _is_blank(v):
                            if v is None or v == "":
                                return True
                            try:
                                if math.isnan(float(v)):
                                    return True
                            except (ValueError, TypeError):
                                pass
                            if str(v).strip().lower() == "nan":
                                return True
                            return False

                        for ri, (_, row_data) in enumerate(df_export.iterrows(), start=2):
                            particular = str(row_data.get("Particulars", ""))
                            currency   = str(row_data.get("Currency", ""))
                            fill, font = row_style_map.get(particular, (None, normal_font))

                            for ci, col_name in enumerate(cols, start=1):
                                val  = row_data[col_name]
                                cell = ws.cell(row=ri, column=ci)

                                if col_name in currency_cols:
                                    if _is_blank(val) or particular in ("Direct Cost", "Indirect Cost", ""):
                                        cell.value = None
                                    elif particular == "FX Rate":
                                        try:
                                            cell.value = round(float(val), 2)
                                            cell.number_format = fx_fmt
                                        except (ValueError, TypeError):
                                            cell.value = None
                                    else:
                                        try:
                                            cell.value = round(float(val), 2)
                                            cell.number_format = usd_fmt if currency == "USD" else inr_fmt
                                        except (ValueError, TypeError):
                                            cell.value = None
                                else:
                                    cell.value = None if _is_blank(val) else str(val)

                                if fill:
                                    cell.fill = fill
                                cell.font      = font
                                cell.alignment = Alignment(
                                    horizontal="right" if col_name in currency_cols else "left",
                                    vertical="center"
                                )
                                cell.border = thin_border

                        # Column widths
                        ws.column_dimensions["A"].width = 30
                        ws.column_dimensions["B"].width = 8
                        for ci in range(3, len(cols) + 1):
                            ws.column_dimensions[
                                ws.cell(row=1, column=ci).column_letter
                            ].width = 16

                        ws.freeze_panes = "C2"
                        return wb

                    # ── "For Month": pick one month ─────────────────
                    if export_format_pnl == "For Month":

                        sel_month = st.session_state.get("pnl_export_month_sel", None)  # ✅ renamed back to sel_month
                        if not sel_month:
                            st.warning("Please select a month above.")
                        else:
                            df_month_export = df_pnl[["Particulars", "Currency", sel_month]].copy()  # ✅ build it here

                            # ── XLSX ─────────────────────────────────
                            wb_m = _build_pnl_xlsx(df_month_export, sheet_title=sel_month)
                            buf_m = io.BytesIO()
                            wb_m.save(buf_m)
                            buf_m.seek(0)
                            st.download_button(
                                label=f"⬇️ Download XLSX – {sel_month}",
                                data=buf_m,
                                file_name=f"PnL_{selected_fy_pnl}_{sel_month}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"pnl_dl_xlsx_month_{sel_month}",
                                use_container_width=True
                            )

                            # ── PDF ──────────────────────────────────
                            from reportlab.platypus import SimpleDocTemplate, Table, Paragraph, Spacer
                            from reportlab.lib.styles import getSampleStyleSheet
                            from reportlab.lib import colors
                            from reportlab.lib.pagesizes import A4, landscape
                            import math

                            def _fmt_pdf_val(v, currency):
                                """Format a cell value for PDF: blank for nan/empty, currency symbol + 2dp for numbers."""
                                if v is None or v == "":
                                    return ""
                                try:
                                    fv = float(v)
                                    if math.isnan(fv):
                                        return ""
                                    formatted = f"{fv:,.2f}"
                                    if currency == "USD":
                                        return f"${formatted}"
                                    elif currency == "INR":
                                        return f"Rs.{formatted}"
                                    return formatted
                                except (ValueError, TypeError):
                                    s = str(v).strip()
                                    return "" if s.lower() == "nan" else s

                            _rcmap = {
                                "Revenue":                  (colors.HexColor("#9370DB"), colors.white),
                                "Revenue INR":              (colors.HexColor("#800080"), colors.white),
                                "Pubs Cost":                (colors.HexColor("#B30000"), colors.white),
                                "Pubs Cost INR":            (colors.HexColor("#670000"), colors.white),
                                "Netted USD":               (colors.HexColor("#008000"), colors.white),
                                "Netted INR":               (colors.HexColor("#1B5E20"), colors.white),
                                "Direct Cost":              (colors.HexColor("#FFD0D7"), colors.black),
                                "Indirect Cost":            (colors.HexColor("#FFD0D7"), colors.black),
                                "Grand Total Cost":         (colors.HexColor("#EF871E"), colors.white),
                                "Net Profit":               (colors.HexColor("#1B5E20"), colors.white),
                                "Total USD":                (colors.HexColor("#003366"), colors.white),
                                "Direct Cost INR":          (colors.HexColor("#003366"), colors.white),
                                "Total Direct Cost INR":    (colors.HexColor("#003366"), colors.white),
                                "Total Indirect Cost INR":  (colors.HexColor("#003366"), colors.white),
                            }

                            buf_pdf_m = io.BytesIO()
                            doc = SimpleDocTemplate(buf_pdf_m, pagesize=landscape(A4),
                                                    leftMargin=20, rightMargin=20, topMargin=30, bottomMargin=20)
                            styles = getSampleStyleSheet()
                            elements = []
                            elements.append(Paragraph(f"P&L Report \u2013 {sel_month} ({selected_fy_pnl}) - PeakAds LLP", styles["Title"]))
                            elements.append(Spacer(1, 10))

                            # Build header row
                            pdf_header = ["Particulars", "Currency", sel_month]
                            pdf_data   = [pdf_header]
                            for _, row_data in df_month_export.iterrows():
                                part = str(row_data.get("Particulars", ""))
                                cur_raw = row_data.get("Currency", "")
                                cur = "" if str(cur_raw).strip().lower() == "nan" else str(cur_raw)
                                val  = row_data.get(sel_month, "")
                                # blank for section headers
                                if part in ("Direct Cost", "Indirect Cost", ""):
                                    disp = ""
                                elif part == "FX Rate":
                                    try:
                                        disp = f"{float(val):.2f}" if (val != "" and pd.notna(val)) else ""
                                    except:
                                        disp = ""
                                else:
                                    disp = _fmt_pdf_val(val, cur)
                                pdf_data.append([part, cur, disp])

                            table_style_m = [
                                ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#003366")),
                                ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
                                ("FONTNAME",    (0, 0), (-1, -1), "Helvetica"),
                                ("FONTSIZE",    (0, 0), (-1, -1), 8),
                                ("GRID",        (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
                                ("ALIGN",       (2, 0), (-1, -1), "RIGHT"),
                                ("ALIGN",       (0, 0), (1, -1), "LEFT"),
                                ("RIGHTPADDING", (2, 0), (-1, -1), 6),
                                ("LEFTPADDING",  (2, 0), (-1, -1), 4),
                                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
                            ]
                            for ri, row_vals in enumerate(pdf_data[1:], start=1):
                                part = str(row_vals[0])
                                if part in _rcmap:
                                    bg, tc = _rcmap[part]
                                    table_style_m.append(("BACKGROUND", (0, ri), (-1, ri), bg))
                                    table_style_m.append(("TEXTCOLOR",  (0, ri), (-1, ri), tc))

                            tbl_m = Table(pdf_data, colWidths=[160, 45, 130])
                            tbl_m.setStyle(table_style_m)
                            elements.append(tbl_m)
                            doc.build(elements)
                            buf_pdf_m.seek(0)

                            st.download_button(
                                label=f"⬇️ Download PDF – {sel_month}",
                                data=buf_pdf_m,
                                file_name=f"PnL_{selected_fy_pnl}_{sel_month}.pdf",
                                mime="application/pdf",
                                key=f"pnl_dl_pdf_month_{sel_month}",
                                use_container_width=True
                            )

                    # ── "For Year (Monthwise)": all months ──────────
                    else:

                        df_year_export = df_pnl.drop(columns=["Group"], errors="ignore").copy()

                        # ── XLSX ─────────────────────────────────────
                        wb_y = _build_pnl_xlsx(df_year_export, sheet_title=f"PnL {selected_fy_pnl}")
                        buf_y = io.BytesIO()
                        wb_y.save(buf_y)
                        buf_y.seek(0)
                        st.download_button(
                            label=f"⬇️ Download XLSX – Full Year {selected_fy_pnl}",
                            data=buf_y,
                            file_name=f"PnL_{selected_fy_pnl}_Monthwise.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="pnl_dl_xlsx_year",
                            use_container_width=True
                        )
                        
                        # ── PDF ──────────────────────────────────────
                        from reportlab.platypus import SimpleDocTemplate, Table, Paragraph, Spacer
                        from reportlab.lib.styles import getSampleStyleSheet
                        from reportlab.lib import colors
                        from reportlab.lib.pagesizes import A4, landscape
                        import math

                        def _fmt_pdf_val_y(v, currency):
                            if v is None or v == "":
                                return ""
                            try:
                                fv = float(v)
                                if math.isnan(fv):
                                    return ""
                                formatted = f"{fv:,.2f}"
                                if currency == "USD":
                                    return f"${formatted}"
                                elif currency == "INR":
                                    return f"Rs.{formatted}"
                                return formatted
                            except (ValueError, TypeError):
                                s = str(v).strip()
                                return "" if s.lower() == "nan" else s

                        _rcmap_y = {
                            "Revenue":                  (colors.HexColor("#9370DB"), colors.white),
                            "Revenue INR":              (colors.HexColor("#800080"), colors.white),
                            "Pubs Cost":                (colors.HexColor("#B30000"), colors.white),
                            "Pubs Cost INR":            (colors.HexColor("#670000"), colors.white),
                            "Netted USD":               (colors.HexColor("#008000"), colors.white),
                            "Netted INR":               (colors.HexColor("#1B5E20"), colors.white),
                            "Direct Cost":              (colors.HexColor("#FFD0D7"), colors.black),
                            "Indirect Cost":            (colors.HexColor("#FFD0D7"), colors.black),
                            "Grand Total Cost":         (colors.HexColor("#EF871E"), colors.white),
                            "Net Profit":               (colors.HexColor("#1B5E20"), colors.white),
                            "Total USD":                (colors.HexColor("#003366"), colors.white),
                            "Direct Cost INR":          (colors.HexColor("#003366"), colors.white),
                            "Total Direct Cost INR":    (colors.HexColor("#003366"), colors.white),
                            "Total Indirect Cost INR":  (colors.HexColor("#003366"), colors.white),
                        }

                        all_cols_y = [c for c in df_year_export.columns if not str(c).startswith("::")]
                        num_cols_y = [c for c in all_cols_y if c not in ("Particulars", "Currency")]

                        buf_pdf_y = io.BytesIO()
                        doc_y = SimpleDocTemplate(buf_pdf_y, pagesize=landscape(A4),
                                                  leftMargin=10, rightMargin=10, topMargin=20, bottomMargin=15)
                        styles_y = getSampleStyleSheet()
                        elements_y = []
                        elements_y.append(Paragraph(f"P&L Report \u2013 Full Year {selected_fy_pnl} (Monthwise) - PeakAds LLP", styles_y["Title"]))
                        elements_y.append(Spacer(1, 8))

                        pdf_data_y = [all_cols_y]
                        for _, row_data in df_year_export.iterrows():
                            part = str(row_data.get("Particulars", ""))
                            cur_raw = row_data.get("Currency", "")
                            cur = "" if str(cur_raw).strip().lower() == "nan" else str(cur_raw)
                            pdf_row = [part, cur]
                            for col_name in num_cols_y:
                                val = row_data.get(col_name, "")
                                if part in ("Direct Cost", "Indirect Cost", ""):
                                    pdf_row.append("")
                                elif part == "FX Rate" or col_name == "FX Rate":
                                    try:
                                        pdf_row.append(f"{float(val):.2f}" if (val != "" and pd.notna(val)) else "")
                                    except:
                                        pdf_row.append("")
                                else:
                                    pdf_row.append(_fmt_pdf_val_y(val, cur))
                            pdf_data_y.append(pdf_row)

                        table_style_y = [
                            ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#003366")),
                            ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
                            ("FONTNAME",    (0, 0), (-1, -1), "Helvetica"),
                            ("FONTSIZE",    (0, 0), (-1, -1), 6),
                            ("GRID",        (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
                            ("ALIGN",       (2, 0), (-1, -1), "RIGHT"),
                            ("ALIGN",       (0, 0), (1, -1), "LEFT"),
                            ("RIGHTPADDING", (2, 0), (-1, -1), 4),
                            ("LEFTPADDING",  (2, 0), (-1, -1), 2),
                            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
                        ]
                        for ri, row_vals in enumerate(pdf_data_y[1:], start=1):
                            part = str(row_vals[0])
                            if part in _rcmap_y:
                                bg, tc = _rcmap_y[part]
                                table_style_y.append(("BACKGROUND", (0, ri), (-1, ri), bg))
                                table_style_y.append(("TEXTCOLOR",  (0, ri), (-1, ri), tc))

                        num_data_cols = len(all_cols_y)
                        fixed_widths = [115, 28] + [52] * (num_data_cols - 2)

                        tbl_y = Table(pdf_data_y, colWidths=fixed_widths)
                        tbl_y.setStyle(table_style_y)
                        elements_y.append(tbl_y)
                        doc_y.build(elements_y)
                        buf_pdf_y.seek(0)

                        st.download_button(
                            label=f"⬇️ Download PDF – Full Year {selected_fy_pnl}",
                            data=buf_pdf_y,
                            file_name=f"PnL_{selected_fy_pnl}_Monthwise.pdf",
                            mime="application/pdf",
                            key="pnl_dl_pdf_year",
                            use_container_width=True
                        )
                    
                    
                # Call the export dialog when button is clicked
                if export_clicked_pnl:
                    st.session_state["_any_dialog_open"] = True
                    st.session_state.open_cost_popup = False  # close any other dialog
                    export_popup()
                    st.session_state["_any_dialog_open"] = False
                
                # ── Grouping ───────────────────────────────────────
                df_pnl["Group"] = ""
                df_pnl.loc[df_pnl["Particulars"].str.contains("Direct Cost"), "Group"] = "Direct Cost"
                df_pnl.loc[df_pnl["Particulars"].str.contains("Indirect Cost"), "Group"] = "Indirect Cost"
                df_pnl["Group"] = df_pnl["Group"].replace("", np.nan).ffill().fillna("")

                # ── AgGrid ─────────────────────────────────────────
                from st_aggrid import AgGrid, GridOptionsBuilder, JsCode

                gb_pnl = GridOptionsBuilder.from_dataframe(df_pnl)
                gb_pnl.configure_column("Group",       rowGroup=True, hide=True)
                gb_pnl.configure_column("Particulars", pinned="left", minWidth=160)
                gb_pnl.configure_column("Currency",    pinned="left", minWidth=70)

                # Hide the auto-generated unique index column if present
                gb_pnl.configure_column("__index_level_0__", hide=True)

                pnl_currency_fmt = JsCode("""
                function(params){
                    if(params.value == null || params.value === '' || isNaN(params.value)) return '';
                    let cur = params.data ? params.data.Currency : '';
                    let num = Number(params.value).toLocaleString('en-IN',{minimumFractionDigits:2, maximumFractionDigits:2});
                    if(cur === 'USD')  return '$'  + num;
                    if(cur === 'INR')  return '\u20b9' + num;
                    return Number(params.value).toFixed(2);
                }
                """)

                for col in month_cols_pnl + ["Annual/FY Total"]:
                    gb_pnl.configure_column(
                        col,
                        type=["numericColumn"],
                        valueFormatter=pnl_currency_fmt,
                        cellStyle={"textAlign": "right"}
                    )

                gb_pnl.configure_default_column(resizable=True, sortable=False)

                pnl_grid_opts = gb_pnl.build()
                pnl_grid_opts["stopEditingWhenCellsLoseFocus"] = True
                pnl_grid_opts["groupDefaultExpanded"] = 0
                pnl_grid_opts["suppressMovableColumns"] = True
                # Suppress auto-generated columns
                if "columnDefs" in pnl_grid_opts:
                    pnl_grid_opts["columnDefs"] = [
                        c for c in pnl_grid_opts["columnDefs"]
                        if not str(c.get("field","")).startswith("::")
                    ]
                _pnl_js = JsCode("""
                function(params){
                    setTimeout(function(){
                        try {
                            if(params.columnApi) params.columnApi.autoSizeAllColumns(false);
                            else if(params.api)  params.api.autoSizeAllColumns(false);
                        } catch(e) {}
                    }, 400);
                }
                """)
                pnl_grid_opts["onGridReady"]         = _pnl_js
                pnl_grid_opts["onFirstDataRendered"] = _pnl_js
                pnl_grid_opts["suppressHorizontalScroll"] = False
                if "columnDefs" in pnl_grid_opts:
                    for _cd in pnl_grid_opts["columnDefs"]:
                        _cd.pop("flex", None)
                        _cd.pop("width", None)
                        _f = _cd.get("field","") or _cd.get("headerName","")
                        if _f in ("Particulars","Description","Category"): _cd["minWidth"]=160
                        elif _f == "Currency":        _cd["minWidth"]=70
                        elif _f == "Annual/FY Total": _cd["minWidth"]=120
                        else:                         _cd["minWidth"]=95

                _pnl_row_style_js = JsCode("""
                function(params){
                    if(!params.data) return;
                    if(params.data.Particulars === 'Net Profit'){
                        return {backgroundColor:'#1B5E20', color:'white', fontWeight:'bold', fontSize:'14px'};
                    }
                    if(params.data.Particulars === 'Revenue'){
                        return {backgroundColor:'#9370db', color:'white', fontWeight:'bold', fontSize:'13px'};
                    }
                    if(params.data.Particulars === 'Revenue INR'){
                        return {backgroundColor:'#800080', color:'white', fontWeight:'bold', fontSize:'13px'};
                    }
                    if(params.data.Particulars === 'Pubs Cost'){
                        return {backgroundColor:'#b30000', color:'white', fontWeight:'bold', fontSize:'13px'};
                    }
                    if(params.data.Particulars === 'Pubs Cost INR'){
                        return {backgroundColor:'#670000', color:'white', fontWeight:'bold', fontSize:'13px'};
                    }
                    if(params.data.Particulars === 'Netted USD'){
                        return {backgroundColor:'#008000', color:'white', fontWeight:'bold', fontSize:'13px'};
                    }
                    if(params.data.Particulars === 'Netted INR'){
                        return {backgroundColor:'#1B5E20', color:'white', fontWeight:'bold', fontSize:'13px'};
                    }
                    if(params.data.Particulars === 'Direct Cost'){
                        return {backgroundColor:'#ffd0d7', color:'black', fontWeight:'bold', fontSize:'14px'};
                    }
                    if(params.data.Particulars === 'Indirect Cost'){
                        return {backgroundColor:'#ffd0d7', color:'black', fontWeight:'bold', fontSize:'14px'};
                    }
                    if(params.data.Particulars === 'Grand Total Cost'){
                        return {backgroundColor:'#ef871e', color:'white', fontWeight:'bold', fontSize:'13px'};
                    }
                    if(
                        params.data.Particulars === 'Total USD' ||
                        params.data.Particulars === 'Direct Cost INR' ||
                        params.data.Particulars === 'Total Direct Cost INR' ||
                        params.data.Particulars === 'Total Indirect Cost INR'
                    ){
                        return {backgroundColor:'#003366', color:'white', fontWeight:'bold'};
                    }
                }
                """)
                pnl_grid_opts["getRowStyle"] = _pnl_row_style_js
                pnl_grid_opts_rs = _pnl_row_style_js

                pnl_custom_css = {
                    ".ag-header": {
                        "background-color": "#003366 !important",
                        "color": "white !important",
                        "font-weight": "bold !important"
                    }
                }

                # Apply view filter BEFORE building grid opts
                if pnl_view == "📊 Annual/FY Total Only":
                    _pnl_view_cols = ["Particulars", "Currency", "Annual/FY Total"]
                elif pnl_view == "📆 Specific Month" and pnl_sel_month and pnl_sel_month in df_pnl.columns:
                    _pnl_view_cols = ["Particulars", "Currency", pnl_sel_month]
                else:
                    _pnl_view_cols = None  # Monthwise — use all cols

                if _pnl_view_cols:
                    df_pnl_display = df_pnl[[c for c in _pnl_view_cols if c in df_pnl.columns]]
                    # Rebuild compact grid opts (3 cols) with same styling
                    _gb3 = GridOptionsBuilder.from_dataframe(df_pnl_display)
                    _gb3.configure_column("Particulars", pinned="left", minWidth=160)
                    _gb3.configure_column("Currency",    pinned="left", minWidth=70)
                    _gb3.configure_default_column(resizable=True, sortable=False)
                    pnl_grid_opts = _gb3.build()
                    pnl_grid_opts["suppressHorizontalScroll"] = False
                    # Preserve same row colouring
                    pnl_grid_opts["getRowStyle"] = pnl_grid_opts_rs
                else:
                    df_pnl_display = df_pnl

                AgGrid(
                    df_pnl_display,
                    gridOptions=pnl_grid_opts,
                    allow_unsafe_jscode=True,
                    height=sc.grid_height(650),
                    fit_columns_on_grid_load=(pnl_view == "📊 Annual/FY Total Only"),
                    custom_css=pnl_custom_css,
                    key="pnl_grid"
                )

                

# ====================================================
# ⚙️ ADMIN CONTROL PANEL (FINAL FIX)
# ====================================================

if "Admin Control" in tabs:
    with tabs["Admin Control"]:
        
        st.markdown('''
                <div style="background:linear-gradient(135deg,#003366 0%,#005599 100%);
                border-radius:10px;padding:18px 24px;margin-bottom:20px;
                box-shadow:0 4px 16px rgba(0,51,102,.2);
                display:flex;align-items:center;height:55px;gap:14px;">
                <div style="font-size:32px;">⚙️</div>
                <div>
                    <div style="color:white;font-size:20px;font-weight:800;">
                        Admin Control Panel</div>
                </div></div>
            ''', unsafe_allow_html=True)

        st.success("Admin Panel Loaded ✅")

        st.write("User:", st.session_state.get("user"))
        st.write("Role:", st.session_state.get("role"))

        try:
            admin_change_password()
        except Exception as e:
            st.error("Error in Admin Panel")
            st.code(str(e))
        
        st.divider()
        
        import dropbox
        from datetime import datetime

        import requests

        APP_KEY = "xn40ddpn2ow57xg"
        APP_SECRET = "9n8shqjaywwigor"

        url = "https://api.dropboxapi.com/oauth2/token"

        data = {
            "code": "AIPgJRZfWnAAAAAAAAABVwiKqaQOYhC4zqno-Gu0sCA",
            "grant_type": "authorization_code"
        }

        auth = (APP_KEY, APP_SECRET)

        response = requests.post(url, data=data, auth=auth)
        print(response.json())

        BACKUP_FOLDER = "/PALLP-backups"


        # ====================================================
        # 🎨 SIDE-BY-SIDE LAYOUT
        # ====================================================
        col1, col2, col3 = st.columns(3)

        # ====================================================
        # 📦 LEFT SIDE → BACKUP
        # ====================================================
        with col1:

            st.subheader("📦 Data Backup")

            if st.button("Backup Now"):

                if not st.session_state.backup_running:

                    st.session_state.backup_running = True

                    try:
                        dbx = get_dropbox_client()

                        folder_path = "/PALLP-backups"

                        # Create folder if not exists
                        try:
                            dbx.files_get_metadata(folder_path)
                        except:
                            dbx.files_create_folder_v2(folder_path)

                        file_path = DB_PATH
                        file_name = f"backup_{datetime.now().strftime('%d%m%Y_%I%M%S%p')}.db"

                        with open(file_path, "rb") as f:
                            dbx.files_upload(
                                f.read(),
                                f"{folder_path}/{file_name}",
                                mode=dropbox.files.WriteMode("overwrite")
                            )

                        st.success("✅ Backup uploaded successfully!")

                    except Exception as e:
                        st.error("❌ Backup failed")
                        st.code(str(e))

                    finally:
                        st.session_state.backup_running = False

            with col2:
                # ====================================================
                # 📊 BACKUP STATUS PANEL
                # ====================================================

                st.subheader("📊 Backup Status")

                try:
                    dbx = get_dropbox_client()
                    files = dbx.files_list_folder(BACKUP_FOLDER).entries

                    if files:
                        # Sort latest first
                        files_sorted = sorted(
                            files,
                            key=lambda x: x.server_modified,
                            reverse=True
                        )

                        latest_file = files_sorted[0]
                        total_backups = len(files_sorted)

                        last_backup_time = latest_file.server_modified
                        time_diff = datetime.utcnow() - last_backup_time.replace(tzinfo=None)

                        hours_ago = round(time_diff.total_seconds() / 3600, 2)

                        st.markdown(f"📦 **Total Backups:** {total_backups}")
                        st.markdown(f"🕒 **Latest Backup:** {latest_file.name}")

                        if hours_ago < 24:
                            st.success(f"🟢 Healthy (Last backup {hours_ago} hrs ago)")
                        else:
                            st.warning(f"🟡 Old Backup ({hours_ago} hrs ago)")

                    else:
                        st.error("No backups found")

                except Exception as e:
                    st.error("Unable to fetch backup info")
                    st.code(str(e))
        
        with col3:
            # ====================================================
            # 📂 RESTORE SECTION
            # ====================================================

            st.subheader("📂 Select Backup")

            try:
                files_sorted = sorted(
                    files,
                    key=lambda x: x.server_modified,
                    reverse=True
                )[:10]

                backup_names = [f.name for f in files_sorted]

                selected_backup = st.selectbox(
                    "Choose Backup",
                    backup_names,
                    label_visibility="collapsed"
                )

            except:
                st.warning("Click Load Backup List first")


            # ====================================================
            # 🎯 ACTION BUTTONS (SIDE BY SIDE)
            # ====================================================
            col1, col2 = st.columns(2)

            with col1:
                if st.button("⬇️ Download Backup", use_container_width=True):

                    try:
                        metadata, res = dbx.files_download(
                            f"{BACKUP_FOLDER}/{selected_backup}"
                        )

                        st.download_button(
                            "Click to Download",
                            data=res.content,
                            file_name=selected_backup,
                            use_container_width=True
                        )

                    except Exception as e:
                        st.error("Download failed")
                        st.code(str(e))


            with col2:
                if st.button("⚠️ Restore Selected Backup", use_container_width=True):

                    try:
                        metadata, res = dbx.files_download(
                            f"{BACKUP_FOLDER}/{selected_backup}"
                        )

                        with open(DB_PATH, "wb") as f:
                            f.write(res.content)

                        st.success("Database restored successfully")

                        # Clear cache
                        load_master_data.clear()
                        load_partner_list.clear()
                        load_cost_centre.clear()

                        st.rerun()

                    except Exception as e:
                        st.error("Restore failed")
                        st.code(str(e))
            
            # ====================================================
            # 🔄 AUTO BACKUP (ONCE PER DAY)
            # ====================================================
            today = datetime.now().strftime("%Y-%m-%d")

            if st.session_state.get("last_backup_date") != today:

                try:
                    dbx = get_dropbox_client()

                    file_name = f"auto_{datetime.now().strftime('%d%m%Y_%I%M%S%p').lower()}.db"

                    with open(DB_PATH, "rb") as f:
                        dbx.files_upload(
                            f.read(),
                            f"{BACKUP_FOLDER}/{file_name}",
                            mode=dropbox.files.WriteMode("overwrite")
                        )

                    # Clean old backups
                    files = dbx.files_list_folder(BACKUP_FOLDER).entries
                    files_sorted = sorted(files, key=lambda x: x.server_modified, reverse=True)

                    for old_file in files_sorted[10:]:
                        dbx.files_delete_v2(old_file.path_lower)

                    st.session_state.last_backup_date = today

                    st.info("Auto backup completed")

                except:
                    pass

# ====================================================
# 🛠️ EDIT DATABASE TAB (FULL SCREEN EDITOR)
# ====================================================

if "Edit Database" in tabs:
    with tabs["Edit Database"]:
        
        st.markdown('''
                <div style="background:linear-gradient(135deg,#003366 0%,#005599 100%);
                border-radius:10px;padding:18px 24px;margin-bottom:20px;
                box-shadow:0 4px 16px rgba(0,51,102,.2);
                display:flex;align-items:center;height:55px;gap:14px;">
                <div style="font-size:32px;">🛠</div>
                <div>
                    <div style="color:white;font-size:20px;font-weight:800;">
                        Edit Database (Direct SQL Editor)</div>
                </div></div>
            ''', unsafe_allow_html=True)

            
        # ==========================================
        # IMPORT DATA SECTION
        # ==========================================
        st.subheader("📥 Import Excel Data")

        col1, col2 = st.columns([1, 1])

        with col1:
            table_list = [
                "master_data",
                "partner_list",
                "dsp_data",
                "ssp_data",
                "cost_centre",
                "invoice_details",
                "bc_report",
                "bc_teams_config",
                "bc_activity_log",
                "login_logs",
            ]
            st.caption("ℹ️ GST Report (All Combined) is a computed report — export only, not importable.")

            selected_table = st.selectbox(
                "Select Table",
                table_list,
                key="import_table"
            )

        with col2:
            st.markdown("<br>", unsafe_allow_html=True)

            # Download sample format
            if selected_table:
                sample_cols = get_table_columns(selected_table)

                # Remove id column
                sample_cols = [c for c in sample_cols if c.lower() != "id"]

                sample_df = pd.DataFrame(columns=sample_cols)

                # ✅ FIRST create buffer (OUTSIDE)
                buffer = io.BytesIO()
                sample_df.to_excel(buffer, index=False, engine="openpyxl")
                buffer.seek(0)

                # ✅ THEN call button
                st.download_button(
                    "📄 Download Sample Format",
                    data=buffer,
                    file_name=f"{selected_table}_sample.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        uploaded_file = st.file_uploader(
            "Upload Excel File (.xlsx)",
            type=["xlsx"],
            key="import_file"
        )

        if uploaded_file:

            try:
                df_import = pd.read_excel(uploaded_file)

                st.write("### 👀 Preview Data")
                st.dataframe(df_import, height=250)

                # VALIDATION
                db_columns = get_table_columns(selected_table)
                db_columns = [c for c in db_columns if c.lower() != "id"]

                if list(df_import.columns) != db_columns:
                    st.error("❌ Column mismatch! Please match exact DB columns.")
                else:

                    if st.button("🚀 Import Data"):

                        # -------------------------
                        # AUTO DATE HANDLING
                        # -------------------------
                        for col in df_import.columns:
                            if "date" in col.lower() or "month" in col.lower():
                                df_import[col] = pd.to_datetime(
                                    df_import[col],
                                    errors="coerce"
                                )

                        rows_inserted = insert_dataframe_to_db(df_import, selected_table)

                        st.success(f"✅ {rows_inserted} rows imported successfully!")

                        # 🔥 AUTO REFRESH ALL DATA
                        load_master_data.clear()
                        load_partner_list.clear()
                        load_dsp_sheet.clear()
                        load_ssp_sheet.clear()
                        load_cost_centre.clear()

                        initialize_session_data()

                        st.rerun()

            except Exception as e:
                st.error(f"Import Failed: {e}")

        st.divider()

        # ====================================================
        # TABLE SELECTOR
        # ====================================================
        table_list = {
            "Master Data":              "master_data",
            "DSP Data":                 "dsp_data",
            "SSP Data":                 "ssp_data",
            "Cost Centre":              "cost_centre",
            "Partner List":             "partner_list",
            "Invoice History":          "invoice_details",
            "GST Report (All Combined)":"_gst_combined_view",
            # ── BC Report tables ─────────────────────────────
            "BC Report Data":           "bc_report",
            "BC Teams Config":          "bc_teams_config",
            "BC Azure Config":          "bc_azure_config",
            "BC Activity Log":          "bc_activity_log",
            # ── System tables ────────────────────────────────
            "Login Logs":               "login_logs",
        }

        sel_col, exp_col = st.columns([2, 1])

        with sel_col:
            selected_table_name = st.selectbox(
                "Select Table",
                list(table_list.keys()),
                key="edb_table_select"
            )

        selected_table = table_list[selected_table_name]

        with exp_col:
            st.markdown("<br>", unsafe_allow_html=True)
            try:
                if selected_table == "_gst_combined_view":
                    import json as _edb_json
                    _edb_conn = get_db_connection()
                    _inv_df   = pd.read_sql(
                        "SELECT * FROM invoice_details WHERE is_deleted=0 ORDER BY invoice_date ASC",
                        _edb_conn)
                    _prt_df   = pd.read_sql("SELECT * FROM partner_list", _edb_conn)
                    _edb_conn.close()
                    _short_col = next((c for c in _prt_df.columns
                                       if "short" in c.lower() or "bidscube" in c.lower()), None)
                    _ST = {
                        "maharashtra":"Maharashtra","pune":"Maharashtra","mumbai":"Maharashtra",
                        "delhi":"Delhi","new delhi":"Delhi","karnataka":"Karnataka",
                        "bangalore":"Karnataka","bengaluru":"Karnataka","telangana":"Telangana",
                        "hyderabad":"Telangana","tamil nadu":"Tamil Nadu","chennai":"Tamil Nadu",
                        "gujarat":"Gujarat","haryana":"Haryana","gurugram":"Haryana",
                        "gurgaon":"Haryana","uttar pradesh":"Uttar Pradesh","noida":"Uttar Pradesh",
                        "west bengal":"West Bengal","kolkata":"West Bengal","rajasthan":"Rajasthan",
                        "kerala":"Kerala","madhya pradesh":"Madhya Pradesh","andhra pradesh":"Andhra Pradesh",
                        "odisha":"Odisha","bihar":"Bihar","punjab":"Punjab","goa":"Goa","assam":"Assam",
                    }
                    def _ep(dsp):
                        if _short_col and not _prt_df.empty:
                            m = _prt_df[_prt_df[_short_col] == dsp]
                            if not m.empty:
                                r = m.iloc[0]
                                return {"gstin": str(r.get("gstin") or r.get("GSTIN") or "").strip(),
                                        "address": str(r.get("address") or r.get("Registered Address") or "").strip(),
                                        "country": str(r.get("country") or r.get("Country") or "").strip()}
                        return {}
                    def _epos(p):
                        addr = str(p.get("address","")).lower()
                        ctry = str(p.get("country","")).lower()
                        for kw, sname in _ST.items():
                            if kw in addr: return sname
                        if ctry and "india" not in ctry: return str(p.get("country","")).strip() or "Outside India"
                        segs = [s.strip() for s in addr.replace("\n",",").split(",") if s.strip()]
                        return segs[-1].title() if segs else "India"
                    def _enm(m):
                        try:
                            _p = pd.to_datetime(str(m).strip(), format="%b-%Y", errors="coerce")
                            if pd.isna(_p): _p = pd.to_datetime(str(m).strip(), errors="coerce")
                            return _p.strftime("%b-%Y") if pd.notna(_p) else str(m).strip()
                        except: return str(m).strip()
                    _pc2 = {}
                    _gst_rows = []
                    for _, inv in _inv_df.iterrows():
                        _dsp = inv.get("dsp_name","")
                        if _dsp not in _pc2: _pc2[_dsp] = _ep(_dsp)
                        _p = _pc2[_dsp]; _gstin = _p.get("gstin","") or ""; _supply = _epos(_p)
                        _inv_no = inv.get("invoice_number",""); _inv_date = inv.get("invoice_date","")
                        try: _inv_date = pd.to_datetime(str(_inv_date), dayfirst=True, errors="coerce").strftime("%d-%b-%Y")
                        except: pass
                        _inv_type = inv.get("invoice_type",""); _cur = str(inv.get("currency","USD")).upper()
                        _is_cn = int(inv.get("is_credit_note",0) or 0); _is_del = int(inv.get("is_deleted",0) or 0)
                        _cat = "Credit Note" if _is_cn else ("Deleted" if _is_del else ("Export" if _cur=="USD" else "Regular"))
                        _status = str(inv.get("status","")); _fx = float(inv.get("fx_rate",0) or 0)
                        _braw = str(inv.get("month_breakdown","") or "")
                        if _cur == "INR":
                            _tax = float(inv.get("taxable_amount",0) or 0); _ig = float(inv.get("igst_amount",0) or 0)
                            _cg = float(inv.get("cgst_amount",0) or 0); _sg = float(inv.get("sgst_amount",0) or 0)
                            _tot = float(inv.get("total_amount",0) or 0); _uamt = float(inv.get("amount",0) or 0)
                            if _tax==0 and _tot>0:
                                _tax = round(_tot/1.18,2)
                                if _ig==0 and _cg==0 and _sg==0:
                                    if _inv_type=="INR_CGST_SGST": _cg=_sg=round(_tax*0.09,2)
                                    elif _inv_type=="INR_IGST": _ig=round(_tax*0.18,2)
                            _tr = "18% (CGST+SGST)" if _inv_type=="INR_CGST_SGST" else "18% (IGST)" if _inv_type=="INR_IGST" else "GST"
                            if _braw.startswith("["):
                                try:
                                    _bd=_edb_json.loads(_braw)
                                    for _brow in _bd:
                                        _bm=_enm(str(_brow.get("month",""))); _bu=float(_brow.get("usd_amount",0) or 0)
                                        _bf=float(_brow.get("fx_rate",0) or 0); _bi=float(_brow.get("inr_amount",0) or 0) or round(_bu*_bf,2)
                                        _r=(_bu/_uamt) if _uamt>0 else (1/len(_bd))
                                        _gst_rows.append({"Category":_cat,"GSTIN":_gstin,"DSP Name":_dsp,
                                            "Invoice No.":f"{_inv_no} [{_bm}]","Month":_bm,"Invoice Date":_inv_date,"Status":_status,
                                            "Place of Supply":_supply,"Tax Rate %":_tr,
                                            "Taxable Value (INR)":round(_tax*_r,2),"IGST (INR)":round(_ig*_r,2),
                                            "CGST (INR)":round(_cg*_r,2),"SGST (INR)":round(_sg*_r,2),
                                            "Total Invoice Value (INR)":round(_tot*_r,2),
                                            "Amount (USD)":round(_bu,2),"FX Rate":round(_bf,4) if _bf>0 else "",
                                            "Amount (INR equiv.)":round(_bi,2)})
                                    continue
                                except: pass
                            _gst_rows.append({"Category":_cat,"GSTIN":_gstin,"DSP Name":_dsp,
                                "Invoice No.":("CN: " if _is_cn else "")+_inv_no,
                                "Month":inv.get("month",""),"Invoice Date":_inv_date,"Status":_status,
                                "Place of Supply":_supply,"Tax Rate %":_tr,
                                "Taxable Value (INR)":round(_tax,2),"IGST (INR)":round(_ig,2),
                                "CGST (INR)":round(_cg,2),"SGST (INR)":round(_sg,2),
                                "Total Invoice Value (INR)":round(_tot,2),
                                "Amount (USD)":_uamt,"FX Rate":round(_fx,4) if _fx>0 else "",
                                "Amount (INR equiv.)":round(_tot,2)})
                        else:
                            _amt=float(inv.get("amount",0) or 0)
                            if _braw.startswith("["):
                                try:
                                    _bd=_edb_json.loads(_braw)
                                    for _brow in _bd:
                                        _bm=_enm(str(_brow.get("month",""))); _bu=float(_brow.get("usd_amount",0) or 0)
                                        _bf=float(_brow.get("fx_rate",0) or 0); _bi=float(_brow.get("inr_amount",0) or 0) or round(_bu*_bf,2)
                                        _gst_rows.append({"Category":_cat,"GSTIN":_gstin,"DSP Name":_dsp,
                                            "Invoice No.":f"{_inv_no} [{_bm}]","Month":_bm,"Invoice Date":_inv_date,"Status":_status,
                                            "Place of Supply":_supply,"Tax Rate %":"0% (Export)",
                                            "Taxable Value (INR)":0.0,"IGST (INR)":0.0,"CGST (INR)":0.0,"SGST (INR)":0.0,
                                            "Total Invoice Value (INR)":0.0,"Amount (USD)":round(_bu,2),
                                            "FX Rate":round(_bf,4) if _bf>0 else "","Amount (INR equiv.)":round(_bi,2)})
                                    continue
                                except: pass
                            _gst_rows.append({"Category":_cat,"GSTIN":_gstin,"DSP Name":_dsp,
                                "Invoice No.":("CN: " if _is_cn else "")+_inv_no,
                                "Month":inv.get("month",""),"Invoice Date":_inv_date,"Status":_status,
                                "Place of Supply":_supply,"Tax Rate %":"0% (Export)",
                                "Taxable Value (INR)":0.0,"IGST (INR)":0.0,"CGST (INR)":0.0,"SGST (INR)":0.0,
                                "Total Invoice Value (INR)":0.0,"Amount (USD)":round(_amt,2),
                                "FX Rate":round(_fx,4) if _fx>0 else "","Amount (INR equiv.)":round(_amt*_fx,2) if _fx>0 else 0.0})
                    df_export = pd.DataFrame(_gst_rows)
                    st.caption(f"GST All Combined: {len(df_export)} rows")
                else:
                    conn = get_db_connection()
                    df_export = pd.read_sql(f"SELECT * FROM {selected_table}", conn)
                    conn.close()

                export_buffer = io.BytesIO()
                with pd.ExcelWriter(export_buffer, engine="openpyxl") as _ew:
                    df_export.to_excel(_ew, index=False, sheet_name=selected_table_name[:31])
                    _ws = _ew.sheets[selected_table_name[:31]]
                    try:
                        from openpyxl.styles import Font, PatternFill, Alignment
                        _hf = PatternFill("solid", fgColor="003366"); _hft = Font(bold=True, color="FFFFFF")
                        for _c in _ws[1]:
                            _c.font=_hft; _c.fill=_hf; _c.alignment=Alignment(horizontal="center")
                        for _cc in _ws.columns:
                            _ws.column_dimensions[_cc[0].column_letter].width = min(
                                max(len(str(_c.value or "")) for _c in _cc)+4, 45)
                    except: pass
                export_buffer.seek(0)
                st.download_button(
                    label=f"📥 Export {selected_table_name}", data=export_buffer,
                    file_name=f"GST_AllCombined_{pd.Timestamp.today().strftime('%Y%m%d')}.xlsx"
                              if selected_table=="_gst_combined_view"
                              else f"{selected_table}_{pd.Timestamp.today().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="edb_export_btn")
            except Exception as e:
                st.error(f"Export failed: {e}")
        
        st.divider()

        # ====================================================
        # LOAD DATA
        # ====================================================
        if selected_table == "_gst_combined_view":
            st.info("ℹ️ GST Report (All Combined) is a computed report — use the **Export** button above to download it. Direct editing is not available.")
            st.stop()

        conn = get_db_connection()
        df_edit = pd.read_sql(f"SELECT * FROM {selected_table}", conn)
        conn.close()

        if df_edit.empty:
            st.warning("No data found in this table")
        else:

            from st_aggrid import GridOptionsBuilder, AgGrid, GridUpdateMode, DataReturnMode, JsCode

            has_id = "id" in df_edit.columns

            # Add a hidden _row_idx column so we can identify rows even without id
            df_edit["_row_idx"] = range(len(df_edit))

            gb = GridOptionsBuilder.from_dataframe(df_edit)

            # Checkbox selection
            gb.configure_selection(
                selection_mode="multiple",
                use_checkbox=True,
                header_checkbox=True,
                pre_selected_rows=[]
            )

            # id column — show but not editable
            if has_id:
                gb.configure_column("id", editable=False, hide=False, width=70, pinned="left")

            # Hide internal index column
            gb.configure_column("_row_idx", hide=True)

            # All other columns editable
            for col in df_edit.columns:
                if col not in ("id", "_row_idx"):
                    gb.configure_column(col, editable=True)

            gb.configure_default_column(resizable=True, filter=True, sortable=True)

            custom_css = {
                ".ag-header": {
                    "background-color": "#003366 !important",
                    "color": "white !important",
                    "font-weight": "bold !important",
                    "font-size": "14px !important"
                },
                ".ag-header-cell-label": {
                    "color": "white !important",
                    "font-weight": "bold !important"
                },
                ".ag-row-selected": {
                    "background-color": "#fdecea !important"
                }
            }

            gridOptions = gb.build()

            grid_response = AgGrid(
                df_edit,
                gridOptions=gridOptions,
                height=sc.grid_height(600),
                fit_columns_on_grid_load=True,
                update_mode=GridUpdateMode.MODEL_CHANGED,
                data_return_mode=DataReturnMode.AS_INPUT,
                allow_unsafe_jscode=True,
                custom_css=custom_css
            )

            updated_df    = pd.DataFrame(grid_response["data"])
            selected_rows = grid_response.get("selected_rows", [])
            if isinstance(selected_rows, pd.DataFrame):
                selected_rows = selected_rows.to_dict("records")

            # Persist selected rows so they survive Streamlit reruns
            if selected_rows:
                st.session_state["_edb_pending_delete"] = selected_rows

            st.divider()

            # Action buttons
            btn1, btn2, btn3 = st.columns([2, 2, 2])

            with btn1:
                save_btn = st.button("💾 Save Changes", key="edb_save")

            with btn2:
                refresh_btn = st.button("🔄 Reload Data", key="edb_refresh")

            with btn3:
                n_sel = len(selected_rows) if selected_rows else 0
                delete_btn = st.button(
                    f"🗑️ Delete Selected ({n_sel})",
                    key="edb_delete",
                    disabled=(n_sel == 0),
                    type="primary" if n_sel > 0 else "secondary"
                )

            # Save logic
            if save_btn:
                try:
                    conn = get_db_connection()
                    cursor = conn.cursor()
                    for _, row in updated_df.iterrows():
                        if not has_id:
                            continue
                        row_id = int(row["id"])
                        update_cols, values = [], []
                        for col in updated_df.columns:
                            if col in ("id", "_row_idx"):
                                continue
                            val = row[col]
                            if isinstance(val, str):
                                try:
                                    val = float(val)
                                except:
                                    pass
                            update_cols.append(f'"{col}"=?')
                            values.append(val)
                        if not update_cols:
                            continue
                        cursor.execute(
                            f'UPDATE {selected_table} SET {", ".join(update_cols)} WHERE id=?',
                            values + [row_id]
                        )
                    conn.commit()
                    conn.close()
                    st.success("✅ Data saved successfully")
                    load_master_data.clear()
                    load_partner_list.clear()
                    load_cost_centre.clear()
                    st.rerun()
                except Exception as e:
                    st.error("Save failed")
                    st.code(str(e))

            # When delete button clicked — store intent in session state
            if delete_btn and selected_rows:
                st.session_state["_edb_delete_triggered"] = True

            # Delete confirmation — rendered every run when state is set
            if st.session_state.get("_edb_delete_triggered"):
                pending = st.session_state.get("_edb_pending_delete", [])
                n = len(pending)
                if n == 0:
                    st.session_state.pop("_edb_delete_triggered", None)
                else:
                    st.warning(f"⚠️ **{n} row(s)** marked for deletion. Click Yes to confirm.")
                    col_a, col_b, col_c = st.columns([3, 2, 2])
                    with col_b:
                        if st.button("✅ Yes, Delete", key="edb_yes_delete", type="primary"):
                            try:
                                conn = get_db_connection()
                                deleted = 0
                                for r in pending:
                                    if has_id and r.get("id") not in (None, "", "None"):
                                        conn.execute(
                                            f"DELETE FROM {selected_table} WHERE id=?",
                                            (int(r["id"]),)
                                        )
                                        deleted += 1
                                    else:
                                        match_cols = [
                                            c for c in r.keys()
                                            if c not in ("id", "_row_idx")
                                            and r[c] not in (None, "", "None")
                                        ]
                                        if not match_cols:
                                            continue
                                        where = " AND ".join([f'"{c}"=?' for c in match_cols])
                                        vals  = [r[c] for c in match_cols]
                                        conn.execute(
                                            f"DELETE FROM {selected_table} WHERE {where}",
                                            vals
                                        )
                                        deleted += 1
                                conn.commit()
                                conn.close()
                                st.success(f"✅ {deleted} row(s) deleted from {selected_table}!")
                                st.session_state.pop("_edb_delete_triggered", None)
                                st.session_state.pop("_edb_pending_delete", None)
                                load_master_data.clear()
                                load_partner_list.clear()
                                load_cost_centre.clear()
                                for key in ["dsp_df", "ssp_df", "data_initialized"]:
                                    if key in st.session_state:
                                        del st.session_state[key]
                                st.rerun()
                            except Exception as e:
                                st.error(f"Delete failed: {e}")
                    with col_c:
                        if st.button("❌ Cancel", key="edb_cancel_delete"):
                            st.session_state.pop("_edb_delete_triggered", None)
                            st.session_state.pop("_edb_pending_delete", None)
                            st.rerun()

            if refresh_btn:
                st.rerun()

# ==========================================
# MODERN KPI COMPONENT
# ==========================================

import altair as alt
import pandas as pd

def render_premium_kpi(title, value, trend_data=None, is_currency=True):

    numeric_value = float(value)

    display_value = (
        f"${numeric_value:,.2f}"
        if is_currency
        else f"{numeric_value:.2f}%"
    )

    # ----------------------------------
    # 🎨 KPI COLOR LOGIC (Nature Based)
    # ----------------------------------

    title_lower = title.lower()

    # Revenue / Receivable → Green
    if "revenue" in title_lower or "outstanding receivable" in title_lower:
        bg = "linear-gradient(135deg, #0f5132, #198754)"
        text_color = "#ffffff"
        
    elif "overdue receivable" in title_lower:
        bg = "linear-gradient(135deg, #646D7E, #3F829D)"
        text_color = "#ffffff"

    # Cost / Payable → Red
    elif "cost" in title_lower or "outstanding payable" in title_lower:
        bg = "linear-gradient(135deg, #842029, #dc3545)"
        text_color = "#ffffff"

    # Cost / Payable → Red
    elif "overdue payable" in title_lower:
        bg = "linear-gradient(135deg, #646D7E, #3F829D)"
        text_color = "#ffffff"
    
    # Profit → Blue
    elif "profit" in title_lower:
        bg = "linear-gradient(135deg, #d765c5, #0d6efd)"
        text_color = "#ffffff"

    # IVT → Orange
    elif "ivt" in title_lower:
        bg = "linear-gradient(135deg, #646D7E, #3F829D)"
        text_color = "#ffffff"

    # % Metrics → Teal
    elif "%" in title_lower:
        bg = "linear-gradient(135deg, #d765c5, #0d6efd)"
        text_color = "#ffffff"

    # Default → Grey
    else:
        bg = "linear-gradient(135deg, #343a40, #6c757d)"
        text_color = "#ffffff"

    st.markdown(f"""
    <div style="
        background: {bg};
        padding:20px;
        border-radius:16px;
        box-shadow:0 6px 18px rgba(0,0,0,0.2);
        margin-bottom:10px;
    ">
        <div style="font-size:18px;font-weight:900;color:#FFEF00;">
            {title}
        </div>
        <div style="font-size:30px;font-weight:900;color:#FFEF00;margin-top:6px;">
            {display_value}
        </div>
    </div>
    """, unsafe_allow_html=True)


# ==========================================
# CASH CONTROL ENGINE
# ==========================================

from datetime import datetime

def calculate_outstanding_metrics(dsp_df, ssp_df):
    today = pd.Timestamp.today()

    dsp_df = dsp_df.copy()
    ssp_df = ssp_df.copy()

    # Compute Outstanding for DSP (Receivable - Received)
    dsp_df["Receivable $"]     = pd.to_numeric(dsp_df["Receivable $"],     errors="coerce").fillna(0)
    dsp_df["Received Amount $"]= pd.to_numeric(dsp_df["Received Amount $"], errors="coerce").fillna(0)
    dsp_df["_outstanding"]     = dsp_df["Receivable $"] - dsp_df["Received Amount $"]
    dsp_df["Due Date"]         = pd.to_datetime(dsp_df["Due Date"], errors="coerce")

    # Compute Outstanding for SSP (Payable - Paid)
    ssp_df["Payable $"]        = pd.to_numeric(ssp_df["Payable $"],        errors="coerce").fillna(0)
    ssp_df["Paid Amount $"]    = pd.to_numeric(ssp_df["Paid Amount $"],    errors="coerce").fillna(0)
    ssp_df["_outstanding"]     = ssp_df["Payable $"] - ssp_df["Paid Amount $"]
    ssp_df["Due Date"]         = pd.to_datetime(ssp_df["Due Date"], errors="coerce")

    # DSP
    total_receivable       = dsp_df["Receivable $"].sum()
    total_received         = dsp_df["Received Amount $"].sum()
    total_outstanding_dsp  = dsp_df["_outstanding"].sum()
    overdue_dsp = dsp_df[
        (dsp_df["_outstanding"] > 0) &
        (dsp_df["Due Date"] < today)
    ]["_outstanding"].sum()

    # SSP
    total_payable          = ssp_df["Payable $"].sum()
    total_paid             = ssp_df["Paid Amount $"].sum()
    total_outstanding_ssp  = ssp_df["_outstanding"].sum()
    overdue_ssp = ssp_df[
        (ssp_df["_outstanding"] > 0) &
        (ssp_df["Due Date"] < today)
    ]["_outstanding"].sum()

    return {
        "total_receivable": total_receivable,
        "total_received": total_received,
        "total_outstanding_dsp": total_outstanding_dsp,
        "overdue_dsp": overdue_dsp,
        "total_payable": total_payable,
        "total_paid": total_paid,
        "total_outstanding_ssp": total_outstanding_ssp,
        "overdue_ssp": overdue_ssp
    }


def calculate_collection_efficiency(dsp_df, ssp_df):

    total_receivable = dsp_df["Receivable $"].sum()
    total_received = dsp_df["Received Amount $"].sum()

    total_payable = ssp_df["Payable $"].sum()
    total_paid = ssp_df["Paid Amount $"].sum()

    collection_pct = (
        (total_received / total_receivable) * 100
        if total_receivable != 0 else 0
    )

    payment_pct = (
        (total_paid / total_payable) * 100
        if total_payable != 0 else 0
    )

    return {
        "collection_pct": collection_pct,
        "payment_pct": payment_pct
    }


        
if "Ageing" in tabs:
    with tabs["Ageing"]:
        if "dsp_df" not in st.session_state or st.session_state.dsp_df is None:
            try:
                st.session_state.dsp_df = load_dsp_final(
                    st.session_state.master_df, st.session_state.partner_df)
            except Exception:
                st.session_state.dsp_df = pd.DataFrame()
        if "ssp_df" not in st.session_state or st.session_state.ssp_df is None:
            try:
                st.session_state.ssp_df = load_ssp_final(
                    st.session_state.master_df, st.session_state.partner_df)
            except Exception:
                st.session_state.ssp_df = pd.DataFrame()
        render_ageing_tab(
            dsp_df=st.session_state.dsp_df,
            ssp_df=st.session_state.ssp_df,
        )
